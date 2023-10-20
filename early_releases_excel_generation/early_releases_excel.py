import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def early_release_creation(data):

    # if early_releases_excel_generation/early_releases.xlsx exists then delete it
    try:
        os.remove("early_releases_excel_generation/early_releases.xlsx")
    except Exception as e:
        print(e)

    # get todays date and format it as YYYY-MM-DD
    today = datetime.now().strftime("%Y-%m-%d")

    # create a new workbook
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # set the title of the worksheet
    ws.title = "Early Releases"
    # set the tab colour of the worksheet
    ws.sheet_properties.tabColor = "1072BA"
    # in A1 insert the text "Early Releases"
    ws['A1'] = "Early Releases"
    # in A1 set the font to bold, size 20 and set the alignment to center
    ws['A1'].font = Font(bold=True, size=15)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "Report Date: " + today
    ws['A2'].font = Font(bold=True, size=12)
    # merge A1 from column A to column K
    ws.merge_cells('A1:K1')
    ws.merge_cells('A2:K2')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = []
    for key in data[0]:
        headers.append(key)
    # insert the headers into the worksheet
    ws.append(headers)

    # loop through the data and insert it into the worksheet
    for row in data:
        # loop through the row and insert the values into a list called values and then append values to the worksheet
        values = []
        for key in row:
            values.append(row[key])
        ws.append(values)

    # make column A wider
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for column in columns:
        ws.column_dimensions[column].width = 20
    # center the text in row 3
    for cell in ws["3:3"]:
        cell.alignment = Alignment(horizontal='center')
    # set the font to bold for the headers
    for cell in ws["3:3"]:
        cell.font = Font(bold=True)

    # put a border around all cells from row 3 to the end of the worksheet
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = thin_border
    # center the text in all columns from row 4 except for columns G & H
    columns_to_center = ['A', 'B', 'C', 'D', 'E', 'F', 'I', 'J', 'K']
    for column in columns_to_center:
        for cell in ws[column]:
            cell.alignment = Alignment(horizontal='center')
    columns_as_currency = ['G', 'H']
    for column in columns_as_currency:
        for cell in ws[column]:
            cell.number_format = 'R#,##0.00'

    # Freeze the top 3 rows
    ws.freeze_panes = 'A4'
    # Apply a filter to the headers
    ws.auto_filter.ref = "A3:K3"

    # make A1 grey
    ws['A1'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')

    ws['A3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['B3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['C3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['D3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['E3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['F3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['G3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['H3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['I3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['J3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')
    ws['K3'].fill = PatternFill(start_color='C2D9FF', end_color='C2D9FF', fill_type='solid')



    # get the max row number
    max_row = ws.max_row

    # =SUBTOTAL(9, G4: G78)
    ws['G' + str(max_row + 2)] = '=SUBTOTAL(9, G4:G' + str(max_row) + ')'
    ws['G' + str(max_row + 2)].font = Font(bold=True)
    # fromat as currency
    ws['G' + str(max_row + 2)].number_format = 'R#,##0.00'

    # =SUBTOTAL(9, H4: H78)
    ws['H' + str(max_row + 2)] = '=SUBTOTAL(9, H4:H' + str(max_row) + ')'
    ws['H' + str(max_row + 2)].font = Font(bold=True)
    # fromat as currency
    ws['H' + str(max_row + 2)].number_format = 'R#,##0.00'

    # save the workbook as early_releases.xlsx to the early_releases_excel_generation folder
    wb.save(filename="early_releases_excel_generation/early_releases.xlsx")

    return {"Done": "Done"}
