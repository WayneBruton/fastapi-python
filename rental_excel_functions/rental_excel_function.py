from datetime import datetime

from openpyxl import Workbook
# import os
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side


def rental_report(data):
    # create a new workbook
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # set the title of the worksheet
    ws.title = "Rentals"
    # set the tab colour of the worksheet
    ws.sheet_properties.tabColor = "1072BA"
    # get today's date and format it as YYYY-MM-DD
    today = datetime.now().strftime("%Y-%m-%d")
    # in A1 insert the text "Rentals"
    ws['A1'] = "Rentals"
    # in A1 set the font to bold, size 20 and set the alignment to center
    ws['A1'].font = Font(bold=True, size=15)
    ws['A1'].alignment = Alignment(horizontal='center')
    # merge A1 from column A to column U
    ws.merge_cells('A1:U1')

    ws['A2'] = "Report Date: " + today
    ws['A2'].font = Font(bold=True, size=12)
    # merge A2 from column A to column U
    ws.merge_cells('A2:U2')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = []
    for key in data[0]:
        headers.append(key)
    # insert the headers into the worksheet
    ws.append(headers)

    # set the fill color of the headers to blue and make them bold and font white and borders
    for cell in ws["1:1"]:
        cell.fill = PatternFill("solid", fgColor="190482")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = Border(bottom=Side(border_style="thin", color="000000"))

    # do the same for row 2 and 3
    for cell in ws["2:2"]:
        cell.fill = PatternFill("solid", fgColor="190482")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = Border(bottom=Side(border_style="thin", color="000000"))

    for cell in ws["3:3"]:
        cell.fill = PatternFill("solid", fgColor="190482")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = Border(bottom=Side(border_style="thin", color="000000"))

    # in row 3, put a border around each cell
    for cell in ws["3:3"]:
        cell.border = Border(bottom=Side(border_style="thin", color="000000"),
                             top=Side(border_style="thin", color="000000"),
                             left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"))


    # loop through the data and insert it into the worksheet
    for row in data:
        # loop through the row and insert the values into a list called values and then append values to the worksheet
        values = []
        for key in row:
            values.append(row[key])
        ws.append(values)

    # make all columns 20 wide
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J','K','L','M','N','O','P','Q','R','S','T','U']
    for column in columns:
        ws.column_dimensions[column].width = 15


    # make columns J to T currency from row 4 till the last row
    for row in ws.iter_rows(min_row=4, min_col=10, max_col=20):
        for cell in row:
            cell.number_format = 'R #,##0.00'



    # make column U a percentage from row 4 till the last row
    for row in ws.iter_rows(min_row=4, min_col=21, max_col=21):
        for cell in row:
            cell.number_format = '0.00%'


    # get the max row number
    max_row = ws.max_row

    # put a border around all cells from row 3 to the end of the worksheet
    thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=3, max_row=max_row):
        for cell in row:
            cell.border = thin_border

    # filter all the data from row 3 to the end of the worksheet
    ws.auto_filter.ref = "A3:U" + str(max_row)

    # Freeze the top 3 rows
    ws.freeze_panes = 'A4'

    # =SUBTOTAL(9, J4: J6)
    # from column J to column T, insert a subtotal for each column
    columns_to_subtotal = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
    for column in columns_to_subtotal:
        ws[column + str(max_row + 2)] = '=SUBTOTAL(9, ' + column + '4:' + column + str(max_row) + ')'
        ws[column + str(max_row + 2)].font = Font(bold=True)
        ws[column + str(max_row + 2)].number_format = 'R #,##0.00'








    # save the workbook
    wb.save("rental_excel_functions/rentals.xlsx")

    return "ok"





