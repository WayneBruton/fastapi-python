# create an Excel file from a list of dictionaries using openpyxl
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


# CREATE THE EXCEL FILE
def create_excel_file(data, file_name):
    wb = Workbook()
    developments = []
    for item in data:
        developments.append(item['development'])
    developments = list(set(developments))
    # sort the developments
    developments.sort()

    for development in developments:
        # filter data by development and create new list of dictionaries using list comprehension
        development_data = [item for item in data if item['development'] == development]
        # if development is first in list then use active worksheet, otherwise create a new worksheet
        if development == developments[0]:
            ws = wb.active
            # make the ws title the development name plus today's date formatted as dd-mm-yyyy
            ws.title = development + " " + datetime.today().strftime('%d-%m-%Y')

        else:
            ws = wb.create_sheet(development + " " + datetime.today().strftime('%d-%m-%Y'))
        # CREATE THE HEADINGS
        ws.append(
            ["UNIT NO", "SALES PRICE", "SALES EX VAT", "Special on Sales price", "PARKING",
             "ADDITIONAL PARKING PURCHASED",
             "PURCHASER", "SPECIALS ON OTP", "AGENT", "PURCHASE DATE", "SECURING DEPOSIT", "DATE OF DEP PAYMENT",
             "FICA",
             "OTP TO BO", "BOND ORIGINATOR", "BOND DOCS SUBMIT DATE", "AIP DATE", "BOND INSTR DATE", "BOND AMOUNT",
             "BANK",
             "TRNSF DOCS SIGNED", "BOND DOCS SIGNED ", "SHORTFALL AMOUNT", "DATE PD",
             "TRANSFER DUTY REQUESTED", "TRANSFER DUTY RECEIVED", "RATES APPLIED", "RATES FIGURES RECEIVED",
             "RATES CLEARANCE RECEIVED ", "BUILDING PLANS SUBMITTED", "NHBRC UNIT ENROLMENT",
             "BUILDERS ALL RISK POLICY SUBMITTED", "UNIT INSURANCE SUBMITTED", "STRUCTURAL ENGINEERS COMPLETION",
             "SLAB CERTIFICATE", "GLAZING CERTIFICATE", "GEYSER COC", "PLUMBING COC", "ELECTRICAL COC", "PC DATE",
             "BPAS CERTIFICATE ", "OCCUPATION COC", "HAPPY LETTER", "CERTIFICATES SUBMITTED TO ATT", "STORE DOC DATE",
             "BOND PROCEED RECEIVED", "POTENTIAL LODGEMENT", "ACTUAL LODGEMENT", "POTENTIAL REG DATE",
             "ACTUAL REG DATE", ])
        # CREATE THE HEADINGS STYLE
        for cell in ws['1:1']:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical="center")
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

        # ADD THE DATA TO THE EXCEL FILE
        for item in development_data:
            # Loop through the dictionary, check if the key is a boolean and if it is set it to Yes or No
            for key, value in item.items():
                if type(value) == bool:
                    if value:
                        item[key] = "Yes"
                    else:
                        item[key] = "No"

            opportunity_specials = ""
            opportunity_extras_not_listed = ""

            # check if item key opportunity_base_price exists, if not set it to item['opportunity_best_price']
            if 'opportunity_base_price' not in item:
                item['opportunity_base_price'] = item['opportunity_best_price']
            if 'opportunity_best_price' in item:
                item['opportunity_base_price'] = item['opportunity_base_price']

            if item['opportunity_base_price'] == "":
                item['opportunity_base_price'] = '0'
            # remove the R from the price
            item['opportunity_base_price'] = str(item['opportunity_base_price']).replace("R", "")
            item['opportunity_bond_amount'] = str(item['opportunity_bond_amount']).replace("R", "")
            item['opportunity_deposite'] = str(item['opportunity_deposite']).replace("R", "")

            if 'opportunity_specials' not in item:
                item['opportunity_specials'] = None
            if 'opportunity_extras_not_listed' not in item:
                item['opportunity_extras_not_listed'] = None
            if 'opportunity_sales_date' not in item:
                item['opportunity_sales_date'] = None
            if item['opportunity_deposite'] is None:
                item['opportunity_deposite'] = 0
            if 'opportunity_companyname' not in item:
                item['opportunity_companyname'] = ""
            if 'opportunity_extra_cost' not in item:
                item['opportunity_extra_cost'] = '0'
            if 'opportunity_parking_base' not in item:
                item['opportunity_parking_base'] = ''

            base_price = float(item['opportunity_base_price']) + float(item['opportunity_parking_cost']) + float(
                item['opportunity_extra_cost']) + float(item['opportunity_stove_cost'])

            if item['opportunity_deposite'] == "" or item['opportunity_deposite'] is None:
                item['opportunity_deposite'] = '0'

            deposit = float(item['opportunity_deposite'])

            bond_amount = float(item['opportunity_bond_amount'])
            if item['opportunity_specials'] is not None:
                for special in item['opportunity_specials']:
                    opportunity_specials += special + ", "
                item['opportunity_specials'] = opportunity_specials
            if item['opportunity_extras_not_listed'] is not None:
                for extra in item['opportunity_extras_not_listed']:
                    if extra['description'] != "":
                        opportunity_extras_not_listed += extra['description'] + ", "
                item['opportunity_extras_not_listed'] = opportunity_extras_not_listed

            if item['opportunity_client_type'] == "Company":
                purchaser = item['opportunity_companyname']
            elif item['opportunity_client_type'] == "Individual" and item['opportunity_client_no'] == "2 Person":
                purchaser = f"{item['opportunity_lastname']} {item['opportunity_firstname'][:1]} &" \
                            f" {item['opportunity_lastname_sec']} {item['opportunity_firstname_sec'][:1]}"
            else:
                purchaser = f"{item['opportunity_lastname']} {item['opportunity_firstname'][:1]}"

            if item['opportunity_pay_type'] == "Cash":
                item['opportunity_otp_to_bo'] = "Cash"
                item['opportunity_bond_originator'] = "Cash"
                item['opportunity_aip_date'] = "Cash"
                item['opportunity_bond_docs_submit_date'] = "Cash"
                bond_amount = "Cash"
                item['opportunity_bond_instruction_date'] = "Cash"
                item['opportunity_bank'] = "N/A"
                item['opportunity_bond_docs_signed'] = "N/A"
                item['opportunity_shortfall_amount'] = 0
            else:
                item['opportunity_shortfall_amount'] = base_price - float(item['opportunity_bond_amount'])

            ws.append([item['opportunity_code'], base_price,
                       base_price / 1.15, item['opportunity_specials'],
                       item['opportunity_originalBayNo'], item['opportunity_parking_base'],
                       purchaser,
                       item['opportunity_extras_not_listed'], item['opportunity_sale_agreement'],
                       item['opportunity_sales_date'], deposit,
                       item['opportunity_deposite_date'], item['opportunity_fica'], item['opportunity_otp_to_bo'],
                       item['opportunity_bond_originator'], item['opportunity_bond_docs_submit_date'],
                       item['opportunity_aip_date'], item['opportunity_bond_instruction_date'],
                       bond_amount, item['opportunity_bank'],
                       item['opportunity_trnsf_docs_signed'], item['opportunity_bond_docs_signed'],
                       float(item['opportunity_shortfall_amount']), item['opportunity_date_paid'],
                       item['opportunity_transfer_duty_requested'], item['opportunity_transfer_duty_received'],
                       item['opportunity_rates_applied'], item['opportunity_rates_figures_received'],
                       item['opportunity_rates_clearance_received'], item['opportunity_building_plans_submitted'],
                       item['opportunity_nhbrc_unit_enrolment'], item['opportunity_builders_all_risk_policy_submitted'],
                       item['opportunity_unit_insurance_submitted'],
                       item['opportunity_structural_engineers_completion'],
                       item['opportunity_slab_cerficate'], item['opportunity_glazing_certicicate'],
                       item['opportunity_geyser_coc'], item['opportunity_plumbering_coc'],
                       item['opportunity_electrical_coc'],
                       item['opportunity_pc_date'], item['opportunity_bpas_certificate'],
                       item['opportunity_occupation_coc'], item['opportunity_happy_letter'],
                       item['opportunity_certificates_submitted_to_attorneys'], item['opportunity_store_doc_date'],
                       item['opportunity_bond_proceed_received'], item['opportunity_potential_lodgement'],
                       item['opportunity_actual_lodgement'], item['opportunity_potential_reg_date'],
                       item['opportunity_actual_reg_date']])

        # STYLING THE CELLS
        # CREATE THE DATA STYLE FOR THE CELLS below Row 1
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'),
                                     top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))

        # style columns with names =  'B', 'C', 'K' , 'S' and 'W' as currency
        for col in ['B', 'C', 'K', 'S', 'W']:
            for cell in ws[col]:
                cell.number_format = '#,##0.00'

        # SET THE WIDTH OF THE COLUMNS
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        # make the last 9 columns have a width of 15
        for col in ['B', 'E', 'M', 'F', 'S', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']:
            ws.column_dimensions[col].width = 15

        # align columns with names =  'B', 'C', 'K' , 'S' and 'W' from row 2 to the last row to the right
        for col in ['B', 'C', 'K', 'S', 'W']:
            for cell in ws[col]:
                cell.alignment = Alignment(horizontal='right', vertical='center')

        # if a cells value is 'Yes' make it light-green, if 'No' make it light-red
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'Yes':
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif cell.value == 'No':
                    cell.fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

        # if a cells value is 'N/A' make it light-brown
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                if cell.value == 'N/A':
                    cell.fill = PatternFill(start_color='D2B48C', end_color='D2B48C', fill_type='solid')

        # the cells in columns 'B', 'C', 'E', 'G' and 'J' from row 2 to the last row have a value of '0.00' or if
        # they are empty, make them Blue
        for col in ['B', 'C', 'E', 'G', 'J']:
            for cell in ws[col]:
                if cell.value == 0.00 or cell.value is None or cell.value == "":
                    cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

        # make row 1 bold, centered and with a light gray background, higher font size, larger row height and wrap text
        for cell in ws[1]:
            ws.row_dimensions[1].height = 50
            cell.font = Font(size=14)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # SAVE THE EXCEL FILE
    wb.save(f"excel_files/{file_name}")
