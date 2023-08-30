from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment

import time
import datetime


def create_draw_down_file(request, app_total):
    start_time = time.time()
    filename = 'Investor Allocations'
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
              'November', 'December']

    wb = Workbook()

    ws = wb.active
    # ws.title = sheet_name
    ws.sheet_properties.tabColor = "1072BA"

    # insert in row 3 "QUINATE CONSULTING (PTY) LTD" and make it bold
    ws['B3'] = "Heron Projects (Pty) Ltd"
    ws['B3'].font = Font(bold=True)
    ws.append(["", "INVESTOR ALLOCATIONS"])
    ws.append(["", f"{request['drawdowns'][0]['Category']} PROJECT - DRAW {request['draw_number']}".upper()])
    # create variable called date and assign it the value of request['date'] formatted as day full month and year as
    # a string like 1 January 2023
    date = request['date']
    # get the month from the date variable and use its full name
    month = int(datetime.datetime.strptime(date, '%Y/%M/%d').strftime('%M'))
    # get the year from the date variable
    year = datetime.datetime.strptime(date, '%Y/%M/%d').strftime('%Y')
    # get the day from the date variable
    day = datetime.datetime.strptime(date, '%Y/%M/%d').strftime('%d')

    ws.append(["", f"DATE OF LAST INVOICE - {day} {months[month - 1]} {year}"])
    ws.append(["", f"DATE OF DRAW - {day} {months[month - 1]} {year}"])
    ws.append([])
    ws.append(["", "UNIT NO", "INVESTOR", "CAPITAL AMOUNT", f"TOTAL BALANCE AS AT {day} {months[month - 1]} {year}",
               "DRAW DOWN AMOUNT",
               "RESIDUAL BALANCE", "NOTES"])

    rows_to_format = [9]
    # for rows in rows to format, format the cells in the row from 'B' to 'H' to have a borders on all sides,
    # be centered, and have a fill color of light blue, and all columns to have a width of 20 except for column 'E'
    # which has a width of 40 and be bold

    for row in rows_to_format:
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            if col != 5:
                ws.column_dimensions[get_column_letter(col)].width = 22
            else:
                ws.column_dimensions[get_column_letter(col)].width = 40

    ws.append([])

    for index, item in enumerate(request['drawdowns']):
        # print("item", item['opportunity_code'])

        ws.append(
            ["", item['opportunity_code'], item['investment_name'], item['draw_down_amount'], item['draw_down_amount'],
             item['draw_down_amount'], "", ""])
        #  get current row number
        if index == 0:
            current_row = ws.max_row
            first_row = ws.max_row
            # column 'B' to 'H' to have a top border, column B to have a left border, column H to have a right border
            for col in range(2, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(top=Side(style='thin'))
                if col == 2:
                    cell.border = Border(left=Side(style='thin'), top=Side(style='thin'))
                if col == 8:
                    cell.border = Border(right=Side(style='thin'), top=Side(style='thin'))
                if col in [3, 4, 5, 6]:
                    # format as currency with 2 decimal places
                    cell.number_format = '#,##0.00'

        else:
            # get current row number
            current_row = ws.max_row
            # column B to have a left border, column H to have a right border
            for col in range(2, 9):
                cell = ws.cell(row=current_row, column=col)
                if col == 2:
                    cell.border = Border(left=Side(style='thin'))
                if col == 8:
                    cell.border = Border(right=Side(style='thin'))
                if col in [3, 4, 5, 6]:
                    # format as currency with 2 decimal places
                    cell.number_format = '#,##0.00'

    ws.append(["", "", "", "", "", "", "", ""])
    # get current row number
    current_row = ws.max_row
    for col in range(2, 9):
        cell = ws.cell(row=current_row, column=col)
        if col == 2:
            cell.border = Border(left=Side(style='thin'))
        if col == 8:
            cell.border = Border(right=Side(style='thin'))

    ws.append(["", "", "", "", "", "", "", ""])
    # get current row number
    current_row = ws.max_row
    for col in range(2, 9):
        cell = ws.cell(row=current_row, column=col)
        if col == 2:
            cell.border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        if col == 8:
            cell.border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        if col != 2 and col != 8:
            cell.border = Border(bottom=Side(style='thin'))

    # get current row less 1
    current_row_less_1 = ws.max_row - 1
    #     # in cell 'F' of current row less 1, add the formula =SUM(F{first_row}:F{current_row_less_1})
    cell = ws.cell(row=current_row_less_1, column=6)
    cell.value = f'=SUM(F{first_row}:F{current_row_less_1 - 1})'

    cell.number_format = '#,##0.00'
    # Make cell bold
    cell.font = Font(bold=True)

    current_draw_row = ws.max_row

    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", f"CURRENT MOMENTUM BALANCE as at {day} {months[month - 1]} {year}", "", 0, "", "", "", ""])
    # get current row number
    current_row = ws.max_row
    for col in range(2, 9):
        cell = ws.cell(row=current_row, column=col)
        if col == 4:
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                 bottom=Side(style='medium'))

    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", f"DRAW {request['draw_number']}", "", f"=F{current_row_less_1}", "", "", "", ""])
    current_row = ws.max_row
    for col in range(2, 9):
        cell = ws.cell(row=current_row, column=col)
        if col == 4:
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                 bottom=Side(style='medium'))

    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "BALANCE AFTER DRAW", "", f"=D{ws.max_row - 3}-D{ws.max_row - 1}", "", "", "", ""])
    current_row = ws.max_row
    for col in range(2, 9):
        cell = ws.cell(row=current_row, column=col)
        if col == 4:
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                 bottom=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])

    ws.append(["", "NOTES", "", "", "", "", "ALL SUPPLIERS PAID ON DRAWS", ""])
    current_row = ws.max_row
    for col in range(2, 9):
        if col == 2:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(left=Side(style='medium'), top=Side(style='medium'))
        if col == 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(right=Side(style='medium'), top=Side(style='medium'))
        if col != 2 and col != 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(top=Side(style='medium'))

    for index, item in enumerate(request['previous_draws']):
        # draw_down_date
        if index == 1:
            ws.append(
                ["", item['draw_number'], item['draw_down_amount'] - 100000, item["planned_draw_date"], item["note"], "", "",
                 ""])
        elif index == 2:
            ws.append(["", item['draw_number'], item['draw_down_amount'] - 100000, item["planned_draw_date"], item["note"], "", "",
                       ""])
        else:
            ws.append(["", item['draw_number'], item['draw_down_amount'], item["planned_draw_date"], item["note"], "", "",
                       ""])
        current_row = ws.max_row

        if index == 0:
            first_row = current_row

        for col in range(2, 9):
            if col == 2:
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(left=Side(style='medium'))
            if col == 8:
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(right=Side(style='medium'))
            if col == 3:
                # format as currency with 2 decimal places
                cell = ws.cell(row=current_row, column=col)
                cell.number_format = '#,##0.00'
            if col == 4 or col == 8:
                # format as date
                cell = ws.cell(row=current_row, column=col)
                # align center
                cell.alignment = Alignment(horizontal='center')

    ws.append(["", "Current Draw", f"=+F{current_draw_row - 1}", "", "", "", "", ""])
    current_row = ws.max_row
    for col in range(2, 9):
        if col == 2:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(left=Side(style='medium'))
        if col == 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(right=Side(style='medium'))
        if col == 3:
            # format as currency with 2 decimal places
            cell = ws.cell(row=current_row, column=col)
            cell.number_format = '#,##0.00'
        if col == 4 or col == 8:
            # format as date
            cell = ws.cell(row=current_row, column=col)
            # align center
            cell.alignment = Alignment(horizontal='center')

    ws.append(["", "", "", "", "", "", "", ""])
    current_row = ws.max_row
    total_draws = ws.max_row
    for col in range(2, 9):
        if col == 2:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(left=Side(style='medium'))
        if col == 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(right=Side(style='medium'))

    ws.append(["", "TOTAL DRAWS TO DATE", f"=sum(C{first_row}:C{current_row - 1})", "", "", "", "", ""])
    current_row = ws.max_row
    for col in range(2, 9):
        if col == 2:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(left=Side(style='medium'))
        if col == 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(right=Side(style='medium'))
        if col == 3:
            cell = ws.cell(row=current_row, column=col)
            # format as currency with 2 decimal places
            cell.number_format = '#,##0.00'
            # add a top border and make cell bold and bottom border of double line
            cell.border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            cell.font = Font(bold=True)

    ws.append(["", "", "", "", "", "", "", ""])
    current_row = ws.max_row
    for col in range(2, 9):
        if col == 2:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(left=Side(style='medium'), bottom=Side(style='medium'))
        if col == 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(right=Side(style='medium'), bottom=Side(style='medium'))
        if col != 2 and col != 8:
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(bottom=Side(style='medium'))

    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "CASHFLOW FINANCE - DERIC", 0, "Dynamic", "", "", "", ""])
    # format as currency with 2 decimal places in column 'C'
    cell = ws.cell(row=ws.max_row, column=3)
    cell.number_format = '#,##0.00'
    start_sum = ws.max_row
    # print("start_sum", start_sum)
    ws.append(["", "GLC NOT ON DRAW", 0, "", "", "", "", ""])
    cell = ws.cell(row=ws.max_row, column=3)
    cell.number_format = '#,##0.00'
    ws.append(["", "ROUNDING", 0, "", "", "", "", ""])
    cell = ws.cell(row=ws.max_row, column=3)
    cell.number_format = '#,##0.00'
    ws.append(["", "", "", "", "", "", "", ""])
    end_sum = ws.max_row
    # print("end_sum", ws.max_row)
    ws.append(["", "TOTAL DRAW", f"=SUM(C{start_sum}:C{end_sum})", f"=C{end_sum + 1}-C{end_sum + 3}-C{total_draws - 1}", "", "", "", ""])
    # put a top border on column 'C' of current row and a bottom border of double line
    cell = ws.cell(row=ws.max_row, column=3)
    cell.border = Border(top=Side(style='medium'), bottom=Side(style='double'))
    cell.number_format = '#,##0.00'
    cell.font = Font(bold=True)

    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "APP", app_total, "", "", "", "", ""])
    app_total_row = ws.max_row
    cell = ws.cell(row=ws.max_row, column=3)
    cell.number_format = '#,##0.00'
    cell.font = Font(bold=True)

    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "Reconciliation", "", "", "", "", "", ""])
    ws.append(["", "Lebusa", 0, "Previous Error", "", "", "", ""])
    ws.append(["", "INTEREST", 137832.56, "Interest received from STBB", "", "", "", ""])
    interest_row = ws.max_row
    cell = ws.cell(row=ws.max_row, column=3)
    cell.number_format = '#,##0.00'
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "CHECK", f"=C{total_draws + 1}-C{total_draws - 1}-C{interest_row}-C{app_total_row}", "", "", "", "", ""])
    cell = ws.cell(row=ws.max_row, column=3)
    cell.border = Border(top=Side(style='medium'), bottom=Side(style='double'))
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'), top=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))

    ws.append(["", "", "CN Morgan", "JW Haywood", "MD van Rooyen", "Leandri Admin",
               "Deric Finance", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'),
                             right=Side(style='medium'))

    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws.append(["", "", "", "", "", "", "", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws.append(["", "", "date", "date", "date", "date", "date", ""])
    for col in [3, 4, 5, 6, 7]:
        cell = ws.cell(row=ws.max_row, column=col)
        cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'),
                             right=Side(style='medium'))

        wb.save(f"excel_files/{filename}.xlsx")

    end_time = time.time()

    elapsed = end_time - start_time
    minutes, seconds = divmod(elapsed, 60)
    hours, minutes = divmod(minutes, 60)
    days, hours = divmod(hours, 24)

    print(f"Elapsed time: {int(days)} days, {int(hours)} hours, {int(minutes)} minutes and {round(seconds, 2)} seconds")

    return filename

    # return f"Elapsed time: {int(days)} days, {int(hours)} hours, {int(minutes)} minutes and {round(seconds, 2)} seconds"
