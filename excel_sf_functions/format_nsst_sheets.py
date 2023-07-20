from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment


def format_nsst(num_sheets, index, sheet, list_to_filter):
    if num_sheets == 3:
        sheet_index = [2]

    else:
        sheet_index = [4, 5, 6]
    if index in sheet_index:
        # print("list_to_filter", len(list_to_filter))
        if index == 2:
            total = sum([x['with_interest'] for x in list_to_filter]) - sum([x['amount'] for x in list_to_filter])
            # get total_sold like above but where 'sold' is True
            total_sold = sum([x['with_interest'] for x in list_to_filter if x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if x['sold']])
            # do the same for total_unsold
            total_unsold = sum([x['with_interest'] for x in list_to_filter if not x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if not x['sold']])
            # sheet[f'B54'] = f"='NSST Heron Fields'!B54"
            # print("total", total)
        if index == 4:
            total = sum([x['with_interest'] for x in list_to_filter]) - sum([x['amount'] for x in list_to_filter])
            total_sold = sum([x['with_interest'] for x in list_to_filter if x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if x['sold']])
            # do the same for total_unsold
            total_unsold = sum([x['with_interest'] for x in list_to_filter if not x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if not x['sold']])
            sheet[f'B54'] = f"='NSST Heron Fields'!B54 + 'NSST Heron View'!B54"
            # print("total", total)
        if index == 5:
            # filter out list_filter this whose unit do not begin with "HF"
            list_to_filter = [x for x in list_to_filter if x['unit'][:2] == "HF"]
            total = sum([x['with_interest'] for x in list_to_filter]) - sum([x['amount'] for x in list_to_filter])
            total_sold = sum([x['with_interest'] for x in list_to_filter if x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if x['sold']])
            # do the same for total_unsold
            total_unsold = sum([x['with_interest'] for x in list_to_filter if not x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if not x['sold']])
            # print("total", total)
        if index == 6:
            # filter out list_filter this whose unit do not begin with "HF"
            list_to_filter = [x for x in list_to_filter if x['unit'][:2] == "HV"]
            total = sum([x['with_interest'] for x in list_to_filter]) - sum([x['amount'] for x in list_to_filter])
            total_sold = sum([x['with_interest'] for x in list_to_filter if x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if x['sold']])
            # do the same for total_unsold
            total_unsold = sum([x['with_interest'] for x in list_to_filter if not x['sold']]) - sum(
                [x['amount'] for x in list_to_filter if not x['sold']])
            # print("total", total)

        gross_income_column_names = ['B', 'C', 'D', 'E']
        # row 43, add row 35 and deduct rows 38 to 42 for the column names in gross_income_column_names
        for column in gross_income_column_names:
            sheet[f'{column}32'] = f'=SUM({column}23)-SUM({column}26:{column}31)'
            # =B36 + B42 + B41
            sheet[f'{column}37'] = f'=SUM({column}36)+SUM({column}41:{column}42)'
            # sheet[f'D37'] = f'=B37'
            sheet[f'B40'] = f'=SUM(B36)-SUM(C36)-SUM(B38)'
            sheet[f'C40'] = f'=0'
            sheet[f'D40'] = f'=D36-D38'
            sheet[f'E40'] = f'=E36'
            sheet[f'D41'] = f'=B41/SUM(D22:E22)*D22'
            sheet[f'E41'] = f'=B41/SUM(D22:E22)*E22'
            # sheet[f'B40'] = f'=+D40+E40'
            sheet[f'B46'] = f'=C45'
            # sheet[f'B49'] = f'=SUM(B43)-SUM(C43)'
            # sheet[f'B49'] = total
            sheet[f'C51'] = f''
            # sheet[f'D49'] = total_sold
            sheet[f'E51'] = f'=B51-D51'
            # =SUM(B32) - SUM(B37) - SUM(B45)
            sheet[f'{column}55'] = f'=SUM({column}32)-SUM({column}37)-SUM({column}45)'

            # =SUM(B52) + SUM(B38) + SUM(B39)
            sheet[f'B56'] = f'=SUM(B52) + SUM(B38) + SUM(B39)'
            sheet[f'C56'] = f'=SUM(C52) + SUM(C38) + SUM(C39)'
            sheet[f'D56'] = f'=SUM(D52) + SUM(D38) + SUM(D39)'
            sheet[f'E56'] = f'=SUM(E52) + SUM(E38) + SUM(E39)'
            sheet[f'D57'] = f'=B57*0.05'
            sheet[f'E57'] = f'=B57-D57'
            sheet[f'{column}57'] = f'=SUM({column}41)+SUM({column}42)'
            sheet[f'{column}59'] = f'=SUM({column}55)+SUM({column}56)+SUM({column}58)-SUM({column}57)'

            # =B36 + B42 + B41

        # if index == 4:
        # sheet[f'B54'] = f"='NSST Heron Fields'!B54 + 'NSST Heron View'!B54"

        # make column 1 30 units wide and columns 2 to 5 15 units wide
        sheet.column_dimensions['A'].width = 55
        for column in gross_income_column_names:
            sheet.column_dimensions[column].width = 18

        rows_to_format_as_currency = [8, 9, 10, 11, 12, 13, 21, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 35, 36, 37,
                                      38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57,
                                      58, 59]
        # Format the rows in rows_to_format_as_currency as currency from column B to F
        for row in rows_to_format_as_currency:
            for column in gross_income_column_names:
                sheet[f'{column}{row}'].number_format = 'R#,##0.00'

        # Center all the text from column B to F
        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=59, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # make all the cells with values as bold
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=59, max_col=6):
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(bold=True)
                    # put a thick border around the cells with values
                    cell.border = Border(top=Side(border_style='medium'),
                                         bottom=Side(border_style='medium'),
                                         left=Side(border_style='medium'),
                                         right=Side(border_style='medium'))
                    cell.alignment = Alignment(horizontal='center')

        # align the cells in column A to the left
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=59, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        rows_for_full_merge = [1, 2, 3, 6, 7, 14, 15, 19, 20, 33, 34, 43, 44, 54]
        # merge the cells in rows_for_full_merge from column A to E
        for row in rows_for_full_merge:
            sheet.merge_cells(f'A{row}:E{row}')
            # fill the merged cells with a green color except for row
            sheet[f'A{row}'].fill = PatternFill(start_color='539165', end_color='539165', fill_type='solid')
            # make the text in the merged cells white
            sheet[f'A{row}'].font = Font(color='FFFFFF')
            # center the text in the merged cells
            sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')

        rows_for_partial_merge = [4, 5, 8, 9, 10, 11, 12, 13, 16, 17, 18]
        # merge the cells in rows_for_partial_merge from column B to E
        for row in rows_for_partial_merge:
            sheet.merge_cells(f'B{row}:E{row}')

        # make rows 2,3, 6,14, 19 and 31 only 5 unit high

        for row in [2, 3, 6, 14, 19, 33, 43, 53]:
            sheet.row_dimensions[row].height = 1

        # make all rows except rows 2,3, 6,14, 19 and 31 15 unit high
        for row in range(1, 60):
            if row not in [2, 3, 6, 14, 19, 33, 43, 53]:
                sheet.row_dimensions[row].height = 20

        # rows to fill with a grey color from column A to E
        rows_to_fill_with_grey = [35, 45, 55]
        for row in rows_to_fill_with_grey:
            for column in ['A', 'B', 'C', 'D', 'E']:
                sheet[f'{column}{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3',
                                                           fill_type='solid')

        # Make cell A1 bold, white and 30 units high and increase the font size to 20
        sheet['A1'].font = Font(bold=True, color='FFFFFF', size=20)
        sheet.row_dimensions[1].height = 30
        # Make cells A7, A15, A20, A34, A41 25 units high and increase the font size to 15
        for row in [7, 15, 20, 34, 44, 54]:
            sheet.row_dimensions[row].height = 25
            # sheet[f'A{row}'].font = Font(size=15)
            # make them bold and white
            sheet[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=18)

        # For all other rows make the font size 12
        for row in range(1, 60):
            if row not in [1, 7, 15, 20, 34, 44, 54]:
                sheet[f'A{row}'].font = Font(size=12)

        # hide row 43
        # sheet.row_dimensions[44].hidden = True
