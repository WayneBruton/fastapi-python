# from datetime import timedelta
from datetime import datetime, timedelta

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter, column_index_from_string

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# from openpyxl.utils import range_boundaries

from excel_sf_functions.create_sales_sheet import create_excel_array
from excel_sf_functions.create_NSST_sheets import create_nsst_sheet
from excel_sf_functions.format_sf_sheets import format_sales_forecast
from excel_sf_functions.format_nsst_sheets import format_nsst
import time
import threading


def create_sales_forecast_file(data, developmentinputdata, pledges, firstName, listData, request):
    start_time = time.time()
    # print(firstName)
    category = developmentinputdata['Category']
    worksheet_names = []
    if len(category) > 1:
        filename = f'Sales Forecast{category[0].split(" ")[0]}'
        sheet_name = f"SF {category[0]} & {category[1]}"

    else:
        sheet_name = f"SF {category[0]}"
        filename = f'Sales Forecast{category[0]}'
        worksheet_names.append(sheet_name)

    wb = Workbook()
    worksheet_data, interest_on_funds_drawn, interest_on_funds_in_momentum = create_excel_array(data)

    ws = wb.active
    ws.title = sheet_name
    ws.sheet_properties.tabColor = "1072BA"

    row1_data = [f"{sheet_name} - {developmentinputdata['date']}", f"{developmentinputdata['date']}"]
    # insert row1_data into the beginning of worksheet_data
    worksheet_data.insert(0, row1_data)
    # loop through worksheet_data and append each item to the worksheet
    for item in worksheet_data:
        ws.append(item)

    if len(category) > 1:
        for index, item in enumerate(category):
            sheet_name = f"SF {item}"
            row1_data = [f"{sheet_name} - {developmentinputdata['date']}"]
            # filter data based on the category into new list using list comprehension
            data_input = [x for x in data if x['Category'] == item]
            worksheet_data, interest_on_funds_drawn, interest_on_funds_in_momentum = create_excel_array(data_input)

            # insert row1_data into the beginning of worksheet_data
            worksheet_data.insert(0, row1_data)
            # create a new worksheet
            ws = wb.create_sheet(sheet_name)
            ws.sheet_properties.tabColor = "1072BA"
            # loop through worksheet_data and append each item to the worksheet
            for item1 in worksheet_data:
                ws.append(item1)

    # LOOP THROUGH EACH SHEET AND FORMAT ACCORDINGLY
    threads = []
    for sheet in wb.worksheets:
        t = threading.Thread(target=format_sales_forecast, args=(sheet,))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    ##################################

    worksheet_data = []

    report_array = request['Category']
    if len(report_array) > 1:
        sheet_name = f"Investor Exit List Heron"
    else:
        sheet_name = f"Investor Exit List {report_array[0]}"

    ws = wb.create_sheet(sheet_name)

    # ws = wb.active
    ws.title = sheet_name

    # make tab color dark red
    ws.sheet_properties.tabColor = "B31312"

    row1_data = [f"{sheet_name} - {request['date']}"]

    # insert row1_data into the beginning of worksheet_data
    worksheet_data.append(row1_data)
    # loop through worksheet_data and append each item to the worksheet
    row2_data = ["Total Units", "Unit No", "Block", "Investor", "Capital Amount", "Fund Release Date",
                 "Unit Sold Status",
                 "Occupation Date", "Estimated Transfer Date", "Actual Transfer Date", "Exit Deadline (712 Days)",
                 "Current Report Date", "Days to Contact Exit", "180 day Threshhold Warning",
                 "Days to Estimated Exit",
                 "Investor Contract expiry exit", "Capital & Interest to be Exited", "Investor Exit Value On Sales",
                 "Exited by Developer",
                 "Date of Exit", "Early Release", "Investor pay Back On transfer", "Developer & Unbonded"]
    # print(row2_data[16])
    worksheet_data.append(row2_data)
    row3_data = ["", "", "", "Investor Capital Deployed", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                 "",
                 "", "", "", ""]
    worksheet_data.append(row3_data)

    ## DEVELOPMENT UNITS
    # remove the last 2 dictionaries from listData and insert them into a new list called development_units
    development_units = listData[-2:]
    # remove the last 2 dictionaries from listData
    listData = listData[:-2]
    # print("development_units", development_units)

    for item in listData:
        row_data = [item['investor_acc_number'], item['opportunity_code'], item['block'], item['investment_name'],
                    float(item['investment_amount']),
                    item['release_date'],
                    item['opportunity_sold'], item['occupation_date'], item['estimated_transfer_date'],
                    item['final_transfer_date'], "", item['report_date'],
                    item['days_to_exit_deadline'], "", "", item['investment_interest'], item['investment_interest'],
                    0, item['exited_by_developer'], item['date_of_exit'], item['early_release'], "", 0]
        # print(row_data[16], row_data[17], row_data[18])
        worksheet_data.append(row_data)

    for item in development_units:
        row_data = [item['investor_acc_number'], item['opportunity_code'], "B", "Dev Unit",
                    float(item.get('opportunity_amount_required', 1100000)),
                    "",
                    False, "", "",
                    "", "", "",
                    0, "", "", 0, 0,
                    0, 0, 0, False, "", 999]
        # print(row_data[16], row_data[17], row_data[18])
        worksheet_data.append(row_data)

    for item in worksheet_data:
        ws.append(item)

    # Format column D to currency
    for cell in ws['E']:
        cell.number_format = '#,##0.00'
    for cell in ws['P']:
        cell.number_format = '#,##0.00'
    for cell in ws['Q']:
        cell.number_format = '#,##0.00'
    for cell in ws['R']:
        cell.number_format = '#,##0.00'
    for cell in ws['S']:
        cell.number_format = '#,##0.00'
    for cell in ws['V']:
        cell.number_format = '#,##0.00'
    for cell in ws['W']:
        cell.number_format = '#,##0.00'

    cols = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
    # Make Column C 20 wide
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    for col in cols:
        ws.column_dimensions[col].width = 15

    # make columns D through N 15 wide

    for col in ws.iter_cols(min_col=1, max_col=23):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')

            cell.font = Font(size=10)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

        # make row 1 bold, center, fot size 14, text color white, background color blue
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='1072BA', end_color='1072BA', fill_type='solid')

    # make row 2 bold, center, font size 12, text color black, background color light grey
    for cell in ws[2]:
        cell.font = Font(bold=True, size=12, color='000000')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')

    # for all rows after row 2, if the value in column F is True, make the row background color light green
    for row in ws.iter_rows(min_row=4):
        if row[6].value:
            for cell in row:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # for all rows after row 2, if the value in column J has a value, make the row background color light red and
    # hide the row
    for row in ws.iter_rows(min_row=4):

        if row[9].value != "":

            for cell in row:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws.row_dimensions[row[0].row].hidden = True

    # loop through cells in column M from row 4 to the end of the worksheet, if the value is <= 90, make the cell
    # background color light red

    # for all rows after 2, set a formula in column j to equal to column E + (365 * 2)
    for row in ws.iter_rows(min_row=4):
        row[10].value = f'=F{row[0].row}+730'
        if row[6].value == True and row[20].value == True and row[9].value == "":
            row[21].value = row[18].value
        else:
            row[21].value = 0
    # format column J as date YYYY-MM-DD
    for cell in ws['K']:
        cell.number_format = 'YYYY-MM-DD'

    for cell in ws['M']:
        cell.number_format = '0'

    # if cells after row 4 in column 'M' are <= 90, make the cell in colmn 'M' have a background color light red
    for row in ws.iter_rows(min_row=4):
        if row[12].value <= 90:
            row[12].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            if row[18].value != 0:
                row[4].value = 0

        else:
            row[15].value = 0
        if row[18].value != 0:
            row[4].value = 0

    sheet1 = wb.sheetnames[0]
    # print("sheets", sheet1)
    ws2 = wb[sheet1]
    # print("ws", ws2)
    # get the column name, for example 'A' or 'B' etc of the last column of ws2
    last_col = get_column_letter(ws2.max_column)
    # print("last_col", last_col)

    # set a formula in colmn L to equal to J - K formatted as a date YYYY-MM-DD
    for row in ws.iter_rows(min_row=4):
        # get row number and print it
        # print("row", row[0].row)

        # =IF(S20 <> 0, 0, IF(J20="", K20 - L20, 0))
        row[12].value = f'=IF(S{row[0].row}<>0,0,IF(J{row[0].row}="",K{row[0].row}-L{row[0].row},0))'
        row[15].value = f'=IF(M{row[0].row}<90,+Q{row[0].row},0)'
        # row[17].value = row[6].value
        if row[6].value:
            row[17].value = row[16].value
        else:
            row[17].value = 0

        # DO SOMETHING HERE

    # format column L as integer

    # in column N, if column I is not blank, set the value to I - K formatted as integer, else set the value to H - K
    # formatted as integer
    for row in ws.iter_rows(min_row=4):
        row[14].value = f'=IF(J{row[0].row}="",I{row[0].row}-L{row[0].row},J{row[0].row}-L{row[0].row})'
    for row in ws.iter_rows(min_row=3, max_row=3):
        row[1].value = f'=COUNTIFS(B4:B{ws.max_row},"<>",J4:J{ws.max_row},"=")'
        row[0].value = listData[0]['count_of_units']
        row[4].value = f'=SUMIFS(E4:E{ws.max_row}, J4:J{ws.max_row}, "")'
        row[15].value = f'=SUM(P4:P{ws.max_row})'
        row[16].value = f'=SUM(Q4:Q{ws.max_row})'
        row[17].value = f'=SUM(R4:R{ws.max_row})'
        # =SUBTOTAL(109,S4:S500)
        row[18].value = f'=SUBTOTAL(109,S4:S{ws.max_row})'
        row[21].value = f'=SUM(V4:V{ws.max_row})'
        row[22].value = f'=SUM(W4:W{ws.max_row})'

    # put the contents of column B, column E and column J into a new list of dictionaries
    list_to_filter = []
    for row in ws.iter_rows(min_row=4):
        list_to_filter.append(
            {'unit': row[1].value, 'amount': row[4].value, 'date': row[9].value, 'with_interest': row[16].value,
             'sold': row[6].value})
        # filter out of list_to_filter where the value of 'date' is equal to ""
        list_to_filter = list(filter(lambda x: x['date'] == "", list_to_filter))

    # total = sum([x['with_interest'] for x in list_to_filter]) - sum([x['amount'] for x in list_to_filter])
    # print("total", total)

    # make row 3 bold, center, font size 12, text color white, background color red
    for cell in ws[3]:
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for cell in ws['O']:
        cell.number_format = '0'

    # get a list of worksheets
    ws3 = wb.sheetnames[0]
    # get the last column of ws3
    last_colw3 = get_column_letter(ws2.max_column)
    # print("ws3", ws3)

    # last_col_ws2 = get_column_letter(ws2.max_column)
    # sheet_formula = ws2.title

    for row in ws.iter_rows(min_row=4):
        # print("row", row[0].row)
        row[
            16].value = f"=IF(A{row[0].row}&B{row[0].row}<>A{row[0].row - 1}&B{row[0].row - 1},SUMIFS('{sheet1}'!$G$78:${last_col}$78, '{sheet1}'!$G$10:${last_col}$10,A{row[0].row}, '{sheet1}'!$G$4:${last_col}$4, B{row[0].row}),0)"
        # =IF(A20 <> A19, IF(Q20 <> 0, SUMIFS(
        #     'SF Endulini'!$G$19:$HE$19, 'SF Endulini'!$G$10:$HE$10, A20, 'SF Endulini'!$G$4:$HE$4, B20), 0), 0)
        row[
            4].value = f"=IF(A{row[0].row}&B{row[0].row}<>A{row[0].row - 1}&B{row[0].row - 1},IF(Q{row[0].row},SUMIFS('{sheet1}'!$G$19:${last_col}$19, '{sheet1}'!$G$10:${last_col}$10,A{row[0].row}, '{sheet1}'!$G$4:${last_col}$4, B{row[0].row}),0),0)"
        row[
            18].value = f"=IF(A{row[0].row}<>A{row[0].row - 1},SUMIFS('{sheet1}'!$G$60:${last_col}$60, '{sheet1}'!$G$10:${last_col}$10,A{row[0].row}, '{sheet1}'!$G$4:${last_col}$4, B{row[0].row}),0)"
        # if row[16].value != 0:
        #     row[15].value = row[16].value
        # print("row[18].value", row[18].value)
        # if row[18].value != 0:
        #     row[12].value = 0
        if row[3].value == "Dev Unit":
            # =SUMIFS( 'SF Heron Fields & Heron View'!$G$64:$UW$64, 'SF Heron Fields & Heron View'!$G$4:$UW$4,
            # 'Investor Exit List Heron'!B431)
            row[22].value = f"=SUMIFS('{ws3}'!$G$64:${last_colw3}$64, '{ws3}'!$G$4:${last_colw3}$4, B{row[0].row})"
        else:
            row[22].value = 0

    # Hide columns N & O
    cols_to_hide = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'T', 'U']
    for col in cols_to_hide:
        ws.column_dimensions[col].hidden = True
    # ws.column_dimensions['N'].hidden = True
    # ws.column_dimensions['O'].hidden = True

    # in row2 set all cells to bold and wrap text
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    rows_for_full_merge = [1]
    for row in rows_for_full_merge:
        ws.merge_cells(f'A{row}:N{row}')
    ##################################

    # CREATE NSST SHEET

    # print(firstName)
    if firstName == 'Wayne' or firstName == 'Wynand' or firstName == 'Nick' or firstName == 'Deric' or \
            firstName == 'Debbie' or firstName == 'Leandri' or firstName == 'Izolda' or firstName == 'Dirk':
        if len(category) > 1:
            sheet_name = f"NSST {category[0].split(' ')[0]}"

        else:
            sheet_name = f"NSST {category[0]}"

        worksheets = wb.sheetnames
        # print(worksheets)

        # if len(worksheets) == 1:
        index = 0

        ws = wb.create_sheet(sheet_name)
        # Make Tab Color of new Sheet 'Green'

        ws.sheet_properties.tabColor = "539165"

        nsst_data = create_nsst_sheet(category, developmentinputdata, pledges, index, sheet_name, worksheets)

        for item in nsst_data:
            ws.append(item)

        if len(worksheets) == 4:
            # filter pledges to only include the category[0]
            new_pledges = [pledge for pledge in pledges if pledge['Category'] == category[0]
                           and pledge['opportunity_code'] != 'Unallocated']
            index = 1
            sheet_name = f"NSST {category[0]}"
            ws = wb.create_sheet(sheet_name)
            # Make Tab Color of new Sheet 'Green'
            ws.sheet_properties.tabColor = "539165"
            nsst_data = create_nsst_sheet(category, developmentinputdata, new_pledges, index, sheet_name, worksheets)

            for item in nsst_data:
                ws.append(item)

            if len(pledges):
                new_pledges = [pledge for pledge in pledges if pledge['Category'] == category[1]
                               and pledge['opportunity_code'] != 'Unallocated']
            else:
                new_pledges = []
            index = 2
            sheet_name = f"NSST {category[1]}"

            ws = wb.create_sheet(sheet_name)
            # Make Tab Color of new Sheet 'Green'
            ws.sheet_properties.tabColor = "539165"
            nsst_data = create_nsst_sheet(category, developmentinputdata, new_pledges, index, sheet_name, worksheets)

            for item in nsst_data:
                ws.append(item)

        num_sheets = len(wb.sheetnames)

        for index, sheet in enumerate(wb.worksheets):
            # print("index", index, "sheet", sheet)
            format_nsst(num_sheets, index, sheet, list_to_filter)

    else:
        # loop through each sheet and delete all rows after row 40
        for sheet in wb.worksheets:
            sheet.delete_rows(40, 1000)

    ## PER BLOCK SHEET
    ws = wb.create_sheet("Per Block")
    ws.sheet_properties.tabColor = "F11A7B"

    worksheet_data = []

    ws2 = wb[worksheets[0]]
    last_col_ws2 = get_column_letter(ws2.max_column)
    sheet_formula = ws2.title
    # print("sheet_formula", sheet_formula)
    criteria_range = []

    AVAILABLE = []
    DRAWN = []
    RAISED = []
    LEFT_TO_RAISE = []
    TO_RAISE = []
    transferred = []
    individual_capital = []
    released_interest = []
    investment_interest = []
    early_release = []
    sales_price = []
    interest_rate = []

    # get the value from column G row 4 until the last column from ws2 and insert into criteria_range list
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=4, max_row=4):
        for item in cell:
            criteria_range.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=13, max_row=13):
        for item in cell:
            AVAILABLE.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=15, max_row=15):
        for item in cell:
            DRAWN.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=14, max_row=14):
        for item in cell:
            RAISED.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=16, max_row=16):
        for item in cell:
            LEFT_TO_RAISE.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=17, max_row=17):
        for item in cell:
            TO_RAISE.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=3, max_row=3):
        for item in cell:
            transferred.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=19, max_row=19):
        for item in cell:
            individual_capital.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=32, max_row=32):
        for item in cell:
            released_interest.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=24, max_row=24):
        for item in cell:
            investment_interest.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=58, max_row=58):
        for item in cell:
            early_release.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=42, max_row=42):
        for item in cell:
            sales_price.append(item.value)
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=11, max_row=11):
        for item in cell:
            interest_rate.append(item.value)

    # using list comprehension, replace None values with 0 in AVAILABLE
    AVAILABLE = [0 if x is None else x for x in AVAILABLE]
    # using list comprehension, replace None values with 0 in DRAWN
    DRAWN = [0 if x is None else x for x in DRAWN]
    # using list comprehension, replace None values with 0 in RAISED
    RAISED = [0 if x is None else x for x in RAISED]
    # using list comprehension, replace None values with 0 in LEFT_TO_RAISE
    LEFT_TO_RAISE = [0 if x is None else x for x in LEFT_TO_RAISE]
    # using list comprehension, replace None values with 0 in TO_RAISE
    TO_RAISE = [0 if x is None else x for x in TO_RAISE]
    # using list comprehension, replace None values with 0 in transferred
    individual_capital = [0 if x is None else x for x in individual_capital]
    # using list comprehension, replace None values with 0 in transferred
    released_interest = [0 if x is None else x for x in released_interest]
    # using list comprehension, replace None values with 0 in transferred
    investment_interest = [0 if x is None else x for x in investment_interest]
    # using list comprehension, replace None values with 0 in transferred
    early_release = [0 if x is None else x for x in early_release]
    # using list comprehension, replace None values with 0 in transferred
    sales_price = [0 if x is None else x for x in sales_price]
    # using list comprehension, replace None values with 0 in transferred
    interest_rate = [0 if x is None else x for x in interest_rate]

    # print("AVAILABLE",AVAILABLE)

    # print("criteria_range",criteria_range) using list  utilise only the 4th last character of each item in
    # criteria_range and insert into new list called final_criteria
    final_criteria = [item[-4:-3] for item in criteria_range]
    # in a new variable called development, remove the last 4 characters from each item in criteria_range and insert
    # into development using list comprehension
    development = [item[:-4] for item in criteria_range]

    # print("development", development)
    # print("final_criteria", final_criteria)
    worksheet_data.append(final_criteria)
    worksheet_data.append(AVAILABLE)
    worksheet_data.append(DRAWN)
    worksheet_data.append(RAISED)
    worksheet_data.append(LEFT_TO_RAISE)
    worksheet_data.append(TO_RAISE)
    worksheet_data.append(transferred)
    worksheet_data.append(interest_rate)
    worksheet_data.append(development)
    worksheet_data.append(individual_capital)
    worksheet_data.append(released_interest)
    worksheet_data.append(investment_interest)
    worksheet_data.append(early_release)
    worksheet_data.append(sales_price)

    row2_data = ["BLOCK", "AVAILABLE", "DRAWN", "RAISED", "LEFT TO RAISE", "TO RAISE", "CAPITAL REPAID",
                 "INTEREST REPAID", "UNITS", "TOTAL INCOME PER SALE AFTER EXPENSES", "FOR SALE INVESTOR REPAYMENT",
                 "FOR SALE INCOME DUE TO COMPANY", "SOLD INVESTOR REPAYMENT", "SOLD INCOME DUE TO COMPANY",
                 "TRANSFERRED DUE TO INVESTOR", "TRANSFERRED DUE TO COMPANY", "TOTAL INCOME DUE TO COMPANY",
                 "ESTIMATED TOTAL REPAYMENT", "INTEREST", "NET INCOME", "% DRAWN","Interest / Day", "CTC_COST", "CTC_PROFIT", "MARGIN",
                 "SECURITY MARGIN"]
    worksheet_data.append(row2_data)

    # ws.append(worksheet_data)
    for item in worksheet_data:
        ws.append(item)

    last_col = get_column_letter(ws.max_column)

    for col_idx, col in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=6, max_row=6), start=1):
        # Set the formula for each cell in the current column in row 6
        for cell in col:
            col_letter = get_column_letter(col_idx)
            cell.value = f'=IF({col_letter}7=TRUE, 0, {col_letter}2 - {col_letter}4)'

    # get only unique items
    blocks = list(set(final_criteria))
    blocks.sort()

    # get max_column of ws
    start_sum = ws.max_row + 1
    # columns = ["A", "B", "C", "D", "E", "F"]
    for block in blocks:
        insert_row = []
        # print("block", block)
        max_row = ws.max_row + 1
        insert_row.append(block)
        formula = f'=SUMIFS(A2:{last_col}2, A1:{last_col}1, A{max_row})'
        insert_row.append(formula)
        formula = f'=SUMIFS(A3:{last_col}3, A1:{last_col}1, A{max_row})'
        insert_row.append(formula)
        formula = f'=SUMIFS(A4:{last_col}4, A1:{last_col}1, A{max_row})'
        insert_row.append(formula)
        formula = f'=SUMIFS(A5:{last_col}5, A1:{last_col}1, A{max_row})'
        insert_row.append(formula)
        formula = f'=SUMIFS(A6:{last_col}6, A1:{last_col}1, A{max_row})'
        insert_row.append(formula)
        # =SUMIFS($A$10:$UQ$10,$A$7:$UQ$7, "TRUE",$A$1:$UQ$1, A15)+SUMIFS($A$10:$UQ$10,$A$13:$UQ$13, "TRUE",
        # $A$1:$UQ$1, A15)
        formula = f'=SUMIFS(A10:{last_col}10, A7:{last_col}7, "TRUE", A1:{last_col}1, A{max_row})+SUMIFS(A10:{last_col}10, A13:{last_col}13, "TRUE", A1:{last_col}1, A{max_row})'
        insert_row.append(formula)
        # =SUMIFS($A$11:$UQ$11,$A$7:$UQ$7, "TRUE",$A$1:$UQ$1, A15)+SUMIFS($A$12:$UQ$12,$A$7:$UQ$7, "TRUE",$A$1:$UQ$1,
        # A15)+SUMIFS($A$11:$UQ$11,$A$13:$UQ$13, "TRUE",$A$1:$UQ$1, A15)+SUMIFS($A$12:$UQ$12,$A$13:$UQ$13, "TRUE",
        # $A$1:$UQ$1, A15)

        formula = (f'=SUMIFS(A11:{last_col}11, A7:{last_col}7, "TRUE", A1:{last_col}1, A{max_row})'
                   f'+SUMIFS(A12:{last_col}12, A7:{last_col}7, "TRUE", A1:{last_col}1, A{max_row})'
                   f'+SUMIFS(A11:{last_col}11, A13:{last_col}13, "TRUE", A1:{last_col}1, A{max_row})'
                   f'+SUMIFS(A12:{last_col}12, A13:{last_col}13, "TRUE", A1:{last_col}1, A{max_row})')
        insert_row.append(formula)
        # =COUNTIFS($A$1:$UQ$1, A16,$A$2:$UQ$2, "<>0")
        formula = f'=COUNTIFS(A1:{last_col}1, A{max_row}, A2:{last_col}2, "<>0")'
        insert_row.append(formula)
        # =SUMIFS(
        # 'ws2'!$G$64:$UW$64, 'ws2'!$G$64:$UW$64, "<>0", 'Per Block'!$A$1:$UQ$1, 'Per Block'!A23)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G64:{last_col_ws2}64, \'{sheet_formula}\'!G64:{last_col_ws2}64, '
                   f'"<>0", A1:{last_col}1, A{max_row})')
        insert_row.append(formula)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G63:{last_col_ws2}63, \'{sheet_formula}\'!G63:{last_col_ws2}63,'
                   f'"<>0", A1:{last_col}1, A{max_row}, \'{sheet_formula}\'!G2:{last_col_ws2}2, FALSE)')
        insert_row.append(formula)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G64:{last_col_ws2}64, \'{sheet_formula}\'!G64:{last_col_ws2}64, '
                   f'"<>0", A1:{last_col}1, A{max_row}, \'{sheet_formula}\'!G2:{last_col_ws2}2, FALSE)')
        insert_row.append(formula)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G63:{last_col_ws2}63, \'{sheet_formula}\'!G63:{last_col_ws2}63, '
                   f'"<>0", A1:{last_col}1, A{max_row}, \'{sheet_formula}\'!G2:{last_col_ws2}'
                   f'2, TRUE, \'{sheet_formula}\'!G3:{last_col_ws2}3, FALSE)')
        insert_row.append(formula)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G64:{last_col_ws2}64, \'{sheet_formula}\'!G64:{last_col_ws2}64, '
                   f'"<>0", A1:{last_col}1, A{max_row}, \'{sheet_formula}\'!G2:{last_col_ws2}'
                   f'2, TRUE, \'{sheet_formula}\'!G3:{last_col_ws2}3, FALSE)')
        insert_row.append(formula)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G63:{last_col_ws2}63, \'{sheet_formula}\'!G63:{last_col_ws2}63, '
                   f'"<>0", A1:{last_col}1, A{max_row}, \'{sheet_formula}\'!G2:{last_col_ws2}'
                   f'2, TRUE, \'{sheet_formula}\'!G3:{last_col_ws2}3, TRUE)')
        insert_row.append(formula)
        formula = (f'=SUMIFS(\'{sheet_formula}\'!G64:{last_col_ws2}64, \'{sheet_formula}\'!G64:{last_col_ws2}64, '
                   f'"<>0", A1:{last_col}1, A{max_row}, \'{sheet_formula}\'!G2:{last_col_ws2}'
                   f'2, TRUE, \'{sheet_formula}\'!G3:{last_col_ws2}3, TRUE)')
        insert_row.append(formula)
        # =SUM(N16, L16)
        formula = f'=SUM(N{max_row}, L{max_row})'
        insert_row.append(formula)
        # =SUM(M16, K16)
        formula = f'=SUM(M{max_row}, K{max_row})'
        insert_row.append(formula)
        # =SUM(R16 - D16)
        formula = f'=SUM(R{max_row} - D{max_row})'
        insert_row.append(formula)
        # =J16-SUM(R16)
        formula = f'=J{max_row}-SUM(R{max_row})'
        insert_row.append(formula)
        # =S16 / C16
        formula = f'=IFERROR(S{max_row} / C{max_row}, 0)'
        insert_row.append(formula)
        # =SUMIFS($A$41:$VT$41,$A$1:$VT$1, A16,$A$7:$VT$7, FALSE)
        formula = f'=SUMIFS(A41:{last_col}41, A1:{last_col}1, A{max_row}, A7:{last_col}7, FALSE)'
        insert_row.append(formula)



        ws.append(insert_row)




    club_house = ["CLUBHOUSE", 0, 0, 0, 0, 0, 0, 0]
    ws.append(club_house)
    other = ["OTHER", 0, 0, 0, 0, 0, 0, 0]
    ws.append(other)
    # append an empty row
    ws.append([])
    max_row = ws.max_row

    # in column z for rows

    # add the following formula in row 41 for columns A to last_col
    # ={column_name}10 * {column_name}8 / 365



    # in a new row, insert the sum of each column from start_sum to max_row for columns B through F
    columns = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
               "V", "W", "X", "Y", "Z"]
    insert = ["TOTAL"]
    for col in columns:
        # ws[f'{col}{max_row}'].value = f'=SUM({col}{start_sum}:{col}{max_row - 1})'
        if col != "U":
            insert.append(f'=SUM({col}{start_sum}:{col}{max_row - 1})')
        else:

            insert.append(f'=IFERROR(S{max_row} / C{max_row}, 0)')

    ws.append(insert)

    ws.append([])
    ws.append(["Heron Fields"])
    # for the above row, in columns B through H, add the formula + the value of the cell in row 15 and row 16
    columns = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
               "V", "W", "X", "Y","Z"]
    for col in columns:
        if col != "U":
            ws[f'{col}37'].value = f'=SUM({col}16+{col}17)'
        else:
            ws[f'{col}37'].value = f'=IFERROR(S{max_row} / C{max_row}, 0)'
    ws.append(["Heron View"])
    # for the above row, in columns B through H, add the value of row 34 less the value of row 15 and row 16
    for col in columns:
        if col != "U":
            ws[f'{col}38'].value = f'=SUM({col}35-{col}37)'
        else:
            ws[f'{col}38'].value = f'=IFERROR(S{max_row} / C{max_row}, 0)'

    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
               "V", "W", "X", "Y", "Z"]
    rows = [16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 35, 37, 38]
    # format all cells in columns B through F to currency format for the rows in the rows list, and set the font size
    # to 10, and set the alignment to center, and set the border to thin, and make the columns autofit
    for col in columns:
        for row in rows:

            if col != "U":
                ws[f'{col}{row}'].number_format = '#,##0.00'
                ws[f'{col}{row}'].font = Font(size=10)
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                ws[f'{col}{row}'].border = Border(left=Side(border_style='thin', color='000000'),
                                                  right=Side(border_style='thin', color='000000'),
                                                  top=Side(border_style='thin', color='000000'),
                                                  bottom=Side(border_style='thin', color='000000'))
                # ws.column_dimensions[col].auto_size = True
            # ws.column_dimensions[col].auto_size = True
            else:
                ws[f'{col}{row}'].number_format = '0.00%'
                ws[f'{col}{row}'].font = Font(size=10)
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                ws[f'{col}{row}'].border = Border(left=Side(border_style='thin', color='000000'),
                                                  right=Side(border_style='thin', color='000000'),
                                                  top=Side(border_style='thin', color='000000'),
                                                  bottom=Side(border_style='thin', color='000000'))
                # ws.column_dimensions[col].auto_size = True

    # for columns in columns list, set the font to bold, set the font size to 12, set the alignment to center,
    # set the background color to light grey, and set the border to thin for row 8
    for col in columns:
        ws[f'{col}15'].font = Font(bold=True, size=12)
        ws[f'{col}15'].alignment = Alignment(horizontal='center')
        ws[f'{col}15'].fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
        ws[f'{col}15'].border = Border(left=Side(border_style='thin', color='000000'),
                                       right=Side(border_style='thin', color='000000'),
                                       top=Side(border_style='thin', color='000000'),
                                       bottom=Side(border_style='thin', color='000000'))
        ws[f'{col}35'].font = Font(bold=True, size=10)
        ws[f'{col}35'].alignment = Alignment(horizontal='center')
        ws[f'{col}35'].fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
        ws[f'{col}35'].border = Border(left=Side(border_style='thin', color='000000'),
                                       right=Side(border_style='thin', color='000000'),
                                       top=Side(border_style='thin', color='000000'),
                                       bottom=Side(border_style='thin', color='000000'))

    # for the cells in row 15, set the font to bold, set the font size to 12, set the alignment to center, and set the
    # background color to light grey, and set the border to thin and make the alignment word wrap
    for col in columns:
        ws[f'{col}15'].font = Font(bold=True, size=12)
        ws[f'{col}15'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws[f'{col}15'].fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
        ws[f'{col}15'].border = Border(left=Side(border_style='thin', color='000000'),
                                       right=Side(border_style='thin', color='000000'),
                                       top=Side(border_style='thin', color='000000'),
                                       bottom=Side(border_style='thin', color='000000'))

    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=41, max_row=41):
        for cell in col:
            col_letter = get_column_letter(cell.column)
            cell.value = f'={col_letter}10 * {col_letter}8 / 365'

    # hide rows 1 through 7
    for row in range(1, 15):
        ws.row_dimensions[row].hidden = True

    # hide row 41
    ws.row_dimensions[41].hidden = True

    # make all columns have a width of 15 except for columns A
    for col in columns:
        ws.column_dimensions[col].width = 15

    ## CASHFLOW SHEET
    ws = wb.create_sheet("Cashflow")
    ws.sheet_properties.tabColor = "16FF00"

    cashflow_data = []
    row1 = ["Unit No.", "Date", "Amount"]
    ws.append(row1)

    transferred = []

    ws2 = wb[worksheets[0]]
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=3, max_row=3):
        for item in cell:
            transferred.append(item.value)

    unit_no = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=4, max_row=4):
        for item in cell:
            unit_no.append(item.value)

    investment_account = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=10, max_row=10):
        for item in cell:
            investment_account.append(item.value)

    project_interest = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=11, max_row=11):
        for item in cell:
            project_interest.append(item.value)

    capital_required = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=13, max_row=13):
        for item in cell:
            capital_required.append(item.value)

    # using list comprehension, replace all None values with 0
    capital_required = [0 if x is None else x for x in capital_required]

    capital_invested = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=19, max_row=19):
        for item in cell:
            capital_invested.append(item.value)

    # using list comprehension, replace all None values with 0
    capital_invested = [0 if x is None else x for x in capital_invested]

    momentum_deposit_date = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=20, max_row=20):
        for item in cell:
            momentum_deposit_date.append(item.value)

    # using list comprehension, replace all None values with 0
    momentum_deposit_date = [0 if x is None else x for x in momentum_deposit_date]

    release_date = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=21, max_row=21):
        for item in cell:
            release_date.append(item.value)

    # using list comprehension, replace all None values with 0
    release_date = [0 if x is None else x for x in release_date]

    end_date = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=22, max_row=22):
        for item in cell:
            end_date.append(item.value)

    # using list comprehension, replace all None values with 0
    end_date = [0 if x is None else x for x in end_date]

    invest_account_interest_earned = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=30, max_row=30):
        for item in cell:
            invest_account_interest_earned.append(item.value)

    # using list comprehension, replace all None values with 0
    invest_account_interest_earned = [0 if x is None else x for x in invest_account_interest_earned]

    released_interest_earned = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=32, max_row=32):
        for item in cell:
            released_interest_earned.append(item.value)

    # using list comprehension, replace all None values with 0
    released_interest_earned = [0 if x is None else x for x in released_interest_earned]

    sales_price = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=42, max_row=42):
        for item in cell:
            sales_price.append(item.value)

    # using list comprehension, replace all None values with 0
    sales_price = [0 if x is None else x for x in sales_price]

    commission = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=45, max_row=45):
        for item in cell:
            commission.append(item.value)

    # using list comprehension, replace all None values with 0
    commission = [0 if x is None else x for x in commission]

    transfer_fees = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=46, max_row=46):
        for item in cell:
            transfer_fees.append(item.value)

    # using list comprehension, replace all None values with 0
    transfer_fees = [0 if x is None else x for x in transfer_fees]

    bond_registration_fees = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=47, max_row=47):
        for item in cell:
            bond_registration_fees.append(item.value)

    # using list comprehension, replace all None values with 0
    bond_registration_fees = [0 if x is None else x for x in bond_registration_fees]

    trust_release_fees = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=48, max_row=48):
        for item in cell:
            trust_release_fees.append(item.value)

    # using list comprehension, replace all None values with 0
    trust_release_fees = [0 if x is None else x for x in trust_release_fees]

    unforseen = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=49, max_row=49):
        for item in cell:
            unforseen.append(item.value)

    # using list comprehension, replace all None values with 0
    unforseen = [0 if x is None else x for x in unforseen]

    discount = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=50, max_row=50):
        for item in cell:
            discount.append(item.value)

    # using list comprehension, replace all None values with 0
    discount = [0 if x is None else x for x in discount]

    early_release = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=58, max_row=58):
        for item in cell:
            early_release.append(item.value)

    # using list comprehension, replace all None values with 0
    early_release = [0 if x is None else x for x in early_release]

    unallocated_interest = []
    for cell in ws2.iter_cols(min_col=7, max_col=ws2.max_column, min_row=146, max_row=146):
        for item in cell:
            unallocated_interest.append(item.value)

    merged_list = []

    for index, item in enumerate(transferred):
        insert = {'transferred': item, 'unit_no': unit_no[index], 'investment_account': investment_account[index],
                  'project_interest': project_interest[index], 'capital_required': capital_required[index],
                  'capital_invested': capital_invested[index], 'momentum_deposit_date': momentum_deposit_date[index],
                  'release_date': release_date[index], 'end_date': end_date[index],
                  'invest_account_interest_earned': invest_account_interest_earned[index],
                  'released_interest_earned': released_interest_earned[index], 'sales_price': sales_price[index],
                  'commission': commission[index], 'transfer_fees': transfer_fees[index],
                  'bond_registration_fees': bond_registration_fees[index],
                  'trust_release_fees': trust_release_fees[index], 'unforseen': unforseen[index],
                  'discount': discount[index], 'early_release': early_release[index],
                  'unallocated_interest': unallocated_interest[index]}

        merged_list.append(insert)

    for item in merged_list:
        merged_filteredA = [x for x in merged_list if x['unit_no'] == item['unit_no']]
        if item['capital_required'] == 0:
            item['capital_required'] = merged_filteredA[0]['capital_required']
            item['sales_price'] = merged_filteredA[0]['sales_price']
            item['commission'] = merged_filteredA[0]['commission']
            item['transfer_fees'] = merged_filteredA[0]['transfer_fees']
            item['bond_registration_fees'] = merged_filteredA[0]['bond_registration_fees']
            item['trust_release_fees'] = merged_filteredA[0]['trust_release_fees']
            item['unforseen'] = merged_filteredA[0]['unforseen']
            item['discount'] = merged_filteredA[0]['discount']
        # get the sum of capital_invested for all items in merged_filteredA
        item['total_capital_invested'] = sum(x['capital_invested'] for x in merged_filteredA)

    # filter out of merged_list all items where the value of the key 'transferred' is False
    merged_list = [x for x in merged_list if x['transferred'] == False]
    # merged_list = [x for x in merged_list if x['early_release'] == False]

    # loop through merged list and convert momentum_deposit_date, release_date and end_date to datetime objects
    for item in merged_list:
        # print("item", item['end_date'], type(item['end_date']), item['unit_no'])
        if type(item['end_date']) == datetime:
            item['end_date'] = ''
        if item['momentum_deposit_date'] != '':
            # replace '-' with '/' in date string
            item['momentum_deposit_date'] = item['momentum_deposit_date'].replace('-', '/')
            item['momentum_deposit_date'] = datetime.strptime(item['momentum_deposit_date'], '%Y/%m/%d')
        if item['release_date'] != '':
            item['release_date'] = item['release_date'].replace('-', '/')
            item['release_date'] = datetime.strptime(item['release_date'], '%Y/%m/%d')
        if item['end_date'] != '':
            item['end_date'] = item['end_date'].replace('-', '/')
            item['end_date'] = datetime.strptime(item['end_date'], '%Y/%m/%d')

        if item['momentum_deposit_date'] != '':
            # calculate the number of days between the momentum_deposit_date and the release_date
            days = item['release_date'] - item['momentum_deposit_date']
            # convert days to an integer
            days = days.days
            # multiply the capital_invested by the 0.0275 and divide by 365 and multiply by days
            interest = (item['capital_invested'] * 0.0275 / 365) * days
            item['invest_account_interest_earned2'] = interest
        else:
            item['invest_account_interest_earned2'] = 0

        # merged_filtered = [x for x in merged_list if x['unit_no'] == item['unit_no']]
        if item['investment_account'] == 'ZZUN01' and item['momentum_deposit_date'] != '' and item['end_date'] != '':
            # calculate the number of days between the momentum_deposit_date and the end_date
            days = item['end_date'] - item['release_date']
            # convert days to an integer
            days = 720 - days.days
            # divide unallocated_interest by days
            interest = item['unallocated_interest'] / days
            # multiply interest by the number of days between release_date and momentum_deposit_date
            interest = interest * (item['release_date'] - item['momentum_deposit_date']).days
            # add to interest the capital_required less total_capital_invested multiplied by the project_interest
            # divided by 365 and multiplied by the number of days between end_date and release_date
            interest = interest + ((item['capital_required'] - item['total_capital_invested']) * item[
                'project_interest'] / 365) * (item['end_date'] - item['release_date']).days
            item['unallocated_interest'] = interest

        item['transfer_income'] = item['sales_price'] - item['commission'] - item['transfer_fees'] - item[
            'bond_registration_fees'] - item['trust_release_fees'] - item['unforseen'] - item['discount']

        if item['investment_account'] == 'ZZUN01' and item['momentum_deposit_date'] != '' and item['end_date'] != '':
            item['due_to_investor'] = (item['capital_required'] - item['total_capital_invested']) + item[
                'unallocated_interest']
        else:
            item['due_to_investor'] = item['capital_invested'] + item['invest_account_interest_earned'] + item[
                'invest_account_interest_earned2'] + item['released_interest_earned']

        if item['early_release']:
            item['due_to_investor'] = 0

    unit_no_list = [x['unit_no'] for x in merged_list]
    unit_no_list = list(dict.fromkeys(unit_no_list))

    final_list = []
    for unit in unit_no_list:
        # create a new list variable from merged_list with only the items that match the unit_no
        merged_filtered = [x for x in merged_list if x['unit_no'] == unit]
        insert = {'unit_no': merged_filtered[0]['unit_no']}

        # insert['date'] = merged_filtered[0]['opportunity_final_transfer_date']
        datafiltered = [x for x in data if x['opportunity_code'] == insert['unit_no']]

        insert['date'] = datafiltered[0]['opportunity_final_transfer_date']

        insert['amount'] = merged_filtered[0]['transfer_income'] - sum(x['due_to_investor'] for x in merged_filtered)

        final_list.append(insert)

    for item in final_list:
        insert = [item['unit_no'], item['date'], item['amount']]

        ws.append(insert)

    ## ROLLOVER SHEET
    ws = wb.create_sheet("Rollover")
    ws.sheet_properties.tabColor = "FF8E00"

    row1 = ["Unit No.", "Investor", "Name", "End Date", "Rollover Date", "Amount"]
    ws.append(row1)

    # get the last column in the first worksheet
    # list the sheets in the workbook
    worksheets = wb.sheetnames
    # get the first sheet in the workbook
    ws2 = wb[worksheets[0]]
    # jst get the name of the first sheet in the workbook
    ws_name = worksheets[0]
    last_col_ws2 = get_column_letter(ws2.max_column)

    for item in data:
        insert = []
        if item['opportunity_transferred'] == False and item['investor_acc_number'] != 'ZZUN01':
            insert.append(item['opportunity_code'])
            insert.append(item['investor_acc_number'])
            insert.append(f"{item['investor_name']} {item['investor_surname']}")
            insert.append(item['opportunity_final_transfer_date'])

            ws.append(insert)

    # in column E, insert the formula =D2+3 and format it as a date
    for row in range(2, ws.max_row + 1):
        ws[f'E{row}'].value = f'=D{row}+3'
        ws[f'E{row}'].number_format = 'yyyy/mm/dd'

    # in column F insert the formula =SUMIFS('SF Heron View'!$G$112:$PS$112,'SF Heron View'!$G$4:$PS$4,A2,
    # 'SF Heron View'!$G$10:$PS$10,B2) and format it as currency
    for row in range(2, ws.max_row + 1):
        ws[
            f'F{row}'].value = f'=SUMIFS(\'{ws_name}\'!$G$112:${last_col_ws2}$112,\'{ws_name}\'!$G$4:${last_col_ws2}$4,A{row},\'{ws_name}\'!$G$10:${last_col_ws2}$10,B{row})'
        ws[f'F{row}'].number_format = '#,##0.00'

    # make all columns 20 wide
    for col in ws.iter_cols(min_col=1, max_col=6):
        ws.column_dimensions[col[0].column_letter].width = 20

    # make col C 30 wide
    ws.column_dimensions['C'].width = 30

    # make first row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # freeze the first row
    ws.freeze_panes = 'A2'

    # SAVE TO FILE
    wb.save(f"excel_files/{filename}.xlsx")

    end_time = time.time()

    elapsed = end_time - start_time
    minutes, seconds = divmod(elapsed, 60)
    hours, minutes = divmod(minutes, 60)
    days, hours = divmod(hours, 24)

    print(f"Elapsed time: {int(days)} days, {int(hours)} hours, {int(minutes)} minutes and {round(seconds, 2)} seconds")


def create_investment_list(data, request):
    wb = Workbook()

    worksheet_data = []

    report_array = request['Category']
    if len(report_array) > 1:
        sheet_name = f"Investor Exit List Heron"
    else:
        sheet_name = f"Investor Exit List {report_array[0]}"

    ws = wb.active
    ws.title = sheet_name
    ws.sheet_properties.tabColor = "1072BA"

    row1_data = [f"{sheet_name} - {request['date']}"]

    # insert row1_data into the beginning of worksheet_data
    worksheet_data.append(row1_data)
    # loop through worksheet_data and append each item to the worksheet
    row2_data = ["Total Units", "Unit No", "Block", "Investor", "Capital Amount", "Fund Release Date",
                 "Unit Sold Status",
                 "Occupation Date", "Estimated Transfer Date", "Actual Transfer Date", "Exit Deadline (712 Days)",
                 "Current Report Date", "Days to Contact Exit", "180 day Threshhold Warning",
                 "Days to Estimated Exit",
                 "Investor Contract expiry exit", "Capital & Interest to be Exited", "Investor Exit Value On Sales",
                 "Exited by Developer",
                 "Date of Exit"]
    worksheet_data.append(row2_data)
    row3_data = ["", "", "", "Investor Capital Deployed", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                 ""]
    worksheet_data.append(row3_data)

    for item in data:
        row_data = [item['investor_acc_number'], item['opportunity_code'], item['block'], item['investment_name'],
                    float(item['investment_amount']),
                    item['release_date'],
                    item['opportunity_sold'], item['occupation_date'], item['estimated_transfer_date'],
                    item['final_transfer_date'], "", item['report_date'],
                    item['days_to_exit_deadline'], "", "", item['investment_interest'], item['investment_interest'],
                    0, item['exited_by_developer'], item['date_of_exit']]
        worksheet_data.append(row_data)

    for item in worksheet_data:
        ws.append(item)

    # Format column D to currency
    for cell in ws['E']:
        cell.number_format = '#,##0.00'
    for cell in ws['P']:
        cell.number_format = '#,##0.00'
    for cell in ws['Q']:
        cell.number_format = '#,##0.00'
    for cell in ws['R']:
        cell.number_format = '#,##0.00'
    for cell in ws['S']:
        cell.number_format = '#,##0.00'

    cols = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T']
    # Make Column C 20 wide
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    for col in cols:
        ws.column_dimensions[col].width = 15

    # make columns D through N 15 wide

    for col in ws.iter_cols(min_col=1, max_col=20):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')

            cell.font = Font(size=10)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

        # make row 1 bold, center, fot size 14, text color white, background color blue
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='1072BA', end_color='1072BA', fill_type='solid')

    # make row 2 bold, center, font size 12, text color black, background color light grey
    for cell in ws[2]:
        cell.font = Font(bold=True, size=12, color='000000')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')

    # for all rows after row 2, if the value in column F is True, make the row background color light green
    for row in ws.iter_rows(min_row=4):
        if row[6].value:
            for cell in row:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # for all rows after row 2, if the value in column J has a value, make the row background color light red and
    # hide the row
    for row in ws.iter_rows(min_row=4):
        if row[9].value != "":
            for cell in row:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws.row_dimensions[row[0].row].hidden = True

    # loop through cells in column M from row 4 to the end of the worksheet, if the value is <= 90, make the cell
    # background color light red

    # for all rows after 2, set a formula in column j to equal to column E + (365 * 2)
    for row in ws.iter_rows(min_row=4):
        row[10].value = f'=F{row[0].row}+730'
    # format column J as date YYYY-MM-DD
    for cell in ws['K']:
        cell.number_format = 'YYYY-MM-DD'

    # if cells after row 4 in column 'M' are <= 90, make the cell in colmn 'M' have a background color light red
    for row in ws.iter_rows(min_row=4):
        if row[12].value <= 90:
            row[12].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            if row[18].value != 0:
                row[4].value = 0

        else:
            row[15].value = 0

    # set a formula in colmn L to equal to J - K formatted as a date YYYY-MM-DD
    for row in ws.iter_rows(min_row=4):
        # =IF(S112 <> 0, 0, IF(J112="", K112 - L112, 0))
        row[12].value = f'=IF(S{row[0].row}<>0,0,IF(J{row[0].row}="",K{row[0].row}-L{row[0].row},0))'
        # row[17].value = row[6].value
        if row[6].value:
            row[17].value = row[16].value
        else:
            row[17].value = 0

    # format column L as integer
    for cell in ws['M']:
        cell.number_format = '0'

    # in column N, if column I is not blank, set the value to I - K formatted as integer, else set the value to H - K
    # formatted as integer
    for row in ws.iter_rows(min_row=4):
        row[14].value = f'=IF(J{row[0].row}="",I{row[0].row}-L{row[0].row},J{row[0].row}-L{row[0].row})'
    for row in ws.iter_rows(min_row=3, max_row=3):
        # =SUM(--(LEN(UNIQUE(FILTER(B4:B192, J4:J192="", ""))) > 0))
        row[0].value = f'=SUM(--(LEN(UNIQUE(FILTER(B4:B{ws.max_row}, J4:J{ws.max_row}, ""))) > 0))'
        # =COUNTIFS($B$4:$B$192, "<>",$J$4:$J$192, "=")
        row[1].value = f'=COUNTIF(B4:B{ws.max_row}, "<>", J4:J{ws.max_row}, "=")'

        row[4].value = f'=SUMIFS(E4:E{ws.max_row}, J4:J{ws.max_row}, "")'
        row[15].value = f'=SUM(P4:P{ws.max_row})'
        row[16].value = f'=SUM(Q4:Q{ws.max_row})'
        row[17].value = f'=SUM(R4:R{ws.max_row})'
        # =SUBTOTAL(109, S4: S448)

        row[18].value = f'=SUBTOTAL(109,S4:S{ws.max_row})'

    # make row 3 bold, center, font size 12, text color white, background color red
    for cell in ws[3]:
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    for cell in ws['O']:
        cell.number_format = '0'

    # Hide columns N & O
    ws.column_dimensions['N'].hidden = True
    ws.column_dimensions['O'].hidden = True

    # in row2 set all cells to bold and wrap text
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    rows_for_full_merge = [1]
    for row in rows_for_full_merge:
        ws.merge_cells(f'A{row}:N{row}')

    wb.save(f"excel_files/{sheet_name}.xlsx")


def create_cash_flow(data, request, other_data):
    print("data", data)
    # print("request", request)

    worksheet_data = []

    if len(request['Category']) > 1:
        heading = 'Heron'
    else:
        heading = request['Category'][0]

    wb = Workbook()

    ws = wb.active
    ws.title = "Cash Flow"

    row1_data = [f"Weekly Cashflow ({heading}) - {request['date']} ", " ", ""]

    row2_data = [" ", " ", ""]

    row3_data = ["Date", "Amount", "Units"]

    worksheet_data.append(row1_data)
    worksheet_data.append(row2_data)
    worksheet_data.append(row3_data)
    worksheet_data.append(row2_data)

    for item in data:
        row_data = [item['date'], item['total_cashflow'], item['units']]
        worksheet_data.append(row_data)

    for item in worksheet_data:
        # print(item)
        ws.append(item)

    # format column B as currency
    for cell in ws['B']:
        cell.number_format = 'R#,##0.00'

    # format column A as date
    for cell in ws['A']:
        cell.number_format = 'YYYY-MM-DD'

    # set column B width to 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['A'].width = 15

    # add a worksheet
    ws2 = wb.create_sheet("Cash Flow Summary")

    worksheet_data = []
    row1_data = ["Units", "Date", "Amount", "Rollover Date", "Rollover Amount"]
    worksheet_data.append(row1_data)

    for item in other_data:
        row_data = [item['opportunity_code'], item['opportunity_end_date'], item['nett_cashflow'],
                    item['rollover_date'], item['rollover_amount']]
        worksheet_data.append(row_data)

    for item in worksheet_data:
        # print(item)
        ws2.append(item)

    for cell in ws2['C']:
        cell.number_format = 'R#,##0.00'

    for cell in ws2['E']:
        cell.number_format = 'R#,##0.00'

    # format column A as date
    for cell in ws2['B']:
        cell.number_format = 'YYYY-MM-DD'

    for cell in ws2['D']:
        cell.number_format = 'YYYY-MM-DD'

    # set column B width to 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['C'].width = 20

    wb.save(f"excel_files/Cashflow {heading}.xlsx")


    return f"Cashflow {heading}.xlsx"

# create_cash_flow([{
#     "date": "2023/06/25",
#     "total_cashflow": 1520710.5982950684,
#     "units": "HVC102,HVC106,HVC105,HVC103,HVC101,HVC104"
# },
#     {
#         "date": "2023/06/27",
#         "total_cashflow": 58374.981506849406,
#         "units": "HFB214"
#     },
#     {
#         "date": "2023/07/18",
#         "total_cashflow": 2381304.6562584257,
#         "units": "HVC305,HVC205,HVC302,HVC204,HVC301,HVC206,HVC304,HVC303,HVC306,HVC201,HVC202,HVC203"
#     },
#     {
#         "date": "2023/08/19",
#         "total_cashflow": 4305378.80258411,
#         "units": "HVP101,HVP102,HVP201,HVP103,HVP203,HVP301,HVP302,HVP303,HVP202"
#     },
#     {
#         "date": "2023/09/19",
#         "total_cashflow": 1311740.05,
#         "units": "HFB315"
#     },
#     {
#         "date": "2023/11/07",
#         "total_cashflow": 2547835.3119508224,
#         "units": "HVD101,HVD102,HVD103,HVD104,HVD202,HVD201,HVD203,HVD204,HVD301,HVD302,HVD303,HVD304"
#     },
#     {
#         "date": "2023/12/05",
#         "total_cashflow": 3449508.8032786306,
#         "units": "HVN101,HVN102,HVN103,HVN104,HVN201,HVN202,HVN203,HVN204,HVN301,HVN302,HVN303,HVN304"
#     }],
#     {
#         "Category": ["Heron Fields", "Heron View"],
#         "date": "2023/05/31",
#         "firstName": "Wayne"
#     }
# )
