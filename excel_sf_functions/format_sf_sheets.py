import time

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment


def format_sales_forecast(sheet):
    # make all column_widths in ws 20 wide
    for col in sheet.columns:
        col_width = 15
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = col_width

    # for rows 6 and 7, if the value in the cell from column 6 is FALSE, set the value to 'No' else set the value
    # to 'Yes'
    for index, row in enumerate(sheet.iter_rows(min_row=6, min_col=7, max_row=7, max_col=sheet.max_column)):
        for cell in row:
            if not cell.value:
                sheet[f'{get_column_letter(cell.column)}{cell.row}'] = 'No'
            else:
                sheet[f'{get_column_letter(cell.column)}{cell.row}'] = 'Yes'

    # In cell B6, count the values of all records from F6 to the last column in row 6
    sheet['B6'] = f'=COUNTIF(G7:{get_column_letter(sheet.max_column)}7,"*")'
    sheet['D6'] = f'=COUNTIF(G7:{get_column_letter(sheet.max_column)}7,"Yes")'
    sheet[
        'E6'] = f'=COUNTIF(G6:{get_column_letter(sheet.max_column)}6,"Yes")-' \
                f'COUNTIF(G7:{get_column_letter(sheet.max_column)}7,"Yes") '
    # In Cell F6, Add the value of B6 less D6 less E6
    sheet['F6'] = f'=B6-D6-E6'

    # create a list of column letters from column 2 to the last column in the sheet
    column_letters_2 = [get_column_letter(x) for x in range(2, sheet.max_column + 1)]

    # splice this list into a new list starting only at column 7
    column_letters_7 = column_letters_2[5:]

    # create a list of rows to add stuff to
    columns_to_add_text = ['A']
    new_rows = [71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 83, 84, 86, 87]

    for letter in columns_to_add_text:
        for row in new_rows:
            if row == 71:
                sheet[f'{letter}71'] = 'EXITED REPAYABLE ON TRANSFER'
            elif row == 72:
                sheet[f'{letter}72'] = ' EXIT ON TRANSFER'
            elif row == 73:
                sheet[f'{letter}73'] = 'EXITED NOT SOLD'
            elif row == 74:
                sheet[f'{letter}74'] = 'STILL TO EXIT NOT SOLD'
            elif row == 75:
                sheet[f'{letter}75'] = 'TOTAL TO EXIT'
            elif row == 76:
                sheet[f'{letter}76'] = 'TOTAL TO EXIT'
            elif row == 77:
                sheet[f'{letter}77'] = 'UNALLOCATED'
            elif row == 78:
                sheet[f'{letter}78'] = 'TOTAL TO STILL EXIT'
            elif row == 79:
                sheet[f'{letter}79'] = 'Exit Repayments due'
            elif row == 80:
                sheet[f'{letter}80'] = ''
            elif row == 81:
                sheet[f'{letter}81'] = ''
            elif row == 83:
                sheet[f'{letter}83'] = 'EXITED REPAYABLE ON TRANSFER - Capital Only'
            elif row == 84:
                sheet[f'{letter}84'] = 'EXITED REPAYABLE ON TRANSFER - Interest Only'
            elif row == 86:
                sheet[f'{letter}86'] = 'EXIT NOT SOLD - Capital Only'
            elif row == 87:
                sheet[f'{letter}87'] = 'EXIT NOT SOLD - Interest Only'

    rows_to_add_formulas = [17, 23, 27, 29, 31, 34, 35, 39, 40, 51, 52, 53, 55, 59, 60, 61, 62, 63, 64, 67, 68, 71, 72,
                            73, 74, 75, 76, 77, 78, 79, 81, 83, 84, 86, 87]

    for letter in column_letters_7:
        for row in rows_to_add_formulas:
            if row == 17:
                # =IF(G3=FALSE, +G13 - G14, 0)
                sheet[f'{letter}{row}'] = f'=if({letter}3=FALSE,SUM({letter}13-{letter}14),0)'
            elif row == 23:
                sheet[f'{letter}{row}'] = f'=SUM({letter}22-{letter}21)'
            elif row == 27:
                sheet[f'{letter}{row}'] = f'=SUM({letter}24+{letter}25+{letter}26)'
            elif row == 29:
                sheet[f'{letter}{row}'] = f'=720-{letter}23'
            elif row == 31:
                # =IF(G19 <> 0, G19 * 0.0275 / 365 * (G21 - G20), 0)
                sheet[f'{letter}{row}'] = f'=IF({letter}19<>0,{letter}19*0.0275/365*({letter}21-{letter}20),0)'
            elif row == 34:
                sheet[f'{letter}{row}'] = f'=SUM({letter}30+{letter}31+{letter}32+{letter}33)'
            elif row == 35:
                sheet[f'{letter}{row}'] = f'=SUM({letter}19+{letter}34)'
            elif row == 39:
                sheet[f'{letter}{row}'] = f'=SUM({letter}37+{letter}38)'
            elif row == 40:
                sheet[f'{letter}{row}'] = f'=SUM({letter}19-{letter}39)'
            elif row == 51:
                sheet[f'{letter}{row}'] = f'={letter}42-SUM({letter}45:{letter}50)'
            elif row == 52:
                sheet[f'{letter}{row}'] = f'=SUMIFS({column_letters_7[0]}35:' \
                                          f'{column_letters_7[len(column_letters_7) - 1]}35,{column_letters_7[0]}4:' \
                                          f'{column_letters_7[len(column_letters_7) - 1]}4, {letter}4)'
            elif row == 53:
                sheet[f'{letter}{row}'] = f'={letter}51-{letter}52'
            elif row == 55:
                # =IFERROR(+D51/D50,0)
                sheet[f'{letter}{row}'] = f'=IFERROR({letter}52/{letter}51,0)'
            elif row == 59:
                sheet[f'{letter}{row}'] = f'=IF({letter}58=TRUE, {letter}22, "")'
            elif row == 60:
                sheet[f'{letter}{row}'] = f'=IF({letter}58=TRUE, {letter}35, 0)'
            elif row == 61:
                # =SUMIFS($G59:$KO59,$G4:$KO4, G4)
                sheet[f'{letter}{row}'] = f'=SUMIFS($G60:' \
                                          f'${column_letters_7[len(column_letters_7) - 1]}60,$G4:' \
                                          f'${column_letters_7[len(column_letters_7) - 1]}4, {letter}4)'
            elif row == 62:
                sheet[f'{letter}{row}'] = f'=SUM({letter}51)'
            elif row == 63:
                sheet[f'{letter}{row}'] = f'=SUM({letter}52-{letter}61)'
            elif row == 64:
                sheet[f'{letter}{row}'] = f'=SUM({letter}62-{letter}63)'
            elif row == 67:
                # =SUMIFS(G19: KO19, G57: KO57, TRUE)
                sheet[f'B{row}'] = f'=SUMIFS(G19:' \
                                   f'{column_letters_7[len(column_letters_7) - 1]}19,G58:' \
                                   f'{column_letters_7[len(column_letters_7) - 1]}58, TRUE)'
            elif row == 68:
                # =SUMIFS(G33: KO33, G57: KO57, TRUE)
                sheet[f'B{row}'] = f'=SUMIFS(G34:' \
                                   f'{column_letters_7[len(column_letters_7) - 1]}34,G58:' \
                                   f'{column_letters_7[len(column_letters_7) - 1]}58, TRUE)'
            elif row == 71:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2 = TRUE, IF({letter}$3 = FALSE, IF({letter}$58 = TRUE, +{letter}$35, 0), 0), 0)'
            elif row == 72:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2=TRUE,IF({letter}$3=FALSE,IF({letter}$58=FALSE,+{letter}$35,0),0),0)'
            elif row == 73:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2=FALSE,IF({letter}$3=FALSE,IF({letter}$58=TRUE,+{letter}$35,0),0),0)'
            elif row == 74:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2=FALSE,IF({letter}$3=FALSE,IF({letter}$58=FALSE,+{letter}$35,0),0),0)'
            elif row == 75:
                sheet[f'{letter}{row}'] = f'=SUM({letter}71:{letter}74)'
            elif row == 76:
                sheet[f'{letter}{row}'] = f'=IF({letter}3=FALSE,{letter}35,0)'
            elif row == 77:
                sheet[f'{letter}{row}'] = f'==IF({letter}2=FALSE,{letter}13-{letter}14,0)'
            elif row == 78:
                # =IF(G10 <> "ZZUN01", IF(G25 <> 0, G76 - G79, IF(G21=$B$1, G76 - G79, 0)), 0)

                sheet[f'{letter}{row}'] = f'=IF({letter}10<>"ZZUN01",IF({letter}25<>0,{letter}76-{letter}79,IF({letter}21=$B$1,{letter}76-{letter}79,0)),0)'
                # sheet[f'{letter}{row}'] = f'={letter}76-{letter}79'
            elif row == 79:
                sheet[f'{letter}{row}'] = f'=SUM({letter}71,{letter}73)'
            elif row == 81:
                sheet[f'{letter}{row}'] = f'={letter}76-{letter}75'
            elif row == 83:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2 = TRUE, IF({letter}$3 = FALSE, IF({letter}$58 = TRUE, +{letter}$19, 0), 0), 0)'
            elif row == 84:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2 = TRUE, IF({letter}$3 = FALSE, IF({letter}$58 = TRUE, +{letter}$34, 0), 0), 0)'
            elif row == 86:
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2=FALSE,IF({letter}$3=FALSE,IF({letter}$58=TRUE,+{letter}$19,0),0),0)'
            elif row == 87:
                # =IF(G$2 = FALSE, IF(G$3 = FALSE, IF(G$57 = TRUE, +G$33, 0), 0), 0)
                sheet[
                    f'{letter}{row}'] = f'=IF({letter}$2=FALSE,IF({letter}$3=FALSE,IF({letter}$58=TRUE,+{letter}$34,0),0),0)'

    # format all rows with data except rows 1 to 11, 20 to 23, 28 and 29 as currency with 2 decimal places and
    # comma every 3 digits, bold and white font, and for row 11 as a percentage with 2 decimal places and comma
    # every 3 digits
    for row in sheet.iter_rows(min_row=12, min_col=2, max_row=64, max_col=sheet.max_column):
        for cell in row:
            if cell.row == 11:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.number_format = '0.00%'
            if cell.row == 55:
                cell.number_format = '0.00%'
            elif cell.row == 20 or cell.row == 21 or cell.row == 22 or cell.row == 23 or cell.row == 28 \
                    or cell.row == 29:
                continue
            else:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.number_format = 'R#,##0.00'

    rows_to_center = [5, 6, 7, 9, 10, 11, 13, 14, 15, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32,
                      33, 34, 35, 37, 38, 39, 40, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 55, 59, 60, 61, 62,
                      63, 64, 71, 72, 73, 74, 75, 76, 77, 78, 79, 81, 83, 84, 86, 87]
    # Loop through the rows_to_format_currency list and align the cells from column 6 to the last column in the
    # centre
    for row in rows_to_center:
        # COLUMNS TO FILL ETC
        if row in [9, 10, 11, 20, 21, 22, 23, 28, 29]:
            col_number = 7
        else:
            col_number = 2
        # DARKER COLOR

        if row in [9, 10, 23, 13, 19, 27, 29, 33, 34, 35, 39, 40, 44]:
            color_is = '3E54AC'
        elif row in [53, 64]:
            color_is = 'E14D2A'
        else:
            color_is = '537FE7'
        for col in sheet.iter_cols(min_row=row, min_col=col_number, max_row=row, max_col=sheet.max_column):
            for cell in col:
                if row == 11:
                    if cell.value == 0.18:
                        color_is = 'EA047E'
                    elif cell.value == 0.16:
                        color_is = 'ED50F1'
                    elif cell.value == 0.15:
                        color_is = 'FF4A4A'
                    elif cell.value == 0.14:
                        color_is = 'FF9E9E'

                    sheet[f'{get_column_letter(cell.column)}11'].number_format = '0.00%'
                    sheet[f'{get_column_letter(cell.column)}11'].alignment = Alignment(horizontal='center')

                if row == 9:
                    if cell.value == 'UnAllocated':
                        color_is = '00B7C2'
                    else:
                        color_is = '3E54AC'

                if row == 10:
                    if cell.value == 'ZZUN01':
                        color_is = '00B7C2'
                    else:
                        color_is = '3E54AC'

                if row == 33:
                    if cell.value != 0:
                        color_is = '00B7C2'
                    else:
                        color_is = '3E54AC'

                # if row == 58:
                #     if cell.value:
                #
                #         color_is = '00B7C2'
                #     else:
                #         color_is = '3E54AC'

                cell.fill = PatternFill(start_color=color_is, end_color=color_is, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.border = Border(left=Side(border_style='medium', color='000000'),
                                     right=Side(border_style='medium', color='000000'),
                                     top=Side(border_style='medium', color='000000'),
                                     bottom=Side(border_style='medium', color='000000'))

    rows_to_hide = [2, 3, 4, 54, 58]
    # Loop through the rows_to_hide list and hide the rows
    for row in rows_to_hide:
        sheet.row_dimensions[row].hidden = True

    true_false = [58]

    for row in true_false:
        for col in sheet.iter_cols(min_row=row, min_col=1, max_row=row, max_col=sheet.max_column):
            for cell in col:
                # If the cell value in row 57 is True then set the color of the cell in row 58 to green
                if cell.value:
                    sheet[f'{get_column_letter(cell.column)}59'].fill = PatternFill(start_color='A555EC',
                                                                                    end_color='A555EC',
                                                                                    fill_type='solid')
                    sheet[f'{get_column_letter(cell.column)}60'].fill = PatternFill(start_color='A555EC',
                                                                                    end_color='A555EC',
                                                                                    fill_type='solid')
                    sheet[f'{get_column_letter(cell.column)}22'].fill = PatternFill(start_color='A555EC',
                                                                                    end_color='A555EC',
                                                                                    fill_type='solid')

    # Set column 1 to 25 units wide and make column 1 bold
    for col in sheet.iter_cols(min_row=1, min_col=1, max_row=sheet.max_row, max_col=5):
        for cell in col:
            sheet.column_dimensions[cell.column_letter].width = 30
            cell.font = Font(bold=True)

    # ROWS TO SUM
    rows_to_sum = [13, 14, 15, 16, 17, 19, 24, 25, 26, 27, 30, 31, 32, 33, 34, 35, 37, 38, 39, 42, 43, 44, 45, 46,
                   47, 48, 49, 50, 51, 52, 53, 55, 60, 61, 62, 63, 64, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 83,
                   84, 86, 87]
    # for each row in rows_to_sum, sum the cells from column 6 to the last column in the sheet in insert the
    # formula in column 'B', with white font and bold and format the cell as currency with 2 decimal places and
    # comma every 3 digits

    sheet['B3'] = f'=COUNTIF({get_column_letter(7)}10:{get_column_letter(sheet.max_column)}10,"<>ZZUN01")'
    # =COUNTIFS(G3: FM3, TRUE, G10: FM10, "<>ZZUN01")
    sheet['D3'] = f'=COUNTIFS({get_column_letter(7)}3:{get_column_letter(sheet.max_column)}3,TRUE,{get_column_letter(7)}10:{get_column_letter(sheet.max_column)}10,"<>ZZUN01")'
    sheet['A54'] = f'LTV'

    columns_to_format = ['B', 'C', 'D', 'E', 'F']

    for row in rows_to_sum:
        sheet[f'B{row}'] = f'=SUM({get_column_letter(7)}{row}:{get_column_letter(sheet.max_column)}{row})'
        for column in columns_to_format:
            sheet[f'{column}{row}'].font = Font(bold=True, color='FFFFFF')
            sheet[f'{column}{row}'].number_format = 'R#,##0.00'
            sheet[f'{column}{row}'].alignment = Alignment(horizontal='center')

            if column == 'D':
                if row == 52:
                    sheet[
                        f'{column}{row}'] = f'=+D35'
                elif row == 53:
                    sheet[
                        f'{column}{row}'] = f'=+D51-D52'
                elif row == 55:
                    sheet[
                        f'{column}{row}'] = f'=IFERROR(+D52/D51,0)'
                    sheet[f'{column}{row}'].number_format = '0.00%'
                    sheet[f'{column}{row}'].font = Font(bold=True, color='000000')

                elif row >= 71:
                    sheet[
                        f'{column}{row}'] = ""

                else:
                    sheet[
                        f'{column}{row}'] = f'=SUMIFS($G{row}:${get_column_letter(sheet.max_column)}{row},' \
                                            f'$G3:${get_column_letter(sheet.max_column)}3,' \
                                            f'TRUE)'

            if column == 'E':
                if row == 52:
                    sheet[
                        f'{column}{row}'] = f'=+E35'
                elif row == 53:
                    sheet[
                        f'{column}{row}'] = f'=+E51-E52'
                elif row == 55:
                    # =IFERROR(+D51 / D50, 0)
                    sheet[
                        f'{column}{row}'] = f'=IFERROR(+E52/E51,0)'
                    sheet[f'{column}{row}'].number_format = '0.00%'
                    sheet[f'{column}{row}'].font = Font(bold=True, color='000000')

                elif row >= 71:
                    sheet[
                        f'{column}{row}'] = ""
                else:
                    sheet[
                        f'{column}{row}'] = f'=SUMIFS($G{row}:${get_column_letter(sheet.max_column)}{row},' \
                                            f'$G2:${get_column_letter(sheet.max_column)}2,' \
                                            f'TRUE)-SUMIFS($G{row}:${get_column_letter(sheet.max_column)}{row},' \
                                            f'$G3:${get_column_letter(sheet.max_column)}3,' \
                                            f'TRUE)'

            if column == 'F':
                if row >= 71:
                    sheet[
                        f'{column}{row}'] = ""
                else:
                    sheet[
                        f'{column}{row}'] = f'=B{row}-D{row}-E{row}'

    # Join rows_to_add_formulas,rows_to_format_currency,rows_to_center and rows_to_sum as one new list ordered
    # and only unique values all_rows_to_format = list(set(rows_to_add_formulas + rows_to_format_currency +
    # rows_to_center + rows_to_sum)) all_rows_to_format.sort()

    # If the values in row 16 from column 2 are greater than 0, then the cells in row 16 from column 2 must be
    # filled with red and the font must be white
    sheet['B16'].fill = PatternFill(start_color='DF7857', end_color='DF7857', fill_type='solid')
    for col in sheet.iter_cols(min_row=16, min_col=7, max_row=16, max_col=sheet.max_column):
        for cell in col:
            if type(cell.value) == float:
                if cell.value > 0:
                    for row in range(16, 17):
                        cell.fill = PatternFill(start_color='DF7857', end_color='DF7857', fill_type='solid')
                        cell.font = Font(color='FFFFFF')

    # from column 7 to the last column in the sheet, if the value in the cell in row 2 is not equal to  0,
    # then the corresponding cell in row 19 must have a red fill and the font white
    for col in sheet.iter_cols(min_row=54, min_col=7, max_row=54, max_col=sheet.max_column):
        for cell in col:
            if cell.value != 0:
                for row in range(19, 20):
                    sheet[f'{get_column_letter(cell.column)}{row}'].fill = PatternFill(start_color='DF7857',
                                                                                       end_color='DF7857',
                                                                                       fill_type='solid')
                    sheet[f'{get_column_letter(cell.column)}{row}'].font = Font(color='FFFFFF')

    sheet['D5'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Make cell E5 fill with dark-green and the font white
    sheet['E5'].fill = PatternFill(start_color='539165', end_color='539165', fill_type='solid')

    # loop through cells in row 5, row 6 and row 7 from column 7. If the value in the cells in row 7 equals 'Yes'
    # then fill cells in row 5, 6 and 7 with red and the font white
    for col in sheet.iter_cols(min_row=7, min_col=7, max_row=7, max_col=sheet.max_column):
        for cell in col:
            # if cell.value == 'Yes' then make cells in row 5, 6 and 7 fill with red and the font white
            if cell.value == 'Yes':
                for row in range(5, 8):
                    sheet[f'{cell.column_letter}{row}'].fill = PatternFill(start_color='FF0000',
                                                                           end_color='FF0000',
                                                                           fill_type='solid')
                    sheet[f'{cell.column_letter}{row}'].font = Font(color='FFFFFF')

    # loop through cells in row 6 and row 7 from column 7. If the value in the cells in row 7 equals 'No' and the
    # cells in row 6 equals 'Yes' then fill cells in row 6 and 7 with dark-green and the font white
    for col in sheet.iter_cols(min_row=6, min_col=7, max_row=6, max_col=sheet.max_column):
        for cell in col:
            # if cell.value == 'Yes' then make cells in row 5, 6 and 7 fill with red and the font white
            if cell.value == 'Yes':
                for row in range(7, 8):
                    if sheet[f'{cell.column_letter}{row}'].value == 'No':
                        for row1 in range(5, 8):
                            sheet[f'{cell.column_letter}{row1}'].fill = PatternFill(start_color='539165',
                                                                                    end_color='539165',
                                                                                    fill_type='solid')
                            sheet[f'{cell.column_letter}{row1}'].font = Font(color='FFFFFF')

    # MERGE CELLS
    # merge_start_time = time.time()

    merge_master = []
    merge_start = []
    merge_end = []

    # CREATE MERGE LISTS Loop through all the cells in row 4 until the last column and insert the column number
    # and the cell value as a dict into the merge_start list
    for index, row in enumerate(sheet.iter_rows(min_row=4, min_col=6, max_row=4, max_col=sheet.max_column)):
        for cell in row:
            merge_master.append({'column': cell.column, 'value': cell.value})

    # Loop through merge_master, from index 1 to the end of the list, if the value of 'value' is different to the
    # value of the previous item in the list, add the column number to the merge_start list
    for index, item in enumerate(merge_master):
        if index > 0:
            if item['value'] != merge_master[index - 1]['value']:
                merge_start.append(item['column'])

        # Loop through merge_master, from index 1 to the end of the list, if the value of 'value' is different to
        # the value of the next item in the list, add the column number to the merge_end list
        if 0 < index < len(merge_master) - 1:
            if item['value'] != merge_master[index + 1]['value']:
                merge_end.append(item['column'])

    merge_end.append(merge_master[len(merge_master) - 1]['column'])
    # start_time = time.time()
    # Create a dictionary to store the start and end columns for each row
    rows_to_merge = [5, 6, 7, 13, 14, 15, 16, 17, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 55, 61, 62, 63, 64]
    merge_dict = {}

    # Loop through the rows_to_merge list and populate merge_dict with start and end columns for each row
    for row in rows_to_merge:
        merge_dict[row] = []
        start_column = None
        end_column = None
        for cell in sheet[row]:
            if cell.column in merge_start:
                start_column = cell.column
            elif cell.column in merge_end:
                end_column = cell.column
            if start_column and end_column:
                merge_dict[row].append((start_column, end_column))
                start_column = None
                end_column = None

    # Merge cells for each row in bulk using the merge_dict
    for row, merge_cols in merge_dict.items():
        for start, end in merge_cols:
            sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)

    # Merge cells in row 6 and 7 in column B
    cells_to_merge = [6]
    columns_affected = ['B', 'C', 'D', 'E', 'F']
    for columns_affected in columns_affected:
        for row in cells_to_merge:
            sheet.merge_cells(f'{columns_affected}{row}:{columns_affected}{row + 1}')

    sheet.merge_cells('B6:B7')

    # Format 'B5' to 'F7' with white font and bold
    for row in sheet.iter_rows(min_row=5, min_col=2, max_row=7, max_col=6):
        for cell in row:
            cell.font = Font(bold=True, color='FFFFFF')

    # Freeze frames at C6
    sheet.freeze_panes = 'C6'
