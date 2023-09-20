# from datetime import timedelta
# from datetime import datetime
import os

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter, column_index_from_string

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# from openpyxl.utils import range_boundaries

# from excel_sf_functions.create_sales_sheet import create_excel_array
# from excel_sf_functions.create_NSST_sheets import create_nsst_sheet
# from excel_sf_functions.format_sf_sheets import format_sales_forecast
# from excel_sf_functions.format_nsst_sheets import format_nsst
# import time
# import threading


def create_draw_history_report(data, pledges, opportunities):
    filename = 'excel_files/current_funds_available.xlsx'
    if os.path.exists(filename):
        os.remove(filename)
        print("File Removed!")
    else:
        print("The file does not exist")

    # create workbook
    wb = Workbook()
    # create sheet
    ws = wb.active
    ws.title = "Draw History Report"
    # tabb color = green
    ws.sheet_properties.tabColor = "1072BA"

    data_to_insert = []
    row1 = ["CURRENT FUNDS AVAILABLE TO DRAW"]
    row2 = []
    row3 = ["Unit Number", "Acc.", "Investor", "Capital Amount", "Investment Date", "Amount Available to Draw",
            "Draw Date", "Total Drawn to Date", "Planned Draw Date", "Planned Draw No:"]

    data_to_insert.append(row1)
    data_to_insert.append(row2)
    data_to_insert.append(row3)

    for draw in data:
        row = [draw['opportunity_code'], draw['investor_acc_number'], draw['investment_name'],
               draw['investment_amount'],
               draw['investment_date'], draw['available_to_draw'], draw['draw_date'],
               draw['drawn_to_date'], draw['planned_draw_date'], draw['draw']]
        data_to_insert.append(row)

    for data_in in data_to_insert:
        ws.append(data_in)

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    for col in range(3, 4):
        ws.column_dimensions[get_column_letter(col)].width = 33

    # merge first row from A1 to H1, then center it and make it bold and 20 font size and blue color and add thick
    # border
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(size=20, bold=True, color="1072BA")
    ws['A1'].border = Border(top=Side(border_style='thick', color='000000'),
                             left=Side(border_style='thick', color='000000'),
                             right=Side(border_style='thick', color='000000'),
                             bottom=Side(border_style='thick', color='000000'))

    # center all cells in row 3, make font bold and add thick border and make background color gray and wrap text
    for cell in ws["3:3"]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(border_style='thick', color='000000'),
                             left=Side(border_style='thick', color='000000'),
                             right=Side(border_style='thick', color='000000'),
                             bottom=Side(border_style='thick', color='000000'))
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

    columns_to_center = ['A', 'B', 'C', 'E', 'G', 'I', 'J']
    # center everything in columns_to_center from row4 until the end, and put a thin border around it
    for col in columns_to_center:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    columns_as_currency = ['D', 'F', 'H']
    # format columns_as_currency as currency and put a thin border around it
    for col in columns_as_currency:
        for cell in ws[col]:
            cell.number_format = 'R#,##0.00'
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    max_row = ws.max_row

    # make columns I & J have a yellow fill
    columns_to_fill = ['I', 'J']

    for col in columns_to_fill:
        for row in range(4, max_row + 1):
            cell = ws[col + str(row)]
            cell.fill = PatternFill(start_color="FAF2D3", end_color="FAF2D3", fill_type="solid")

    # put a filter in place in row 3 for data from row 4 until the end
    ws.auto_filter.ref = "A3:J" + str(max_row)

    # in columns D, F, H, add a sum of the column, and format it as currency as above, and fill it with gray and put
    # a thick border around each one
    for col in columns_as_currency:
        cell = ws[col + str(max_row + 2)]
        # =SUBTOTAL(9,D4:D707)
        cell.value = f"=SUBTOTAL(9, {col}4:{col}{max_row})"
        cell.number_format = 'R#,##0.00'
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        cell.border = Border(top=Side(border_style='medium', color='000000'),
                             left=Side(border_style='medium', color='000000'),
                             right=Side(border_style='medium', color='000000'),
                             bottom=Side(border_style='medium', color='000000'))

    # freeze panes at row 4
    ws.freeze_panes = "A4"

    # add a sheet
    ws2 = wb.create_sheet("Pledges")
    # tabb color = green
    ws2.sheet_properties.tabColor = "004225"

    data_to_insert = []
    row1 = ["CURRENT PLEDGES IN SYSTEM"]
    row2 = []
    row3 = ["Unit Number", "Acc.", "Investor", "Capital Amount"]

    data_to_insert.append(row1)
    data_to_insert.append(row2)
    data_to_insert.append(row3)

    for pledge in pledges:
        row = [pledge['opportunity_code'], pledge['investor_acc_number'], pledge['investment_name'],
               pledge['investment_amount'],
               ]
        data_to_insert.append(row)

    for data_in in data_to_insert:
        ws2.append(data_in)

    for col in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    for col in range(3, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 33

    # merge first row from A1 to H1, then center it and make it bold and 20 font size and blue color and add thick
    # border
    ws2.merge_cells('A1:D1')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2['A1'].font = Font(size=20, bold=True, color="1072BA")
    ws2['A1'].border = Border(top=Side(border_style='thick', color='000000'),
                              left=Side(border_style='thick', color='000000'),
                              right=Side(border_style='thick', color='000000'),
                              bottom=Side(border_style='thick', color='000000'))

    # center all cells in row 3, make font bold and add thick border and make background color gray and wrap text
    for cell in ws2["3:3"]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(border_style='thick', color='000000'),
                             left=Side(border_style='thick', color='000000'),
                             right=Side(border_style='thick', color='000000'),
                             bottom=Side(border_style='thick', color='000000'))
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

    columns_to_center = ['A', 'B', 'C']
    # center everything in columns_to_center from row4 until the end, and put a thin border around it
    for col in columns_to_center:
        for cell in ws2[col]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    columns_as_currency = ['D']
    # format columns_as_currency as currency and put a thin border around it
    for col in columns_as_currency:
        for cell in ws2[col]:
            cell.number_format = 'R#,##0.00'
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    max_row = ws2.max_row

    # put a filter in place in row 3 for data from row 4 until the end
    ws2.auto_filter.ref = "A3:D" + str(max_row)

    # in columns D, F, H, add a sum of the column, and format it as currency as above, and fill it with gray and put
    # a thick border around each one
    for col in columns_as_currency:
        cell = ws2[col + str(max_row + 2)]
        # =SUBTOTAL(9,D4:D707)
        cell.value = f"=SUBTOTAL(9, {col}4:{col}{max_row})"
        cell.number_format = 'R#,##0.00'
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        cell.border = Border(top=Side(border_style='medium', color='000000'),
                             left=Side(border_style='medium', color='000000'),
                             right=Side(border_style='medium', color='000000'),
                             bottom=Side(border_style='medium', color='000000'))

    # freeze panes at row 4
    ws2.freeze_panes = "A4"

    # add a sheet
    ws3 = wb.create_sheet("Shortfall")
    # tabb color = green
    ws3.sheet_properties.tabColor = "D83F31"

    data_to_insert = []
    row1 = ["CURRENT FUNDING"]
    row2 = []
    row3 = ["Development", "Unit Number", "Investment Required", "Actual Investment", "Shortfall", "Pledged",
            "Net Shortfall"]

    data_to_insert.append(row1)
    data_to_insert.append(row2)
    data_to_insert.append(row3)

    for opportunity in opportunities:
        row = [opportunity['Category'], opportunity['opportunity_code'],
               float(opportunity['opportunity_amount_required'])]
        data_to_insert.append(row)

    for opportunity_in in data_to_insert:
        ws3.append(opportunity_in)

    for col in range(1, ws3.max_column + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    ws3.merge_cells('A1:G1')
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3['A1'].font = Font(size=20, bold=True, color="1072BA")
    ws3['A1'].border = Border(top=Side(border_style='thick', color='000000'),
                              left=Side(border_style='thick', color='000000'),
                              right=Side(border_style='thick', color='000000'),
                              bottom=Side(border_style='thick', color='000000'))

    for cell in ws3["3:3"]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(border_style='thin', color='000000'),
                             left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

    columns_to_center = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    # # center everything in columns_to_center from row4 until the end, and put a thin border around it
    for col in columns_to_center:
        for cell in ws3[col]:
            # cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    max_ws3_row = ws3.max_row

    max_ws2_row = ws2.max_row - 2

    max_ws1_row = ws.max_row - 2

    ws2_sheet = 'Pledges'
    ws1_sheet = 'Draw History Report'

    # in column D from row 4 until the end, add a formula =SUMIFS('Draw History Report'!$H$3:$H$723,'Draw History
    # Report'!$A$3:$A$723,Shortfall!B4) and format it as currency as above
    for row in range(4, max_ws3_row + 1):
        cell = ws3['D' + str(row)]
        cell.value = f"=SUMIFS('{ws1_sheet}'!$H$3:$H${max_ws1_row},'{ws1_sheet}'!$A$3:$A${max_ws1_row},{'B' + str(row)})"
        cell.number_format = 'R#,##0.00'

    # in column E insert =C4-D4 and format it as currency as above
    for row in range(4, max_ws3_row + 1):
        cell = ws3['E' + str(row)]
        cell.value = f"=C{row}-D{row}"
        cell.number_format = 'R#,##0.00'

    # in column F insert =SUMIFS(Pledges!$D$3:$D$31,Pledges!$A$3:$A$31,Shortfall!B4) and format it as currency as above
    for row in range(4, max_ws3_row + 1):
        cell = ws3['F' + str(row)]
        cell.value = f"=SUMIFS('{ws2_sheet}'!$D$3:$D${max_ws2_row},'{ws2_sheet}'!$A$3:$A${max_ws2_row},{'B' + str(row)})"
        cell.number_format = 'R#,##0.00'

    # in column G insert =E4-F4 and format it as currency as above
    for row in range(4, max_ws3_row + 1):
        cell = ws3['G' + str(row)]
        cell.value = f"=E{row}-F{row}"
        cell.number_format = 'R#,##0.00'

    columns_as_currency = ['C', 'D', 'E', 'F', 'G']
    # format columns_as_currency as currency and put a thin border around it
    for col in columns_as_currency:
        for cell in ws3[col]:
            cell.number_format = 'R#,##0.00'
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    max_row = ws3.max_row

    # put a filter in place in row 3 for data from row 4 until the end
    ws3.auto_filter.ref = "A3:G" + str(max_row)

    # in columns D, F, H, add a sum of the column, and format it as currency as above, and fill it with gray and put
    # a thick border around each one
    subtotal_columns = ['C', 'D', 'E', 'F', 'G']

    for col in subtotal_columns:
        cell = ws3[col + str(max_row + 2)]
        # =SUBTOTAL(9,D4:D707)
        cell.value = f"=SUBTOTAL(9, {col}4:{col}{max_row})"
        cell.number_format = 'R#,##0.00'
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        cell.border = Border(top=Side(border_style='medium', color='000000'),
                             left=Side(border_style='medium', color='000000'),
                             right=Side(border_style='medium', color='000000'),
                             bottom=Side(border_style='medium', color='000000'))

    # freeze panes at row 4
    ws3.freeze_panes = "A4"

    # save workbook
    wb.save('excel_files/current_funds_available.xlsx')

    return "excel_files/current_funds_available.xlsx"
