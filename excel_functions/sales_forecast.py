from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.styles.borders import Border
from openpyxl.styles import borders
from openpyxl.styles.alignment import Alignment


def create_sales_forecast_file(data, developmentinputdata):
    # print(developmentinputdata)
    filename = 'Sales Forecast'
    # sheet_name = "Sales Forecast"
    # print(filename)

    # CREATE WORKBOOK AND WORKSHEET/S
    wb = Workbook()
    ws = wb.active
    ws.title = filename

    # CREATE EMPTY LISTS TO BE POPULATED FROM DATA BROUGHT IN (POSSIBLY COULD BE DONE AT THE ROUTE LEVEL)
    opportunity_code, opportunity_amount_required, opportunity_end_date, opportunity_final_transfer_date, \
        opportunity_sale_price, opportunity_sold, opportunity_transferred, report_date, trust_release_fee, \
        raising_commission, structuring_fee, commission, transfer_fees, bond_registration, unforseen, \
        investment_amount, investor_acc_number, investor_name, deposit_date, release_date, investment_number, \
        released_investment_amount, investment_end_date, investment_release_date, \
        released_investment_number, investment_interest_rate, rollover_amount, rollover_date, \
        investment_interest_total, investment_interest_to_date, released_interest_total, \
        released_interest_to_date = [], [], [], [], [], [], \
                                    [], [], [], [], [], [], [], \
                                    [], [], [], [], [], [], [], \
                                    [], [], [], [], [], [], [], \
                                    [], [], [], [], []
    final_insert_data = []

    cell_labels = [
        "opportunity_transferred",
        "opportunity_sold",
        "Opportunity",
        "Opportunity",
        "opportunity_transferred",
        "opportunity_sold",
        "investor_name",
        "investor_acc_number",
        "investment_interest_rate",
        "opportunity_amount_required",
        "investment_amount",
        "investment_amount",
        "deposit_date",
        "release_date",
        "opportunity_final_transfer_date",
        "investment_interest_to_date",
        "released_interest_to_date",
        "investment_interest_total",
        "released_interest_total",
        "raising_commission",
        "structuring_fee",
        "opportunity_sale_price",
        "VAT",
        "Gross",
        "commission",
        "transfer_fees",
        "bond_registration",
        "trust_release_fee",
        "unforseen",
        "rollover_amount",
        "rollover_date",

        # "opportunity_end_date",
        # "report_date",
        # "investment_number",
        # "released_investment_amount",
        # "investment_end_date",
        # "investment_release_date",
        # "released_investment_number",
    ]

    # APPEND DATA
    for info in data:
        opportunity_code.append(info['opportunity_code'])
        opportunity_amount_required.append(info['opportunity_amount_required'])
        opportunity_end_date.append(info['opportunity_end_date'])
        opportunity_final_transfer_date.append(info['opportunity_final_transfer_date'])
        opportunity_sale_price.append(info['opportunity_sale_price'])
        opportunity_sold.append(info['opportunity_sold'])
        opportunity_transferred.append(info['opportunity_transferred'])
        report_date.append(info['report_date'])
        trust_release_fee.append(info['trust_release_fee'])
        raising_commission.append(info['raising_commission'])
        structuring_fee.append(info['structuring_fee'])
        commission.append(info['commission'])
        transfer_fees.append(info['transfer_fees'])
        bond_registration.append(info['bond_registration'])
        unforseen.append(info['unforseen'])
        investment_amount.append(info['investment_amount'])
        investor_acc_number.append(info['investor_acc_number'])
        investor_name.append(info['investor_name'])
        deposit_date.append(info['deposit_date'])
        release_date.append(info['release_date'])
        investment_number.append(info['investment_number'])
        released_investment_amount.append(info['released_investment_amount'])
        investment_end_date.append(info['investment_end_date'])
        investment_release_date.append(info['investment_release_date'])
        released_investment_number.append(info['released_investment_number'])
        investment_interest_rate.append(info['investment_interest_rate'] / 100)
        rollover_amount.append(info['rollover_amount'])
        rollover_date.append(info['rollover_date'])
        investment_interest_total.append(info['investment_interest_total'])
        investment_interest_to_date.append(info['investment_interest_to_date'])
        released_interest_total.append(info['released_interest_total'])
        released_interest_to_date.append(info['released_interest_to_date'])

    # PERFORM FUNCTIONS TO GET VALUES (RECREATE TO FILL VALUES AND NOT PERCENTAGES)
    raising_commission = [x * y for x, y in zip(investment_amount, raising_commission)]
    structuring_fee = [x * y for x, y in zip(investment_amount, structuring_fee)]
    vat = [x / 1.15 * .15 for x in opportunity_sale_price]
    gross = [x / 1.15 for x in opportunity_sale_price]
    commission = [x * y for x, y in zip(gross, commission)]
    unforseen = [x * y for x, y in zip(opportunity_sale_price, unforseen)]

    # INSERT DATA INTO A SINGLE LIST
    final_insert_data.append(opportunity_transferred)
    final_insert_data.append(opportunity_sold)
    final_insert_data.append(opportunity_code)
    final_insert_data.append(opportunity_code)
    final_insert_data.append(opportunity_transferred)
    final_insert_data.append(opportunity_sold)
    final_insert_data.append(investor_name)
    final_insert_data.append(investor_acc_number)
    final_insert_data.append(investment_interest_rate)
    final_insert_data.append(opportunity_amount_required)
    final_insert_data.append(investment_amount)
    final_insert_data.append(investment_amount)
    final_insert_data.append(deposit_date)
    final_insert_data.append(release_date)
    final_insert_data.append(opportunity_final_transfer_date)
    final_insert_data.append(investment_interest_to_date)
    final_insert_data.append(released_interest_to_date)
    final_insert_data.append(investment_interest_total)
    final_insert_data.append(released_interest_total)
    final_insert_data.append(raising_commission)
    final_insert_data.append(structuring_fee)
    final_insert_data.append(opportunity_sale_price)
    final_insert_data.append(vat)
    final_insert_data.append(gross)
    final_insert_data.append(commission)
    final_insert_data.append(transfer_fees)
    final_insert_data.append(bond_registration)
    final_insert_data.append(trust_release_fee)
    final_insert_data.append(unforseen)

    # final_insert_data.append(opportunity_end_date)
    # final_insert_data.append(report_date)
    # final_insert_data.append(investment_number)
    # final_insert_data.append(released_investment_amount)
    # final_insert_data.append(investment_end_date)
    # final_insert_data.append(investment_release_date)
    # final_insert_data.append(released_investment_number)

    final_insert_data.append(rollover_amount)
    final_insert_data.append(rollover_date)

    # APPEND SINGLE LIST INTO EXCEL
    for row in final_insert_data:
        ws.append(row)

    # MOVE DATA FROM TOP OF FILE TO MA MORE PLIABLE POSITION
    ws.move_range(f"A1:{get_column_letter(ws.max_column)}{str(ws.max_row)}", rows=5, cols=6)

    # PUT LABELS INTO COLUMN A
    for x in range(6, len(cell_labels) + 6):
        ws[f"A{x}"].value = cell_labels[x - 6]

    ws["B9"].value = "Total"
    ws["C9"].value = "Transferred"
    ws["D9"].value = "Sold"
    ws["E9"].value = "Remaining"
    ws["F9"].value = "Check"

    # VARIABLES USED THROUGHOUT THE FUNCTION
    last_column = get_column_letter(ws.max_column)
    last_column_number = ws.max_column

    # COLUMN WIDTHS
    ws.column_dimensions['A'].width = 30
    for x in range(2, 7):
        ws.column_dimensions[f'{get_column_letter(x)}'].width = 15
    for x in range(7, last_column_number + 1):
        ws.column_dimensions[f'{get_column_letter(x)}'].width = 13

    # INSERT ROWS - ENHANCE LOOK AND FEEL
    ws.insert_rows(12)
    ws.insert_rows(16)
    ws.insert_rows(19, 7)
    ws.insert_rows(30)
    ws.insert_rows(33, 2)
    ws.insert_rows(37, 2)
    ws.insert_rows(39)
    ws.insert_rows(42, 3)
    ws.insert_rows(53, 9)

    # CREATE GENERAL FORMULAE
    formula_row = [10, 11, 18, 19, 20, 21, 22, 23, 24, 33, 37, 38, 41, 42, 46, 52, 53, 54, 55]
    for row in formula_row:
        for col in range(7, last_column_number + 1):
            col_name = get_column_letter(col)
            # INVESTMENTS RECEIVED
            if row == 18:
                ws[f'{col_name}18'] = f'=SUMIFS($G$26:${last_column}$26,$G$8:${last_column}$8, {col_name}8)'
            # LTSV
            if row == 19:
                ws[f'{col_name}19'] = f'={col_name}18/{col_name}45'
            ws["A19"].value = "LTSV"
            # LTRV
            if row == 20:
                ws[f'{col_name}20'] = f'={col_name}18/{col_name}17'
            ws["A20"].value = "LTRV"
            # FUNDS AVAILABLE
            if row == 21:
                ws[f'{col_name}21'] = f'={col_name}17-{col_name}18'
            ws["A21"].value = "Funds Available"
            # UNIT INCOME EXCL VAT
            if row == 22:
                ws[f'{col_name}22'] = f'={col_name}53'
            ws["A22"].value = "Unit Income Excl VAT"
            # REPAYMENT VALUE PREDICTED
            if row == 23:
                ws[f'{col_name}23'] = f'={col_name}54'
            ws["A23"].value = "Repayment Value - Predicted"
            # PROFIT LOS EXCL VAT
            if row == 24:
                ws[f'{col_name}24'] = f'={col_name}55'
            ws["A24"].value = "Profit / Loss Excl VAT"
            # ADD INTEREST TO DATE
            if row == 33:
                ws[f'{col_name}33'] = f'={col_name}31+{col_name}32'
            ws["A33"].value = "Total Interest to date"
            # TOTAL INTEREST
            if row == 37:
                ws[f'{col_name}37'] = f'={col_name}35+{col_name}36'
            ws["A37"].value = "Total Interest"
            # TOTAL INTEREST
            if row == 38:
                ws[f'{col_name}38'] = f'={col_name}37+{col_name}26'
            ws["A38"].value = "Due to Investors"
            # TOTAL FEES
            if row == 41:
                ws[f'{col_name}41'] = f'={col_name}39+{col_name}40'
            ws["A41"].value = "Total Fees"
            # AVAILABLE FOR CONSTRUCTION
            if row == 42:
                ws[f'{col_name}42'] = f'={col_name}26-{col_name}41'
            ws["A42"].value = "Available for construction"
            # VAT CALC
            if row == 46:
                ws[f'{col_name}46'] = f'={col_name}45 / 1.15 * 0.15'
            ws["A46"].value = "VAT"
            # GROSS CALC
            if row == 47:
                ws[f'{col_name}47'] = f'={col_name}45-{col_name}46'
            ws["A47"].value = "GROSS"
            # TRANSFER INCOME
            if row == 53:
                ws[f'{col_name}53'] = f'={col_name}45-SUM({col_name}48:{col_name}52)'
            ws["A53"].value = "Transfer Income"
            # DUE TO INVESTORS
            if row == 54:
                ws[f'{col_name}54'] = f'=SUMIFS($G$38:${last_column}$38,$G$8:${last_column}$8,{col_name}8)'
            ws["A54"].value = "Due to Investors"
            # SURPLUS / DEFICIT
            if row == 55:
                ws[f'{col_name}55'] = f'={col_name}53-{col_name}54'
            ws["A55"].value = "Surplus / Deficit"

    # TOTALS FORMULAE
    # UNIT COUNT
    ws['B10'] = f'=COUNTA(G9:{last_column}9)'
    ws['C10'] = f'=COUNTIFS(G10:{last_column}10, "Yes")'
    ws['D10'] = f'=COUNTIFS(G11:{last_column}11, "Yes")-COUNTIFS(G10:{last_column}10, "Yes")'
    ws['E10'] = f'=COUNTIFS(G11:{last_column}11, "No")'
    ws['F10'] = f'=B10-C10-D10-E10'
    # TOTALS SUMS
    total_sum_rows = [17, 18, 21, 22, 23, 24, 26, 31, 32, 33, 35, 36, 37, 38, 40, 41, 42, 45, 46, 47, 48, 49, 50, 51,
                      52, 53, 54, 55]
    for row in total_sum_rows:
        ws[f'B{row}'] = f'=SUM(G{row}:{last_column}{row})'
        ws[f'C{row}'] = f'=SUMIFS(G{row}: {last_column}{row}, $G$6:${last_column}$6, {True})'
        ws[f'D{row}'] = f'=SUMIFS(G{row}: {last_column}{row}, $G$6:${last_column}$6, {False}, G7:{last_column}7,{True})'
        ws[f'E{row}'] = f'=SUMIFS(G{row}: {last_column}{row}, $G$7:${last_column}$7, {False})'
        ws[f'F{row}'] = f'=B{row}-C{row}-D{row}-E{row}'

    # FORMAT CELLS DEFINE PARAMETERS
    border1 = borders.Side(style=None, color=Color(indexed=0), border_style='medium')
    medium = Border(left=border1, right=border1, bottom=border1, top=border1)
    rows_to_format_from_g = [9, 10, 11, 13, 14, 15, 17, 18, 19, 20, 21, 22, 23, 24, 26, 27, 28, 29, 31,
                             32, 33, 35, 36, 37, 38, 40, 41, 42, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55]
    rows_to_format_from_b_to_f = [9, 10, 11, 17, 18, 21, 22, 23, 24, 26, 31, 32, 33, 35, 36, 37, 38, 40, 41,
                                  42, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55]

    # FORMAT CELLS IN ROWS FROM COLUMN B TO G
    for specific_summary_row in rows_to_format_from_b_to_f:
        for row in ws.iter_rows(min_row=specific_summary_row, min_col=2, max_row=specific_summary_row, max_col=6):
            for i, cell in enumerate(row):
                cell.fill = PatternFill("solid", start_color="000066CC")
                cell.border = medium
                cell.font = Font(color="00FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                if specific_summary_row == 10 or specific_summary_row == 11:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if specific_summary_row == 55 or specific_summary_row == 24:
                    cell.fill = PatternFill("solid", start_color="00FF6600")
                if specific_summary_row in [17, 18, 18, 21, 22, 23, 24, 26, 31, 32, 33, 35, 36, 37, 38, 40, 41, 42,
                                            45, 46, 47, 48, 48, 49, 50, 51, 52, 53, 54, 55]:
                    cell.number_format = '"R"#,##0.00_);[White]("R"#,##0.00)'
                if specific_summary_row == 33 or specific_summary_row == 38 \
                        or specific_summary_row == 42 or specific_summary_row == 47 or specific_summary_row == 53 \
                        or specific_summary_row == 54:
                    cell.fill = PatternFill("solid", start_color="00666699")

    # FORMAT CELLS IN ROWS FROM COLUMN G (THIS LOOP AND THE ONE ABOVE COULD BE REFACTORED INTO ONE LOOP - I THINK
    # THIS IS MORE READABLE THOUGH)
    for specific_row in rows_to_format_from_g:
        for row in ws.iter_rows(min_row=specific_row, min_col=7, max_row=specific_row, max_col=last_column_number):
            for i, cell in enumerate(row):
                cell.fill = PatternFill("solid", start_color="000066CC")
                cell.border = medium
                cell.font = Font(color="00FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                if specific_row in [17, 18, 18, 21, 22, 23, 24, 26, 31, 32, 33, 35, 36, 37, 38, 40, 41, 42,
                                    45, 46, 47, 48, 48, 49, 50, 51, 52, 53, 54, 55]:
                    cell.number_format = '"R"#,##0.00_);[White]("R"#,##0.00)'
                if specific_row == 15:
                    cell.number_format = '0%'
                    if cell.value == 0.18:
                        cell.fill = PatternFill("solid", start_color="00FF00FF")
                    else:
                        cell.fill = PatternFill("solid", start_color="00339966")
                if specific_row == 55 or specific_row == 24:
                    cell.fill = PatternFill("solid", start_color="00FF6600")
                if specific_row == 19 or specific_row == 20:
                    cell.number_format = '0%'
                if specific_row == 10:
                    if cell.value:
                        cell.value = "Yes"
                    else:
                        cell.value = "No"
                if specific_row == 11:
                    if cell.value:
                        cell.value = "Yes"
                    else:
                        cell.value = "No"
                if specific_row == 9:
                    test = [cell.value]
                    if specific_row == 10:
                        test.append(cell.value)
                if specific_row == 19 or specific_row == 20 or specific_row == 33 or specific_row == 38 \
                        or specific_row == 42 or specific_row == 47 or specific_row == 53 or specific_row == 54:
                    cell.fill = PatternFill("solid", start_color="00666699")

    # COLOR CELLS FROM COLUMN G - DEPENDING ON SOLD OR TRANSFERRED
    for x in range(7, last_column_number + 1):
        col = get_column_letter(x)
        if ws[f"{col}6"].value:
            ws[f"{col}9"].fill = PatternFill("solid", start_color="00FF0000")
            ws[f"{col}10"].fill = PatternFill("solid", start_color="00FF0000")
            ws[f"{col}11"].fill = PatternFill("solid", start_color="00FF0000")
        if not ws[f"{col}6"].value and ws[f"{col}7"].value:
            ws[f"{col}9"].fill = PatternFill("solid", start_color="0099CC00")
            ws[f"{col}10"].fill = PatternFill("solid", start_color="0099CC00")
            ws[f"{col}11"].fill = PatternFill("solid", start_color="0099CC00")

    # FILL TOTALS TRANSFERRED AND SOLD TO MATCH RELEVANT COLORS
    ws['C9'].fill = PatternFill("solid", start_color="00FF0000")
    ws['D9'].fill = PatternFill("solid", start_color="0099CC00")

    # PREPARE FOR CELL MERGING
    merge_cells_start = []
    merge_cells_end = []
    rows_for_merging = [9, 10, 11, 17, 18, 19, 20, 21, 22, 23, 24, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55]

    for row in ws.iter_rows(min_row=8, min_col=7, max_row=8, max_col=last_column_number):
        for i, cell in enumerate(row):
            if row[i].value != row[i - 1].value:
                end_cell = f"{get_column_letter(i + 6)}"
                start_cell = f"{get_column_letter(i + 7)}"
                merge_cells_start.append(start_cell)
                merge_cells_end.append(end_cell)
    merge_cells_end.append(f"{last_column}")
    merge_cells_end = merge_cells_end[1:]

    # MERGE CELLS - HORIZONTAL MERGE (UNIT DETAILS)
    for row_number in rows_for_merging:
        for i, start in enumerate(merge_cells_start):
            merge_range = f"{start}{row_number}:{merge_cells_end[i]}{row_number}"
            ws.merge_cells(merge_range)

    # VERTICAL MERGE - UNIT TOTALS
    for cell in range(2, 7):
        col_number = get_column_letter(cell)
        ws.merge_cells(f'{col_number}10:{col_number}11')

    # HIDE ROWS REQUIRED FOR FORMULAE BUT NOT FOR THE READER
    ws.row_dimensions[6].hidden = True
    ws.row_dimensions[7].hidden = True
    ws.row_dimensions[8].hidden = True

    # SAVE TO FILE
    wb.save(f"excel_files/{filename}.xlsx")
