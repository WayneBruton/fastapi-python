# import copy
# import time
# from datetime import datetime, timedelta
# from multiprocessing import Pool
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment, numbers
# from openpyxl.styles import Font, numbers

from excel_functions.create_sales_sheet import create_excel_array
from excel_functions.create_NSST_sheets import create_nsst_sheet


def create_sales_forecast_file(data, developmentinputdata, pledges):
    category = developmentinputdata['Category']
    worksheet_names = []
    if len(category) > 1:
        filename = f'Sales Forecast{category[0].split(" ")[0]}'
        sheet_name = f"SF {category[0]} & {category[1]}"

    else:
        sheet_name = f"SF {category[0]}"
        filename = f'Sales Forecast{category[0]}'
        worksheet_names.append(sheet_name)

    wb = Workbook()
    worksheet_data, interest_on_funds_drawn, interest_on_funds_in_momentum = create_excel_array(data)

    ws = wb.active
    ws.title = sheet_name
    ws.sheet_properties.tabColor = "1072BA"

    row1_data = [f"{sheet_name} - {developmentinputdata['date']}"]
    # insert row1_data into the beginning of worksheet_data
    worksheet_data.insert(0, row1_data)
    # loop through worksheet_data and append each item to the worksheet
    for item in worksheet_data:
        ws.append(item)

    if len(category) > 1:
        for index, item in enumerate(category):
            sheet_name = f"SF {item}"
            row1_data = [f"{sheet_name} - {developmentinputdata['date']}"]
            # filter data based on the category into new list using list comprehension
            data_input = [x for x in data if x['Category'] == item]
            worksheet_data, interest_on_funds_drawn, interest_on_funds_in_momentum = create_excel_array(data_input)

            # insert row1_data into the beginning of worksheet_data
            worksheet_data.insert(0, row1_data)
            # create a new worksheet
            ws = wb.create_sheet(sheet_name)
            ws.sheet_properties.tabColor = "1072BA"
            # loop through worksheet_data and append each item to the worksheet
            for item1 in worksheet_data:
                ws.append(item1)

    # LOOP THROUGH EACH SHEET AND FORMAT ACCORDINGLY
    for sheet in wb.worksheets:

        # make all column_widths in ws 20 wide
        for col in sheet.columns:
            col_width = 15
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = col_width

        # for rows 6 and 7, if the value in the cell from column 6 is FALSE, set the value to 'No' else set the value
        # to 'Yes'
        for index, row in enumerate(sheet.iter_rows(min_row=6, min_col=7, max_row=7, max_col=sheet.max_column)):
            for cell in row:
                if not cell.value:
                    sheet[f'{get_column_letter(cell.column)}{cell.row}'] = 'No'
                else:
                    sheet[f'{get_column_letter(cell.column)}{cell.row}'] = 'Yes'

        # In cell B6, count the values of all records from F6 to the last column in row 6
        sheet['B6'] = f'=COUNTIF(G7:{get_column_letter(sheet.max_column)}7,"*")'
        sheet['D6'] = f'=COUNTIF(G7:{get_column_letter(sheet.max_column)}7,"Yes")'
        sheet[
            'E6'] = f'=COUNTIF(G6:{get_column_letter(sheet.max_column)}6,"Yes")-' \
                    f'COUNTIF(G7:{get_column_letter(sheet.max_column)}7,"Yes") '
        # In Cell F6, Add the value of B6 less D6 less E6
        sheet['F6'] = f'=B6-D6-E6'

        # create a list of column letters from column 2 to the last column in the sheet
        column_letters_2 = [get_column_letter(x) for x in range(2, sheet.max_column + 1)]

        # splice this list into a new list starting only at column 7
        column_letters_7 = column_letters_2[5:]

        rows_to_add_formulas = [17, 23, 27, 29, 33, 34, 38, 39, 50, 51, 52, 54]

        for letter in column_letters_7:
            for row in rows_to_add_formulas:
                if row == 17:
                    sheet[f'{letter}17'] = f'=SUM({letter}13-{letter}14)'
                elif row == 23:
                    sheet[f'{letter}23'] = f'=SUM({letter}22-{letter}21)'
                elif row == 27:
                    sheet[f'{letter}27'] = f'=SUM({letter}24+{letter}25+{letter}26)'
                elif row == 29:
                    sheet[f'{letter}29'] = f'=720-{letter}23'
                elif row == 33:
                    sheet[f'{letter}33'] = f'=SUM({letter}30+{letter}31+{letter}32)'
                elif row == 34:
                    sheet[f'{letter}34'] = f'=SUM({letter}19+{letter}33)'
                elif row == 38:
                    sheet[f'{letter}38'] = f'=SUM({letter}36+{letter}37)'
                elif row == 39:
                    sheet[f'{letter}39'] = f'=SUM({letter}19-{letter}38)'
                elif row == 50:
                    sheet[f'{letter}50'] = f'={letter}41-SUM({letter}44:{letter}49)'
                elif row == 51:
                    sheet[f'{letter}51'] = f'=SUMIFS({column_letters_7[0]}34:' \
                                           f'{column_letters_7[len(column_letters_7) - 1]}34,{column_letters_7[0]}4:' \
                                           f'{column_letters_7[len(column_letters_7) - 1]}4, {letter}4)'
                elif row == 52:
                    sheet[f'{letter}52'] = f'={letter}50-{letter}51'
                    # =SUMIFS(G34: I34, G4: I4, G4)
                elif row == 54:
                    sheet[f'{letter}54'] = f'=SUM({letter}51/{letter}50)'

        # format all rows with data except rows 1 to 11, 20 to 23, 28 and 29 as currency with 2 decimal places and
        # comma every 3 digits, bold and white font, and for row 11 as a percentage with 2 decimal places and comma
        # every 3 digits
        for row in sheet.iter_rows(min_row=12, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                if cell.row == 11:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.number_format = '0.00%'
                if cell.row == 54:
                    cell.number_format = '0.00%'
                elif cell.row == 20 or cell.row == 21 or cell.row == 22 or cell.row == 23 or cell.row == 28 or cell.row == 29:
                    continue
                else:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.number_format = 'R#,##0.00'

        # create a list of rows to format as currency

        rows_to_center = [5, 6, 7, 9, 10, 11, 13, 14, 15, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31,
                          32, 33, 34, 36, 37, 38, 39, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 54]
        # Loop through the rows_to_format_currency list and align the cells from column 6 to the last column in the
        # centre
        for row in rows_to_center:
            # COLUMNS TO FILL ETC
            if row in [9, 10, 11, 20, 21, 22, 23, 28, 29]:
                col_number = 7
            else:
                col_number = 2
            # DARKER COLOR

            if row in [23, 13, 19, 27, 29, 33, 34, 38, 39, 43]:
                color_is = '3E54AC'
            elif row in [52]:
                color_is = 'E14D2A'
            else:
                color_is = '537FE7'
            for col in sheet.iter_cols(min_row=row, min_col=col_number, max_row=row, max_col=sheet.max_column):
                for cell in col:
                    if row == 11:
                        if cell.value == 0.18:
                            color_is = 'EA047E'
                        elif cell.value == 0.16:
                            color_is = 'ED50F1'
                        elif cell.value == 0.15:
                            color_is = 'FF4A4A'
                        elif cell.value == 0.14:
                            color_is = 'FF9E9E'
                        sheet[f'{get_column_letter(cell.column)}11'].number_format = '0.00%'
                        sheet[f'{get_column_letter(cell.column)}11'].alignment = Alignment(horizontal='center')

                    cell.fill = PatternFill(start_color=color_is, end_color=color_is, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.border = Border(left=Side(border_style='medium', color='000000'),
                                         right=Side(border_style='medium', color='000000'),
                                         top=Side(border_style='medium', color='000000'),
                                         bottom=Side(border_style='medium', color='000000'))

        rows_to_hide = [2, 3, 4, 53]
        # Loop through the rows_to_hide list and hide the rows
        for row in rows_to_hide:
            sheet.row_dimensions[row].hidden = True

        # Set column 1 to 25 units wide and make column 1 bold
        for col in sheet.iter_cols(min_row=1, min_col=1, max_row=ws.max_row, max_col=5):
            for cell in col:
                sheet.column_dimensions[cell.column_letter].width = 30
                cell.font = Font(bold=True)

        # ROWS TO SUM
        rows_to_sum = [13, 14, 15, 16, 17, 19, 24, 25, 26, 27, 30, 31, 32, 33, 34, 36, 37, 38, 39, 41, 42, 43, 44, 45,
                       46, 47, 48, 49, 50, 51, 52, 54]
        # for each row in rows_to_sum, sum the cells from column 6 to the last column in the sheet in insert the
        # formula in column 'B', with white font and bold and format the cell as currency with 2 decimal places and
        # comma every 3 digits

        sheet['B3'] = f'=COUNTIF({get_column_letter(7)}10:{get_column_letter(sheet.max_column)}10,"<>ZZUN01")'
        sheet['D3'] = f'=COUNTIF({get_column_letter(7)}3:{get_column_letter(sheet.max_column)}3,TRUE)'
        sheet['A54'] = f'LTV'

        columns_to_format = ['B', 'C', 'D', 'E', 'F']

        for row in rows_to_sum:
            sheet[f'B{row}'] = f'=SUM({get_column_letter(7)}{row}:{get_column_letter(sheet.max_column)}{row})'
            for column in columns_to_format:
                sheet[f'{column}{row}'].font = Font(bold=True, color='FFFFFF')
                sheet[f'{column}{row}'].number_format = 'R#,##0.00'
                sheet[f'{column}{row}'].alignment = Alignment(horizontal='center')

                if column == 'D':
                    if row == 51:
                        sheet[
                            f'{column}{row}'] = f'=+D34'
                    elif row == 52:
                        sheet[
                            f'{column}{row}'] = f'=+D50-D51'
                    elif row == 54:
                        sheet[
                            f'{column}{row}'] = f'=+D51/D50'
                        sheet[f'{column}{row}'].number_format = '0.00%'
                        sheet[f'{column}{row}'].font = Font(bold=True, color='000000')


                    else:
                        sheet[
                            f'{column}{row}'] = f'=SUMIFS($G{row}:${get_column_letter(sheet.max_column)}{row},' \
                                                f'$G3:${get_column_letter(sheet.max_column)}3,' \
                                                f'TRUE)'

                if column == 'E':
                    if row == 51:
                        sheet[
                            f'{column}{row}'] = f'=+E34'
                    elif row == 52:
                        sheet[
                            f'{column}{row}'] = f'=+E50-E51'
                    elif row == 54:
                        sheet[
                            f'{column}{row}'] = f'=+E51/E50'
                        sheet[f'{column}{row}'].number_format = '0.00%'
                        sheet[f'{column}{row}'].font = Font(bold=True, color='000000')
                    else:
                        sheet[
                            f'{column}{row}'] = f'=SUMIFS($G{row}:${get_column_letter(sheet.max_column)}{row},' \
                                                f'$G2:${get_column_letter(sheet.max_column)}2,' \
                                                f'TRUE)-SUMIFS($G{row}:${get_column_letter(sheet.max_column)}{row},' \
                                                f'$G3:${get_column_letter(sheet.max_column)}3,' \
                                                f'TRUE)'

                if column == 'F':
                    sheet[
                        f'{column}{row}'] = f'=B{row}-D{row}-E{row}'

        # Join rows_to_add_formulas,rows_to_format_currency,rows_to_center and rows_to_sum as one new list ordered and only unique values
        # all_rows_to_format = list(set(rows_to_add_formulas + rows_to_format_currency + rows_to_center + rows_to_sum))
        # all_rows_to_format.sort()

        # If the values in row 16 from column 2 are greater than 0, then the cells in row 16 from column 2 must be
        # filled with red and the font must be white
        sheet['B16'].fill = PatternFill(start_color='DF7857', end_color='DF7857', fill_type='solid')
        for col in sheet.iter_cols(min_row=16, min_col=7, max_row=16, max_col=sheet.max_column):
            for cell in col:
                if type(cell.value) == float:
                    if cell.value > 0:
                        for row in range(16, 17):
                            cell.fill = PatternFill(start_color='DF7857', end_color='DF7857', fill_type='solid')
                            cell.font = Font(color='FFFFFF')

        # from column 7 to the last column in the sheet, if the value in the cell in row 2 is not equal to  0,
        # then the corresponding cell in row 19 must have a red fill and the font white
        for col in sheet.iter_cols(min_row=53, min_col=7, max_row=53, max_col=sheet.max_column):
            for cell in col:
                if cell.value != 0:
                    for row in range(19, 20):
                        sheet[f'{get_column_letter(cell.column)}{row}'].fill = PatternFill(start_color='DF7857',
                                                                                           end_color='DF7857',
                                                                                           fill_type='solid')
                        sheet[f'{get_column_letter(cell.column)}{row}'].font = Font(color='FFFFFF')

        sheet['D5'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Make cell E5 fill with dark-green and the font white
        sheet['E5'].fill = PatternFill(start_color='539165', end_color='539165', fill_type='solid')

        # loop through cells in row 5, row 6 and row 7 from column 7. If the value in the cells in row 7 equals 'Yes'
        # then fill cells in row 5, 6 and 7 with red and the font white
        for col in sheet.iter_cols(min_row=7, min_col=7, max_row=7, max_col=sheet.max_column):
            for cell in col:
                # if cell.value == 'Yes' then make cells in row 5, 6 and 7 fill with red and the font white
                if cell.value == 'Yes':
                    for row in range(5, 8):
                        sheet[f'{cell.column_letter}{row}'].fill = PatternFill(start_color='FF0000',
                                                                               end_color='FF0000',
                                                                               fill_type='solid')
                        sheet[f'{cell.column_letter}{row}'].font = Font(color='FFFFFF')

        # loop through cells in row 6 and row 7 from column 7. If the value in the cells in row 7 equals 'No' and the
        # cells in row 6 equals 'Yes' then fill cells in row 6 and 7 with dark-green and the font white
        for col in sheet.iter_cols(min_row=6, min_col=7, max_row=6, max_col=sheet.max_column):
            for cell in col:
                # if cell.value == 'Yes' then make cells in row 5, 6 and 7 fill with red and the font white
                if cell.value == 'Yes':
                    for row in range(7, 8):
                        if sheet[f'{cell.column_letter}{row}'].value == 'No':
                            for row in range(5, 8):
                                sheet[f'{cell.column_letter}{row}'].fill = PatternFill(start_color='539165',
                                                                                       end_color='539165',
                                                                                       fill_type='solid')
                                sheet[f'{cell.column_letter}{row}'].font = Font(color='FFFFFF')

        # MERGE CELLS

        merge_master = []
        merge_start = []
        merge_end = []

        # CREATE MERGE LISTS Loop through all the cells in row 4 until the last column and insert the column number
        # and the cell value as a dict into the merge_start list
        for index, row in enumerate(sheet.iter_rows(min_row=4, min_col=6, max_row=4, max_col=sheet.max_column)):
            for cell in row:
                merge_master.append({'column': cell.column, 'value': cell.value})

        # Loop through merge_master, from index 1 to the end of the list, if the value of 'value' is different to the
        # value of the previous item in the list, add the column number to the merge_start list
        for index, item in enumerate(merge_master):
            if index > 0:
                if item['value'] != merge_master[index - 1]['value']:
                    merge_start.append(item['column'])

            # Loop through merge_master, from index 1 to the end of the list, if the value of 'value' is different to
            # the value of the next item in the list, add the column number to the merge_end list
            if 0 < index < len(merge_master) - 1:
                if item['value'] != merge_master[index + 1]['value']:
                    merge_end.append(item['column'])

        merge_end.append(merge_master[len(merge_master) - 1]['column'])

        # Create a dictionary to store the start and end columns for each row
        rows_to_merge = [5, 6, 7, 13, 14, 15, 16, 17, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 54]
        merge_dict = {}

        # Loop through the rows_to_merge list and populate merge_dict with start and end columns for each row
        for row in rows_to_merge:
            merge_dict[row] = []
            start_column = None
            end_column = None
            for cell in sheet[row]:
                if cell.column in merge_start:
                    start_column = cell.column
                elif cell.column in merge_end:
                    end_column = cell.column
                if start_column and end_column:
                    merge_dict[row].append((start_column, end_column))
                    start_column = None
                    end_column = None

        # Merge cells for each row in bulk using the merge_dict
        for row, merge_cols in merge_dict.items():
            for start, end in merge_cols:
                sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)

        #
        # rows_to_merge = [5, 6, 7, 13, 14, 15, 16, 17, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52]
        # # Loop through the rows_to_merge list and merge the cells in the merge_start and merge_end lists
        # for row in rows_to_merge:
        #     for index, item in enumerate(merge_start):
        #         sheet.merge_cells(start_row=row, start_column=item, end_row=row, end_column=merge_end[index])

        # Merge cells in row 6 and 7 in column B
        cells_to_merge = [6]
        columns_affected = ['B', 'C', 'D', 'E', 'F']
        for columns_affected in columns_affected:
            for row in cells_to_merge:
                sheet.merge_cells(f'{columns_affected}{row}:{columns_affected}{row + 1}')

        sheet.merge_cells('B6:B7')

        # Format 'B5' to 'F7' with white font and bold
        for row in sheet.iter_rows(min_row=5, min_col=2, max_row=7, max_col=6):
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFF')

        # Format 'B8' to 'F8' with white font and bold

    # CREATE NSST SHEET
    if len(category) > 1:
        sheet_name = f"NSST {category[0].split(' ')[0]}"

    else:
        sheet_name = f"NSST {category[0]}"

    worksheets = wb.sheetnames

    # if len(worksheets) == 1:
    index = 0

    ws = wb.create_sheet(sheet_name)
    # Make Tab Color of new Sheet 'Green'

    ws.sheet_properties.tabColor = "539165"

    # if len(worksheets) == 1:

    nsst_data = create_nsst_sheet(category, developmentinputdata, pledges, index, sheet_name, worksheets)

    for item in nsst_data:
        ws.append(item)

    if len(worksheets) == 3:
        # filter pledges to only include the category[0]
        new_pledges = [pledge for pledge in pledges if pledge['Category'] == category[0]
                       and pledge['opportunity_code'] != 'Unallocated']
        index = 1
        sheet_name = f"NSST {category[0]}"
        ws = wb.create_sheet(sheet_name)
        # Make Tab Color of new Sheet 'Green'
        ws.sheet_properties.tabColor = "539165"
        nsst_data = create_nsst_sheet(category, developmentinputdata, new_pledges, index, sheet_name, worksheets)

        for item in nsst_data:
            ws.append(item)

        if len(pledges):
            new_pledges = [pledge for pledge in pledges if pledge['Category'] == category[1]
                           and pledge['opportunity_code'] != 'Unallocated']
        else:
            new_pledges = []
        index = 2
        sheet_name = f"NSST {category[1]}"

        ws = wb.create_sheet(sheet_name)
        # Make Tab Color of new Sheet 'Green'
        ws.sheet_properties.tabColor = "539165"
        nsst_data = create_nsst_sheet(category, developmentinputdata, new_pledges, index, sheet_name, worksheets)

        for item in nsst_data:
            ws.append(item)

    num_sheets = len(wb.sheetnames)

    for index, sheet in enumerate(wb.worksheets):

        if num_sheets == 2:
            sheet_index = [1]
        else:
            sheet_index = [3, 4, 5]
        if index in sheet_index:

            gross_income_column_names = ['B', 'C', 'D', 'E']
            # row 43, add row 35 and deduct rows 38 to 42 for the column names in gross_income_column_names
            for column in gross_income_column_names:
                # sheet[f'B16'] = f'=0'
                # sheet[f'B17'] = f'=0'
                # sheet[f'B18'] = f'=0'
                sheet[f'{column}32'] = f'=SUM({column}23)-SUM({column}26:{column}31)'
                sheet[f'B37'] = f'=SUM(B36)-SUM(C36)'
                sheet[f'C37'] = f'=0'
                sheet[f'D37'] = f'=D36'
                sheet[f'E37'] = f'=E36'
                sheet[f'B43'] = f'=C43'
                sheet[f'B47'] = f'=SUM(B42)-SUM(C42)'
                sheet[f'C47'] = f''
                sheet[f'D47'] = f'=D42'
                sheet[f'E47'] = f'=E42-E46'
                sheet[f'{column}50'] = f'=SUM({column}32)-SUM({column}35)-SUM({column}42)'
                sheet[f'D52'] = f'=B52*0.05'
                sheet[f'E52'] = f'=B52-D52'
                sheet[f'{column}53'] = f'=SUM({column}38)+SUM({column}39)'
                sheet[f'{column}54'] = f'=SUM({column}50)+SUM({column}51)+SUM({column}53)-SUM({column}52)'


            if index == 3:
                sheet[f'B52'] = f"='NSST Heron Fields'!B52 + 'NSST Heron View'!B52"

            # make column 1 30 units wide and columns 2 to 5 15 units wide
            sheet.column_dimensions['A'].width = 55
            for column in gross_income_column_names:
                sheet.column_dimensions[column].width = 18

            rows_to_format_as_currency = [8, 9, 10, 11, 12, 13, 21, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 35, 36, 37,
                                          38, 39, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55]
            # Format the rows in rows_to_format_as_currency as currency from column B to F
            for row in rows_to_format_as_currency:
                for column in gross_income_column_names:
                    sheet[f'{column}{row}'].number_format = 'R#,##0.00'

            # Center all the text from column B to F
            for row in sheet.iter_rows(min_row=1, min_col=2, max_row=55, max_col=6):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # make all the cells with values as bold
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=55, max_col=6):
                for cell in row:
                    if cell.value is not None:
                        cell.font = Font(bold=True)
                        # put a thick border around the cells with values
                        cell.border = Border(top=Side(border_style='medium'),
                                             bottom=Side(border_style='medium'),
                                             left=Side(border_style='medium'),
                                             right=Side(border_style='medium'))
                        cell.alignment = Alignment(horizontal='center')

            # align the cells in column A to the left
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=55, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')

            rows_for_full_merge = [1, 2, 3, 6, 7, 14, 15, 19, 20, 33, 34, 40, 41, 48, 49]
            # merge the cells in rows_for_full_merge from column A to E
            for row in rows_for_full_merge:
                sheet.merge_cells(f'A{row}:E{row}')
                # fill the merged cells with a green color except for row
                sheet[f'A{row}'].fill = PatternFill(start_color='539165', end_color='539165', fill_type='solid')
                # make the text in the merged cells white
                sheet[f'A{row}'].font = Font(color='FFFFFF')
                # center the text in the merged cells
                sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')

            rows_for_partial_merge = [4, 5, 8, 9, 10, 11, 12, 13, 16, 17, 18]
            # merge the cells in rows_for_partial_merge from column B to E
            for row in rows_for_partial_merge:
                sheet.merge_cells(f'B{row}:E{row}')

            # make rows 2,3, 6,14, 19 and 31 only 5 unit high

            for row in [2, 3, 6, 14, 19, 33, 40, 48]:
                sheet.row_dimensions[row].height = 1

            # make all rows except rows 2,3, 6,14, 19 and 31 15 unit high
            for row in range(1, 56):
                if row not in [2, 3, 6, 14, 19, 33, 40, 48]:
                    sheet.row_dimensions[row].height = 20

            # rows to fill with a grey color from column A to E
            rows_to_fill_with_grey = [35, 42, 50]
            for row in rows_to_fill_with_grey:
                for column in ['A', 'B', 'C', 'D', 'E']:
                    sheet[f'{column}{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3',
                                                               fill_type='solid')

            # Make cell A1 bold, white and 30 units high and increase the font size to 20
            sheet['A1'].font = Font(bold=True, color='FFFFFF', size=20)
            sheet.row_dimensions[1].height = 30
            # Make cells A7, A15, A20, A34, A41 25 units high and increase the font size to 15
            for row in [7, 15, 20, 34, 41, 49]:
                sheet.row_dimensions[row].height = 25
                # sheet[f'A{row}'].font = Font(size=15)
                # make them bold and white
                sheet[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=18)

            # For all other rows make the font size 12
            for row in range(1, 56):
                if row not in [1, 7, 15, 20, 34, 41, 49]:
                    sheet[f'A{row}'].font = Font(size=12)

            # hide row 43
            sheet.row_dimensions[43].hidden = True

    # SAVE TO FILE
    wb.save(f"excel_files/{filename}.xlsx")

    return filename
