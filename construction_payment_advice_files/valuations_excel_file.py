# from datetime import timedelta
from datetime import datetime
import os

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter, column_index_from_string

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment, NumberFormatDescriptor
from openpyxl.utils import get_column_letter, column_index_from_string

# from openpyxl.utils import range_boundaries

from excel_sf_functions.create_sales_sheet import create_excel_array
from excel_sf_functions.create_NSST_sheets import create_nsst_sheet
from excel_sf_functions.format_sf_sheets import format_sales_forecast
from excel_sf_functions.format_nsst_sheets import format_nsst
import time
import threading


def create_valuations_file(data, subcontractors):
    # print(subcontractors)
    # if excel_files/valuations.xlsx exists then delete it

    # print(data[0]['block'])

    filename = f"excel_files/Master {data[0]['block']}.xlsx"
    print("filename",filename)
    if os.path.exists(filename):
        os.remove(filename)
        print("File Removed!")
    else:
        print("The file does not exist")

    wb = Workbook()

    for index, subcontractor in enumerate(subcontractors):
        # add worksheet for each subcontractor
        if index == 0:
            ws = wb.active
            ws.title = subcontractor[:31]
        else:
            ws = wb.create_sheet(subcontractor[:31])

        data_to_insert = []

        row1 = ['Development', 'Block', 'Unit', 'FloorType', 'Description', 'UnitMeasure', 'Qty', 'Rate', 'Amount',
                'Measure to date', 'valuation to date', 'Terms']
        data_to_insert.append(row1)
        color_terms = ""
        for index, valuation in enumerate(data):
            if index == 0:
                color_terms = valuation['terms']
            if valuation['subcontractor'] == subcontractor and valuation.get('taskCategory', 'Standard') != 'ATCV':
                insert = [valuation['development'], valuation['block'], valuation['unit'], valuation['floorType'],
                          valuation['description'], valuation['unitMeasure'], valuation['qty'], valuation['rate'],
                          valuation['amount'], valuation['percent_complete'], valuation['amount_complete'],
                          valuation['terms']]
                for item in valuation['tasks']:
                    insert.append(item['paymentAdviceNumber'])
                    insert.append(item['approved'])
                    current_progress = float(item['currentProgress'] - item['initialProgress'])
                    insert.append(current_progress)
                    current_valuation = float(valuation['amount']) * (current_progress / 100)
                    insert.append(current_valuation)
                    if valuation['retention'] is None:
                        insert.append(0)
                    else:
                        insert.append(valuation['retention'])
                    # insert.append("")

                # print(len(valuation['tasks']))

                data_to_insert.append(insert)
        sum_normal_row = len(data_to_insert) + 1

        # print("sum_normal_row", sum_normal_row)
        blank_row = []
        for num in range(3):
            data_to_insert.append(blank_row)
        data_to_insert.append(['', '', '', '', 'ATCV', '', '', '', '', '', '', ''])
        for valuation in data:
            if valuation['subcontractor'] == subcontractor and valuation.get('taskCategory', 'Standard') == 'ATCV':
                # data_to_insert.append(['', '', '', '', 'ATCV', '', '', '', '', '', '', ''])
                insert = ["", "", "", "",
                          valuation['description'], valuation['unitMeasure'], valuation['qty'], valuation['rate'],
                          valuation['amount'], valuation['percent_complete'], valuation['amount_complete'],
                          valuation['terms']]
                for index, item in enumerate(valuation['tasks']):
                    insert.append(item['paymentAdviceNumber'])
                    insert.append(item['approved'])
                    current_progress = float(item['currentProgress'] - item['initialProgress'])
                    insert.append(current_progress)
                    current_valuation = float(valuation['amount']) * (current_progress / 100)
                    insert.append(current_valuation)
                    insert.append(valuation['retention'])
                    # insert.append("")

                # print(len(valuation['tasks']))

                data_to_insert.append(insert)

        if len(data_to_insert) > sum_normal_row + 3:
            sum_atcv_row = len(data_to_insert) + 1
            start_atcv_row = sum_normal_row + 3
        else:
            sum_atcv_row = 0
            start_atcv_row = 0

        for num in range(2):
            data_to_insert.append(blank_row)

        data_to_insert.append([])

        data_to_insert.append(['', '', '', '', 'Total normal and ATCV', '', '', '', '', '', '', ''])
        total_normal_ATCV = len(data_to_insert)
        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'Retention', '', '', '', '', '', '', ''])
        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'Gross Amount Certified (Non Cumulative)', '', '', '', '', '', '', ''])

        gross_amount_certified_row = len(data_to_insert)

        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'Gross Amount Certified', '', '', '', '', '', '', ''])

        cumulative_gross_amount_certified_row = len(data_to_insert)

        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'Previous Amount Paid - Incl of ATCV', '', '', '', '', '', '', ''])

        previous_amount_paid = len(data_to_insert)

        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'Nett amount certified (VAT excl)', '', '', '', '', '', '', ''])

        nett_before_vat = len(data_to_insert)

        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'VAT', '', '', '', '', '', '', ''])

        vat_amount_row = len(data_to_insert)

        data_to_insert.append([])
        data_to_insert.append(['', '', '', '', 'Amount Due Inclusive of VAT', '', '', '', '', '', '', ''])

        amount_due_incl_vat = len(data_to_insert)

        for row in data_to_insert:
            ws.append(row)

        # column I is the amount column, sum from row2 to sum_normal_row - 1 in column I in row sum_normal_row
        ws['I' + str(sum_normal_row)] = '=SUM(I2:I' + str(sum_normal_row - 1) + ')'
        # the same for column K
        ws['K' + str(sum_normal_row)] = '=SUM(K2:K' + str(sum_normal_row - 1) + ')'
        # do above for column P
        ws['P' + str(sum_normal_row)] = '=SUM(P2:P' + str(sum_normal_row - 1) + ')'

        # get the last column number in the sheet
        last_column = ws.max_column
        last_column = 100

        # col I, K and P for row sum_normal_row, make it bold
        for col in range(9, last_column, 5):
            ws.cell(row=sum_normal_row, column=col).font = Font(bold=True)
            # do the same for column K
        for col in range(11, last_column, 5):
            ws.cell(row=sum_normal_row, column=col).font = Font(bold=True)
            # do the same for column P
        for col in range(16, last_column, 5):
            ws.cell(row=sum_normal_row, column=col).font = Font(bold=True)

        # col E for row sum_normal_row, make it bold
        if start_atcv_row > 0:
            for col in range(5, last_column, 5):
                ws.cell(row=start_atcv_row, column=col).font = Font(bold=True)

            # do the same for column H, K & P
            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV - 2, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV - 2, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV - 2, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 6, column=col).font = Font(bold=True)

            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 6, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 8, column=col).font = Font(bold=True)

            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 8, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 10, column=col).font = Font(bold=True)

            for col in range(16, 17):
                ws.cell(row=total_normal_ATCV + 10, column=col).font = Font(bold=True)
                # put a top thin border on row total_normal_ATCV + 10 and a double bottom border
                ws.cell(row=total_normal_ATCV + 10, column=col).border = Border(top=Side(style='thin'),
                                                                                bottom=Side(style='double'))

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 12, column=col).font = Font(bold=True)

            for col in range(16, 17):
                ws.cell(row=total_normal_ATCV + 12, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 14, column=col).font = Font(bold=True)

            for col in range(16, 17):
                ws.cell(row=total_normal_ATCV + 14, column=col).font = Font(bold=True)
                # put a top thin border on row total_normal_ATCV + 10 and a double bottom border
                ws.cell(row=total_normal_ATCV + 14, column=col).border = Border(top=Side(style='thin'),
                                                                                bottom=Side(style='double'))
        else:
            for col in range(5, last_column, 5):
                ws.cell(row=sum_normal_row + 3, column=col).font = Font(bold=True)

            # do the same for column H, K & P
            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV - 2, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV - 2, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV - 2, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 2, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)

            for col in range(9, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)
                # do the same for column K
            for col in range(11, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)
                # do the same for column P
            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 4, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 6, column=col).font = Font(bold=True)

            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 6, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 8, column=col).font = Font(bold=True)

            for col in range(16, last_column, 5):
                ws.cell(row=total_normal_ATCV + 8, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 10, column=col).font = Font(bold=True)

            for col in range(16, 17):
                ws.cell(row=total_normal_ATCV + 10, column=col).font = Font(bold=True)
                # put a top thin border on row total_normal_ATCV + 10 and a double bottom border
                ws.cell(row=total_normal_ATCV + 10, column=col).border = Border(top=Side(style='thin'),
                                                                                bottom=Side(style='double'))

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 12, column=col).font = Font(bold=True)

            for col in range(16, 17):
                ws.cell(row=total_normal_ATCV + 12, column=col).font = Font(bold=True)

            for col in range(5, last_column, 5):
                ws.cell(row=total_normal_ATCV + 14, column=col).font = Font(bold=True)

            for col in range(16, 17):
                ws.cell(row=total_normal_ATCV + 14, column=col).font = Font(bold=True)
                # put a top thin border on row total_normal_ATCV + 10 and a double bottom border
                ws.cell(row=total_normal_ATCV + 14, column=col).border = Border(top=Side(style='thin'),
                                                                                bottom=Side(style='double'))

        for col in range(15, last_column, 5):

            if ws[f'{get_column_letter(col - 2)}2'].value is not None:
                ws[f'{get_column_letter(col)}1'] = f'={get_column_letter(col - 2)}2'

        for col in range(13, last_column, 5):
            ws.column_dimensions[get_column_letter(col)].hidden = True
        # do the same for column N
        for col in range(14, last_column, 5):
            ws.column_dimensions[get_column_letter(col)].hidden = True
            # do the same for column Q
        for col in range(17, last_column, 5):
            ws.column_dimensions[get_column_letter(col)].hidden = True
        # do the same for column V
        for col in range(22, last_column, 5):
            ws.column_dimensions[get_column_letter(col)].hidden = True

        if sum_atcv_row > 0:
            ws['I' + str(sum_atcv_row + 1)] = '=SUM(I' + str(start_atcv_row) + ':I' + str(sum_atcv_row - 1) + ')'
            ws['K' + str(sum_atcv_row + 1)] = '=SUM(K' + str(start_atcv_row) + ':K' + str(sum_atcv_row - 1) + ')'
            ws['P' + str(sum_atcv_row + 1)] = '=SUM(P' + str(start_atcv_row) + ':P' + str(sum_atcv_row - 1) + ')'

            ws['I' + str(total_normal_ATCV)] = '=I' + str(sum_normal_row) + '+I' + str(sum_atcv_row + 1)
            ws['K' + str(total_normal_ATCV)] = '=K' + str(sum_normal_row) + '+K' + str(sum_atcv_row + 1)
            ws['P' + str(total_normal_ATCV)] = '=P' + str(sum_normal_row) + '+P' + str(sum_atcv_row + 1)

            ws['I' + str(total_normal_ATCV + 2)] = ('=(I' + str(sum_normal_row) + ')*(Q'
                                                    + str(sum_normal_row - 1) + '/100)')
            ws['K' + str(total_normal_ATCV + 2)] = ('=(K' + str(sum_normal_row) + ')*(Q'
                                                    + str(sum_normal_row - 1) + '/100)')
            # do the same for P
            ws['P' + str(total_normal_ATCV + 2)] = ('=(P' + str(sum_normal_row) + ')*(Q'
                                                    + str(sum_normal_row - 1) + '/100)')

            ws['I' + str(gross_amount_certified_row)] = '=I' + str(total_normal_ATCV) + '-I' + str(
                total_normal_ATCV + 2)

            ws['K' + str(gross_amount_certified_row)] = '=K' + str(total_normal_ATCV) + '-K' + str(
                total_normal_ATCV + 2)

            ws['P' + str(gross_amount_certified_row)] = '=P' + str(total_normal_ATCV) + '-P' + str(
                total_normal_ATCV + 2)

            ws['P' + str(cumulative_gross_amount_certified_row)] = ('=P' + str(gross_amount_certified_row) + '+K'
                                                                    + str(gross_amount_certified_row))

            ws['P' + str(previous_amount_paid)] = '=K' + str(gross_amount_certified_row)

            ws['P' + str(nett_before_vat)] = ('=P' + str(cumulative_gross_amount_certified_row) + '-P'
                                              + str(previous_amount_paid))

            if valuation['vatable'] == 'Yes':
                vat_rate = 0.15
            else:
                vat_rate = 0

            ws['P' + str(vat_amount_row)] = ('=P' + str(nett_before_vat) + '*' + str(vat_rate))

            ws['P' + str(amount_due_incl_vat)] = ('=P' + str(nett_before_vat) + '+P' + str(vat_amount_row))

        else:
            ws['I' + str(total_normal_ATCV)] = '=I' + str(sum_normal_row)
            ws['K' + str(total_normal_ATCV)] = '=K' + str(sum_normal_row)
            ws['P' + str(total_normal_ATCV)] = '=P' + str(sum_normal_row)
            ws['I' + str(total_normal_ATCV + 2)] = ('=(I' + str(sum_normal_row) + ')*(Q'
                                                    + str(sum_normal_row - 1) + '/100)')
            ws['K' + str(total_normal_ATCV + 2)] = ('=(K' + str(sum_normal_row) + ')*(Q'
                                                    + str(sum_normal_row - 1) + '/100)')
            # do the same for P
            ws['P' + str(total_normal_ATCV + 2)] = ('=(P' + str(sum_normal_row) + ')*(Q'
                                                    + str(sum_normal_row - 1) + '/100)')

            ws['I' + str(gross_amount_certified_row)] = '=I' + str(total_normal_ATCV) + '-I' + str(
                total_normal_ATCV + 2)

            ws['K' + str(gross_amount_certified_row)] = '=K' + str(total_normal_ATCV) + '-K' + str(
                total_normal_ATCV + 2)

            ws['P' + str(gross_amount_certified_row)] = '=P' + str(total_normal_ATCV) + '-P' + str(
                total_normal_ATCV + 2)

            ws['P' + str(cumulative_gross_amount_certified_row)] = '=P' + str(
                gross_amount_certified_row) + '+K' + str(gross_amount_certified_row)

            ws['P' + str(previous_amount_paid)] = '=K' + str(gross_amount_certified_row)

            ws['P' + str(nett_before_vat)] = ('=P' + str(cumulative_gross_amount_certified_row) + '-P'
                                              + str(previous_amount_paid))

            if valuation['vatable'] == 'Yes':
                vat_rate = 0.15
            else:
                vat_rate = 0

            ws['P' + str(vat_amount_row)] = ('=P' + str(nett_before_vat) + '*' + str(vat_rate))

            ws['P' + str(amount_due_incl_vat)] = ('=P' + str(nett_before_vat) + '+P' + str(vat_amount_row))

        # ws.sheet_properties.tabColor = "1072BA"
            print(color_terms)
            if color_terms == '30 Days':
                # tab color = blue else tab color = green
                ws.sheet_properties.tabColor = "4F709C"
            else:
                ws.sheet_properties.tabColor = "00DFA2"






        columns_to_format_as_currency = ['H', 'I', 'K', 'P']
        # get last row
        last_row = ws.max_row
        # from row 2 to last row format columns H, I , K and P to be currency with 2 decimal places and comma
        # separator for thousands and millions and billions and have simbol R
        for col in columns_to_format_as_currency:
            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=column_index_from_string(col),
                                    max_col=column_index_from_string(col)):
                for cell in row:
                    cell.number_format = 'R# ###.00'

        # make all columns have a width of 15 except columns D & E which should have a width of 20
        for col in range(1, ws.max_column + 1):
            if col == 4 or col == 5:
                ws.column_dimensions[get_column_letter(col)].width = 28
            else:
                ws.column_dimensions[get_column_letter(col)].width = 12

        # make row 1 bold
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

    ## 30 DAY SHEET
    thirty_days_subcontractors = [item for item in data if item['terms'] == '30 Days']
    if len(thirty_days_subcontractors) > 0:
        ws = wb.create_sheet("Master(30 Days)", 0)
        ws.sheet_properties.tabColor = "FF0000"
        data_to_insert = []
        row1 = [f"{data[0]['block']} 30 Day Subcontractors"]
        data_to_insert.append(row1)
        row_2 = ["", "Subcontractor", "Trade", "Contract Value", "ATCV", "New Contract Value", "Previous Amount Certified",
                 "Outstanding on Contract Value", "Upcoming Payment", "Settlement Discount(%)", "Settlement Discount",
                 "Final Payment", "Balance Due after Payment", "% Outstanding"]
        data_to_insert.append(row_2)

        for row in data_to_insert:
            ws.append(row)

        # make row 1 bold
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).font = Font(bold=True)

        # make all columns have a width of 18
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        for col in range(2, 3):
            ws.column_dimensions[get_column_letter(col)].width = 35

        for col in range(1, 2):
            ws.column_dimensions[get_column_letter(col)].width = 7

        # Make row 2 center aligned and wrap text
        for col in range(1, ws.max_column + 1):
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # create a valiable list called 30_days_subcontractors and filter the data list to only include
        # dictionaries  with a payment term of 30 days


        thirst_day_subbies = []
        for index, contract in enumerate(thirty_days_subcontractors, 1):
            contract['taskCategory'] = contract.get('taskCategory', 'Standard')
            thirst_day_subbies.append(contract['subcontractor'])

        thirst_day_subbies = list(dict.fromkeys(thirst_day_subbies))

        for index, subbie in enumerate(thirst_day_subbies, 1):
            insert = [index, subbie]
            # filter the list thirty_days_subcontractors to only include dictionaries with the subcontractor name
            # equal to the current subcontractor in the loop
            subbie_contracts = [item for item in thirty_days_subcontractors if item['subcontractor'] == subbie]
            # print(subbie_contracts)
            insert.append(subbie_contracts[0].get('works','Nothing Shown'))
            # sum amount column in subbie_contracts where taskCategory is Standard
            sum_standard = 0
            sum_atcv = 0
            amount_previously_certified = 0
            upcoming_payment = 0
            for contract in subbie_contracts:
                if contract['retention'] is None:
                    contract['retention'] = 0
                unapproved_tasks = [item for item in contract['tasks'] if item['approved'] == False]
                if len(unapproved_tasks) > 0:
                    if contract['vatable'] == 'Yes':
                        # print("Retention", contract['retention'])
                        upcoming_payment += ((float(contract['amount']) * ((unapproved_tasks[0]['currentProgress'] -
                                                                            unapproved_tasks[0][
                                                                                'initialProgress']) / 100))) * (
                                                        (100 - contract['retention']) / 100) * 1.15
                    else:
                        upcoming_payment += (float(contract['amount']) * ((unapproved_tasks[0]['currentProgress'] -
                                                                           unapproved_tasks[0][
                                                                               'initialProgress']) / 100)) * (
                                                        (100 - contract['retention']) / 100)

                if contract['vatable'] == 'Yes':
                    amount_previously_certified += float(contract['amount_complete']) * 1.15
                else:
                    amount_previously_certified += float(contract['amount_complete'])

                if contract['taskCategory'] == 'Standard':
                    if contract['vatable'] == 'Yes':
                        sum_standard += float(contract['amount']) * 1.15
                    else:
                        sum_standard += float(contract['amount'])
                else:
                    if contract['vatable'] == 'Yes':
                        sum_atcv += float(contract['amount']) * 1.15
                    else:
                        sum_atcv += float(contract['amount'])
            new_contract_value = sum_standard + sum_atcv
            outstanding_on_contract_value = new_contract_value - amount_previously_certified
            insert.append(sum_standard)
            insert.append(sum_atcv)
            insert.append(new_contract_value)
            insert.append(amount_previously_certified)
            insert.append(outstanding_on_contract_value)
            insert.append(upcoming_payment)
            insert.append(0)
            ws.append(insert)

        # get last row
        last_row = ws.max_row
        columns_to_format_as_currency = ['D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M']
        # from row 3 to last_row format columns in columns_to_format_as_currency to be currency with 2 decimal places and
        # comma separator for thousands and millions and billions and have symbol R
        for col in columns_to_format_as_currency:
            for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=column_index_from_string(col),
                                    max_col=column_index_from_string(col)):
                for cell in row:
                    cell.number_format = 'R# ###.00'

        columns_to_format_as_percentage = ['J', 'N']
        # from row 3 to last_row format columns in columns_to_format_as_percentage to be percentage with 2 decimal places
        for col in columns_to_format_as_percentage:
            for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=column_index_from_string(col),
                                    max_col=column_index_from_string(col)):
                for cell in row:
                    cell.number_format = '0.00%'

        # in column F from row 3 to last_row apply the following formula
        # =D3 + E3
        for row in range(3, last_row + 1):
            ws['F' + str(row)] = '=D' + str(row) + '+E' + str(row)
        # in column H from row 3 to last_row apply the following formula
        # =F3 - G3
        for row in range(3, last_row + 1):
            ws['H' + str(row)] = '=F' + str(row) + '-G' + str(row)
            # in column K from row 3 to last_row apply the following formula
            # =I3 * J3
        for row in range(3, last_row + 1):
            ws['K' + str(row)] = '=I' + str(row) + '*J' + str(row)
            # in column L from row 3 to last_row apply the following formula
            # =I3-K3
        for row in range(3, last_row + 1):
            ws['L' + str(row)] = '=I' + str(row) + '-K' + str(row)
            # in column M from row 3 to last_row apply the following formula
            # =H3-I3
        for row in range(3, last_row + 1):
            ws['M' + str(row)] = '=H' + str(row) + '-I' + str(row)
            # in column N from row 3 to last_row apply the following formula
            # =M3/H3
        for row in range(3, last_row + 1):
            ws['N' + str(row)] = '=M' + str(row) + '/H' + str(row)

        # merge cells in row 1 from column A to column N
        ws.merge_cells('A1:N1')
        # above merged cells alighn center and make font size 16
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(size=16, bold=True)
        # fill in light grey the merged cells
        ws['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        # make row 1 bold


        # put a thin border around all cells from row 1 to last_row and from column A to column N
        for row in range(1, last_row + 1):
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                             left=Side(style='thin'), right=Side(style='thin'))
        columns_to_fill = ['I', 'J', 'K', 'L']
        # fill columns in columns_to_fill with light green from row 3 to last_row
        for col in columns_to_fill:
            for row in range(3, last_row + 1):
                ws[col + str(row)].fill = PatternFill(start_color='D4E2D4', end_color='D4E2D4', fill_type='solid')

        columns_to_sum = ['I', 'K', 'L']
        # sum formula for columns in columns_to_sum add a row at the end of the sheet and apply the sum formula
        for col in columns_to_sum:
            ws[col + str(last_row + 1)] = '=SUM(' + col + '3:' + col + str(last_row) + ')'
            # make them bold, fill light grey and apply a top thin border
            ws[col + str(last_row + 1)].font = Font(bold=True)
            ws[col + str(last_row + 1)].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws[col + str(last_row + 1)].border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                        left=Side(style='thin'), right=Side(style='thin'))
            # apply currency format to the sum with thousands separator and 2 decimal places and R symbol
            ws[col + str(last_row + 1)].number_format = 'R# ###.00'


    ## EOM SHEET
    eom_subcontractors = [item for item in data if item['terms'] == 'End of Month']
    if len(eom_subcontractors) > 0:
        ws = wb.create_sheet("Master (Month End)", 0)
        ws.sheet_properties.tabColor = "FF0000"
        data_to_insert = []
        row1 = [f"{data[0]['block']} Month End Subcontractors"]
        data_to_insert.append(row1)
        row_2 = ["", "Subcontractor", "Trade", "Contract Value", "ATCV", "New Contract Value",
                 "Previous Amount Certified",
                 "Outstanding on Contract Value", "Upcoming Payment", "Settlement Discount(%)", "Settlement Discount",
                 "Final Payment", "Balance Due after Payment", "% Outstanding"]
        data_to_insert.append(row_2)

        for row in data_to_insert:
            ws.append(row)

        # make row 1 bold
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).font = Font(bold=True)

        # make all columns have a width of 18
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        for col in range(2, 3):
            ws.column_dimensions[get_column_letter(col)].width = 35

        for col in range(1, 2):
            ws.column_dimensions[get_column_letter(col)].width = 7

        # Make row 2 center aligned and wrap text
        for col in range(1, ws.max_column + 1):
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # create a valiable list called 30_days_subcontractors and filter the data list to only include
        # dictionaries  with a payment term of 30 days


        eom_subbies = []
        for index, contract in enumerate(eom_subcontractors, 1):
            contract['taskCategory'] = contract.get('taskCategory', 'Standard')
            eom_subbies.append(contract['subcontractor'])

        eom_subbies = list(dict.fromkeys(eom_subbies))

        for index, subbie in enumerate(eom_subbies, 1):
            insert = [index, subbie]
            # filter the list thirty_days_subcontractors to only include dictionaries with the subcontractor name
            # equal to the current subcontractor in the loop
            subbie_contracts = [item for item in eom_subcontractors if item['subcontractor'] == subbie]
            insert.append(subbie_contracts[0]['works'])
            # sum amount column in subbie_contracts where taskCategory is Standard
            sum_standard = 0
            sum_atcv = 0
            amount_previously_certified = 0
            upcoming_payment = 0
            for contract in subbie_contracts:
                if contract['retention'] is None:
                    contract['retention'] = 0
                unapproved_tasks = [item for item in contract['tasks'] if item['approved'] == False]
                if len(unapproved_tasks) > 0:
                    if contract['vatable'] == 'Yes':
                        # print("Retention", contract['retention'])
                        upcoming_payment += ((float(contract['amount']) * ((unapproved_tasks[0]['currentProgress'] -
                                                                            unapproved_tasks[0][
                                                                                'initialProgress']) / 100))) * (
                                                    (100 - contract['retention']) / 100) * 1.15
                    else:
                        upcoming_payment += (float(contract['amount']) * ((unapproved_tasks[0]['currentProgress'] -
                                                                           unapproved_tasks[0][
                                                                               'initialProgress']) / 100)) * (
                                                    (100 - contract['retention']) / 100)

                if contract['vatable'] == 'Yes':
                    amount_previously_certified += float(contract['amount_complete']) * 1.15
                else:
                    amount_previously_certified += float(contract['amount_complete'])

                if contract['taskCategory'] == 'Standard':
                    if contract['vatable'] == 'Yes':
                        sum_standard += float(contract['amount']) * 1.15
                    else:
                        sum_standard += float(contract['amount'])
                else:
                    if contract['vatable'] == 'Yes':
                        sum_atcv += float(contract['amount']) * 1.15
                    else:
                        sum_atcv += float(contract['amount'])
            new_contract_value = sum_standard + sum_atcv
            outstanding_on_contract_value = new_contract_value - amount_previously_certified
            insert.append(sum_standard)
            insert.append(sum_atcv)
            insert.append(new_contract_value)
            insert.append(amount_previously_certified)
            insert.append(outstanding_on_contract_value)
            insert.append(upcoming_payment)
            insert.append(0)
            ws.append(insert)

        # get last row
        last_row = ws.max_row
        columns_to_format_as_currency = ['D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M']
        # from row 3 to last_row format columns in columns_to_format_as_currency to be currency with 2 decimal places and
        # comma separator for thousands and millions and billions and have symbol R
        for col in columns_to_format_as_currency:
            for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=column_index_from_string(col),
                                    max_col=column_index_from_string(col)):
                for cell in row:
                    cell.number_format = 'R# ###.00'

        columns_to_format_as_percentage = ['J', 'N']
        # from row 3 to last_row format columns in columns_to_format_as_percentage to be percentage with 2 decimal places
        for col in columns_to_format_as_percentage:
            for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=column_index_from_string(col),
                                    max_col=column_index_from_string(col)):
                for cell in row:
                    cell.number_format = '0.00%'

        # in column F from row 3 to last_row apply the following formula
        # =D3 + E3
        for row in range(3, last_row + 1):
            ws['F' + str(row)] = '=D' + str(row) + '+E' + str(row)
        # in column H from row 3 to last_row apply the following formula
        # =F3 - G3
        for row in range(3, last_row + 1):
            ws['H' + str(row)] = '=F' + str(row) + '-G' + str(row)
            # in column K from row 3 to last_row apply the following formula
            # =I3 * J3
        for row in range(3, last_row + 1):
            ws['K' + str(row)] = '=I' + str(row) + '*J' + str(row)
            # in column L from row 3 to last_row apply the following formula
            # =I3-K3
        for row in range(3, last_row + 1):
            ws['L' + str(row)] = '=I' + str(row) + '-K' + str(row)
            # in column M from row 3 to last_row apply the following formula
            # =H3-I3
        for row in range(3, last_row + 1):
            ws['M' + str(row)] = '=H' + str(row) + '-I' + str(row)
            # in column N from row 3 to last_row apply the following formula
            # =M3/H3
        for row in range(3, last_row + 1):
            ws['N' + str(row)] = '=M' + str(row) + '/H' + str(row)

        # merge cells in row 1 from column A to column N
        ws.merge_cells('A1:N1')
        # above merged cells alighn center and make font size 16
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(size=16, bold=True)
        # fill in light grey the merged cells
        ws['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        # make row 1 bold

        # put a thin border around all cells from row 1 to last_row and from column A to column N
        for row in range(1, last_row + 1):
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                             left=Side(style='thin'), right=Side(style='thin'))
        columns_to_fill = ['I', 'J', 'K', 'L']
        # fill columns in columns_to_fill with light green from row 3 to last_row
        for col in columns_to_fill:
            for row in range(3, last_row + 1):
                ws[col + str(row)].fill = PatternFill(start_color='D4E2D4', end_color='D4E2D4', fill_type='solid')

        columns_to_sum = ['I', 'K', 'L']
        # sum formula for columns in columns_to_sum add a row at the end of the sheet and apply the sum formula
        for col in columns_to_sum:
            ws[col + str(last_row + 1)] = '=SUM(' + col + '3:' + col + str(last_row) + ')'
            # make them bold, fill light grey and apply a top thin border
            ws[col + str(last_row + 1)].font = Font(bold=True)
            ws[col + str(last_row + 1)].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws[col + str(last_row + 1)].border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                        left=Side(style='thin'), right=Side(style='thin'))
            # apply currency format to the sum with thousands separator and 2 decimal places and R symbol
            ws[col + str(last_row + 1)].number_format = 'R# ###.00'

    # wb.save('excel_files/valuations.xlsx')
    wb.save(filename)
    return filename
