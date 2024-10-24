import os
from openpyxl import Workbook, formatting, formula
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.utils import get_column_letter

import re
from datetime import datetime
import time


def calculate_vat_due(sale_date):
    vat_periods = {
        1: "03/31", 3: "05/31", 5: "07/31",
        7: "09/30", 9: "11/30", 11: "01/31",
    }

    sale_date = datetime.strptime(sale_date.replace("-", "/"), '%Y/%m/%d')
    sale_month = sale_date.month
    sale_year = sale_date.year

    # Adjust for months not directly mapped in vat_periods
    adjusted_month = sale_month if sale_month % 2 != 0 else sale_month - 1

    vat_year = sale_year + 1 if sale_month > 10 else sale_year
    vat_date = f"{vat_year}/{vat_periods[adjusted_month]}"
    return vat_date


# res1 = calculate_vat_due("2022-02-01")
# print(res1)

def format_header(ws, header_data):
    header = [key.replace("_", " ").title() for key in header_data]
    ws.append([])
    ws.append([])
    ws.append(header)

    for cell in ws[4]:
        cell.font = Font(bold=True)


def format_columns(ws, column_widths, number_format_ranges):
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = 30
    for col_range, number_format in number_format_ranges.items():
        for row in ws.iter_rows(min_row=5, **col_range):
            for cell in row:
                cell.number_format = number_format


def format_cells(ws, last_row):
    for row in ws.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def create_sheet(wb, title, data, column_widths, number_format_ranges):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = "1072BA"
    ws['A1'] = title
    format_header(ws, data[0])
    for item in data:
        ws.append([item[key] for key in item])
    format_columns(ws, column_widths, number_format_ranges)
    last_row = ws.max_row
    format_cells(ws, last_row)
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{last_row}"
    ws.freeze_panes = "A5"

    return ws


def cashflow_projections(invest, construction, sales, operational_costs, xero, opportunities, investor_exit, momentum,
                         report_date):
    # cashflow_p&l_files/cashflow_projection.xlsx exists, then delete it
    # if os.path.exists("cashflow_p&l_files/cashflow_projection.xlsx"):
    #     os.remove("cashflow_p&l_files/cashflow_projection.xlsx")
    start_time = time.time()
    for sale in sales:
        # print("Sale:::",sale)
        sale['refanced_value'] = float(sale['profit_loss'])

    try:
        report_date = datetime.strptime(report_date, '%Y-%m-%d')

        wb = Workbook()
        ws = create_sheet(wb, 'Investors', invest, {'B': 40}, {})

        last_row = ws.max_row
        columns = ['M', 'Q', 'R', 'S', 'T']

        for col in columns:
            ws[f"{col}2"] = f"=SUBTOTAL(9,{col}5:{col}{last_row})"
            ws[f"{col}2"].number_format = '#,##0.00'

        ws1 = wb.create_sheet('Opportunities')
        ws1.sheet_properties.tabColor = "1072BA"
        ws1['A1'] = "Opportunities"
        ws1.append([])
        ws1.append([])
        ws1.append([])
        headers = [item.replace("_", " ").title() for item in opportunities[0]]
        ws1.append(headers)

        [ws1.append([item[key] for key in item]) for item in opportunities]
        ws1.auto_filter.ref = f"A5:{get_column_letter(ws1.max_column)}{ws1.max_row}"
        ws1.freeze_panes = "A6"

        for cell in ['E3', 'H3']:
            ws1[cell] = f"=SUBTOTAL(9,{cell[0]}5:{cell[0]}{ws1.max_row})"
            ws1[cell].number_format = '#,##0.00'

        ws2 = wb.create_sheet('Construction')

        ws2.sheet_properties.tabColor = "1072BA"

        ws2['A1'] = "Construction"
        ws2['A2'] = ""
        ws2['A3'] = ""

        headers = []

        for item in construction:

            if item['Blocks'][:6] == "Block ":
                item['Renamed_block'] = item['Blocks'][6]
            else:
                item['Renamed_block'] = item['Blocks']

        for key in construction[0]:
            headers.append(key.replace("_", " ").title())
        ws2.append(headers)

        for item in construction:

            row = []
            for key in item:
                row.append(item[key])

            ws2.append(row)

        ws2.auto_filter.ref = f"A4:{get_column_letter(ws2.max_column)}{ws2.max_row}"
        ws2.freeze_panes = "A5"

        subtotal_columns = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']

        subtotal_columns = [get_column_letter(col) for col in range(9, ws2.max_column + 1)]

        for col in range(9, ws2.max_column + 1):
            ws2[f"{get_column_letter(col)}5"].number_format = '#,##0.00'

        for col in subtotal_columns:
            ws2[f"{col}2"] = f"=SUBTOTAL(9,{col}5:{col}{ws2.max_row})"
            ws2[f"{col}2"].number_format = '#,##0.00'

        for col in ws2.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws2.column_dimensions[column].width = adjusted_width

        # loop through all cells and if the cell in column A is True, then make the row green
        for row in ws2.iter_rows(min_row=5, max_row=ws2.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == True:
                    for inner_row in ws2.iter_rows(min_row=row[0].row, max_row=row[0].row, min_col=1,
                                                   max_col=ws2.max_column):
                        for cells in inner_row:
                            cells.fill = PatternFill(start_color="4CCD99", end_color="4CCD99", fill_type="solid")

        original_list = []
        converted_dicts = []

        for index, item in enumerate(construction):
            original_dict = item
            original_list.append(original_dict)

        amount_pattern = r"R\s(\d{1,3}(?:\s\d{3})*,\d{2})"

        for ind, original_dict in enumerate(original_list):

            for key, value in original_dict.items():

                if key.endswith("Actual") or re.match(r"\d{2}-[A-Za-z]{3}-\d{2}", key):
                    date = key.split()[0] if key.endswith("Actual") else key
                    amount = value
                    if key.endswith("Actual"):
                        actual = True
                    else:
                        actual = False
                    # print()
                    actual_dict = {
                        "Whitebox-Able": original_dict["Whitebox-Able"],
                        "Blocks": original_dict["Blocks"],
                        "Complete Build": original_dict["Complete Build"],
                        "Actual": actual,
                        "Date": date,
                        "Amount": amount,
                        "Renamed_block": original_dict["Renamed_block"],
                    "Development": original_dict["Development"]
                    }

                    converted_dicts.append(actual_dict)

        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        for i, d in enumerate(converted_dicts):

            d['Date'] = d['Date'].replace("-", "/")

            for index, month in enumerate(months, 1):
                d['Date'] = d['Date'].replace(month, f"{index:02d}")

            try:

                if not isinstance(d['Date'], datetime):
                    d['Date'] = datetime.strptime(d['Date'], '%Y/%m/%d')


            except Exception as e:

                if not isinstance(d['Date'], datetime):
                    d['Date'] = datetime.strptime(d['Date'], '%d/%m/%y')

        for item in converted_dicts:

            date_calc = item['Date'].strftime('%Y/%m/%d')

            result = calculate_vat_due(date_calc)
            item['Vat_due'] = result
            if not isinstance(item['Vat_due'], datetime):
                item['Vat_due'] = datetime.strptime(item['Vat_due'], '%Y/%m/%d')

        ws2b = wb.create_sheet('Updated Construction')

        ws2b.sheet_properties.tabColor = "1072BA"

        ws2b['A1'] = "Updated Construction"

        headers = ["Whitebox-Able", "Blocks", "Complete Build", "Actual", "Date", "Amount", "Renamed_block","Development",
                   "Vat Due", ]
        ws2b.append(headers)
        for item in converted_dicts:
            ws2b.append([item[key] for key in item])

        ws3 = create_sheet(wb, 'Sales', sales, {'B': 44, 'F': 44, 'G': 44, 'H': 44}, {})

        last_row = ws3.max_row
        columns = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
        # in row 2 add a subtotal for each column
        for col in columns:
            ws3[f"{col}2"] = f"=SUBTOTAL(9,{col}5:{col}{last_row})"
            ws3[f"{col}2"].number_format = '#,##0.00'



        for row in ws3.iter_rows(min_row=5, max_row=ws3.max_row, min_col=19, max_col=19):
            for cell in row:
                cell.value = f"=IF(T{cell.row}=FALSE,Q{cell.row}-R{cell.row},+V{cell.row})"
                cell.number_format = '#,##0.00'

        for row in ws3.iter_rows(min_row=5, max_row=ws3.max_row, min_col=22, max_col=22):
            for cell in row:
                cell.number_format = '#,##0.00'

        ws4 = create_sheet(wb, 'Operational Costs', operational_costs, {'A': 44, 'B': 44, 'F': 44, 'G': 44, 'H': 44},
                           {})

        last_row = ws4.max_row
        columns = ['N']

        for col in columns:
            ws4[f"{col}2"] = f"=SUBTOTAL(9,{col}5:{col}{last_row})"
            ws4[f"{col}2"].number_format = '#,##0.00'

        ws5 = create_sheet(wb, 'Xero', xero, {'A': 44, 'B': 44, 'F': 44, 'G': 44, 'H': 44}, {})

        last_row = ws5.max_row
        columns = ['F', 'G']

        for col in columns:
            ws5[f"{col}2"] = f"=SUBTOTAL(9,{col}5:{col}{last_row})"
            ws5[f"{col}2"].number_format = '#,##0.00'
            # format from row 5 to last row as currency

        # ws5.auto_filter.ref = f"A4:{get_column_letter(ws5.max_column)}{last_row}"
        # ws5.freeze_panes = "A5"
        # filter column D to only show 8400/000, 8440/000 and 8400/010

        ws5c = wb.create_sheet('Xero_bank')
        xero_bank = list(
            filter(
                lambda x: x['AccountCode']
                          in [
                              '8400/000',
                              '8440/000',
                              '8400/010',
                          ],
                xero,
            )
        )
        xero_bank = list(
            filter(
                lambda x: x['ReportTitle']
                            in ["Heron Fields (Pty) Ltd", "Purple Blok Projects (Pty) Ltd"],
                xero_bank,
            )
        )

        ws5c.sheet_properties.tabColor = "1072BA"
        ws5c['A1'] = "Xero Bank"
        ws5c.append([])
        ws5c.append([])
        ws5c.append(["ReportTitle", "ReportDate", "Category", "AccountCode", "AccountName", "Current", "Ytd"])
        for item in xero_bank:
            ws5c.append([item[key] for key in item])
        # filter from row 4
        ws5c.auto_filter.ref = f"A4:{get_column_letter(ws5c.max_column)}{ws5c.max_row}"
        ws5c.freeze_panes = "A5"










        ws5b = wb.create_sheet('Other Costs')

        ws5b.sheet_properties.tabColor = "1072BA"
        ws5b.title = "Other Costs"
        ws5b['A1'] = "Other Costs"
        ws5b['A2'] = ""
        ws5b['A3'] = ""

        headers_other_costs = []
        for item in other_costs[0]:
            headers.append(item.replace("_", " ").title())
        ws5b.append(headers_other_costs)

        for item in other_costs:

            if not isinstance(item['ReportDate'], datetime):
                item['ReportDate'] = datetime.strptime(item['ReportDate'], '%Y-%m-%d')

            row = []
            for key in item:
                row.append(item[key])
            ws5b.append(row)



        ws8 = wb.create_sheet('Heron')

        # make tab color blue
        ws8.sheet_properties.tabColor = "0070C0"

        ws8['A1'] = "HERON P&L - Based on Cash Projection and Cashflow"
        ws8["A1"].font = Font(bold=True, color="31304D", size=14)
        ws8.append([])
        ws8.append([])
        ws8.append(["Heron Fields (Pty) Ltd", "Heron Fields"])
        fields = ws8.max_row
        ws8.append(["Heron View (Pty) Ltd", "Heron View"])
        view = ws8.max_row
        ws8.append(["Cape Projects Construction (Pty) Ltd"])
        cpc = ws8.max_row
        ws8.append([])
        ws8.append([])

        reporting_months = []

        # filter Xero and only include where ReportTitle = "Heron Fields (Pty) Ltd" or ReportTitle = "Heron View (Pty) Ltd" using list comprehension
        xero_heron = [x for x in xero if
                      x['ReportTitle'] == "Heron Fields (Pty) Ltd" or x['ReportTitle'] == "Heron View (Pty) Ltd"]
        # xero_heron to only include where Category = "Income" or Category = "Expenses"
        xero_heron = [x for x in xero_heron if x['Category'] == "Revenue" or x['Category'] == "Expenses"]
        # print("Xero Heron", xero_heron)
        income_accounts = []
        expense_accounts = []
        # print("Xero Heron", xero_heron[0])
        # print()
        for item in xero_heron:
            reporting_months.append(item['ReportDate'])
            if item['Category'] == "Revenue":
                income_accounts.append(item['AccountName'])
            elif item['Category'] == "Expenses":
                expense_accounts.append(item['AccountName'])
        reporting_months = list(set(reporting_months))

        income_accounts = list(set(income_accounts))
        # sort the income accounts
        income_accounts.sort()
        expense_accounts = list(set(expense_accounts))
        # sort the expense accounts
        expense_accounts.sort()

        expense_accounts.append("Heron View Land")
        # print("Expense Accounts", expense_accounts)

        # sort the reporting months
        reporting_months.sort()
        # Insert "Month", "","" at the beginning of the list
        reporting_months.insert(0, "Month")
        reporting_months.insert(1, "")
        reporting_months.insert(2, "")
        # add 12 months to the list

        ws8.append(reporting_months)
        months_report = ws8.max_row
        # make months report row bold
        for i in range(1, ws8.max_column + 1):
            ws8[f"{get_column_letter(i)}{months_report}"].font = Font(bold=True, size=14)
            ws8[f"{get_column_letter(i)}{months_report}"].alignment = Alignment(horizontal='center', vertical='center')
            ws8[f"{get_column_letter(i)}{months_report}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                             fill_type="solid")
            if i > 3:
                # format as "YYYY-MM-DD"
                ws8[f"{get_column_letter(i)}{months_report}"].number_format = 'yyyy-mm-dd'

        for i in range(ws8.max_column + 1, ws8.max_column + 19):
            # f"=EOMONTH(EDATE($B$2, 0), 1)"
            ws8[
                f"{get_column_letter(i)}{months_report}"] = f"=EOMONTH(EDATE({get_column_letter(i - 1)}{months_report}, 0),1)"
            ws8[f"{get_column_letter(i)}{months_report}"].number_format = 'yyyy-mm-dd'
            ws8[f"{get_column_letter(i)}{months_report}"].font = Font(bold=True, size=14)
            ws8[f"{get_column_letter(i)}{months_report}"].alignment = Alignment(horizontal='center', vertical='center')
            ws8[f"{get_column_letter(i)}{months_report}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                             fill_type="solid")

        for index, i in enumerate(range(ws8.max_column + 1, ws8.max_column + 3)):
            if index == 0:
                ws8[f"{get_column_letter(i)}{months_report}"] = "Total"
            else:
                ws8[f"{get_column_letter(i)}{months_report}"] = "NSST"
            ws8[f"{get_column_letter(i)}{months_report}"].font = Font(bold=True, size=14)
            ws8[f"{get_column_letter(i)}{months_report}"].alignment = Alignment(horizontal='center', vertical='center')
            ws8[f"{get_column_letter(i)}{months_report}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                             fill_type="solid")

        ws8.append([])
        ws8.append(['Income'])
        # make the income row bold and a bit bigger
        ws8[f"A{ws8.max_row}"].font = Font(bold=True, size=14)
        ws8.append([])
        for index, item in enumerate(income_accounts):
            ws8.append([item])
            ws8[f"A{ws8.max_row}"].font = Font(bold=False, size=12)
            if index == 0:
                income_start = ws8.max_row
        income_end = ws8.max_row
        # print("Income Start", income_start)
        # print("Income End", income_end)

        ws8.append([])
        ws8.append(['Total Income'])
        ws8[f"A{ws8.max_row}"].font = Font(bold=True, size=14)

        for i in range(4, ws8.max_column + 1):
            ws8[
                f"{get_column_letter(i)}{ws8.max_row}"] = f"=sum({get_column_letter(i)}{income_start}:{get_column_letter(i)}{income_end})"
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].number_format = '#,##0.00'
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].font = Font(bold=True, size=14)
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].border = Border(top=Side(style='medium'),
                                                                        bottom=Side(style='double'))

        ws8.append([])
        ws8.append(['Expenses'])
        # make the income row bold and a bit bigger
        ws8[f"A{ws8.max_row}"].font = Font(bold=True, size=14)
        ws8.append([])

        for index, item in enumerate(expense_accounts):
            ws8.append([item])
            ws8[f"A{ws8.max_row}"].font = Font(bold=False, size=12)
            if index == 0:
                expense_start = ws8.max_row

        cpc_costs = ["COS - Heron Fields - Printing & Stationary", "COS - Heron Fields - Construction",
                     "COS - Heron Fields - P & G", "COS - Heron View - P&G", "COS - Heron View - Construction",
                     "COS - Heron View - Printing & Stationary"]
        # sort the cpc costs
        cpc_costs.sort()
        for index, item in enumerate(cpc_costs):
            ws8.append([item])
            ws8[f"A{ws8.max_row}"].font = Font(bold=False, size=12)
            if index == 0:
                cpc_start = ws8.max_row
        cpc_end = ws8.max_row

        other_costs_headers = []
        for item in other_costs:
            other_costs_headers.append(item['AccountName'])

        other_costs_headers = list(set(other_costs_headers))
        other_costs_headers.sort()

        for index, item in enumerate(other_costs_headers):
            # print("Item", item)
            ws8.append([item])
            ws8[f"A{ws8.max_row}"].font = Font(bold=False, size=12)
            if index == 0:
                other_costs_start = ws8.max_row
        other_costs_end = ws8.max_row

        expense_end = ws8.max_row
        print("Expense Start", expense_start)
        print("Expense End", expense_end)

        ws8.append([])
        ws8.append(['Total Expenses'])
        ws8[f"A{ws8.max_row}"].font = Font(bold=True, size=14)

        for i in range(4, ws8.max_column + 1):
            ws8[
                f"{get_column_letter(i)}{ws8.max_row}"] = f"=sum({get_column_letter(i)}{expense_start}:{get_column_letter(i)}{expense_end})"
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].number_format = '#,##0.00'
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].font = Font(bold=True, size=14)
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].border = Border(top=Side(style='medium'),
                                                                        bottom=Side(style='double'))

        ws8.append([])
        ws8.append([])
        ws8.append(['Net Profit'])
        ws8[f"A{ws8.max_row}"].font = Font(bold=True, size=14)

        for i in range(4, ws8.max_column + 1):
            ws8[
                f"{get_column_letter(i)}{ws8.max_row}"] = f"={get_column_letter(i)}{income_end + 2}-{get_column_letter(i)}{expense_end + 2}"
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].number_format = '#,##0.00'
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].font = Font(bold=True, size=14)
            ws8[f"{get_column_letter(i)}{ws8.max_row}"].border = Border(top=Side(style='medium'),
                                                                        bottom=Side(style='double'))

        for i in range(1, ws8.max_column + 1):
            if i == 1:
                ws8.column_dimensions[get_column_letter(i)].width = 35
            elif i > 3:
                ws8.column_dimensions[get_column_letter(i)].width = 18

        for i in range(4, ws8.max_column - 1):
            for row in range(income_start, income_end + 1):
                # print(ws8[f"A{row}"].value)
                if ws8[f"A{row}"].value == "Sales - Heron View Sales":
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=-SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})-SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})+SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$A:$A,Heron!$B$5,Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                elif ws8[f"A{row}"].value == "Sales - Heron Fields":
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=-SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})-SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})+SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$A:$A,Heron!$B$4,Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                else:
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=-SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})-SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
            for row in range(expense_start, cpc_start):
                if ws8[f"A{row}"].value == "COS - Commission HV Units":
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})+SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})+SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$A:$A,Heron!$B$5,Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)/1.15"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)

                elif ws8[f"A{row}"].value == "COS - Commission HF Units":
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})+SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})+SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$A:$A,Heron!$B$4,Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)/1.15"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                elif ws8[f"A{row}"].value == "COS - Legal Fees":
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})+SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})+SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\",Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)+SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\",Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)+SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\",Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)+SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\",Sales!$H:$H,\"<=\"&Heron!{get_column_letter(i)}$9,Sales!$H:$H,\">\"&Heron!{get_column_letter(i - 1)}$9,Sales!$F:$F,1)"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                elif ws8[f"A{row}"].value == "Interest Paid - Investors @ 18%":
                    # "+SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!D$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A73)+SUMIFS(Investors!$S:$S,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$P:$P,FALSE,Investors!$K:$K,"<="&Heron!D$9,Investors!$K:$K,">"&Heron!C$9)"
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})+SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})+SUMIFS(Investors!$S:$S,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$P:$P,FALSE,Investors!$K:$K,\"<=\"&Heron!{get_column_letter(i)}$9,Investors!$K:$K,\">\"&Heron!{get_column_letter(i - 1)}$9)"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                else:
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$4,Xero!$E:$E,Heron!$A{row})+SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$5,Xero!$E:$E,Heron!$A{row})"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)

            for row in range(cpc_start, cpc_end + 1):
                if ws8[f"A{row}"].value == "COS - Heron View - Construction":
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$6,Xero!$E:$E,Heron!$A{row})+SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$C:$C,1,'Updated Construction'!$E:$E,\"<=\"&Heron!{get_column_letter(i)}$9,'Updated Construction'!$E:$E,\">\"&Heron!{get_column_letter(i - 1)}$9,'Updated Construction'!$D:$D,FALSE)"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                    # fill with light grey
                    ws8[f"{get_column_letter(i)}{row}"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3",
                                                                           fill_type="solid")
                else:
                    ws8[
                        f"{get_column_letter(i)}{row}"].value = f"=SUMIFS(Xero!$F:$F,Xero!$B:$B,Heron!{get_column_letter(i)}$9,Xero!$A:$A,Heron!$A$6,Xero!$E:$E,Heron!$A{row})"
                    ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                    ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                    # fill with light grey
                    ws8[f"{get_column_letter(i)}{row}"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3",
                                                                           fill_type="solid")

            for row in range(other_costs_start, other_costs_end + 1):
                "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,Heron!$A106,'Other Costs'!$C:$C," <= "&Heron!D$9,'Other Costs'!$C:$C," > "&Heron!C$9)"
                ws8[
                    f"{get_column_letter(i)}{row}"].value = f"=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,Heron!$A{row},'Other Costs'!$C:$C,\"<=\"&Heron!{get_column_letter(i)}$9,'Other Costs'!$C:$C,\">\"&Heron!{get_column_letter(i - 1)}$9)"
                ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)
                # fill with light orange
                ws8[f"{get_column_letter(i)}{row}"].fill = PatternFill(start_color="FFD966", end_color="FFD966",
                                                                       fill_type="solid")

        for i in range(ws8.max_column - 1, ws8.max_column):
            # print("I", i)
            for row in range(income_start, income_end + 1):
                ws8[
                    f"{get_column_letter(i)}{row}"].value = f"=sum(D{row}:{get_column_letter(i - 1)}{row})"
                ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)

            for row in range(expense_start, expense_end + 1):
                ws8[
                    f"{get_column_letter(i)}{row}"].value = f"=sum(D{row}:{get_column_letter(i - 1)}{row})"
                ws8[f"{get_column_letter(i)}{row}"].number_format = '#,##0.00'
                ws8[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=12)

        # print(opportunities[0])
        # using list comprehension, put Category from opportunities into a new list
        opportunity_categories = [x['Category'] for x in opportunities]
        # remove duplicates
        opportunity_categories = list(set(opportunity_categories))

        print(opportunity_categories)

        reports_by_project = ["Consolidated"]
        for opp in opportunity_categories:
            # if the first 5 characters of the opportunity are "Heron", append "Heron" to reports_by_project
            if opp[:5] == "Heron":
                reports_by_project.append("Heron")
            else:
                reports_by_project.append(opp)

        # remove duplicates from reports_by_project
        reports_by_project = list(set(reports_by_project))
        # sort the list
        reports_by_project.sort()
        # print(reports_by_project)
        tab_colors = ["FF204E", "399918", "FF8225", "7C00FE", "F6FB7A", "36C2CE"]
        # print("Got this far")

        # global block_cost_range_start, block_cost_range_end
        block_cost_range_start = []
        block_cost_range_end = []

        refinanced_units_start_range = []
        refinanced_units_end_range = []


        for idx, project in enumerate(reports_by_project):


            if idx == 0:
                ws6 = wb.create_sheet('Cashflow Projection')
            else:
                ws6 = wb.create_sheet(f'Cashflow Projection - {project}')
                # ws7 = wb.create_sheet(f'Cashflow - {project}')
            # make tab color red
            ws6.sheet_properties.tabColor = tab_colors[idx]

            # CASH PROJECTION

            # ws6 = wb.create_sheet('Cashflow Projection')
            # # make tab color red
            # ws6.sheet_properties.tabColor = "FF204E"

            ws6['A1'] = "Cashflow Projection"
            ws6["A1"].font = Font(bold=True, color="0C0C0C", size=28)
            # Merge A1 to B1
            ws6.merge_cells('A1:B1')

            ws6['A2'] = 'Date'
            ws6["A2"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6['B2'] = report_date.strftime('%d-%b-%Y')
            ws6["B2"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6['A3'] = 'NSST Cashflow Projection'
            ws6["A3"].font = Font(bold=True, color="0C0C0C", size=26)
            # fill with olive green
            ws6['A3'].fill = PatternFill(start_color="7F9F80", end_color="7F9F80", fill_type="solid")
            ws6['A3'].border = Border(left=Side(style='medium'),
                                      right=Side(style='medium'),
                                      top=Side(style='medium'),
                                      bottom=Side(style='medium'))

            ws6['A4'] = ''
            sale_info = []

            # sale_profit_info = []
            for sale in sales:
                if not sale['transferred']:
                    insert = {
                        "Development": sale['Category'],
                        "Block": sale['block'],
                        "Complete Build": sale['complete_build'],
                        "Totals": "",
                    }
                    sale_info.append(insert)
            sale_info = [dict(t) for t in {tuple(d.items()) for d in sale_info}]
            sale_info = sorted(sale_info, key=lambda x: (x['Development'], x['Block']))
            row = []

            for key in sale_info[0]:
                row.append(key)
            ws6.append(row)

            ws6['A6'] = 'SALES'
            ws6['A6'].font = Font(bold=True, color="0C0C0C", size=22)
            ws6['A5'] = 'PROJECTION'
            ws6['B5'] = ''
            ws6['C5'] = ''
            ws6['D5'] = 'TOTAL'
            # center D5
            ws6['D5'].alignment = Alignment(horizontal='center', vertical='center')

            ws6['A5'].font = Font(bold=True, color="0C0C0C", size=22)
            ws6['D5'].font = Font(bold=True, color="0C0C0C", size=22)

            toggles_start = ws6.max_row + 1
            row = []
            # print("Sale Info", sale_info[10])
            if project == "Consolidated":
                sale_info_filtered = sale_info
            elif project == "Heron":
                sale_info_filtered = [sale for sale in sale_info if sale['Development'][:5] == "Heron"]
            else:
                sale_info_filtered = [sale for sale in sale_info if sale['Development'] == project]

            for sale in sale_info_filtered:
                for key in sale:
                    row.append(sale[key])
                ws6.append(row)

                row = []

            toggles_end = ws6.max_row
            print("toggles_start", toggles_start)
            print("toggles_end", toggles_end)

            month_headings = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
                              'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']

            ws6.merge_cells(f"A3:{month_headings[len(month_headings) - 1]}3")

            ws6['A3'].alignment = Alignment(horizontal='center', vertical='center')
            if project != "Goodwood":
                ws6[f"{month_headings[len(month_headings) - 1]}1"] = 'C.3.e'
            else:
                ws6[f"{month_headings[len(month_headings) - 1]}1"] = 'C.5.1'
            ws6[f"{month_headings[len(month_headings) - 1]}1"].font = Font(bold=True, size=28)
            # center the above
            ws6[f"{month_headings[len(month_headings) - 1]}1"].alignment = Alignment(horizontal='center',
                                                                                     vertical='center')
            # put a border around the above cell
            ws6[f"{month_headings[len(month_headings) - 1]}1"].border = Border(left=Side(style='medium'),
                                                                               right=Side(style='medium'),
                                                                               top=Side(style='medium'),
                                                                               bottom=Side(style='medium'))

            for row in ws6.iter_rows(min_row=toggles_start, max_row=toggles_end, min_col=1, max_col=5):
                for cell in row:
                    cell.font = Font(bold=True, color="0C0C0C", size=22)
                    # cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

            for row in ws6.iter_rows(min_row=toggles_start, max_row=toggles_end, min_col=9, max_col=ws6.max_column):
                for cell in row:
                    cell.font = Font(bold=True, color="0C0C0C", size=22)
                    # cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

            ws6.append([])
            ws6.append([])
            ws6.append(['VAT INCOME ON SALES', "", 1])
            vat_income = ws6.max_row
            ws6[f"A{vat_income}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{vat_income}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{vat_income}"].number_format = 'R#,##0'
            ws6[f"D{vat_income}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{vat_income}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C", fill_type="solid")
            ws6[f"D{vat_income}"].border = ws6[f"A{vat_income}"].border + Border(left=Side(style='medium'),
                                                                                 right=Side(style='medium'),
                                                                                 top=Side(style='medium'),
                                                                                 bottom=Side(style='medium'))

            ws6.append(['CPC INVOICES STILL DUE'])
            CPC_INVOICES_STILL_DUE = ws6.max_row
            ws6.append(['VAT DEDUCTION FROM INVOICES'])
            Vat_deduction_from_invoices = ws6.max_row

            # ws6.append([])
            ws6.append(['VAT PAYABLE ON SALES', "", 1])
            vat_payable_on_sales = ws6.max_row
            print("vat_payable_on_sales", vat_payable_on_sales)

            ws6[f"A{vat_payable_on_sales}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{vat_payable_on_sales}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{vat_payable_on_sales}"].number_format = 'R#,##0'
            ws6[f"D{vat_payable_on_sales}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{vat_payable_on_sales}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                               fill_type="solid")
            ws6[f"D{vat_payable_on_sales}"].border = ws6[f"A{vat_payable_on_sales}"].border + Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

            ws6.append([])
            ws6.append(['VAT RECOVERY WHEN REFINANCED', "", 1])
            vat_recovery_when_refinanced = ws6.max_row
            print("vat_recovery_when_refinanced", vat_recovery_when_refinanced)

            print("vat_income", vat_income)

            ws6[f"A{CPC_INVOICES_STILL_DUE}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{CPC_INVOICES_STILL_DUE}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{CPC_INVOICES_STILL_DUE}"].number_format = 'R#,##0'
            ws6[f"D{CPC_INVOICES_STILL_DUE}"].font = Font(bold=True, color="0C0C0C", size=22)

            ws6[f"A{Vat_deduction_from_invoices}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{Vat_deduction_from_invoices}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{Vat_deduction_from_invoices}"].number_format = 'R#,##0'
            ws6[f"D{Vat_deduction_from_invoices}"].font = Font(bold=True, color="0C0C0C", size=22)

            vat_row = ws6.max_row
            ws6[f"A{vat_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{vat_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{vat_row}"].number_format = 'R#,##0'
            ws6[f"D{vat_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            # fill with light green
            ws6[f"D{vat_row}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C", fill_type="solid")
            ws6[f"D{vat_row}"].border = ws6[f"A{vat_row}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

            print("vat_row", vat_row)

            ws6.append([])
            ws6.append(["Income (Profit on Sale)"])
            profit_on_sale = ws6.max_row
            print("profit_on_sale", profit_on_sale)
            ws6[f"A{profit_on_sale}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{profit_on_sale}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{profit_on_sale}"].number_format = 'R#,##0'
            ws6[f"D{profit_on_sale}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{profit_on_sale}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C", fill_type="solid")
            ws6[f"D{profit_on_sale}"].border = ws6[f"A{profit_on_sale}"].border + Border(left=Side(style='medium'),

                                                                                         right=Side(style='medium'),
                                                                                         top=Side(style='medium'),
                                                                                         bottom=Side(style='medium'))
            ws6.append([])
            ws6.append(["FUNDS AVAILABLE"])
            funds_available_start = ws6.max_row
            print("funds_available_start", funds_available_start)
            ws6[f"A{funds_available_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6.append(["Sales Income"])
            ws6.append(["Momentum"])
            ws6.append(["Bank"])
            ws6.append(["Early Exit_Investor Roll Over funding"])
            ws6.append(["Developer Contribution towards exit of Investors"])
            ws6.append(["Developer Exit Payment of Investors"])
            ws6.append(["New Investors Income"])
            ws6.append(["Re-Financed Income"])
            ws6.append(["Rollover Income from Investors on Sales Transfers"])
            ws6.append(["Deposits"])
            ws6.append(["INCOME"])
            funds_available_end = ws6.max_row

            ws6[f'D{funds_available_start}'].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                                fill_type="solid")
            ws6[f'D{funds_available_start}'].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f'D{funds_available_start}'].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))

            ws6[f'D{funds_available_end}'].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                              fill_type="solid")
            ws6[f'D{funds_available_end}'].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f'D{funds_available_end}'].border = Border(left=Side(style='medium'),
                                                           right=Side(style='medium'),
                                                           top=Side(style='medium'),
                                                           bottom=Side(style='medium'))

            for row in ws6.iter_rows(min_row=funds_available_start + 1, max_row=funds_available_end - 1, min_col=4,
                                     max_col=4):
                for cell in row:
                    cell.number_format = 'R#,##0'
                    cell.font = Font(bold=True, color="0C0C0C", size=22)
                    # cell.fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C", fill_type="solid")
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

            for row in ws6.iter_rows(min_row=funds_available_start, max_row=funds_available_end, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True, color="0C0C0C", size=22)

            # ws6['A38'].font = Font(bold=True, color="0C0C0C", size=22)

            ws6.append([])
            ws6.append(["Costs To Complete"])
            costs_to_complete_start = ws6.max_row
            ws6[f"A{costs_to_complete_start}"].font = Font(bold=True, color="0C0C0C", size=22)

            # ws6.append(["UNITS"])
            ws6.append(["CONSTRUCTION COSTS"])
            units = ws6.max_row
            ws6[f"A{units}"].font = Font(bold=True, color="0C0C0C", size=22)

            block_costs_start = ws6.max_row + 1

            projectA = project
            # block_costs_start = 1000

            # Dynamically create a global variable
            globals()[f"block_costs_start_{projectA}"] = block_costs_start

            # Access the dynamically created variable
            # print(f"Block Costs Start - {projectA}", globals()[f"block_costs_start_{projectA}"])



            if project == "Consolidated":
                construction_filtered_block = construction
            else:
                construction_filtered_block = [item for item in construction if item['Development'] == project]



            construction_blocks = []
            if len(construction_filtered_block) > 0:
                for item in construction_filtered_block:
                    construction_blocks.append(item['Renamed_block'])
            else:
                construction_blocks.append("No Blocks")
            construction_blocks = list(set(construction_blocks))
            construction_blocks = sorted(construction_blocks)

            for block in construction_blocks:

                filtered_construction = [item for item in construction if
                                         item['Renamed_block'] == block and item['Whitebox-Able'] == True]

                if len(filtered_construction) > 0:
                    value = filtered_construction[0]['Complete Build']
                    ws6.append([f"{block}", value])
                else:
                    ws6.append([f"{block}", 1])
            block_costs_end = ws6.max_row

            globals()[f"block_costs_end_{projectA}"] = block_costs_end

            # Access the dynamically created variable


            print()
            print(f"Block Costs Start - {projectA}", globals()[f"block_costs_start_{projectA}"])
            print(f"Block Costs End - {projectA}", globals()[f"block_costs_end_{projectA}"])
            print()

            block_cost_range_start.append(globals()[f"block_costs_start_{projectA}"])

            block_cost_range_end.append(globals()[f"block_costs_end_{projectA}"])

            print("block_cost_range_start", block_cost_range_start)
            print("block_cost_range_end", block_cost_range_end)
            print("block_costs_start", block_costs_start)

            print("block_costs_end", block_costs_end)

            for row in ws6.iter_rows(min_row=block_costs_start, max_row=block_costs_end, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = 'R#,##0'
                    cell.font = Font(bold=True, color="0C0C0C", size=22)
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

            for row in ws6.iter_rows(min_row=block_costs_start, max_row=block_costs_end, min_col=9,
                                     max_col=ws6.max_column):
                for cell in row:
                    cell.number_format = 'R#,##0'
                    cell.font = Font(bold=True, color="0C0C0C", size=22)
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

            ws6.append([])
            ws6.append([])
            ws6.append([])

            if project != "Goodwood":
                ws6.append(["VAT ON CONSTRUCTION", "", 1])
            else:
                ws6.append(["VAT ON CONSTRUCTION", "", 0])
            vat_construction = ws6.max_row
            ws6[f"A{vat_construction}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{vat_construction}"].font = Font(bold=True, color="0C0C0C", size=22)

            ws6[f"C{vat_income}"].value = f"=+C{vat_construction}"
            ws6[f"C{vat_payable_on_sales}"] = f"=+C{vat_construction}"
            ws6[f"C{vat_row}"] = f"=+C{vat_construction}"

            # fill C{vat_construction} with yellow
            ws6[f"C{vat_construction}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            print("vat_construction", vat_construction)

            ws6.append([])
            ws6.append(["OPERATING EXPENSES", "", 0.55])
            operating_expenses = ws6.max_row
            ws6[f"A{operating_expenses}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{operating_expenses}"].font = Font(bold=True, color="0C0C0C", size=22)

            # format the above cell as a percentage
            ws6[f"C{operating_expenses}"].number_format = '0%'

            ws6.append([])
            ws6.append(["MONTHLY"])
            monthly = ws6.max_row
            ws6[f"A{monthly}"].font = Font(bold=True, color="0C0C0C", size=22)
            # ws6[f"C{monthly}"].font = Font(bold=True, color="31304D", size=14)

            ws6.append([])
            ws6.append(["BANK BALANCE"])

            running = ws6.max_row
            variables = {}
            variables[f"{project}_running"] = running
            # if project == "Heron":
            #     heron_running = running
            # elif project == "Goodwood":
            #     goodwood_running = running
            ws6[f"A{running}"].font = Font(bold=True, color="0C0C0C", size=22)
            print("running", running)
            ws6.append([])
            ws6.append(["INVESTOR EXIT ROLLOVER BALANCE"])
            investor_exited = ws6.max_row
            ws6[f"A{investor_exited}"].font = Font(bold=True, color="0C0C0C", size=22)
            print("investor_exited", investor_exited)

            ws6.append([])
            ws6.append(['ROLLOVER'])
            rollover = ws6.max_row
            print("rollover", rollover)
            ws6[f"A{rollover}"].font = Font(bold=True, color="0C0C0C", size=22)

            ws6[f'D{rollover}'].fill = PatternFill(start_color="3ABEF9", end_color="3ABEF9", fill_type="solid")
            ws6[f'D{rollover}'].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f'D{rollover}'].border = Border(left=Side(style='medium'),
                                                right=Side(style='medium'),
                                                top=Side(style='medium'),
                                                bottom=Side(style='medium'))

            ws6.append([])
            ws6.append(['PROJECT INCOME'])

            project_income = ws6.max_row
            print("project_income", project_income)

            ws6.append([])
            ws6.append(['Roll Over_Re-Finance'])
            roll_over_refinance = ws6.max_row
            ws6[f"A{roll_over_refinance}"].font = Font(bold=True, color="0C0C0C", size=22)
            print("roll_over_refinance", roll_over_refinance)
            ws6.append(['Roll Over_Re-Invest'])
            roll_over_reinvest = ws6.max_row
            print("roll_over_reinvest", roll_over_reinvest)
            ws6[f"A{roll_over_reinvest}"].font = Font(bold=True, color="0C0C0C", size=22)

            block_finance_category = []
            if project == "Consolidated":
                for sale in sales:
                    if not sale['transferred'] and not sale['sold']:
                        insert = {
                            "Development": sale['Category'],
                            "Block": sale['block'],
                        }
                        block_finance_category.append(insert)
            elif project == "Heron":
                for sale in sales:
                    if not sale['transferred'] and not sale['sold'] and sale['Category'][:5] == "Heron":
                        insert = {
                            "Development": sale['Category'],
                            "Block": sale['block'],
                        }
                        block_finance_category.append(insert)
            else:
                for sale in sales:
                    if not sale['transferred'] and not sale['sold'] and sale['Category'] == project:
                        insert = {
                            "Development": sale['Category'],
                            "Block": sale['block'],
                        }
                        block_finance_category.append(insert)

            # filter out duplicates
            block_finance_category = [dict(t) for t in {tuple(d.items()) for d in block_finance_category}]
            block_finance_category = sorted(block_finance_category, key=lambda x: (x['Development'], x['Block']))
            # print("block_finance_category", block_finance_category)

            block_finance = ['C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

            ws6.append([""])
            block_finance_start = ws6.max_row

            for i in range(len(block_finance)):
                ws6.append([])
                ws6[f"A{block_finance_start + i}"] = f"BLOCK {block_finance[i]}"
                ws6[f"A{block_finance_start + i}"].font = Font(bold=True, color="0C0C0C", size=22)
                # apply borders
                ws6[f"A{block_finance_start + i}"].border = Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                ws6[f"B{block_finance_start + i}"] = f"{block_finance[i]}"
                ws6[f"B{block_finance_start + i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"B{block_finance_start + i}"].border = Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                ws6[f"C{block_finance_start + i}"] = f"Heron View"
                ws6[f"C{block_finance_start + i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"C{block_finance_start + i}"].border = Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                ws6[f"E{block_finance_start + i}"] = f"=+E{roll_over_refinance}"
                ws6[f"E{block_finance_start + i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"E{block_finance_start + i}"].border = Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                # format D as %
                ws6[f"E{block_finance_start + i}"].number_format = '0%'

                formula_start = "="
                # count_formula_start = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start += f"+{col}{block_finance_start + i}"

                    ws6[f"D{block_finance_start + i}"] = f"{formula_start}"
                    ws6[f"D{block_finance_start + i}"].number_format = 'R#,##0'
                    ws6[f"D{block_finance_start + i}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"D{block_finance_start + i}"].border = Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))

            block_finance_end = ws6.max_row

            for row in ws6.iter_rows(min_row=block_finance_start, max_row=block_finance_end, min_col=1, max_col=1):
                for cell in row:
                    # apply borders
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))

            print("block_finance_start", block_finance_start)
            print("block_finance_end", block_finance_end)
            ws6.append(["", "", "", f"=sum(D{block_finance_start}:D{block_finance_end})"])
            ws6[f"D{block_finance_end + 2}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{block_finance_end + 2}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))

            ws6[f"A{project_income}"].font = Font(bold=True, color="0C0C0C", size=22)

            ws6[f"D{profit_on_sale}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{profit_on_sale}"].number_format = 'R#,##0'

            ws6[f"D{profit_on_sale}"].border = Border(left=Side(style='medium'),
                                                      right=Side(style='medium'),
                                                      top=Side(style='medium'),
                                                      bottom=Side(style='medium'))

            ws6.append([])

            ws6.append(["Other Income - Goodwood"])
            ws6.append(["Purple Blok", 4279223.84, 1])

            purple_blok_start = ws6.max_row
            ws6.append(["Other Income", 0, 0])
            # make B currency
            # ws6[f"B{purple_blok_start + 1}"].number_format = 'R#,##0'
            # make font bold and
            ws6.append(["Repayment"])
            ws6.append(["Total"])

            ws6[f"A{purple_blok_start - 1}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"A{purple_blok_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"B{purple_blok_start}"].number_format = 'R#,##0'
            ws6[f"B{purple_blok_start + 1}"].number_format = 'R#,##0'
            ws6[f"B{purple_blok_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{purple_blok_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{purple_blok_start}"].number_format = '0'
            # ws6[f"D{purple_blok_start}"].number_format = 'R#,##0'
            # ws6[f"D{purple_blok_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            purple_blok_end = ws6.max_row

            for row in ws6.iter_rows(min_row=purple_blok_start, max_row=purple_blok_end, min_col=1, max_col=3):
                for cell in row:
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))
                    cell.font = Font(bold=True, color="0C0C0C", size=22)

            print("purple_blok_start", purple_blok_start)
            print("purple_blok_end", purple_blok_end)

            ws6.append([])
            ws6.append(["FINANCE WATERFALL - EXIT PROCESS"])
            # make font bold, 22 and RED
            ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="FF0000", size=22)

            finance_waterfall_start = ws6.max_row
            ws6.append(["Income"])
            wf1 = ws6.max_row
            ws6.append(["Endulini & Heron Fields_New Debenture_P/Block"])
            wf2 = ws6.max_row
            ws6.append(["Vat _ Late"])
            wf3 = ws6.max_row
            ws6.append(["New Raise"])
            wf4 = ws6.max_row
            ws6.append(["Developer Early Exit"])
            wf5 = ws6.max_row
            ws6.append(["Early Exit Investor Roll Over funding"])
            ws6.append(["Developer Contribution"])
            wf6 = ws6.max_row
            ws6.append(["Cash"])
            wf7 = ws6.max_row
            ws6.append(["Sales Income"])
            ws6.append(["New Investor Income"])
            wf8 = ws6.max_row
            ws6.append(["Momentum & Deposits & VAT & Other"])
            wf9 = ws6.max_row
            ws6.append([""])
            wf10 = ws6.max_row
            ws6.append(["CPC + Operating Expenses & VAT on Cycle"])
            wf11 = ws6.max_row
            ws6.append([""])
            wf12 = ws6.max_row
            ws6.append([""])
            wf13 = ws6.max_row
            finance_waterfall_end = ws6.max_row
            print("finance_waterfall_start", finance_waterfall_start)
            print("finance_waterfall_end", finance_waterfall_end)

            for row in ws6.iter_rows(min_row=finance_waterfall_start, max_row=finance_waterfall_end, min_col=1,
                                     max_col=ws6.max_column):
                for cell in row:
                    if row == finance_waterfall_start:
                        cell.font = Font(bold=True, color="FF0000", size=22)
                    else:
                        cell.font = Font(bold=True, color="0C0C0C", size=22)
                    # fill in light green
                    cell.fill = PatternFill(start_color="B5C18E", end_color="B5C18E", fill_type="solid")
                    # cell.border = Border(left=Side(style='medium'),
                    #                      right=Side(style='medium'),
                    #                      top=Side(style='medium'),
                    #                      bottom=Side(style='medium'))

            ws6[f"A{finance_waterfall_start}"].font = Font(bold=True, color="FF0000", size=22)

            ws6.append([])
            ws6.append(["DEVELOPER REFINANCE PROPERTIES"])
            # make bold and RED
            ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="FF0000", size=22)
            ws6.append([""])
            refinanced_units_start = ws6.max_row
            if project == "Heron":
                refinanced_units_start_heron = refinanced_units_start

            # projectA = project
            # block_costs_start = 1000

            # Dynamically create a global variable
            globals()[f"refinanced_units_start_{projectA}"] = refinanced_units_start + 1



            # print("test",sales[0])

            # let units_a equals opportunity_code where sold is false
            units_unfiltered = []
            for sale in sales:
                if not sale['sold']:
                    # units_unfiltered.append(sale)
                    insert = {
                        "Category": sale['Category'],
                        "opportunity_code": sale['opportunity_code'],
                    }
                    units_unfiltered.append(insert)

            units = []

            for unit in units_unfiltered:
                if project == "Consolidated":
                    units.append(unit['opportunity_code'])
                elif project == "Heron":
                    if unit['Category'][:5] == "Heron":
                        units.append(unit['opportunity_code'])
                else:
                    if unit['Category'] == project:
                        units.append(unit['opportunity_code'])

            # units = ["HVD302", "HVD303", "HVD304", "HVC202", "HVC204", "HVC302", "HVC304", "HVC305", "HVJ101", "HVJ102",
            #          "HVJ103",
            #          "HVJ201", "HVJ202", "HVJ203", "HVJ301", "HVJ302", "HVJ303", "HVJ401", "HVJ402", "HVJ403", "HVM101",
            #          "HVM102",
            #          "HVM103", "HVM104", "HVM201", "HVM202", "HVM203", "HVM204", "HVN302", "HVN303", "HVN304", "HVO205",
            #          "HVO305",
            #          "HVP203", "HVP303"]

            # units = sorted(units)
            # sort the units only after the 5th element
            # units = units[:5] + sorted(units[5:])

            # print("LEN",len(units))
            for index, unit in enumerate(units):
                "=IF(F117<>0,1,0)"
                ws6.append(["", "", "", unit, f"=IF(F{refinanced_units_start + index + 1}<>0,1,0)"])
                # print("index",index)

                # Put a border around Units and make font 22 and bold and fill in light green
                ws6[f"D{refinanced_units_start + index + 1}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"D{refinanced_units_start + index + 1}"].fill = PatternFill(start_color="B5C18E",
                                                                                 end_color="B5C18E",
                                                                                 fill_type="solid")
                ws6[f"D{refinanced_units_start + index + 1}"].border = Border(left=Side(style='medium'),
                                                                              right=Side(style='medium'),
                                                                              top=Side(style='medium'),
                                                                              bottom=Side(style='medium'))
                ws6[f"E{refinanced_units_start + index + 1}"].font = Font(bold=True, color="0C0C0C", size=22)
                # ws6[f"E{refinanced_units_start + index + 1}"].fill = PatternFill(start_color="B5C18E", end_color="B5C18E",
                #                                                                  fill_type="solid")
                ws6[f"E{refinanced_units_start + index + 1}"].border = Border(left=Side(style='medium'),
                                                                              right=Side(style='medium'),
                                                                              top=Side(style='medium'),
                                                                              bottom=Side(style='medium'))
                # if index == len(units) - 1:
                #     ws6.append([""])
            refinanced_units_end = ws6.max_row
            if project == "Heron":
                refinanced_units_end_heron = refinanced_units_end

            globals()[f"refinanced_units_end_{projectA}"] = refinanced_units_end

            refinanced_units_start_range.append(globals()[f"refinanced_units_start_{projectA}"])
            refinanced_units_end_range.append(globals()[f"refinanced_units_end_{projectA}"])
            print()
            print("refinanced_units_start_range", refinanced_units_start_range)
            print("refinanced_units_end_range", refinanced_units_end_range)
            print()

            # "=SUM(E117:E149)-E115"
            ws6[
                f"E{refinanced_units_start}"] = f"=SUM(E{refinanced_units_start + 1}:E{refinanced_units_end})-E{refinanced_units_start - 1}"
            ws6[f"E{refinanced_units_start}"].font = Font(bold=True, color="0C0C0C", size=22)

            print("refinanced_units_start", refinanced_units_start)
            print("refinanced_units_end", refinanced_units_end)

            ws6.append([])
            ws6.append(["DEBENTURE TRANSACTION", 0, f"=E{refinanced_units_start - 1}"])
            ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="FF0000", size=22)
            ws6[f"B{ws6.max_row}"].font = Font(bold=True, color="FF0000", size=22)
            ws6[f"C{ws6.max_row}"].font = Font(bold=True, color="FF0000", size=22)
            # put borders around the above rows
            ws6[f"A{ws6.max_row}"].border = Border(left=Side(style='medium'),
                                                   right=Side(style='medium'),
                                                   top=Side(style='medium'),
                                                   bottom=Side(style='medium'))
            ws6[f"B{ws6.max_row}"].border = Border(left=Side(style='medium'),
                                                   right=Side(style='medium'),
                                                   top=Side(style='medium'),
                                                   bottom=Side(style='medium'))
            ws6[f"C{ws6.max_row}"].border = Border(left=Side(style='medium'),
                                                   right=Side(style='medium'),
                                                   top=Side(style='medium'),
                                                   bottom=Side(style='medium'))
            # fill B in yellow
            ws6[f"B{ws6.max_row}"].fill = PatternFill(start_color="FFF67E", end_color="FFF67E", fill_type="solid")
            # format B as currency
            ws6[f"B{ws6.max_row}"].number_format = 'R#,##0'

            debenture_transaction_start = ws6.max_row
            print("debenture_transaction_start", debenture_transaction_start)

            ws6.append([""])
            ws6.append([""])

            ws6.append(["ROLL OVER REFINANCE PROPERTIES"])
            ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="FF0000", size=22)
            roll_over_refinance_properties_start = ws6.max_row
            # print("roll_over_refinance_properties_start", roll_over_refinance_properties_start)
            for i in range(85):
                ws6.append([i])
                ws6[f"H{ws6.max_row}"].font = Font(bold=True, color="000000", size=22)
                if i == 2:
                    # print("test", roll_over_refinance_properties_start + 89)
                    ws6[f"H{ws6.max_row}"].font = Font(bold=True, color="000000", size=22)

                    ws6[
                        f"H{ws6.max_row}"].value = f"=_xlfn.IFERROR(_xlfn.UNIQUE(_xlfn.FILTER($B${roll_over_refinance_properties_start + 89}:$B${roll_over_refinance_properties_start + 89 + 800},($C${roll_over_refinance_properties_start + 89}:$C${roll_over_refinance_properties_start + 89 + 800}<=I5)*($C${roll_over_refinance_properties_start + 89}:$C${roll_over_refinance_properties_start + 89 + 800}>G5)*($F${roll_over_refinance_properties_start + 89}:$F${roll_over_refinance_properties_start + 89 + 800}=0))),\"\")"
            roll_over_refinance_properties_end = ws6.max_row
            print("roll_over_refinance_properties_start", roll_over_refinance_properties_start)
            print("roll_over_refinance_properties_end", roll_over_refinance_properties_end)
            ws6[f"B{roll_over_refinance_properties_start + 1}"].font = Font(bold=True, color="000000", size=22)
            ws6[f"B{roll_over_refinance_properties_start + 1}"].value = "Developer Contribution"
            ws6[f"C{roll_over_refinance_properties_start + 1}"].value = "25%"
            ws6[f"C{roll_over_refinance_properties_start + 1}"].font = Font(bold=True, color="000000", size=22)
            # fill in yellow
            ws6[f"C{roll_over_refinance_properties_start + 1}"].fill = PatternFill(start_color="FFF67E",
                                                                                   end_color="FFF67E",
                                                                                   fill_type="solid")
            ws6[f"B{roll_over_refinance_properties_start + 1}"].font = Font(bold=True, color="000000", size=22)

            ws6[f"B{roll_over_refinance_properties_start + 2}"].font = Font(bold=True, color="000000", size=22)
            ws6[f"B{roll_over_refinance_properties_start + 2}"].value = "Roll Over - Investors"

            ws6.append([""])
            ws6.append([""])
            ws6.append(["block", "unique", "date", "value", "unit", "unit_used", "sold", ])
            # make the above row bold and font 22
            ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"B{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"E{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"F{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"G{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)

            if project == "Consolidated":

                ws6.append([f"=mid(E{ws6.max_row + 1}, 3, 1)",
                            "=_xlfn.UNIQUE(_xlfn.FILTER('Investor Exit List'!$AB:$AB,('Investor Exit List'!$I:$I=FALSE)*('Investor Exit List'!$AE:$AE=1)*('Investor Exit List'!$U:$U=FALSE)*('Investor Exit List'!$F:$F<>0)))",
                            f"=SUMIFS('Investor Exit List'!$AC:$AC,'Investor Exit List'!$AB:$AB,B{ws6.max_row + 1})",
                            f"=SUMIFS('Investor Exit List'!$Q:$Q,'Investor Exit List'!$AB:$AB,B{ws6.max_row + 1})",
                            f"=LEFT(B{ws6.max_row + 1},6)"])
            else:
                ws6.append([f"=mid(E{ws6.max_row + 1}, 3, 1)",
                            f"=_xlfn.UNIQUE(_xlfn.FILTER('Investor Exit List - {project}'!$AB:$AB,('Investor Exit List - {project}'!$I:$I=FALSE)*('Investor Exit List - {project}'!$AE:$AE=1)*('Investor Exit List - {project}'!$U:$U=FALSE)*('Investor Exit List - {project}'!$F:$F<>0)))",
                            f"=SUMIFS('Investor Exit List - {project}'!$AC:$AC,'Investor Exit List - {project}'!$AB:$AB,B{ws6.max_row + 1})",
                            f"=SUMIFS('Investor Exit List - {project}'!$Q:$Q,'Investor Exit List - {project}'!$AB:$AB,B{ws6.max_row + 1})",
                            f"=LEFT(B{ws6.max_row + 1},6)"])

            ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"B{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"C{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"E{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"F{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"G{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)

            calculate_exit_dates_start = ws6.max_row
            ws6[
                f"G{calculate_exit_dates_start}"].value = f"=_xlfn.UNIQUE(_xlfn.FILTER(E{calculate_exit_dates_start}:E{calculate_exit_dates_start + 800},(F{calculate_exit_dates_start}:F{calculate_exit_dates_start + 800}=0)))"
            print("calculate_exit_dates_start", calculate_exit_dates_start)
            ws6[
                f"F{calculate_exit_dates_start}"].value = f"=IF(SUMIFS($E${refinanced_units_start + 1}:$E${refinanced_units_end},$D${refinanced_units_start + 1}:$D${refinanced_units_end},E{ws6.max_row + 1})<>0,E{ws6.max_row + 1},0)"
            for i in range(350):
                ws6.append([f"=mid(E{ws6.max_row + 1}, 3, 1)",
                            "",
                            f"=SUMIFS('Investor Exit List'!$AC:$AC,'Investor Exit List'!$AB:$AB,B{ws6.max_row + 1})",
                            f"=SUMIFS('Investor Exit List'!$Q:$Q,'Investor Exit List'!$AB:$AB,B{ws6.max_row + 1})",
                            f"=LEFT(B{ws6.max_row + 1},6)",
                            f"=IF(SUMIFS($E${refinanced_units_start + 1}:$E${refinanced_units_end},$D${refinanced_units_start + 1}:$D${refinanced_units_end},E{ws6.max_row + 1})<>0,E{ws6.max_row + 1},0)"])
                # make font bold and 22
                ws6[f"A{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"B{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"C{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"D{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"E{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"F{ws6.max_row}"].font = Font(bold=True, color="0C0C0C", size=22)
                # format C as date and E as currency
                ws6[f"C{ws6.max_row}"].number_format = 'dd-mmm-yy'
                ws6[f"E{ws6.max_row}"].number_format = 'R#,##0'
            calculate_exit_dates_end = ws6.max_row
            print("calculate_exit_dates_end", calculate_exit_dates_end)

            for row in ws6.iter_rows(min_row=refinanced_units_start + 1, max_row=refinanced_units_end, min_col=3,
                                     max_col=3):
                for cell in row:
                    cell.border = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))
                    cell.font = Font(bold=True, color="0C0C0C", size=22)
                    # format as date
                    cell.number_format = 'dd-mm-yyyy'
                    # print("row", cell.row)
                    value = ""
                    "=IF(H117<>"",I$5,IF(J117<>"",K$5,IF(L117<>"",M$5,IF(N117<>"",O$5,IF(P117<>"",Q$5,IF(R117<>"",S$5,IF(T117<>"",U$5,IF(V117<>"",W$5,IF(X117<>"",Y$5,IF(Z117<>"",AA$5,IF(AB117<>"",AC$5,IF(AD117<>"",AE$5,""))))))))))))"
                    cell.value = f"=IF(H{cell.row}<>\"\",I$5,IF(J{cell.row}<>\"\",K$5,IF(L{cell.row}<>\"\",M$5,IF(N{cell.row}<>\"\",O$5,IF(P{cell.row}<>\"\",Q$5,IF(R{cell.row}<>\"\",S$5,IF(T{cell.row}<>\"\",U$5,IF(V{cell.row}<>\"\",W$5,IF(X{cell.row}<>\"\",Y$5,IF(Z{cell.row}<>\"\",AA$5,IF(AB{cell.row}<>\"\",AC$5,IF(AD{cell.row}<>\"\",AE$5,\"\"))))))))))))"
                    # cell.value = "XX"



            ws3["W4"].value = "New Forecast Date"
            ws3["X4"].value = "New VAT Date"
            ws3["Y4"].value = "VAT Recovery Date"
            ws3["Z4"].value = "Sales Toggle"

            for i in range(5, ws3.max_row + 1):
                ws3[f"Z{i}"].value = False
                if project == "Consolidated":
                    ws3[
                        f"F{i}"].value = f"=IF(OR(Z{i}=TRUE,SUMIFS('Cashflow Projection'!$E${refinanced_units_start + 1}:$E${refinanced_units_end},'Cashflow Projection'!$D${refinanced_units_start + 1}:$D${refinanced_units_end},Sales!C{i})=1),0,SUMIFS('Cashflow Projection'!$C${toggles_start}:$C${toggles_end},'Cashflow Projection'!$B${toggles_start}:$B${toggles_end},Sales!B{i},'Cashflow Projection'!$A${toggles_start}:$A${toggles_end},Sales!A{i}))"
                    ws3[
                        f"W{i}"].value = f"=IF(SUMIFS('Cashflow Projection'!$E${refinanced_units_start + 1}:$E${refinanced_units_end},'Cashflow Projection'!$D${refinanced_units_start + 1}:$D${refinanced_units_end},Sales!C{i})=1,SUMIFS('Cashflow Projection'!$C${refinanced_units_start + 1}:$C${refinanced_units_end},'Cashflow Projection'!$D${refinanced_units_start + 1}:$D${refinanced_units_end},Sales!C{i}),+Sales!H{i})"
                    ws3[
                        f'T{i}'].value = f"=IF(SUMIFS('Cashflow Projection'!$E${refinanced_units_start + 1}:$E${refinanced_units_end},'Cashflow Projection'!$D${refinanced_units_start + 1}:$D${refinanced_units_end},Sales!C{i})<>0,TRUE,FALSE)"
                    ws3[f"J{i}"].value = f"=IF(A{i}=\"Goodwood\",0,I{i}/115*15)"
                    ws3[f"K{i}"].value = f"=I{i}-J{i}"
                # format as date
                ws3[f"W{i}"].number_format = 'dd-mm-yyyy'
                "=IF(MOD(MONTH(W24), 2) <> 0, EOMONTH(W24, 2), EOMONTH(W24, 1))"
                ws3[f"X{i}"].value = f"=IF(MOD(MONTH(W{i}), 2) <> 0, EOMONTH(W{i}, 2), EOMONTH(W{i}, 1))"
                ws3[f"X{i}"].number_format = 'dd-mm-yyyy'
                "=EOMONTH(X24,1)"
                ws3[f"Y{i}"].value = f"=EOMONTH(X{i},1)"
                ws3[f"Y{i}"].number_format = 'dd-mm-yyyy'
                "=K5-SUM(L5:P5)"
                ws3[f"Q{i}"].value = f"=K{i}-SUM(L{i}:P{i})"
                ws3[f"Q{i}"].number_format = 'R#,##0'



            sales_columns = []

            variables_array = [vat_construction, operating_expenses, monthly, running, investor_exited]

            for variable in variables_array:
                ws6[f"D{variable}"].border = Border(left=Side(style='medium'),
                                                    right=Side(style='medium'),
                                                    top=Side(style='medium'),
                                                    bottom=Side(style='medium'))

                ws6[f"D{variable}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"D{variable}"].number_format = 'R#,##0'

                ws6.conditional_formatting.add(f"D{variable}",
                                               formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                          fill=PatternFill(
                                                                              start_color="FFC7CE",
                                                                              end_color="FFC7CE",
                                                                              fill_type="solid")))
                ws6.conditional_formatting.add(f"D{variable}",
                                               formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                          fill=PatternFill(
                                                                              start_color="BFEA7C",
                                                                              end_color="BFEA7C",
                                                                              fill_type="solid")))
                ws6.conditional_formatting.add(f"D{variable}",
                                               formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                          fill=PatternFill(
                                                                              start_color="FFF67E",
                                                                              end_color="FFF67E",
                                                                              fill_type="solid")))

            ws6[f"D{toggles_start - 1}"] = f"=SUM(D{toggles_start}:D{toggles_end})"
            ws6[f"D{toggles_start - 1}"].number_format = 'R#,##0'
            ws6[f"D{toggles_start - 1}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"E{toggles_start - 1}"] = f"=SUM(E{toggles_start}:E{toggles_end})"
            # center the text in E
            ws6[f"E{toggles_start - 1}"].alignment = Alignment(horizontal='center', vertical='center')
            ws6[f"E{toggles_start - 1}"].number_format = 'R#,##0'
            ws6[f"E{toggles_start - 1}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{toggles_start - 1}"].fill = PatternFill(start_color="7F9F80", end_color="7F9F80", fill_type="solid")
            ws6[f"D{toggles_start - 1}"].border = Border(left=Side(style='medium'),
                                                         right=Side(style='medium'),
                                                         top=Side(style='medium'),
                                                         bottom=Side(style='medium'))

            ws6[f"E{toggles_start - 1}"].border = Border(left=Side(style='medium'),
                                                         right=Side(style='medium'),
                                                         top=Side(style='medium'),
                                                         bottom=Side(style='medium'))

            ws6[f"D{block_costs_start - 1}"] = f"=SUM(D{block_costs_start}:D{block_costs_end})"
            ws6[f"D{block_costs_start - 1}"].font = Font(bold=True, color="0C0C0C", size=22)
            ws6[f"D{block_costs_start - 1}"].number_format = 'R#,##0'
            ws6[f"D{block_costs_start - 1}"].fill = PatternFill(start_color="7F9F80", end_color="7F9F80",
                                                                fill_type="solid")
            ws6[f"D{block_costs_start - 1}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))

            ws6[f"B{refinanced_units_start}"] = "24 Months"
            ws6[f"B{refinanced_units_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            # fill in blue
            ws6[f"B{refinanced_units_start}"].fill = PatternFill(start_color="3FA2F6", end_color="3FA2F6",
                                                                 fill_type="solid")
            # put a border around the cell
            ws6[f"B{refinanced_units_start}"].border = Border(left=Side(style='medium'),
                                                              right=Side(style='medium'),
                                                              top=Side(style='medium'),
                                                              bottom=Side(style='medium'))

            ws6[f"C{refinanced_units_start}"] = "EXIT"
            ws6[f"C{refinanced_units_start}"].font = Font(bold=True, color="0C0C0C", size=22)
            # fill in blue
            ws6[f"C{refinanced_units_start}"].fill = PatternFill(start_color="3FA2F6", end_color="3FA2F6",
                                                                 fill_type="solid")
            # put a border around the cell
            ws6[f"C{refinanced_units_start}"].border = Border(left=Side(style='medium'),
                                                              right=Side(style='medium'),
                                                              top=Side(style='medium'),
                                                              bottom=Side(style='medium'))

            for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                ws6[f"B{i}"] = f"=_xlfn.MINIFS('Investor Exit List'!$L:$L,'Investor Exit List'!$C:$C,D{i})"
                ws6[f"B{i}"].number_format = 'dd-mm-yyyy'
                # put a border around the cell
                ws6[f"B{i}"].border = Border(left=Side(style='medium'),
                                             right=Side(style='medium'),
                                             top=Side(style='medium'),
                                             bottom=Side(style='medium'))
                # make the font bold and 18
                ws6[f"B{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                # if index % 2 != 0:
                #     ws6[f"{col}{i}"].fill = PatternFill(start_color="FFFF80", end_color="FFFF80", fill_type="solid")

            for i in range(toggles_start, toggles_end + 1):
                formula_start = "="
                count_formula_start = "="
                for index, col in enumerate(month_headings):
                    if index % 2 == 0:
                        formula_start += f"+{col}{i}"
                    else:
                        count_formula_start += f"+{col}{i}"

                ws6[f"D{i}"] = f"{formula_start}"
                ws6[f"D{i}"].number_format = 'R#,##0'
                ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                ws6[f"E{i}"] = f"{count_formula_start}"
                ws6[f"E{i}"].number_format = 'R#,##0'
                ws6[f"E{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                # center the text in E
                ws6[f"E{i}"].alignment = Alignment(horizontal='center', vertical='center')

            for i in range(vat_row, vat_row + 1):
                formula_start_vat = "="
                for index, col in enumerate(month_headings):
                    formula_start = "="
                    if index % 2 == 0:
                        formula_start_vat += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start_vat}"

            for i in range(vat_income, vat_income + 1):
                formula_start_vat = "="
                for index, col in enumerate(month_headings):
                    formula_start = "="
                    if index % 2 == 0:
                        formula_start_vat += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start_vat}"

            for i in range(profit_on_sale, profit_on_sale + 1):
                formula_start_vat = "="
                for index, col in enumerate(month_headings):
                    formula_start = "="
                    if index % 2 == 0:
                        formula_start_vat += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start_vat}"

            for i in range(block_costs_start, block_costs_end + 1):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(vat_construction, vat_construction + 1):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(vat_payable_on_sales, vat_payable_on_sales + 1):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(CPC_INVOICES_STILL_DUE, CPC_INVOICES_STILL_DUE + 2):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # put a border around both cells
                        ws6[f"D{i}"].border = Border(left=Side(style='medium'),
                                                     right=Side(style='medium'),
                                                     top=Side(style='medium'),
                                                     bottom=Side(style='medium'))

                        # make the cell yellow if row is CPC_INVOICES_STILL_DUE and blue if row is CPC_INVOICES_STILL_DUE + 1
                        if i == CPC_INVOICES_STILL_DUE:
                            ws6[f"D{i}"].fill = PatternFill(start_color="FFF67E", end_color="FFF67E", fill_type="solid")
                        else:
                            ws6[f"D{i}"].fill = PatternFill(start_color="7F9F80", end_color="7F9F80", fill_type="solid")

            for i in range(operating_expenses, operating_expenses + 1):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            #

            for i in range(funds_available_start + 1, funds_available_end + 1):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(monthly, monthly + 1):
                formula_start_block = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start_block += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start_block}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(investor_exited, investor_exited + 1):
                formula_start = "="
                # count_formula_start = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start += f"+{col}{i}"
                    # else:
                    #     count_formula_start += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start}"
                    ws6[f"D{i}"].number_format = 'R#,##0'
                    ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                    # ws6[f"E{i}"] = f"{count_formula_start}"
                    # ws6[f"E{i}"].number_format = '0'
                    # ws6[f"E{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(rollover, rollover + 1):
                formula_start = "="
                # count_formula_start = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start}"
                    ws6[f"D{i}"].number_format = 'R#,##0'
                    ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(roll_over_refinance, roll_over_refinance + 1):
                formula_start = "="
                # count_formula_start = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start}"
                    ws6[f"D{i}"].number_format = 'R#,##0'
                    ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"E{i}"].value = 0.6
                    ws6[f"E{i}"].number_format = '0%'
                    ws6[f"E{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                    # fill the cell in yellow
                    ws6[f"E{i}"].fill = PatternFill(start_color="FFF67E", end_color="FFF67E", fill_type="solid")
                    # apply borders to the cell
                    ws6[f"E{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))
                    ws6[f"D{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))

            for i in range(roll_over_reinvest, roll_over_reinvest + 1):
                formula_start = "="
                # count_formula_start = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start}"
                    ws6[f"D{i}"].number_format = 'R#,##0'
                    ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"D{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))
                    ws6[f"E{i}"] = f"=E{roll_over_refinance}"
                    ws6[f"E{i}"].number_format = '0%'
                    ws6[f"E{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"E{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))

            for i in range(purple_blok_start, purple_blok_end + 1):
                formula_start = "="
                # count_formula_start = "="
                for index, col in enumerate(month_headings):

                    if index % 2 == 0:
                        formula_start += f"+{col}{i}"

                    ws6[f"D{i}"] = f"{formula_start}"
                    ws6[f"D{i}"].number_format = 'R#,##0'
                    ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"D{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))

                for i in range(finance_waterfall_start + 1, finance_waterfall_end + 1):
                    formula_start = "="
                    # count_formula_start = "="
                    for index, col in enumerate(month_headings):

                        if index % 2 == 0:
                            formula_start += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"D{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

                for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                    formula_start = "="
                    # count_formula_start = "="
                    for index, col in enumerate(month_headings):

                        if index % 2 == 0:
                            formula_start += f"+{col}{i}"

                        ws6[f"F{i}"] = f"{formula_start}"
                        ws6[f"F{i}"].number_format = 'R#,##0'
                        ws6[f"F{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # fill in light green
                        ws6[f"F{i}"].fill = PatternFill(start_color="B5C18E", end_color="B5C18E", fill_type="solid")
                        ws6[f"F{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

                for i in range(refinanced_units_start - 1, refinanced_units_start):
                    formula_start = "="
                    count_formula_start = f"=+H{i}"
                    for index, col in enumerate(month_headings):

                        if index % 2 == 0:
                            formula_start += f"+{col}{i}"
                        else:
                            count_formula_start += f"+{col}{i}"

                        ws6[f"F{i}"] = f"{formula_start}"
                        ws6[f"F{i}"].number_format = 'R#,##0'
                        ws6[f"F{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # fill in light green
                        ws6[f"F{i}"].fill = PatternFill(start_color="B5C18E", end_color="B5C18E", fill_type="solid")
                        ws6[f"F{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

                        ws6[f"E{i}"] = f"{count_formula_start}"
                        ws6[f"E{i}"].number_format = '0'
                        ws6[f"E{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # fill in light green
                        ws6[f"E{i}"].fill = PatternFill(start_color="B5C18E", end_color="B5C18E", fill_type="solid")
                        ws6[f"E{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

                for i in range(debenture_transaction_start, debenture_transaction_start + 1):
                    formula_start = "="
                    # count_formula_start = "="
                    for index, col in enumerate(month_headings):

                        if index % 2 == 0:
                            formula_start += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="FFFFFF", size=22)
                        # fill in light green
                        ws6[f"D{i}"].fill = PatternFill(start_color="3572EF", end_color="3572EF", fill_type="solid")
                        ws6[f"D{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

                for i in range(roll_over_refinance_properties_start, roll_over_refinance_properties_start + 3):
                    formula_start = "="
                    # count_formula_start = "="
                    for index, col in enumerate(month_headings):

                        if index % 2 == 0:
                            formula_start += f"+{col}{i}"

                        ws6[f"D{i}"] = f"{formula_start}"
                        ws6[f"D{i}"].number_format = 'R#,##0'
                        ws6[f"D{i}"].font = Font(bold=True, color="FFFFFF", size=22)
                        # fill in light green
                        ws6[f"D{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws6[f"D{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                           right=Side(style='medium'),
                                                                           top=Side(style='medium'),
                                                                           bottom=Side(style='medium'))

            for i in range(running, running + 1):
                ws6[f"D{i}"] = f"=SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, $B$2, Xero_bank!$D:$D, \"84*\")+8823977"
                ws6[f"D{i}"].number_format = 'R#,##0'
                ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)
            # "=IF(SUMIFS($B$50:$B$64,$E$50:$E$64,"Goodwood")>0,Cashflow!B14,'Cashflow - Heron'!B14)"
            # ws6[f"D{funds_available_start + 3}"] = "=Cashflow!B14"
            if project == "Consolidated":
                ws6[
                    f"D{funds_available_start + 3}"] = f"=IF(SUMIFS($B${block_costs_start}:$B${block_costs_end},$E${block_costs_start}:$E${block_costs_end},\"Goodwood\")>0,Cashflow!B14,'Cashflow - Heron'!B14)"
            else:
                ws6[
                    f"D{funds_available_start + 3}"] = f"='Cashflow - {project}'!B14"
            for index, col in enumerate(month_headings):

                ws6[f"{col}{toggles_start - 1}"] = f"=SUM({col}{toggles_start}:{col}{toggles_end})"
                ws6[f"{col}{toggles_start - 1}"].number_format = 'R#,##0'
                ws6[f"D{toggles_start - 1}"].fill = PatternFill(start_color="D4E7C5", end_color="D4E7C5",
                                                                fill_type="solid")
                ws6[f"{col}{toggles_start - 1}"].font = Font(bold=True, color="0C0C0C", size=22)

                # apply borders to the cell
                ws6[f"{col}{toggles_start - 1}"].border = ws6[f"A{toggles_start - 1}"].border + Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium'))

                ws6[f"{col}{toggles_start - 1}"].alignment = Alignment(horizontal='center', vertical='center')

                ws6[f"{col}{roll_over_refinance - 1}"] = f"=SUM({col}{roll_over_refinance}:{col}{block_finance_end})"
                ws6[f"{col}{roll_over_refinance - 1}"].number_format = 'R#,##0'
                ws6[f"{col}{roll_over_refinance - 1}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                          fill_type="solid")
                ws6[f"{col}{roll_over_refinance - 1}"].font = Font(bold=True, color="0C0C0C", size=22)

                ws6[f"{col}{roll_over_refinance - 1}"].border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium'))

                for i in range(finance_waterfall_start + 1, finance_waterfall_end + 1):
                    ws6[f"{col}{i}"].number_format = 'R#,##0'
                    ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                    # apply borders around the cell
                    ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))

                    ws6[f"H{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))

                    # fill the cell in yellow
                    # print("month heading index:", index, month_headings[index])
                    if index % 2 == 0:
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFF67E", end_color="FFF67E", fill_type="solid")
                    else:
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="B5C18E", end_color="B5C18E", fill_type="solid")
                    ws6[f"H{i}"].fill = PatternFill(start_color="B5C18E", end_color="B5C18E", fill_type="solid")

                for i in range(finance_waterfall_start + 1, finance_waterfall_start + 2):

                    # ws6[f"{col}{i}"].fill = PatternFill(start_color="FFF455", end_color="FFF455", fill_type="solid")
                    # ws6[f"H{i}"].fill = PatternFill(start_color="FFF455", end_color="FFF455", fill_type="solid")

                    try:
                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="94FFD8",
                                                                                      end_color="94FFD8",
                                                                                      fill_type="solid")))

                        ws6[f"H{i}"].number_format = 'R#,##0'
                        ws6[f"H{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"H{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))
                    except:
                        pass

                for i in range(finance_waterfall_start + 12, finance_waterfall_start + 13):
                    try:
                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="94FFD8",
                                                                                      end_color="94FFD8",
                                                                                      fill_type="solid")))

                        ws6[f"H{i}"].number_format = 'R#,##0'
                        ws6[f"H{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"H{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="94FFD8",
                                                                                      end_color="94FFD8",
                                                                                      fill_type="solid")))
                    except:
                        pass

                for i in range(finance_waterfall_end, finance_waterfall_end + 1):
                    try:
                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{month_headings[index - 1]}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="94FFD8",
                                                                                      end_color="94FFD8",
                                                                                      fill_type="solid")))

                        ws6[f"H{i}"].number_format = 'R#,##0'
                        ws6[f"H{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"H{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"H{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="94FFD8",
                                                                                      end_color="94FFD8",
                                                                                      fill_type="solid")))
                    except:
                        pass

                    ws6[
                        f"H{finance_waterfall_start + 1}"] = f"=sum(H{finance_waterfall_start + 2}:H{finance_waterfall_start + 8})"
                    ws6[f"H{finance_waterfall_start + 1}"].number_format = 'R#,##0'
                    ws6[f"H{finance_waterfall_start + 1}"].font = Font(bold=True, color="0C0C0C", size=22)

                    ws6[
                        f"{col}{finance_waterfall_start + 1}"] = f"=sum({col}{finance_waterfall_start + 2}:{col}{finance_waterfall_start + 8})"
                    ws6[f"{col}{finance_waterfall_start + 1}"].number_format = 'R#,##0'
                    ws6[f"{col}{finance_waterfall_start + 1}"].font = Font(bold=True, color="0C0C0C", size=22)

                if index % 2 == 0:
                    # for i in range(finance_waterfall_start + 2, finance_waterfall_start + 3):
                    if index == 0:
                        if project == "Consolidated":
                            ws6[f"H{finance_waterfall_start + 2}"] = 13500000
                        else:
                            ws6[f"H{finance_waterfall_start + 2}"] = 0
                        ws6[f"H{funds_available_start + 2}"].number_format = 'R#,##0'

                        ws6[f"{col}{finance_waterfall_start + 2}"] = f"=H{finance_waterfall_start + 2}"
                        ws6[f"{col}{finance_waterfall_start + 2}"].number_format = 'R#,##0'
                        if project == "Consolidated":
                            ws6[f"H{finance_waterfall_start + 3}"] = f"=Cashflow!B16"
                        else:
                            ws6[f"H{finance_waterfall_start + 3}"] = f"='Cashflow - {project}'!B16"
                        ws6[f"H{finance_waterfall_start + 3}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 3}"] = f"=H{finance_waterfall_start + 3}"
                        ws6[f"{col}{finance_waterfall_start + 3}"].number_format = 'R#,##0'

                        ws6[f"H{finance_waterfall_start + 4}"] = f"={col}{debenture_transaction_start}"
                        ws6[f"H{finance_waterfall_start + 4}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 4}"] = f"=H{finance_waterfall_start + 4}"
                        ws6[f"{col}{finance_waterfall_start + 4}"].number_format = 'R#,##0'

                        # ws6[f"H{finance_waterfall_start + 5}"] = f"={col}{funds_available_start + 6}"
                        ws6[f"H{finance_waterfall_start + 5}"] = 0
                        ws6[f"H{finance_waterfall_start + 5}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 5}"] = f"=H{finance_waterfall_start + 5}"
                        ws6[f"{col}{finance_waterfall_start + 5}"].number_format = 'R#,##0'

                        ws6[f"H{finance_waterfall_start + 6}"] = 0
                        ws6[f"H{finance_waterfall_start + 6}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 6}"] = 0
                        ws6[f"{col}{finance_waterfall_start + 6}"].number_format = 'R#,##0'

                        ws6[f"H{finance_waterfall_start + 7}"] = f"=-{col}{roll_over_refinance_properties_start + 1}"
                        ws6[f"H{finance_waterfall_start + 7}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 7}"] = f"=H{finance_waterfall_start + 7}"
                        ws6[f"{col}{finance_waterfall_start + 7}"].number_format = 'R#,##0'
                        if project == "Consolidated":
                            ws6[f"H{finance_waterfall_start + 8}"] = f"=Cashflow!B14+Cashflow!B13"
                        else:
                            ws6[
                                f"H{finance_waterfall_start + 8}"] = f"='Cashflow - {project}'!B14+'Cashflow - {project}'!B13"
                        ws6[f"H{finance_waterfall_start + 8}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 8}"] = f"=H{finance_waterfall_start + 8}"
                        ws6[f"{col}{finance_waterfall_start + 8}"].number_format = 'R#,##0'

                        ws6[
                            f"H{finance_waterfall_start + 9}"] = f"=SUMIFS({col}{toggles_start}:{col}{toggles_end},$A${toggles_start}:$A${toggles_end},\"<>\"&\"Endulini\")"
                        ws6[f"H{finance_waterfall_start + 9}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 9}"] = f"=H{finance_waterfall_start + 9}"
                        ws6[f"{col}{finance_waterfall_start + 9}"].number_format = 'R#,##0'

                        ws6[f"{col}{CPC_INVOICES_STILL_DUE}"] = 0
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE + 9}"].number_format = 'R#,##0'
                        # put borders around and fill in yellow
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE}"].fill = PatternFill(start_color="FFF67E",
                                                                                 end_color="FFF67E",
                                                                                 fill_type="solid")
                        # make the font size 22
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE}"].font = Font(bold=True, color="0C0C0C", size=22)

                        ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"] = f"={col}{CPC_INVOICES_STILL_DUE} / 115 * 15"
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].number_format = 'R#,##0'
                        # put borders around and fill in blue
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].fill = PatternFill(start_color="3572EF",
                                                                                     end_color="3572EF",
                                                                                     fill_type="solid")
                        # make the font size 22
                        ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].font = Font(bold=True, color="0C0C0C", size=22)

                        ws6[
                            f"H{finance_waterfall_start + 10}"] = f"={col}{funds_available_start + 7}"
                        ws6[f"H{finance_waterfall_start + 10}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 10}"] = f"=H{finance_waterfall_start + 10}"
                        ws6[f"{col}{finance_waterfall_start + 10}"].number_format = 'R#,##0'

                        ws6[
                            f"H{finance_waterfall_start + 11}"] = f"={col}{vat_income}+{col}{purple_blok_start + 1}+{col}{block_costs_end + 4}"
                        ws6[f"H{finance_waterfall_start + 11}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 11}"] = f"=H{finance_waterfall_start + 11}"
                        ws6[f"{col}{finance_waterfall_start + 11}"].number_format = 'R#,##0'

                        ws6[
                            f"H{finance_waterfall_start + 12}"] = f"={col}{finance_waterfall_start + 12}-{col}{block_costs_start - 4}"
                        ws6[f"H{finance_waterfall_start + 12}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 12}"] = f"=sum({col}{finance_waterfall_start + 2}:{col}{finance_waterfall_start + 11})"
                        ws6[f"{col}{finance_waterfall_start + 12}"].number_format = 'R#,##0'

                        ws6[f"{col}{finance_waterfall_start + 13}"] = f"={col}{running - 2}"
                        ws6[f"{col}{finance_waterfall_start + 13}"].number_format = 'R#,##0'

                        ws6[f"H{finance_waterfall_start + 15}"] = f"={col}{finance_waterfall_start + 15}-{col}{running}"
                        ws6[f"H{finance_waterfall_start + 15}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 15}"] = f"={col}{finance_waterfall_start + 12}+{col}{finance_waterfall_start + 13}"
                        ws6[f"{col}{finance_waterfall_start + 15}"].number_format = 'R#,##0'


                    else:

                        # "=SUMIFS(K7:K23,$A$7:$A$23,"<>"&"Heron View")"
                        "=H102-I102"
                        if index < 3:

                            ws6[
                                f"{month_headings[index - 1]}{finance_waterfall_start + 2}"] = f"=H{finance_waterfall_start + 2}-{month_headings[index - 2]}{finance_waterfall_start + 2}"
                            ws6[f"{month_headings[index - 1]}{funds_available_start + 2}"].number_format = 'R#,##0'
                        else:
                            ws6[
                                f"{month_headings[index - 1]}{finance_waterfall_start + 2}"] = f"={month_headings[index - 3]}{finance_waterfall_start + 2}-{month_headings[index - 2]}{finance_waterfall_start + 2}"
                            ws6[f"{month_headings[index - 1]}{funds_available_start + 2}"].number_format = 'R#,##0'

                        ws6[
                            f"{col}{finance_waterfall_start + 2}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 2}"
                        ws6[f"{col}{finance_waterfall_start + 2}"].number_format = 'R#,##0'

                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 3}"] = 0
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 3}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 3}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 3}"
                        ws6[f"{col}{finance_waterfall_start + 3}"].number_format = 'R#,##0'

                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 4}"] = 0
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 4}"].number_format = 'R#,##0'
                        # ws6[f"{col}{finance_waterfall_start + 4}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 4}"
                        ws6[
                            f"{col}{finance_waterfall_start + 4}"] = f"={col}{debenture_transaction_start}"
                        ws6[f"{col}{finance_waterfall_start + 4}"].number_format = 'R#,##0'

                        # ws6[
                        #     f"{month_headings[index - 1]}{finance_waterfall_start + 5}"] = f"={col}{funds_available_start + 6}"
                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 5}"] = 0
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 5}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 5}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 5}"
                        ws6[f"{col}{finance_waterfall_start + 5}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 6}"] = f"={col}{funds_available_start + 4}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 6}"].number_format = 'R#,##0'
                        ws6[f"{col}{finance_waterfall_start + 6}"] = 0
                        ws6[f"{col}{finance_waterfall_start + 6}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 7}"] = f"=-{col}{roll_over_refinance_properties_start + 1}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 7}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 7}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 7}"
                        ws6[f"{col}{finance_waterfall_start + 7}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 8}"] = f"={month_headings[index - 2]}{finance_waterfall_start + 15}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 8}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 8}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 8}"
                        ws6[f"{col}{finance_waterfall_start + 8}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 9}"] = f"=SUMIFS({col}{toggles_start}:{col}{toggles_end},$A${toggles_start}:$A${toggles_end},\"<>\"&\"Endulini\")"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 9}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 9}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 9}"
                        ws6[f"{col}{finance_waterfall_start + 9}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 10}"] = f"={col}{funds_available_start + 7}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 10}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 10}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 10}"
                        ws6[f"{col}{finance_waterfall_start + 10}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 11}"] = f"={col}{vat_income}+{col}{purple_blok_start + 1}+{col}{block_costs_end + 4}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 11}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 11}"] = f"={month_headings[index - 1]}{finance_waterfall_start + 11}"
                        ws6[f"{col}{finance_waterfall_start + 11}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 12}"] = f"={col}{finance_waterfall_start + 12}-{col}{block_costs_start - 4}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 12}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 12}"] = f"=sum({col}{finance_waterfall_start + 2}:{col}{finance_waterfall_start + 11})"
                        ws6[f"{col}{finance_waterfall_start + 12}"].number_format = 'R#,##0'

                        ws6[f"{col}{finance_waterfall_start + 13}"] = f"={col}{running - 2}"
                        ws6[f"{col}{finance_waterfall_start + 13}"].number_format = 'R#,##0'

                        ws6[
                            f"{month_headings[index - 1]}{finance_waterfall_start + 15}"] = f"={col}{finance_waterfall_start + 15}-{col}{running}"
                        ws6[f"{month_headings[index - 1]}{finance_waterfall_start + 15}"].number_format = 'R#,##0'
                        ws6[
                            f"{col}{finance_waterfall_start + 15}"] = f"={col}{finance_waterfall_start + 12}+{col}{finance_waterfall_start + 13}"
                        ws6[f"{col}{finance_waterfall_start + 15}"].number_format = 'R#,##0'

                    # "=SUMIFS(I7:I23,$A$7:$A$23,"Heron View")"
                    ws6[
                        f"{col}{funds_available_start + 1}"] = f"=SUMIFS({col}{toggles_start}:{col}{toggles_end},$A${toggles_start}:$A${toggles_end},\"<>\"&\"Endulini\")"
                    ws6[f"{col}{funds_available_start + 1}"].number_format = 'R#,##0'

                    "=I72"
                    try:
                        ws6[f"{col}{funds_available_start + 3}"] = f"={month_headings[index - 2]}{running}"
                    except:
                        continue

                    "=-I156"
                    ws6[f"{col}{funds_available_start + 4}"] = f"=-{col}{roll_over_refinance_properties_start + 2}"
                    ws6[f"{col}{funds_available_start + 4}"].number_format = 'R#,##0'
                    "=I155"
                    ws6[f"{col}{funds_available_start + 5}"] = f"=-{col}{roll_over_refinance_properties_start + 1}"
                    ws6[f"{col}{funds_available_start + 5}"].number_format = 'R#,##0'

                    "=I115"
                    # ws6[f"{col}{funds_available_start + 6}"] = f"={col}{refinanced_units_start - 1}"
                    ws6[f"{col}{funds_available_start + 6}"] = f"=0"
                    ws6[f"{col}{funds_available_start + 6}"].number_format = 'R#,##0'

                    "=I102"
                    ws6[
                        f"{col}{funds_available_start + 8}"] = f"={col}{finance_waterfall_start + 2}+{col}{debenture_transaction_start}"
                    ws6[f"{col}{funds_available_start + 8}"].number_format = 'R#,##0'
                    # "=SUM(I80:I90)+I95+I97+J79"
                    "=+I79+I95+I97+J79"

                    ws6[
                        f"{col}{funds_available_start + 9}"] = f"=+{col}{roll_over_refinance - 1}+{col}{purple_blok_start}+{col}{purple_blok_start + 2}+{month_headings[index + 1]}{roll_over_refinance - 1}"
                    ws6[f"{col}{funds_available_start + 9}"].number_format = 'R#,##0'
                    # "=I96+I103+I26"
                    "=I96+I103+I26+I66"

                    ws6[
                        f"{col}{funds_available_start + 10}"] = f"={col}{purple_blok_start + 1}+{col}{finance_waterfall_start + 3}+{col}{vat_income}+{col}{block_costs_end + 4}"
                    ws6[f"{col}{funds_available_start + 10}"].number_format = 'R#,##0'
                    "=+I35-I39+I43-I44-I102"
                    "=I35-I39-I42-I44"
                    "=SUM(I36: I45)-I44"
                    ws6[
                        f"{col}{funds_available_start + 11}"] = f"=sum({col}{funds_available_start + 1}:{col}{funds_available_start + 10})-{col}{funds_available_start + 9}-{col}{funds_available_start + 4}"
                    ws6[
                        f"D{funds_available_start + 11}"] = f"=sum(D{funds_available_start + 1}:D{funds_available_start + 10})-D{funds_available_start + 9}-D{funds_available_start + 4}"
                    # ws6[
                    #     f"{col}{funds_available_start + 11}"] = "XXX"
                    ws6[f"{col}{funds_available_start + 11}"].number_format = 'R#,##0'

                    ws6[
                        f"{col}{vat_construction}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,((SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$E:$E,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),'Updated Construction'!$E:$E,\">\"&EOMONTH(EDATE({col}$5, 0), -3),'Updated Construction'!$C:$C,1)/1.15*0.15)+(-{col}{vat_construction + 2}*1.4)/1.15*0.15))*C{vat_construction}"
                    ws6[f"{col}{vat_construction}"].number_format = 'R#,##0'

                    "=$E$80*-I115"
                    ws6[f"{col}{roll_over_refinance}"] = f"=$E${roll_over_refinance}*-{col}{refinanced_units_start - 1}"
                    ws6[f"{col}{roll_over_refinance}"].number_format = 'R#,##0'
                    # make fornt 22
                    ws6[f"{col}{roll_over_refinance}"].font = Font(bold=True, color="0C0C0C", size=22)
                    # place  border around
                    ws6[f"{col}{roll_over_refinance}"].border = Border(left=Side(style='medium'),
                                                                       right=Side(style='medium'),
                                                                       top=Side(style='medium'),
                                                                       bottom=Side(style='medium'))
                    "=$E$81*I154"
                    ws6[
                        f"{col}{roll_over_reinvest}"] = f"=$E${roll_over_reinvest}*{col}{roll_over_refinance_properties_start}"
                    ws6[f"{col}{roll_over_refinance + 1}"].number_format = 'R#,##0'
                    # make fornt 22
                    ws6[f"{col}{roll_over_reinvest}"].font = Font(bold=True, color="0C0C0C", size=22)
                    # place  border around
                    ws6[f"{col}{roll_over_reinvest}"].border = Border(left=Side(style='medium'),
                                                                      right=Side(style='medium'),
                                                                      top=Side(style='medium'),
                                                                      bottom=Side(style='medium'))

                    for i in range(block_finance_start, block_finance_end + 1):
                        if index == 0:

                            ws6[
                                f"{col}{i}"] = f"=SUMIFS(Sales!$R:$R,Sales!$W:$W,\"<=\"&{col}5,Sales!$W:$W,\">\"&B2,Sales!$E:$E,FALSE,Sales!$B:$B,$B{i},Sales!$A:$A,$C{i})*$E{i}"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                            ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))
                        else:

                            ws6[
                                f"{col}{i}"] = f"=SUMIFS(Sales!$R:$R,Sales!$W:$W,\"<=\"&{col}5,Sales!$W:$W,\">\"&{month_headings[index - 2]}5,Sales!$E:$E,FALSE,Sales!$B:$B,$B{i},Sales!$A:$A,$C{i})*$E{i}"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                            ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))

                    ws6[f"{col}{toggles_start - 1}"].fill = PatternFill(start_color="D4E7C5", end_color="D4E7C5",
                                                                        fill_type="solid")
                    # align the text to the right
                    ws6[f"{col}{toggles_start - 1}"].alignment = Alignment(horizontal='right', vertical='center')

                    ws6[f"{col}{block_costs_start - 1}"] = f"=SUM({col}{block_costs_start}:{col}{block_costs_end})"
                    ws6[f"{col}{block_costs_start - 1}"].number_format = 'R#,##0'
                    ws6[f"{col}{block_costs_start - 1}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                            fill_type="solid")

                    ws6[f"D{block_costs_start - 1}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                        fill_type="solid")

                    ws6[f"{col}{block_costs_start - 1}"].font = Font(bold=True, color="0C0C0C", size=22)

                    # center the text
                    ws6[f"{col}{block_costs_start - 1}"].alignment = Alignment(horizontal='center', vertical='center')

                    # apply borders to the cell
                    ws6[f"{col}{block_costs_start - 1}"].border = ws6[f"A{block_costs_start - 1}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))

                    for i in range(roll_over_refinance_properties_start + 1, roll_over_refinance_properties_start + 2):
                        "=SUM(I243:I593)"
                        ws6[f"{col}{i}"] = f"=SUM({col}{calculate_exit_dates_start}:{col}{calculate_exit_dates_end})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                         right=Side(style='medium'),
                                                         top=Side(style='medium'),
                                                         bottom=Side(style='medium'))

                    for i in range(calculate_exit_dates_start, calculate_exit_dates_end + 1):
                        if index == 0:

                            ws6[
                                f"{col}{i}"] = f"=IF(AND(SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0,$E${calculate_exit_dates_start}:$E${calculate_exit_dates_end},$G{i},$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\"<=\"&B$2)<1000000,SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0,$E${calculate_exit_dates_start}:$E${calculate_exit_dates_end},$G{i},$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\"<=\"&{col}5)>=1000000),SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0,$E${calculate_exit_dates_start}:$E${calculate_exit_dates_end},$G{i})*$C${roll_over_refinance_properties_start + 1} ,0)"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                            ws6[f"G{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                            ws6[f"G{i}"].border = Border(left=Side(style='medium'),
                                                         right=Side(style='medium'),
                                                         top=Side(style='medium'),
                                                         bottom=Side(style='medium'))
                            ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))
                        else:
                            ws6[
                                f"{col}{i}"] = f"=IF(AND(SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0,$E${calculate_exit_dates_start}:$E${calculate_exit_dates_end},$G{i},$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\"<=\"&{month_headings[index - 2]}$5)<1000000,SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0,$E${calculate_exit_dates_start}:$E${calculate_exit_dates_end},$G{i},$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\"<=\"&{col}5)>=1000000),SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0,$E${calculate_exit_dates_start}:$E${calculate_exit_dates_end},$G{i})*$C${roll_over_refinance_properties_start + 1} ,0)"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                            ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             bottom=Side(style='medium'))

                if index % 2 != 0:
                    if index == 1:
                        for i in range(block_finance_start, block_finance_end + 1):
                            ws6[
                                f"{col}{i}"] = f"=SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$A${calculate_exit_dates_start}:$A${calculate_exit_dates_end},B{i},$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\"<=\"&{month_headings[index - 1]}5,$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\">\"&B2,$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0)*$E{i}"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                            ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))
                    else:
                        for i in range(block_finance_start, block_finance_end + 1):
                            ws6[
                                f"{col}{i}"] = f"=SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$A${calculate_exit_dates_start}:$A${calculate_exit_dates_end},B{i},$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\"<=\"&{month_headings[index - 1]}5,$C${calculate_exit_dates_start}:$C${calculate_exit_dates_end},\">\"&{month_headings[index - 3]}5,$F${calculate_exit_dates_start}:$F${calculate_exit_dates_end},0)*$E{i}"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                            ws6[f"{col}{i}"].border = Border(left=Side(style='medium'),
                                                             right=Side(style='medium'),
                                                             top=Side(style='medium'),
                                                             bottom=Side(style='medium'))

                if index == 0:
                    if project == "Consolidated":
                        ws6[f"{col}{funds_available_start + 2}"] = f"=Cashflow!B13"

                        ws6[f"{col}{funds_available_start + 3}"] = f"=Cashflow!B14"

                        ws6[f"{col}{funds_available_start + 7}"] = f"=Cashflow!B15"
                    else:
                        ws6[f"{col}{funds_available_start + 2}"] = f"='Cashflow - {project}'!B13"

                        ws6[f"{col}{funds_available_start + 3}"] = f"='Cashflow - {project}'!B14"

                        ws6[f"{col}{funds_available_start + 7}"] = f"='Cashflow - {project}'!B15"
                    # I AM HERE
                    ws6[f"{col}{funds_available_start + 2}"].number_format = 'R#,##0'
                    ws6[f"{col}{funds_available_start + 3}"].number_format = 'R#,##0'
                    ws6[f"{col}{funds_available_start + 7}"].number_format = 'R#,##0'

                    # ws6[f"{col}{vat_income}"] = f"=SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W," \
                    #                             f"\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&$B$2)*C{vat_income}"

                    if project == "Consolidated":
                        ws6[f"{col}{vat_income}"] = f"=SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W," \
                                                    f"\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&$B$2)*C{vat_income}"
                    elif project == "Heron":
                        ws6[f"{col}{vat_income}"] = f"=(SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W," \
                                                    f"\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&$B$2,Sales!$A:$A,\"Heron View\")+SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W," \
                                                    f"\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&$B$2,Sales!$A:$A,\"Heron Fields\"))*C{vat_income}"
                    else:
                        ws6[f"{col}{vat_income}"] = f"=SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W," \
                                                    f"\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&$B$2,Sales!$A:$A,\"{project}\")*C{vat_income}"

                    ws6[f"{col}{vat_income}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"{col}{vat_income}"].border = ws6[f"A{vat_income}"].border + Border(left=Side(style='medium'),
                                                                                             right=Side(style='medium'),
                                                                                             top=Side(style='medium'),
                                                                                             bottom=Side(
                                                                                                 style='medium'))

                    ws6[f"{col}{vat_income}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                                 fill_type="solid")
                    ws6[f"{col}{vat_income}"].number_format = 'R#,##0'

                    ws6[
                        f"{col}{vat_row}"] = 0

                    ws6[f"{col}{vat_row}"].font = Font(bold=True, color="0C0C0C", size=22)

                    ws6[f"{col}{vat_row}"].border = ws6[f"A{vat_row}"].border + Border(left=Side(style='medium'),
                                                                                       right=Side(style='medium'),
                                                                                       top=Side(style='medium'),
                                                                                       bottom=Side(style='medium'))
                    # fill equals light green
                    ws6[f"{col}{vat_row}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                              fill_type="solid")

                    ws6[f"{col}{vat_row}"].number_format = 'R#,##0'

                    if project == "Consolidated":
                        ws6[
                            f"{col}{profit_on_sale}"] = (
                            f"=SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&$B$2)")
                    elif project == "Heron":
                        ws6[
                            f"{col}{profit_on_sale}"] = (
                            f"=SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&$B$2,Sales!$A:$A,\"Heron View\")+SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&$B$2,Sales!$A:$A,\"Heron Fields\")")
                    else:
                        ws6[
                            f"{col}{profit_on_sale}"] = (
                            f"=SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&$B$2,Sales!$A:$A,\"{project}\")")

                    ws6[f"{col}{profit_on_sale}"].border = ws6[f"A{profit_on_sale}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))
                    ws6[f"{col}{profit_on_sale}"].font = Font(bold=True, color="0C0C0C", size=22)

                    ws6[f"{col}{profit_on_sale}"].number_format = 'R#,##0'

                    ws6.conditional_formatting.add(f"{col}{profit_on_sale}",
                                                   formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFC7CE",
                                                                                  end_color="FFC7CE",
                                                                                  fill_type="solid")))

                    ws6.conditional_formatting.add(f"{col}{profit_on_sale}",
                                                   formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="BFEA7C",
                                                                                  end_color="BFEA7C",
                                                                                  fill_type="solid")))

                    ws6.conditional_formatting.add(f"{col}{profit_on_sale}",
                                                   formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFF67E",
                                                                                  end_color="FFF67E",
                                                                                  fill_type="solid")))

                    ws6[f"{col}{toggles_start - 2}"] = f"=EOMONTH(EDATE($B$2, 0), 1)"
                    ws6[f"{col}{toggles_start - 2}"].number_format = 'dd-mmm-yy'
                    ws6[f"{col}{toggles_start - 2}"].font = Font(bold=True, color="0C0C0C", size=22)

                    for i in range(vat_payable_on_sales, vat_payable_on_sales + 1):
                        # print("month_headings",month_headings[index])

                        if project == "Heron":
                            # I AM HERE - CHANGE REFERENCE 28 & 31
                            # "=(IF(MOD(MONTH(I5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,"<="&EOMONTH(EDATE(I$5,0),-1),Sales!$W:$W,">"&EOMONTH(EDATE(I$5,0),-3),Sales!$A:$A,"Heron View")-SUMIFS(Sales!$J:$J,Sales!$W:$W,"<="&EOMONTH(EDATE(I$5,0),-1),Sales!$W:$W,">"&EOMONTH(EDATE(I$5,0),-3),Sales!$A:$A,"Heron Fields")*$C25))"
                            ws6[
                                f"{col}{i}"] = f"=(IF(MOD(MONTH({col}5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"Heron View\")-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"Heron Fields\")*$C{vat_payable_on_sales}))"
                        elif project == "Consolidated":
                            ws6[
                                f"{col}{i}"] = f"=(IF(MOD(MONTH({col}5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3))-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3))*$C{vat_payable_on_sales}))"
                        else:
                            ws6[
                                f"{col}{i}"] = f"=(IF(MOD(MONTH({col}5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"{project}\")-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"{project}\")*$C{vat_payable_on_sales}))"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # fil in light red
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        ws6[f"{col}{i}"].border = ws6[f"A{vat_payable_on_sales}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(
                                style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(
                                style='medium'))

                    for i in range(toggles_start, toggles_end + 1):
                        if project == "Consolidated":
                            ws6[
                                f"{col}{i}"] = (f"=SUMIFS(Sales!$S:$S,Sales!$W:$W,\"<=\"&I$5,"
                                                f"Sales!$E:$E,FALSE,Sales!$F:$F,$C{i},Sales!$A:$A,"
                                                f"$A{i},Sales!$B:$B,$B{i},"
                                                f"Sales!$W:$W,\">\"&B$2)*$C{i}")
                        elif project == "Heron":
                            ws6[
                                f"{col}{i}"] = (f"=(SUMIFS(Sales!$S:$S,Sales!$W:$W,\"<=\"&I$5,"
                                                f"Sales!$E:$E,FALSE,Sales!$F:$F,$C{i},Sales!$A:$A,"
                                                f"$A{i},Sales!$B:$B,$B{i},"
                                                f"Sales!$W:$W,\">\"&B$2,Sales!$A:A,\"Heron View\")+SUMIFS(Sales!$S:$S,Sales!$W:$W,\"<=\"&I$5,"
                                                f"Sales!$E:$E,FALSE,Sales!$F:$F,$C{i},Sales!$A:$A,"
                                                f"$A{i},Sales!$B:$B,$B{i},"
                                                f"Sales!$W:$W,\">\"&B$2,Sales!$A:A,\"Heron Fields\"))*$C{i}")
                        else:
                            ws6[
                                f"{col}{i}"] = (f"=SUMIFS(Sales!$S:$S,Sales!$W:$W,\"<=\"&I$5,"
                                                f"Sales!$E:$E,FALSE,Sales!$F:$F,$C{i},Sales!$A:$A,"
                                                f"$A{i},Sales!$B:$B,$B{i},"
                                                f"Sales!$W:$W,\">\"&B$2, Sales!$A:$A,\"{project}\")*$C{i}")

                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                    for i in range(block_costs_start, block_costs_end + 1):
                        # I AM HERE

                        if project == "Goodwood":
                            ws6[
                                f"{col}{i}"] = (
                                f"=-(SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$A:$A,FALSE,"
                                f"'Updated Construction'!$G:$G,$A{i},"
                                f"'Updated Construction'!$E:$E,\"<=\"&"
                                f"{month_headings[index]}$5,'Updated Construction'!$E:$E,\">\"&$B$2)"
                                f"+(SUMIFS('Updated Construction'!$F:$F,"
                                f"'Updated Construction'!$A:$A,TRUE,'Updated Construction'!$G:$G,"
                                f"$A{i},'Updated Construction'!$E:$E,\"<=\"&"
                                f"{month_headings[index]}$5,'Updated Construction'!$E:$E,"
                                f"\">\"&$B$2)*$B{i}))")
                        else:

                            ws6[
                                f"{col}{i}"] = (
                                f"=-(SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$A:$A,FALSE,"
                                f"'Updated Construction'!$G:$G,$A{i},"
                                f"'Updated Construction'!$E:$E,\"<=\"&"
                                f"{month_headings[index]}$5,'Updated Construction'!$E:$E,\">\"&$B$2)"
                                f"+(SUMIFS('Updated Construction'!$F:$F,"
                                f"'Updated Construction'!$A:$A,TRUE,'Updated Construction'!$G:$G,"
                                f"$A{i},'Updated Construction'!$E:$E,\"<=\"&"
                                f"{month_headings[index]}$5,'Updated Construction'!$E:$E,"
                                f"\">\"&$B$2)*$B{i}))*1.15")

                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                    for i in range(vat_construction, vat_construction + 1):
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                            fill_type="solid")

                        ws6[f"{col}{i}"].border = ws6[f"A{vat_construction}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(operating_expenses, operating_expenses + 1):
                        ws6[
                            f"{col}{i}"] = f"=-'Operational Costs'!$N$2*$C{i}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                            fill_type="solid")

                        ws6[f"{col}{i}"].border = ws6[f"A{vat_construction}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(monthly, monthly + 1):
                        ws6[
                            f"{col}{i}"] = f"={col}{vat_row}+{col}{block_costs_start - 1}+{col}{vat_construction}+{col}{operating_expenses}+{col}{vat_payable_on_sales}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"{col}{i}"].border = ws6[f"A{vat_construction}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(running, running + 1):
                        "=D53+F51"
                        ws6[
                            f"{col}{i}"] = f"=+{col}{monthly}+{col}{funds_available_end}"
                        # ws6[
                        #     f"{col}{i}"] = f"=D{running}+{col}{monthly}+{col}{funds_available_end}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{running}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(investor_exited, investor_exited + 1):
                        ws6[
                            f"{col}{i}"] = f"={col}{funds_available_start + 4}+{col}{funds_available_start + 9}-{col}{investor_exited + 2}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(rollover, rollover + 1):
                        ws6[
                            f"{col}{i}"] = f"=0"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="3ABEF9", end_color="3ABEF9", fill_type="solid")

                    for i in range(project_income, project_income + 1):
                        ws6[
                            f"{col}{i}"] = f"SALES"
                        # ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="749BC2", end_color="749BC2", fill_type="solid")

                        for i in range(purple_blok_start, purple_blok_end):
                            ws6[
                                f"{col}{i}"] = f"=$B${i}*{month_headings[index + 1]}{i}"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                            # apply borders around the cell
                            ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                                left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))

                        for i in range(purple_blok_end, purple_blok_end + 1):
                            # "=SUBTOTAL(9,H95:H97)"
                            ws6[
                                f"{col}{i}"] = f"=SUM({col}{purple_blok_start}:{col}{purple_blok_end - 1})"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                            # apply borders around the cell
                            ws6[f"{col}{i}"].border = Border(
                                left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))

                            ws6.conditional_formatting.add(f"{col}{i}",
                                                           formatting.rule.CellIsRule(operator='lessThan',
                                                                                      formula=['0'],
                                                                                      fill=PatternFill(
                                                                                          start_color="FFC7CE",
                                                                                          end_color="FFC7CE",
                                                                                          fill_type="solid")))

                            ws6.conditional_formatting.add(f"{col}{i}",
                                                           formatting.rule.CellIsRule(operator='greaterThan',
                                                                                      formula=['0'],
                                                                                      fill=PatternFill(
                                                                                          start_color="BFEA7C",
                                                                                          end_color="BFEA7C",
                                                                                          fill_type="solid")))

                            ws6.conditional_formatting.add(f"{col}{i}",
                                                           formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                      fill=PatternFill(
                                                                                          start_color="FFF67E",
                                                                                          end_color="FFF67E",
                                                                                          fill_type="solid")))

                        for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                            if project == "Consolidated":
                                ws6[
                                    f"{col}{i}"] = f"=SUMIFS('Investor Exit List'!$Q:$Q, 'Investor Exit List'!$C:$C, H{i})"
                            else:
                                ws6[
                                    f"{col}{i}"] = f"=SUMIFS('Investor Exit List - {project}'!$Q:$Q, 'Investor Exit List - {project}'!$C:$C, H{i})"

                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                            # apply borders around the cell
                            ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                                left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))

                        for i in range(refinanced_units_start - 1, refinanced_units_start):
                            ws6[
                                f"{col}{i}"] = f"=-SUM({col}{refinanced_units_start}:{col}{refinanced_units_end})"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                            # apply borders around the cell
                            ws6[f"{col}{i}"].border = Border(
                                left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))

                        for i in range(debenture_transaction_start, debenture_transaction_start + 1):
                            "=$B$151*J151"
                            ws6[
                                f"{col}{i}"] = f"=$B{debenture_transaction_start}*{month_headings[index + 1]}{debenture_transaction_start}"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                            # apply borders around the cell
                            ws6[f"{col}{i}"].border = Border(
                                left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))

                            # fill in blue
                            ws6[f"{col}{i}"].fill = PatternFill(start_color="3572EF", end_color="3572EF",
                                                                fill_type="solid")

                    for i in range(roll_over_refinance_properties_start, roll_over_refinance_properties_start + 1):
                        ws6[
                            f"{col}{i}"] = f"=sum({col}{roll_over_refinance_properties_start + 3}:{col}{roll_over_refinance_properties_end})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        # fill in blue
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start + 2, roll_over_refinance_properties_start + 3):
                        ws6[
                            f"{col}{i}"] = f"={col}{roll_over_refinance_properties_start}-{col}{roll_over_refinance_properties_start + 1}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        # fill in blue
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start + 3, roll_over_refinance_properties_end):
                        ws6[
                            f"{col}{i}"].value = f"=SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$B${calculate_exit_dates_start}:$B${calculate_exit_dates_end},H{i})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="000000", size=22)
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))



                elif index == 1:

                    for i in range(toggles_start, toggles_end + 1):
                        ws6[
                            f"{col}{i}"] = (f"=COUNTIFS(Sales!$H:$H,\"<=\"&I$5,Sales!$H:$H,"
                                            f"\">\"&B$2,Sales!$A:$A,$A{i},"
                                            f"Sales!$B:$B,$B{i},Sales!$F:$F,'Cashflow "
                                            f"Projection'!$C{i},Sales!$E:$E,FALSE)*$C{i}")

                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # Center the text in the cell
                        ws6[f"{col}{i}"].alignment = Alignment(horizontal='center', vertical='center')

                    for i in range(project_income, project_income + 1):
                        ws6[
                            f"{col}{i}"] = f"EXIT & ROLL"
                        # ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="749BC2", end_color="749BC2", fill_type="solid")

                    for i in range(purple_blok_start, purple_blok_end):
                        ws6[
                            f"{col}{i}"] = 0
                        ws6[f"{col}{i}"].number_format = '0%'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        # fill in bright yellow
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFFF80", end_color="FFFF80", fill_type="solid")
                        ws6[f"H{i}"].fill = PatternFill(start_color="FFFF80", end_color="FFFF80", fill_type="solid")

                    for i in range(refinanced_units_start - 1, refinanced_units_start):
                        ws6[
                            f"{col}{i}"] = f"=COUNTA({col}{refinanced_units_start}:{col}{refinanced_units_end})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(debenture_transaction_start, debenture_transaction_start + 1):
                        ws6[
                            f"{col}{i}"] = f"=$H{refinanced_units_start - 1}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="3572EF", end_color="3572EF", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start + 3, roll_over_refinance_properties_start + 4):
                        ws6[
                            f"{col}{i}"].value = f"=_xlfn.IFERROR(_xlfn.UNIQUE(_xlfn.FILTER($B${calculate_exit_dates_start}:$B${calculate_exit_dates_end},($C${calculate_exit_dates_start}:$C${calculate_exit_dates_end}<={month_headings[index + 1]}$5)*($C${calculate_exit_dates_start}:$C${calculate_exit_dates_end}>{month_headings[index - 1]}$5)*($F${calculate_exit_dates_start}:$F${calculate_exit_dates_end}=0))),\"\")"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="000000", size=22)
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(roll_over_refinance_properties_start + 3, roll_over_refinance_properties_end):
                        ws6[f"{col}{i}"].font = Font(bold=True, color="000000", size=22)
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))




                elif index > 1 and index % 2 == 0:

                    ws6[f"{col}{CPC_INVOICES_STILL_DUE}"] = 0
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE + 9}"].number_format = 'R#,##0'

                    ws6[f"{col}{CPC_INVOICES_STILL_DUE}"].border = ws6[f"A{investor_exited}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE}"].fill = PatternFill(start_color="FFF67E", end_color="FFF67E",
                                                                             fill_type="solid")
                    # make the font size 22
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE}"].font = Font(bold=True, color="0C0C0C", size=22)

                    ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"] = f"={col}{CPC_INVOICES_STILL_DUE} / 115 * 15"
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].number_format = 'R#,##0'
                    # put borders around and fill in blue
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].border = ws6[f"A{investor_exited}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].fill = PatternFill(start_color="3572EF",
                                                                                 end_color="3572EF",
                                                                                 fill_type="solid")
                    # make the font size 22
                    ws6[f"{col}{CPC_INVOICES_STILL_DUE + 1}"].font = Font(bold=True, color="0C0C0C", size=22)
                    if project == "Consolidated":
                        ws6[f"{col}{vat_income}"] = (
                            f"=SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W,\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&{month_headings[index - 2]}$5)*C{vat_income}")
                    elif project == "Heron":
                        ws6[f"{col}{vat_income}"] = (
                            f"=(SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W,\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&{month_headings[index - 2]}$5,Sales!$A:$A,\"Heron Fields\")+SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W,\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&{month_headings[index - 2]}$5,Sales!$A:$A,\"Heron View\"))*C{vat_income}")
                    else:
                        ws6[f"{col}{vat_income}"] = (
                            f"=SUMIFS(Sales!$J:$J,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$W:$W,\"<=\"&{month_headings[index]}$5,Sales!$W:$W,\">\"&{month_headings[index - 2]}$5,Sales!$A:$A,\"{project}\")*C{vat_income}")

                    ws6[f"{col}{vat_income}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"{col}{vat_income}"].border = ws6[f"A{vat_income}"].border + Border(left=Side(style='medium'),
                                                                                             right=Side(style='medium'),
                                                                                             top=Side(style='medium'),
                                                                                             bottom=Side(
                                                                                                 style='medium'))
                    # light green
                    ws6[f"{col}{vat_income}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                                 fill_type="solid")
                    ws6[f"{col}{vat_income}"].number_format = 'R#,##0'

                    for i in range(vat_payable_on_sales, vat_payable_on_sales + 1):

                        if index > 2:

                            if project == "Consolidated":
                                ws6[
                                    f"{col}{i}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,(-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$F:$F,1,Sales!$E:$E,FALSE))+{col}{vat_payable_on_sales - 1}-{col}{vat_payable_on_sales + 2})*C{vat_payable_on_sales}"
                            elif project == "Heron":
                                ws6[
                                    f"{col}{i}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,((-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$F:$F,1,Sales!$E:$E,FALSE,Sales!$A:$A,\"Heron Fields\")-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$F:$F,1,Sales!$E:$E,FALSE,Sales!$A:$A,\"Heron View\")))+{col}{vat_payable_on_sales - 1}-{col}{vat_payable_on_sales + 2})*C{vat_payable_on_sales}"
                            else:
                                ws6[
                                    f"{col}{i}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,(-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$F:$F,1,Sales!$E:$E,FALSE,Sales!$A:$A,\"{project}\"))+{col}{vat_payable_on_sales - 1}-{col}{vat_payable_on_sales + 2})*C{vat_payable_on_sales}"
                        else:
                            if project == "Heron":
                                # "=IF(MOD(MONTH(K5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,"<="&EOMONTH(EDATE(K$5, 0), -2),Sales!$W:$W,">"&EOMONTH(EDATE(K$5, 0), -3),Sales!$A:$A,"Heron Fields")-SUMIFS(Sales!$J:$J,Sales!$W:$W,"<="&EOMONTH(EDATE(K$5, 0), -2),Sales!$W:$W,">"&EOMONTH(EDATE(K$5, 0), -3),Sales!$A:$A,"Heron View")-SUMIFS(Sales!$J:$J,Sales!$W:$W,"<="&EOMONTH(EDATE(K$5, 0), -1),Sales!$W:$W,">"&EOMONTH(EDATE(K$5, 0), -2),Sales!$A:$A,"Heron Fields",Sales!$F:$F,1)-SUMIFS(Sales!$J:$J,Sales!$W:$W,"<="&EOMONTH(EDATE(K$5, 0), -1),Sales!$W:$W,">"&EOMONTH(EDATE(K$5, 0), -2),Sales!$A:$A,"Heron View",Sales!$F:$F,1)+K24-K27)*C25"
                                ws6[
                                    f"{col}{i}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"Heron Fields\")-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"Heron View\")-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$A:$A,\"Heron Fields\",Sales!$F:$F,1)-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$A:$A,\"Heron View\",Sales!$F:$F,1)+{col}{vat_payable_on_sales - 1}-{col}{vat_payable_on_sales + 2})*C{vat_payable_on_sales}"
                            elif project == "Consolidated":
                                ws6[
                                    f"{col}{i}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3))-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$F:$F,1)+{col}{vat_payable_on_sales - 1}-{col}{vat_payable_on_sales + 2})*C{vat_payable_on_sales}"
                            else:
                                ws6[
                                    f"{col}{i}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -3),Sales!$A:$A,\"{project}\")-SUMIFS(Sales!$J:$J,Sales!$W:$W,\"<=\"&EOMONTH(EDATE({col}$5, 0), -1),Sales!$W:$W,\">\"&EOMONTH(EDATE({col}$5, 0), -2),Sales!$A:$A,\"{project}\",Sales!$F:$F,1)+{col}{vat_payable_on_sales - 1}-{col}{vat_payable_on_sales + 2})*C{vat_payable_on_sales}"

                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # fil in light red
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        ws6[f"{col}{i}"].border = ws6[f"A{vat_payable_on_sales}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(
                                style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(
                                style='medium'))

                    ws6[
                        f"{col}{vat_row}"] = f"=IF(MOD(MONTH({col}5),2)=0,0,((+{month_headings[index - 3]}{refinanced_units_start - 1}+{month_headings[index - 1]}{refinanced_units_start - 1})*156521.739130435)*C{vat_row})"

                    ws6[f"{col}{vat_row}"].font = Font(bold=True, color="0C0C0C", size=22)

                    ws6[f"{col}{vat_row}"].border = ws6[f"A{vat_row}"].border + Border(left=Side(style='medium'),
                                                                                       right=Side(style='medium'),
                                                                                       top=Side(style='medium'),
                                                                                       bottom=Side(style='medium'))
                    # fill equals light green
                    ws6[f"{col}{vat_row}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                              fill_type="solid")

                    ws6[f"{col}{vat_row}"].number_format = 'R#,##0'

                    if project == "Consolidated":

                        ws6[
                            f"{col}{profit_on_sale}"] = (
                            f"=SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&"
                            f"{month_headings[index - 2]}$5)")
                    elif project == "Heron":
                        ws6[
                            f"{col}{profit_on_sale}"] = (
                            f"=SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&"
                            f"{month_headings[index - 2]}$5,Sales!$A:$A,\"Heron Fields\")+SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&"
                            f"{month_headings[index - 2]}$5,Sales!$A:$A,\"Heron View\")")
                    else:
                        ws6[
                            f"{col}{profit_on_sale}"] = (
                            f"=SUMIFS(Sales!$S:$S,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,"
                            f"\"<=\"&{month_headings[index]}$5,"
                            f"Sales!$H:$H,\">\"&"
                            f"{month_headings[index - 2]}$5,Sales!$A:$A,\"{project}\")")

                    ws6[f"{col}{profit_on_sale}"].font = Font(bold=True, color="0C0C0C", size=22)

                    ws6[f"{col}{profit_on_sale}"].number_format = 'R#,##0'
                    ws6[f"{col}{profit_on_sale}"].border = ws6[f"A{profit_on_sale}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))

                    ws6.conditional_formatting.add(f"{col}{profit_on_sale}",
                                                   formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFC7CE",
                                                                                  end_color="FFC7CE",
                                                                                  fill_type="solid")))

                    ws6.conditional_formatting.add(f"{col}{profit_on_sale}",
                                                   formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="BFEA7C",
                                                                                  end_color="BFEA7C",
                                                                                  fill_type="solid")))

                    ws6.conditional_formatting.add(f"{col}{profit_on_sale}",
                                                   formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFF67E",
                                                                                  end_color="FFF67E",
                                                                                  fill_type="solid")))

                    ws6[
                        f"{col}{toggles_start - 2}"] = (
                        f"=EOMONTH(EDATE({month_headings[index - 2]}${toggles_start - 2}, "
                        f"0), 1)")
                    ws6[f"{col}{toggles_start - 2}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"{col}{toggles_start - 2}"].number_format = 'dd-mmm-yy'

                    for i in range(toggles_start, toggles_end + 1):
                        ws6[
                            f"{col}{i}"] = f"=SUMIFS(Sales!$S:$S,Sales!$W:$W,\"<=\"&{month_headings[index]}$5,Sales!$E:$E,FALSE,Sales!$F:$F,$C{i},Sales!$A:$A,$A{i},Sales!$B:$B,$B{i},Sales!$W:$W,\">\"&{month_headings[index - 2]}$5)*$C{i}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"{col}{vat_row}"].border = ws6[f"A{vat_row}"].border + Border(left=Side(style='medium'),
                                                                                           right=Side(style='medium'),
                                                                                           top=Side(style='medium'),
                                                                                           bottom=Side(style='medium'))

                    for i in range(block_costs_start, block_costs_end + 1):
                        if project == "Goodwood":
                            ws6[
                                f"{col}{i}"] = f"=-(SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$A:$A,FALSE,'Updated Construction'!$G:$G,$A{i},'Updated Construction'!$E:$E," \
                                               f"\"<=\"&{month_headings[index]}$5,'Updated Construction'!$E:$E,\">\"&{month_headings[index - 2]}$5)+(SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$A:$A,TRUE,'Updated Construction'!$G:$G,$A{i},'Updated Construction'!$E:$E," \
                                               f"\"<=\"&{month_headings[index]}$5,'Updated Construction'!$E:$E,\">\"&{month_headings[index - 2]}$5)*$B{i}))"
                        else:
                            ws6[
                                f"{col}{i}"] = f"=-(SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$A:$A,FALSE,'Updated Construction'!$G:$G,$A{i},'Updated Construction'!$E:$E," \
                                               f"\"<=\"&{month_headings[index]}$5,'Updated Construction'!$E:$E,\">\"&{month_headings[index - 2]}$5)+(SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$A:$A,TRUE,'Updated Construction'!$G:$G,$A{i},'Updated Construction'!$E:$E," \
                                               f"\"<=\"&{month_headings[index]}$5,'Updated Construction'!$E:$E,\">\"&{month_headings[index - 2]}$5)*$B{i}))*1.15"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                    for i in range(vat_construction, vat_construction + 1):
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                            fill_type="solid")

                        ws6[f"{col}{i}"].border = ws6[f"A{vat_construction}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(operating_expenses, operating_expenses + 1):
                        "=-'Operational Costs'!$N$2*$C$50"
                        ws6[
                            f"{col}{i}"] = f"=-'Operational Costs'!$N$2*$C{i}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                            fill_type="solid")

                        ws6[f"{col}{i}"].border = ws6[f"A{vat_construction}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(monthly, monthly + 1):
                        # "=F27+F29+F32+F48+F50"
                        ws6[
                            f"{col}{i}"] = f"={col}{vat_row}+{col}{block_costs_start - 1}+{col}{vat_construction}+{col}{operating_expenses}+{col}{vat_payable_on_sales}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        ws6[f"{col}{i}"].border = ws6[f"A{vat_construction}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(running, running + 1):
                        "=F53+H51"
                        ws6[
                            f"{col}{i}"] = f"=+{col}{monthly}+{col}{funds_available_end}"

                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{running}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(investor_exited, investor_exited + 1):
                        "=I81-I160-I78"
                        ws6[
                            f"{col}{i}"] = f"={col}{funds_available_start + 4}+{col}{funds_available_start + 9}-{col}{investor_exited + 2} + {month_headings[index - 2]}{investor_exited}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))
                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(rollover, rollover + 1):
                        # "=-SUMIFS('Investor Exit List'!$Q:$Q,'Investor Exit List'!$Z:$Z,\"Release\",'Investor Exit List'!$L:$L,\"<=\"&H$5,'Investor Exit List'!$L:$L,\">\"&F$5)"
                        ws6[
                            f"{col}{i}"] = ws6[
                            f"{col}{i}"] = f"=0"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="3ABEF9", end_color="3ABEF9", fill_type="solid")

                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(project_income, project_income + 1):
                        ws6[
                            f"{col}{i}"] = f"SALES"
                        # ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="749BC2", end_color="749BC2", fill_type="solid")

                    for i in range(purple_blok_start, purple_blok_end):
                        ws6[
                            f"{col}{i}"] = f"=$B${i}*{month_headings[index + 1]}{i}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(purple_blok_end, purple_blok_end + 1):
                        # "=SUBTOTAL(9,H95:H97)"
                        ws6[
                            f"{col}{i}"] = f"=SUBTOTAL(9,{col}{purple_blok_start}:{col}{purple_blok_end - 1})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFC7CE",
                                                                                      end_color="FFC7CE",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="BFEA7C",
                                                                                      end_color="BFEA7C",
                                                                                      fill_type="solid")))

                        ws6.conditional_formatting.add(f"{col}{i}",
                                                       formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                                  fill=PatternFill(
                                                                                      start_color="FFF67E",
                                                                                      end_color="FFF67E",
                                                                                      fill_type="solid")))

                    for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                        # I AM HERE
                        if project == "Consolidated":
                            ws6[
                                f"{col}{i}"] = f"=IF(VLOOKUP($D{i},Sales!$C:$D,2,FALSE)=TRUE,0,SUMIFS('Investor Exit List'!$Q:$Q, 'Investor Exit List'!$C:$C, {month_headings[index - 1]}{i}))"
                        else:
                            ws6[
                                f"{col}{i}"] = f"=IF(VLOOKUP($D{i},Sales!$C:$D,2,FALSE)=TRUE,0,SUMIFS('Investor Exit List - {project}'!$Q:$Q, 'Investor Exit List - {project}'!$C:$C, {month_headings[index - 1]}{i}))"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(refinanced_units_start - 1, refinanced_units_start):
                        ws6[
                            f"{col}{i}"] = f"=-SUM({col}{refinanced_units_start}:{col}{refinanced_units_end})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(debenture_transaction_start, debenture_transaction_start + 1):
                        ws6[
                            f"{col}{i}"] = f"=$B{debenture_transaction_start}*{month_headings[index + 1]}{debenture_transaction_start}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        # fill in blue
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="3572EF", end_color="3572EF", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start, roll_over_refinance_properties_start + 1):
                        "=$B$151*J151"
                        ws6[
                            f"{col}{i}"] = f"=sum({col}{roll_over_refinance_properties_start + 3}:{col}{roll_over_refinance_properties_end})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        # fill in blue
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start + 2, roll_over_refinance_properties_start + 3):
                        ws6[
                            f"{col}{i}"] = f"={col}{roll_over_refinance_properties_start}-{col}{roll_over_refinance_properties_start + 1}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        # fill in blue
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start + 3, roll_over_refinance_properties_end):
                        ws6[
                            f"{col}{i}"].value = f"=SUMIFS($D${calculate_exit_dates_start}:$D${calculate_exit_dates_end},$B${calculate_exit_dates_start}:$B${calculate_exit_dates_end},{month_headings[index - 1]}{i})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="000000", size=22)
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))


                elif index > 2 and index % 2 != 0:

                    for i in range(toggles_start, toggles_end + 1):
                        if project == "Consolidated":
                            ws6[
                                f"{col}{i}"] = (
                                f"=COUNTIFS(Sales!$W:$W,\"<=\"&{month_headings[index - 1]}"
                                f"$5,Sales!$W:$W,\">\"&{month_headings[index - 3]}$5,"
                                f"Sales!$A:$A,$A{i},Sales!$B:$B,'Cashflow "
                                f"Projection'!$B{i},Sales!$F:$F,$C{i},Sales!$E:$E,"
                                f"FALSE)*$C{i}")
                        else:
                            ws6[
                                f"{col}{i}"] = (
                                f"=COUNTIFS(Sales!$W:$W,\"<=\"&{month_headings[index - 1]}"
                                f"$5,Sales!$W:$W,\">\"&{month_headings[index - 3]}$5,"
                                f"Sales!$A:$A,$A{i},Sales!$B:$B,'Cashflow "
                                f"Projection - {project}'!$B{i},Sales!$F:$F,$C{i},Sales!$E:$E,"
                                f"FALSE)*$C{i}")
                        # elif project == "Heron":
                        #     ws6[
                        #         f"{col}{i}"] = (
                        #         f"=(COUNTIFS(Sales!$W:$W,\"<=\"&{month_headings[index - 1]}"
                        #         f"$5,Sales!$W:$W,\">\"&{month_headings[index - 3]}$5,"
                        #         f"Sales!$A:$A,$A{i},Sales!$B:$B,'Cashflow "
                        #         f"Projection'!$B{i},Sales!$F:$F,$C{i},Sales!$E:$E,"
                        #         f"FALSE,Sales!$A:$A,\"Heron Fields\")+COUNTIFS(Sales!$W:$W,\"<=\"&{month_headings[index - 1]}"
                        #         f"$5,Sales!$W:$W,\">\"&{month_headings[index - 3]}$5,"
                        #         f"Sales!$A:$A,$A{i},Sales!$B:$B,'Cashflow "
                        #         f"Projection'!$B{i},Sales!$F:$F,$C{i},Sales!$E:$E,"
                        #         f"FALSE,Sales!$A:$A,\"Heron View\"))*$C{i}")
                        # else:
                        #     ws6[
                        #         f"{col}{i}"] = (
                        #         f"=COUNTIFS(Sales!$W:$W,\"<=\"&{month_headings[index - 1]}"
                        #         f"$5,Sales!$W:$W,\">\"&{month_headings[index - 3]}$5,"
                        #         f"Sales!$A:$A,$A{i},Sales!$B:$B,'Cashflow "
                        #         f"Projection'!$B{i},Sales!$F:$F,$C{i},Sales!$E:$E,"
                        #         f"FALSE,Sales!$A:$A,\"{project}\")*$C{i}")
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # Center the text in the cell
                        ws6[f"{col}{i}"].alignment = Alignment(horizontal='center', vertical='center')

                    for i in range(project_income, project_income + 1):
                        ws6[
                            f"{col}{i}"] = f"EXIT & ROLL"
                        # ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="749BC2", end_color="749BC2", fill_type="solid")

                    for i in range(roll_over_refinance, roll_over_refinance + 1):
                        ws6[
                            f"{col}{i}"] = 0
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(roll_over_reinvest, roll_over_reinvest + 1):
                        ws6[
                            f"{col}{i}"] = 0
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(purple_blok_start, purple_blok_end):
                        ws6[
                            f"{col}{i}"] = 0
                        ws6[f"{col}{i}"].number_format = '0%'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        # for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                        ws6[f"{col}{i}"].fill = PatternFill(start_color="FFFF80", end_color="FFFF80", fill_type="solid")

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(refinanced_units_start - 1, refinanced_units_start):
                        ws6[
                            f"{col}{i}"] = f"=COUNTA({col}{refinanced_units_start}:{col}{refinanced_units_end})"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = ws6[f"A{investor_exited}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                    for i in range(debenture_transaction_start, debenture_transaction_start + 1):
                        ws6[
                            f"{col}{i}"] = f"=${month_headings[index - 2]}{refinanced_units_start - 1}"
                        ws6[f"{col}{i}"].number_format = 'R#,##0'
                        ws6[f"{col}{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                        # apply borders around the cell
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

                        ws6[f"{col}{i}"].fill = PatternFill(start_color="3572EF", end_color="3572EF", fill_type="solid")

                    for i in range(roll_over_refinance_properties_start + 3, roll_over_refinance_properties_start + 4):
                        try:

                            ws6[
                                f"{col}{i}"].value = f"=_xlfn.IFERROR(_xlfn.UNIQUE(_xlfn.FILTER($B${calculate_exit_dates_start}:$B${calculate_exit_dates_end},($C${calculate_exit_dates_start}:$C${calculate_exit_dates_end}<={month_headings[index + 1]}$5)*($C${calculate_exit_dates_start}:$C${calculate_exit_dates_end}>{month_headings[index - 1]}$5)*($F${calculate_exit_dates_start}:$F${calculate_exit_dates_end}=0))),\"\")"
                            ws6[f"{col}{i}"].number_format = 'R#,##0'
                            ws6[f"{col}{i}"].font = Font(bold=True, color="000000", size=22)
                            ws6[f"{col}{i}"].border = Border(
                                left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))
                        except Exception as e:
                            continue

                    for i in range(roll_over_refinance_properties_start + 3, roll_over_refinance_properties_end):

                        ws6[f"{col}{i}"].font = Font(bold=True, color="000000", size=22)
                        ws6[f"{col}{i}"].border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))

            for i in range(refinanced_units_start - 1, refinanced_units_start):
                "=COUNTA(H117:H149)"
                ws6[
                    f"H{i}"] = f"=COUNTA(H{refinanced_units_start}:H{refinanced_units_end})"
                ws6[f"H{i}"].number_format = 'R#,##0'
                ws6[f"H{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                # apply borders around the cell
                ws6[f"H{i}"].border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium'))

            for i in range(refinanced_units_start, refinanced_units_end + 1):
                ws6[f"H{i}"].number_format = 'R#,##0'
                ws6[f"H{i}"].font = Font(bold=True, color="0C0C0C", size=22)

                # apply borders around the cell
                ws6[f"H{i}"].border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium'))

            for i in range(toggles_start, toggles_end + 1):
                ws6[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
                ws6[f"C{i}"].alignment = Alignment(horizontal='center', vertical='center')
                # make font bold and 22 in columns A,B & C
                ws6[f"A{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"B{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"C{i}"].font = Font(bold=True, color="0C0C0C", size=22)

            for i in range(block_costs_start, block_costs_end + 1):
                ws6[f"A{i}"].border = ws6[f"A{i}"].border + Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                ws6[f"B{i}"].border = ws6[f"B{i}"].border + Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                ws6[f"C{i}"].border = ws6[f"C{i}"].border + Border(left=Side(style='medium'),
                                                                   right=Side(style='medium'),
                                                                   top=Side(style='medium'),
                                                                   bottom=Side(style='medium'))
                # make font bold and 22
                ws6[f"A{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"B{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"C{i}"].font = Font(bold=True, color="FFFFFF", size=22)

                if ws6[f"A{i}"].value not in ["D", "H", "N"]:
                    ws6[f"A{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    ws6[f"B{i}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    if ws6[f"A{i}"].value == "E":
                        ws6[f"C{i}"].value = f"=12*B{i}"
                    elif ws6[f"A{i}"].value == "F":
                        ws6[f"C{i}"].value = f"=8*B{i}"
                    elif ws6[f"A{i}"].value == "G":
                        ws6[f"C{i}"].value = f"=12*B{i}"
                    elif ws6[f"A{i}"].value == "I":
                        ws6[f"C{i}"].value = f"=8*B{i}"
                    elif ws6[f"A{i}"].value == "J":
                        ws6[f"C{i}"].value = f"=12*B{i}"
                    elif ws6[f"A{i}"].value == "K":
                        ws6[f"C{i}"].value = f"=24*B{i}"
                    elif ws6[f"A{i}"].value == "L":
                        ws6[f"C{i}"].value = f"=8*B{i}"
                    elif ws6[f"A{i}"].value == "M":
                        ws6[f"C{i}"].value = f"=8*B{i}"
                    elif ws6[f"A{i}"].value == "O":
                        ws6[f"C{i}"].value = f"=15*B{i}"

                # center the text in A, B & C
                ws6[f"A{i}"].alignment = Alignment(horizontal='center', vertical='center')
                ws6[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
                ws6[f"C{i}"].alignment = Alignment(horizontal='center', vertical='center')
                # fill C with red
                ws6[f"C{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                if i == block_costs_start:
                    ws6[f"A{i - 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws6[f"B{i - 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws6[f"C{i - 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws6[f"C{i - 1}"].font = Font(bold=True, color="FFFFFF", size=22)
                    ws[f"C{i - 1}"].border = ws6[f"A{i - 1}"].border + Border(left=Side(style='medium'),
                                                                              right=Side(style='medium'),
                                                                              top=Side(style='medium'),
                                                                              bottom=Side(style='medium'))
                    # Center the text in A, B & C
                    ws6[f"A{i - 1}"].alignment = Alignment(horizontal='center', vertical='center')
                    ws6[f"B{i - 1}"].alignment = Alignment(horizontal='center', vertical='center')
                    ws6[f"C{i - 1}"].alignment = Alignment(horizontal='center', vertical='center')
                    # "=SUM(C33:C45)"
                    ws6[f"C{i - 1}"] = f"=SUM(C{block_costs_start}:C{block_costs_end})"

            ws6[f"A{block_costs_end + 1}"] = "REMAINING"
            ws6[f"A{block_costs_end + 1}"].font = Font(bold=True, color="FFFFFF", size=22)

            # Merge A and B
            ws6.merge_cells(f"A{block_costs_end + 1}:B{block_costs_end + 1}")
            ws6[f"A{block_costs_end + 1}"].alignment = Alignment(horizontal='center', vertical='center')
            ws6[f"A{block_costs_end + 1}"].border = ws6[f"A{block_costs_end + 1}"].border + Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

            ws6[f"C{block_costs_end + 1}"] = f"=107-C{block_costs_start - 1}"
            ws6[f"C{block_costs_end + 1}"].font = Font(bold=True, color="FFFFFF", size=22)
            ws6[f"C{block_costs_end + 1}"].alignment = Alignment(horizontal='center', vertical='center')
            ws6[f"C{block_costs_end + 1}"].border = ws6[f"A{block_costs_end + 1}"].border + Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium'))

            # fill A, B & C with Red
            ws6[f"A{block_costs_end + 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                              fill_type="solid")
            ws6[f"B{block_costs_end + 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                              fill_type="solid")
            ws6[f"C{block_costs_end + 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                              fill_type="solid")
            ws6.append([])
            # print("WS 6 Max Row: ",ws6.max_row)

            for row in ws2b.iter_rows(min_row=5, max_row=ws2b.max_row, min_col=3, max_col=3):
                for cell in row:
                    # I AM HERE
                    if project == "Consolidated":
                        cell.value = f"=IF(A{cell.row}=FALSE, 1,SUMIFS('Cashflow Projection'!$C${block_costs_start}:$C${block_costs_end}, 'Cashflow Projection'!$B$33:$B$45, 'Updated Construction'!G{cell.row}))"
                    else:
                        cell.value = f"=IF(A{cell.row}=FALSE, 1,SUMIFS('Cashflow Projection - {project}'!$C${block_costs_start}:$C${block_costs_end}, 'Cashflow Projection - {project}'!$B$33:$B$45, 'Updated Construction'!G{cell.row}))"

            for row in ws6.iter_rows(min_row=toggles_start, max_row=toggles_end, min_col=3, max_col=3):
                for cell in row:
                    "=IF(ISERROR(VLOOKUP($B7,$A$50:$B$63,2,FALSE)),0,SUMIFS($B$50:$B$63,$A$50:$A$63,B7))"
                    cell.value = f"=IF(ISERROR(VLOOKUP($B{cell.row},$A{block_costs_start}:$B{block_costs_end},2,FALSE)),1,IF(SUMIFS($B{block_costs_start}:$B{block_costs_end},$A{block_costs_start}:$A{block_costs_end},B{cell.row})>0,1,0))"

            for i in range(ws6.max_row + 1, ws6.max_row + 2):
                for x in range(1, ws6.max_column + 1):
                    if x == 1:
                        ws6[f"{get_column_letter(x)}{i}"].border = Border(left=Side(style='medium'),
                                                                          bottom=Side(style='medium'))
                    elif x == ws6.max_column:
                        ws6[f"{get_column_letter(x)}{i}"].border = Border(right=Side(style='medium'),
                                                                          bottom=Side(style='medium'))
                    else:
                        ws6[f"{get_column_letter(x)}{i}"].border = Border(bottom=Side(style='medium'))

            columns_for_funds_available = month_headings

            for index, col in enumerate(columns_for_funds_available):
                if index % 2 == 0:
                    ws6[f"{col}{funds_available_start}"].border = ws6[f"A{funds_available_start}"].border + Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))
                    ws6[f"{col}{funds_available_start}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                                            fill_type="solid")
                    ws6[
                        f"{col}{funds_available_start}"].value = f"=sum({col}{funds_available_start + 1}:{col}{funds_available_end - 1})"
                    ws6[f"{col}{funds_available_start}"].number_format = 'R#,##0'
                    ws6[f"{col}{funds_available_start}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"{col}{funds_available_end}"].number_format = 'R#,##0'
                    ws6[f"{col}{funds_available_end}"].font = Font(bold=True, color="0C0C0C", size=22)
                    ws6[f"{col}{funds_available_end}"].fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C",
                                                                          fill_type="solid")

                    # ws6[f"{col}{funds_available_end}"].value = f"=+{col}{funds_available_start}+{col}{profit_on_sale}"

                    # conditional formatting for funds_available_start and funds_available_end
                    ws6.conditional_formatting.add(f"{col}{funds_available_start}",
                                                   formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFC7CE",
                                                                                  end_color="FFC7CE",
                                                                                  fill_type="solid")))
                    ws6.conditional_formatting.add(f"{col}{funds_available_start}",
                                                   formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="BFEA7C",
                                                                                  end_color="BFEA7C",
                                                                                  fill_type="solid")))
                    ws6.conditional_formatting.add(f"{col}{funds_available_start}",
                                                   formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFF67E",
                                                                                  end_color="FFF67E",
                                                                                  fill_type="solid")))
                    ws6.conditional_formatting.add(f"{col}{funds_available_end}",
                                                   formatting.rule.CellIsRule(operator='lessThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFC7CE",
                                                                                  end_color="FFC7CE",
                                                                                  fill_type="solid")))
                    ws6.conditional_formatting.add(f"{col}{funds_available_end}",
                                                   formatting.rule.CellIsRule(operator='greaterThan', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="BFEA7C",
                                                                                  end_color="BFEA7C",
                                                                                  fill_type="solid")))
                    ws6.conditional_formatting.add(f"{col}{funds_available_end}",
                                                   formatting.rule.CellIsRule(operator='equal', formula=['0'],
                                                                              fill=PatternFill(
                                                                                  start_color="FFF67E",
                                                                                  end_color="FFF67E",
                                                                                  fill_type="solid")))

                    for i in range(funds_available_start, funds_available_end + 1):
                        ws6[f"{col}{i}"].border = ws6[f"{col}{i}"].border + Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
                        ws6[f"{col}{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                        ws6[f"{col}{i}"].number_format = 'R#,##0'

            for i in range(funds_available_start + 1, funds_available_end + 1):
                ws6[f"D{i}"].border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                             top=Side(style='medium'), bottom=Side(style='medium'))
                # ws6[f"B{i}"].border = ws6[f"B{i}"].border + Border(left=Side(style='medium'), right=Side(style='medium'),
                #                                                    top=Side(style='medium'), bottom=Side(style='medium'))
                # ws6[f"C{i}"].border = ws6[f"C{i}"].border + Border(left=Side(style='medium'), right=Side(style='medium'),
                #                                                    top=Side(style='medium'), bottom=Side(style='medium'))
                ws6[f"D{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"D{i}"].number_format = 'R#,##0'

                # ws6[f"B{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                # ws6[f"C{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                # ws6[f"C{i}"].number_format = 'R#,##0'

            # filter the data for columns A to E from calculate_exit_dates_start - 1 to calculate_exit_dates_end

            ws6.auto_filter.ref = f"A{calculate_exit_dates_start - 1}:E{calculate_exit_dates_end}"

            ws6[f"C{vat_row}"].value = f"=C{vat_construction}"

            for i in range(toggles_start, toggles_end + 1):
                # fill column G in Red and make the font white
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)
                ws6[f"G{i}"] = "DJ / IK"

            ws6[f"G{vat_payable_on_sales}"] = "DJ"
            ws6[f"G{vat_payable_on_sales}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                               fill_type="solid")
            ws6[f"G{vat_payable_on_sales}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{vat_income}"] = "DJ"
            ws6[f"G{vat_income}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws6[f"G{vat_income}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{vat_recovery_when_refinanced}"] = "DJ"
            ws6[f"G{vat_recovery_when_refinanced}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                                       fill_type="solid")
            ws6[f"G{vat_recovery_when_refinanced}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{profit_on_sale}"] = "DJ"
            ws6[f"G{profit_on_sale}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                         fill_type="solid")
            ws6[f"G{profit_on_sale}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(funds_available_start, funds_available_start + 3):
                ws6[f"G{i}"] = "DJ / IK"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(funds_available_start + 3, funds_available_end + 1):
                ws6[f"G{i}"] = "DJ / LK"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(block_costs_start - 1, block_costs_end + 1):
                ws6[f"G{i}"] = "DJ / PR"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{vat_construction}"] = "DJ"

            ws6[f"G{vat_construction}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                           fill_type="solid")
            ws6[f"G{vat_construction}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{operating_expenses}"] = "DJ"
            ws6[f"G{operating_expenses}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                             fill_type="solid")
            ws6[f"G{operating_expenses}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{monthly}"] = "DJ"
            ws6[f"G{monthly}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                  fill_type="solid")
            ws6[f"G{monthly}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{running}"] = "DJ"
            ws6[f"G{running}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                  fill_type="solid")
            ws6[f"G{running}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{investor_exited}"] = "DJ / LK"
            ws6[f"G{investor_exited}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                          fill_type="solid")
            ws6[f"G{investor_exited}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{rollover}"] = "DJ / LK"
            ws6[f"G{rollover}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                   fill_type="solid")
            ws6[f"G{rollover}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(roll_over_refinance, block_finance_end + 1):
                ws6[f"G{i}"] = "DJ / LK"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(purple_blok_start, purple_blok_end + 1):
                ws6[f"G{i}"] = "DJ / LK"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(finance_waterfall_start + 1, finance_waterfall_end + 1):
                ws6[f"G{i}"] = "DJ"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(finance_waterfall_start + 2, finance_waterfall_start + 3):
                ws6[f"G{i}"] = "DJ / WH"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(finance_waterfall_start + 4, finance_waterfall_start + 5):
                ws6[f"G{i}"] = "DJ / LK"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{refinanced_units_start - 1}"] = "DJ / LK"
            ws6[f"G{refinanced_units_start - 1}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                                     fill_type="solid")
            ws6[f"G{refinanced_units_start - 1}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(refinanced_units_start + 1, refinanced_units_end + 1):
                ws6[f"G{i}"] = "DJ / WH"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G{debenture_transaction_start}"] = "DJ / LK"
            ws6[f"G{debenture_transaction_start}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                                      fill_type="solid")
            ws6[f"G{debenture_transaction_start}"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(roll_over_refinance_properties_start, roll_over_refinance_properties_end + 1):
                ws6[f"G{i}"] = "DJ / LK"
                ws6[f"G{i}"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                fill_type="solid")
                ws6[f"G{i}"].font = Font(bold=True, color="FFFFFF", size=22)

            ws6[f"G2"] = "CHECK"
            ws6[f"G2"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                          fill_type="solid")
            ws6[f"G2"].font = Font(bold=True, color="FFFFFF", size=22)

            for i in range(block_costs_start, block_costs_end + 1):
                # "=XLOOKUP(A53,$B$7:$B$26,$A$7:$A$26,"Heron View")"
                ws6[
                    f"E{i}"] = f"=_xlfn.XLOOKUP(A{i},$B${toggles_start}:$B${toggles_end},$A${toggles_start}:$A${toggles_end},\"Heron View\")"
                ws6[f"E{i}"].font = Font(bold=True, color="0C0C0C", size=22)
                ws6[f"E{i}"].border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                             top=Side(style='medium'), bottom=Side(style='medium'))

            for index, col in enumerate(month_headings):
                if index % 2 == 0:
                    ws6[f"{col}2"].value = f"CHECK"
                    ws6[f"{col}2"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                      fill_type="solid")

                    ws6[f"{col}2"].font = Font(bold=True, color="FFFFFF", size=22)

            for col in ["E", "J", "L", "N", "P", "R", "T", "V", "X", "Z", "AB", "AD", "AF"]:
                for i in range(toggles_start - 1, toggles_end + 1):
                    ws6[f"{col}{i}"].number_format = '0'


            for i in range(1, ws6.max_column + 1):
                if i == 1:
                    ws6.column_dimensions[get_column_letter(i)].width = 45.5
                elif i == 2:
                    ws6.column_dimensions[get_column_letter(i)].width = 20
                elif i == 3:
                    ws6.column_dimensions[get_column_letter(i)].width = 20
                elif i > 3 and i % 2 == 0:
                    ws6.column_dimensions[get_column_letter(i)].width = 20
                elif i > 3 and i % 2 != 0:
                    ws6.column_dimensions[get_column_letter(i)].width = 20

            # freeze panes at D7
            ws6.freeze_panes = ws6["D6"]

            if idx == 0:
                ws7 = wb.create_sheet('Cashflow')
            else:
                ws7 = wb.create_sheet(f'Cashflow - {project}')
            # make tab color red
            ws7.sheet_properties.tabColor = tab_colors[idx]
            if project != "Goodwood":
                ws7['B1'] = "C.3.f"
            else:
                ws7['B1'] = "C.5.3"
            ws7["B1"].font = Font(bold=True, color="31304D", size=22)
            # center the text
            ws7['B1'].alignment = Alignment(horizontal='center', vertical='center')
            # put a border around the cell
            ws7['B1'].border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                      bottom=Side(style='medium'))

            ws7['A2'] = "CASHFLOW DASHBOARD"
            ws7['A2'].font = Font(bold=True, color="31304D", size=22)
            ws7['A2'].border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                      bottom=Side(style='medium'))

            # merge the cells in the first row
            ws7.merge_cells('A2:B2')
            # put a medium border around the merged cell

            # center the text
            ws7['A2'].alignment = Alignment(horizontal='center', vertical='center')

            ws7['C2'].border = Border(right=Side(style='medium'), top=Side(style='medium'),
                                      bottom=Side(style='medium'))
            # fill with light grey
            ws7['A2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws7['B2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws7['C2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            for i in range(1, ws7.max_column + 1):
                if i == 1:
                    ws7.column_dimensions[get_column_letter(i)].width = 55
                    # set the font to bold and size 13
                    # ws7[f"{get_column_letter(i)}1"].font = Font(bold=False, size=13)
                else:
                    ws7.column_dimensions[get_column_letter(i)].width = 25
                    # set the font to bold and size 13
                    # ws7[f"{get_column_letter(i)}1"].font = Font(bold=True, size=13)
                    # set as currency
                    # ws7[f"{get_column_letter(i)}1"].number_format = 'R #,##0'

            ws7.append(["", "Projected", "Actual"])
            start = ws7.max_row
            for i in range(start, start + 1):
                # center the text in the cells
                ws7[f"A{i}"].alignment = Alignment(horizontal='center', vertical='center')
                ws7[f"A{i}"].font = Font(bold=True, size=22)
                ws7[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
                ws7[f"B{i}"].font = Font(bold=True, size=22)
                ws7[f"C{i}"].alignment = Alignment(horizontal='center', vertical='center')
                ws7[f"C{i}"].font = Font(bold=True, size=22)
                # Underline the text in B and C
                ws7[f"B{i}"].font = Font(underline="single", bold=True, size=22)
                ws7[f"C{i}"].font = Font(underline="single", bold=True, size=22)

            ws7.append([])
            # deep copy reports_by_project
            # reports_by_project_copy = reports_by_project.copy()
            # for project_copy in reports_by_project_copy:

            if project == "Consolidated":
                ws7.append([f"Actual Transfer Income to hit FNB - {project} Sold",
                            '=+SUMIFS(Sales!$S:$S,Sales!$D:$D,TRUE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection\'!B2,Sales!$H:$H,"<="&\'Cashflow Projection\'!AE5)'])
            elif project == "Heron":
                # "=+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"Heron Fields",Sales!$D:$D,TRUE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&'Cashflow Projection - Heron'!B2)+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"Heron View",Sales!$D:$D,TRUE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&'Cashflow Projection - Heron'!B2)"
                ws7.append([f"Actual Transfer Income to hit FNB - {project} Sold",
                            f'=+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"Heron Fields",Sales!$D:$D,TRUE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection - Heron\'!B2,Sales!$H:$H,"<="&\'Cashflow Projection - Heron\'!AE5)+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"Heron View",Sales!$D:$D,TRUE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection - Heron\'!B2,Sales!$H:$H,"<="&\'Cashflow Projection - Heron\'!AE5)'])
            else:
                ws7.append([f"Actual Transfer Income to hit FNB - {project} Sold",
                            f'=+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"{project}",Sales!$D:$D,TRUE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection - {project}\'!B2,Sales!$H:$H,"<="&\'Cashflow Projection - {project}\'!AE5,Sales!$H:$H,"<="&\'Cashflow Projection - {project}\'!AE5)'])
            # format as currency

            transfer_endulini = ws7.max_row

            transfer_heron = ws7.max_row
            if project == "Consolidated":
                ws7.append([f"Projected Transfer Income  - {project}  Not yet Sold (Profit)",
                            f'=+SUMIFS(Sales!$S:$S,Sales!$D:$D,FALSE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection\'!B2, Sales!$H:$H,"<="&\'Cashflow Projection - Heron\'!AE5)'])
            elif project == "Heron":

                ws7.append([f"Projected Transfer Income  - {project}  Not yet Sold (Profit)",
                            '=+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"Heron Fields",Sales!$D:$D,FALSE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection - Heron\'!B2, Sales!$H:$H,"<="&\'Cashflow Projection - Heron\'!AE5)++SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"Heron View",Sales!$D:$D,FALSE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection - Heron\'!B2,Sales!$H:$H,"<="&\'Cashflow Projection - Heron\'!AE5)'])
            else:
                ws7.append([f"Projected Transfer Income  - {project}  Not yet Sold (Profit)",
                            f'=+SUMIFS(Sales!$S:$S,Sales!$A:$A,"="&"{project}",Sales!$D:$D,FALSE,Sales!$E:$E,FALSE,Sales!$F:$F,1,Sales!$H:$H,">"&\'Cashflow Projection - {project}\'!B2,Sales!$H:$H,"<="&\'Cashflow Projection - {project}\'!AE5)'])

            ws7.append([])
            ws7.append([])
            ws7.append([])

            not_allocated = ws7.max_row
            ws7.append([])
            ws7.append(["Total Income", f"=sum(B{transfer_endulini}:B{not_allocated})",
                        f"=sum(C{transfer_endulini}:C{not_allocated})"])
            total_income = ws7.max_row
            # Put a border around the cells B & C in total_income

            ws7.append([])
            if project == "Consolidated" or project == "Heron":

                ws7.append(["Momentum funds available to draw",
                            "=SUMIFS(Xero_bank!$G:$G,Xero_bank!$E:$E,\"Momentum Investors Account RU502229930\",Xero_bank!$B:$B,'Cashflow Projection'!$B$2)",
                            "=SUMIFS(Xero_bank!$G:$G,Xero_bank!$E:$E,\"Momentum Investors Account RU502229930\",Xero_bank!$B:$B,'Cashflow Projection'!$B$2)"])
            else:
                ws7.append([f"Momentum funds available to draw", 0, 0])
            momentum_funds = ws7.max_row
            # "=SUMIFS(Xero!$G:$G, Xero!$B:$B, 'Cashflow Projection'!$B$2, Xero!$D:$D, "84*")-SUMIFS(Xero!$G:$G, Xero!$B:$B, 'Cashflow Projection'!$B$2, Xero!$D:$D, "8480*")-B13"
            if project == "Consolidated":
                ws7.append(["FNB Bank",
                            f"=SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"84*\")-SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"8480*\")-B{momentum_funds}",
                            f"=SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"84*\")-B{momentum_funds}"])
            elif project == "Heron":
                ws7.append(["FNB Bank",
                            f"=SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"84*\",Xero_bank!$A:$A,\"<>Purple Blok Projects (Pty) Ltd\")-SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"8480*\")-B{momentum_funds}",
                            f"=SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"84*\")-B{momentum_funds}"])
            elif project == "Goodwood":
                ws7.append(["FNB Bank",
                            f"=SUMIFS(Xero_bank!$G:$G, Xero_bank!$B:$B, 'Cashflow Projection'!$B$2, Xero_bank!$D:$D, \"84*\",Xero_bank!$A:$A,\"Purple Blok Projects (Pty) Ltd\")"])

            fnb_bank = ws7.max_row
            ws7.append(["New Investors", 0, 0])
            ws7.append(["Deposit made", 0, 0])
            deposit_made = ws7.max_row
            ws7.append(["Momentum interest to be earned", 0, 0])
            momentum_interest = ws7.max_row
            ws7.append(["Investor funds to hit Momentum Account to draw", 0, 0])
            if project == "Consolidated":
                # "='Cashflow Projection - Heron'!D37+'Cashflow Projection - Heron'!D38+'Cashflow Projection - Heron'!D39+'Cashflow Projection - Heron'!D40+'Cashflow Projection - Heron'!D42"
                ws7.append([f"Deposit made",
                            f"='Cashflow Projection'!D{funds_available_start + 5}+'Cashflow Projection'!D{funds_available_start + 6}+'Cashflow Projection'!D{funds_available_start + 7}+'Cashflow Projection'!D{funds_available_start + 8}+'Cashflow Projection'!D{funds_available_start + 10}",
                            0])
            elif project == "Heron":
                ws7.append([f"Deposit made",
                            f"='Cashflow Projection - Heron'!D{funds_available_start + 5}+'Cashflow Projection - Heron'!D{funds_available_start + 6}+'Cashflow Projection - Heron'!D{funds_available_start + 7}+'Cashflow Projection - Heron'!D{funds_available_start + 8}+'Cashflow Projection - Heron'!D{funds_available_start + 10}",
                            0])
            else:
                ws7.append([f"Deposit made", 0, 0])
            ws7.append([])
            investor_funds = ws7.max_row

            ws7.append(
                ["Total Cash", f"=sum(B{momentum_funds}:B{investor_funds})",
                 f"=sum(B{momentum_funds}:B{investor_funds})"])
            total_funds = ws7.max_row
            ws7.append([])
            ws7.append(
                ["Total funds available to draw", f"=B{total_funds}+B{total_income}",
                 f"=C{total_funds}+C{total_income}"])
            total_funds_draw = ws7.max_row
            ws7.append([])
            if project == "Consolidated":
                ws7.append(["Cost to complete Heron Projects (CPC)", f"='Cashflow Projection'!D{block_costs_start - 1}",
                            f"='Cashflow Projection'!D{block_costs_start - 1}"])
            elif project == "Heron":
                ws7.append(["Cost to complete Heron Projects (CPC)",
                            f"=SUMIFS('Cashflow Projection - Heron'!$D${block_costs_start}:$D${block_costs_end},'Cashflow Projection - Heron'!$E${block_costs_start}:$E${block_costs_end},\"Heron View\")",
                            0])
            else:
                ws7.append(["Cost to complete Heron Projects (CPC)",
                            f"=SUMIFS('Cashflow Projection - {project}'!$D${block_costs_start}:$D${block_costs_end},'Cashflow Projection - {project}'!$E${block_costs_start}:$E${block_costs_end},\"{project}\")",
                            0])
            cpc = ws7.max_row
            if project == "Consolidated":
                ws7.append(["Company Running Costs", f"='Cashflow Projection'!D{operating_expenses}",0])
                company_running_costs = ws7.max_row
                ws7.append(["VAT Payable",
                            f"='Cashflow Projection'!D{vat_payable_on_sales}+'Cashflow Projection'!D{vat_row}+'Cashflow Projection'!D{vat_construction}",0])
                vat_payable = ws7.max_row

            elif project == "Heron":
                ws7.append(["Company Running Costs", f"='Cashflow Projection - Heron'!D{operating_expenses}",
                            0])
                company_running_costs = ws7.max_row
                # "='Cashflow Projection - Heron'!D26+'Cashflow Projection - Heron'!D28+'Cashflow Projection - Heron'!D64"
                ws7.append(["VAT Payable",
                            f"='Cashflow Projection - Heron'!D{vat_payable_on_sales}+'Cashflow Projection - Heron'!D{vat_row}+'Cashflow Projection - Heron'!D{vat_construction}",
                            0])
                vat_payable = ws7.max_row

            else:
                ws7.append(["Company Running Costs", f"='Cashflow Projection - {project}'!D{operating_expenses}", 0])
                company_running_costs = ws7.max_row
                ws7.append(["VAT Payable", f"='Cashflow Projection - {project}'!D{vat_payable_on_sales}+'Cashflow Projection - {project}'!D{vat_row}+'Cashflow Projection - {project}'!D{vat_construction}", 0])
                vat_payable = ws7.max_row

            ws7.append([])
            ws7.append(["Project Costs", f"=B{cpc}+B{company_running_costs}+B{vat_payable}",
                        f"=C{cpc}+C{company_running_costs}+C{vat_payable}"])
            project_costs = ws7.max_row
            ws7.append([])
            ws7.append(
                ["NETT EFFECT", f"=B{total_funds_draw}+B{project_costs}", f"=C{total_funds_draw}+C{project_costs}"])
            nett_effect = ws7.max_row
            if project != "Consolidated":
                "=B31-'Cashflow Projection - Heron'!D69"
                ws7[f"D{nett_effect}"] = f"=B{nett_effect}-'Cashflow Projection - {project}'!D{running}"

            style = 'R #,##0'
            rows_chosen = ['A', 'B', 'C']
            for i in range(transfer_endulini, ws7.max_row + 1):
                for index, row in enumerate(rows_chosen):
                    if index == 0:
                        ws7[f"{row}{i}"].font = Font(bold=False, size=18)
                    elif index == 1 or index == 2:
                        ws7[f"{row}{i}"].font = Font(bold=False, size=18)
                        ws7[f"{row}{i}"].number_format = style

            ws7[f"A{total_income}"].font = Font(bold=True, size=18)
            ws7[f"B{total_income}"].font = Font(bold=True, size=18)
            ws7[f"B{total_income}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            ws7[f"C{total_income}"].font = Font(bold=True, size=18)
            ws7[f"C{total_income}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))

            ws7[f"A{total_funds}"].font = Font(bold=True, size=18)
            ws7[f"B{total_funds}"].font = Font(bold=True, size=18)
            ws7[f"B{total_funds}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            ws7[f"C{total_funds}"].font = Font(bold=True, size=18)
            ws7[f"C{total_funds}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))

            ws7[f"A{total_funds_draw}"].font = Font(bold=True, size=18)
            ws7[f"B{total_funds_draw}"].font = Font(bold=True, size=18)
            ws7[f"B{total_funds_draw}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            ws7[f"C{total_funds_draw}"].font = Font(bold=True, size=18)
            ws7[f"C{total_funds_draw}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))

            ws7[f"A{project_costs}"].font = Font(bold=True, size=18)
            ws7[f"B{project_costs}"].font = Font(bold=True, size=18)
            ws7[f"B{project_costs}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            ws7[f"C{project_costs}"].font = Font(bold=True, size=18)
            ws7[f"C{project_costs}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))

            ws7[f"A{nett_effect}"].font = Font(bold=True, size=18)
            ws7[f"B{nett_effect}"].font = Font(bold=True, size=18)
            ws7[f"B{nett_effect}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))
            ws7[f"C{nett_effect}"].font = Font(bold=True, size=18)
            ws7[f"C{nett_effect}"].border = Border(top=Side(style='medium'), bottom=Side(style='double'))

            # hide column C
            ws7.column_dimensions['C'].hidden = True

            # make column A 77.5 wide
            ws7.column_dimensions['A'].width = 77.5
            ws7.column_dimensions['B'].width = 26

            # hide rows 17 & 18
            ws7.row_dimensions[15].hidden = True
            ws7.row_dimensions[16].hidden = True
            ws7.row_dimensions[17].hidden = True
            ws7.row_dimensions[18].hidden = True
            # ws7.row_dimensions[5].hidden = True
            ws7.row_dimensions[9].hidden = True

            for i in range(2, ws7.max_row + 3):
                ws7[f'A{i}'].border = Border(left=Side(style='medium'))
                if i in [11, 21, 23, 29, 31]:
                    ws7[f"B{i}"].border = Border(right=Side(style='medium'), top=Side(style='medium'),
                                                 bottom=Side(style='double'))
                else:
                    ws7[f'B{i}'].border = Border(right=Side(style='medium'))

            for i in range(ws7.max_row + 1, ws7.max_row + 2):
                ws7[f"A{i}"].border = Border(left=Side(style='medium'), bottom=Side(style='medium'))
                ws7[f"B{i}"].border = Border(right=Side(style='medium'), bottom=Side(style='medium'))

            ws['T4'] = "Interest to Date"

            # "=IF(J5<>"",M5*N5/365/100*('Cashflow Projection'!$B$2-Investors!J5),0)"
            inv_last = ws.max_row
            for i in range(5, inv_last + 1):
                ws[f"T{i}"] = f"=IF(J{i}<>\"\",M{i}*N{i}/365/100*('Cashflow Projection'!$B$2-Investors!J{i}),0)"
                ws[f"T{i}"].number_format = '#,##0'
                # "=IF(J5<>"",M5*N5/100/365*(K5-J5),M5*N5/100/365*(K5-(I5+30)))"
                ws[f"R{i}"] = f"=IF(J{i}<>\"\",M{i}*N{i}/100/365*(K{i}-J{i}),M{i}*N{i}/100/365*(K{i}-(I{i}+30)))"
                # ws[f"S{i}"].number_format = '#,##0'
                # "=R5+Q5"
                ws[f"S{i}"] = f"=R{i}+Q{i}"
                # ws[f"S{i}"].number_format = '#,##0'
            "=+D72+D46"
            ws6[f"D{running}"] = f"=D{running - 2}+D{block_costs_start - 4}"

            ws6.column_dimensions.group('F', 'H', hidden=False)
            ws6.row_dimensions.group(toggles_start, toggles_end, hidden=False)
            ws6.row_dimensions.group(vat_income - 1, vat_payable_on_sales - 1, hidden=False)
            ws6.row_dimensions.group(funds_available_start, funds_available_end - 1, hidden=False)
            ws6.row_dimensions.group(block_costs_start, block_costs_end + 1, hidden=False)
            # ws6.row_dimensions.group(block_costs_start + 1, block_costs_end + 1, hidden=True)
            # ws6.row_dimensions.group(block_costs_start, block_costs_end, hidden=True)

            ws6.print_area = f'A1:AF{running}'

            # Set page orientation to landscape
            ws6.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            # Set scale to fit 1 page wide by 1 pages tall
            ws6.page_setup.scale = 50
            ws6.page_setup.fitToWidth = 1
            ws6.page_setup.fitToHeight = 1

            # Set page margins
            ws6.page_margins.left = 0.5
            ws6.page_margins.right = 0.5
            ws6.page_margins.top = 2.75
            ws6.page_margins.bottom = 0.75



            if idx == 0:
                ws9 = wb.create_sheet('NSST Print')
            else:
                ws9 = wb.create_sheet(f'NSST Print - {project}')
            # make tab color red
            ws9.sheet_properties.tabColor = tab_colors[idx]
            if project != "Goodwood":
                ws9.append(["", "", "", "", "C.3_d"])
            else:
                ws9.append(["", "", "", "", "C.5.2"])
            # Merger the first 4 cells in the first row
            ws9.merge_cells('A1:D1')
            # make D1 bold, make the font 20 and put a border around it
            ws9["E1"].font = Font(bold=True, size=22)
            # center the text
            ws9["E1"].alignment = Alignment(horizontal='center', vertical='center')
            ws9["E1"].border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'),
                                      right=Side(style='medium'))


            ws9.append([f"NSST {project.upper()} REPORT"])
            ws9.append(["Report Date", report_date])

            ws9["B3"].number_format = 'dd MMMM yyyy'
            # print(invest[0])

            if project == "Heron":
                ws9.append(["Development", "Heron Fields and Heron View Open Projects"])
            else:
                ws9.append(["Development", f"{project.title()} Open Projects"])
            ws9.append(["CAPITAL"])
            if project == "Consolidated":
                ws9.append(["Total Investment capital to be raised (Estimated)",
                            f"=SUMIFS(Opportunities!$E:$E,Opportunities!$A:$A,\"<>\"&\"\")"])
            elif project == "Heron":
                ws9.append(["Total Investment capital to be raised (Estimated)",
                            f"=SUMIFS(Opportunities!$E:$E,Opportunities!$A:$A,\"Heron Fields\")+SUMIFS(Opportunities!$E:$E,Opportunities!$A:$A,\"Heron View\")"])
            else:
                ws9.append(["Total Investment capital to be raised (Estimated)",
                            f"=SUMIFS(Opportunities!$E:$E,Opportunities!$A:$A,\"{project}\")"])

            if project == "Consolidated":
                "=SUMIFS(Investors!$M:$M,Investors!$I:$I," <= "&'NSST Print'!$B$3)"
                ws9.append(["Total Investment capital received",
                            "=SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3)"])
            elif project == "Heron":

                ws9.append(["Total Investment capital received",
                            "=SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$D:$D,\"Heron Fields\")+SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$D:$D,\"Heron View\")"])
                ws9.append(["Total Investment capital received - Re-invested",0])

            else:
                ws9.append(["Total Investment capital received",
                            f"=SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$D:$D,\"{project}\")"])
            if project == "Consolidated":

                ws9.append(["Total Funds Drawn Down into Development",
                            "=SUMIFS(Investors!$M:$M,Investors!$J:$J,\"<=\"&'NSST Print'!B3)"])
                ws9.append(["Momentum Investment Account", "=Cashflow!B13"])
                ws9.append(["Capital not Raised", "=B6-B7-B11"])

            elif project == "Heron":

                ws9.append(["Total Funds Drawn Down into Development",
                            "=SUMIFS(Investors!$M:$M,Investors!$J:$J,\"<=\"&'NSST Print'!B3,Investors!$D:$D,\"Heron Fields\")+SUMIFS(Investors!$M:$M,Investors!$J:$J,\"<=\"&'NSST Print'!B3,Investors!$D:$D,\"Heron View\")"])
                ws9.append(["Momentum Investment Account", f"='Cashflow - Heron'!B13"])
                # ws9.append(["Capital not Raised", "=B6-B7-B11"])

            else:
                ws9.append(["Total Funds Drawn Down into Development",
                            f"=SUMIFS(Investors!$M:$M,Investors!$J:$J,\"<=\"&'NSST Print'!B3,Investors!$D:$D,\"{project}\")"])
                ws9.append(["Momentum Investment Account", f"='Cashflow - {project}'!B13"])
                ws9.append(["Capital not Raised", "=B6-B7-B11"])
            ws9.append(["Available to be raised (Estimated)",
                        "=IF((B6-B7)<=0,0,B6-B7)"])

            # I AM HERE
            if project == "Consolidated":

                ws9.append(["Capital repaid",
                            "=SUMIFS(Investors!$M:$M,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3)"])

                ws9.append(["Current Investor Capital deployed",
                            "=SUMIFS(Investors!$M:$M,Investors!$K:$K,\">\"&'NSST Print'!$B$3,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$P:$P,FALSE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)"])
            elif project == "Heron":

                ws9.append(["Capital repaid",
                            "=SUMIFS(Investors!$M:$M,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron Fields\")+SUMIFS(Investors!$M:$M,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron View\")"])

                ws9.append(["Current Investor Capital deployed",
                            f"=SUMIFS(Investors!$M:$M,Investors!$K:$K,\">\"&'NSST Print - Heron'!$B$3,Investors!$I:$I,\"<=\"&'NSST Print - Heron'!$B$3,Investors!$P:$P,FALSE,Investors!$D:$D,\"Heron Fields\",Investors!$I:$I,\"<=\"&'NSST Print - Heron'!$B$3)+SUMIFS(Investors!$M:$M,Investors!$K:$K,\">\"&'NSST Print - Heron'!$B$3,Investors!$P:$P,FALSE,Investors!$D:$D,\"Heron View\",Investors!$I:$I,\"<=\"&'NSST Print - Heron'!$B$3)"])
            else:

                ws9.append(["Capital repaid",
                            f"=SUMIFS(Investors!$M:$M,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"{project}\")"])

                ws9.append(["Current Investor Capital deployed",
                            f"=SUMIFS(Investors!$M:$M,Investors!$K:$K,\">\"&'NSST Print - {project}'!$B$3,Investors!$I:$I,\"<=\"&'NSST Print - {project}'!$B$3,Investors!$P:$P,FALSE,Investors!$D:$D,\"{project}\",Investors!$I:$I,\"<=\"&'NSST Print - {project}'!$B$3)"])

            ws9.append(["INVESTMENTS"])
            if project == "Consolidated":

                ws9.append(
                    ["No. of Capital Investments received",
                     "=COUNTIFS(Investors!$G:$G,FALSE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)+COUNTIFS(Investors!$G:$G,TRUE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)"])
                # "=COUNTIFS(Investors!$G:$G,TRUE,Investors!$K:$K,"<="&'NSST Print'!$B$3)+COUNTIFS(Investors!$G:$G,FALSE,Investors!$K:$K,"<="&'NSST Print'!$B$3)"]
                ws9.append(["No. Investments exited to date",
                            "=COUNTIFS(Investors!$G:$G,TRUE,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3)+COUNTIFS(Investors!$G:$G,FALSE,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3)"])
            elif project == "Heron":
                # "=COUNTIFS(Investors!$G:$G,FALSE,Investors!$I:$I,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron Fields")+COUNTIFS(Investors!$G:$G,TRUE,Investors!$I:$I,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron Fields")+COUNTIFS(Investors!$G:$G,FALSE,Investors!$I:$I,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron View")+COUNTIFS(Investors!$G:$G,TRUE,Investors!$I:$I,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron View")"
                ws9.append(
                    ["No. of Capital Investments received",
                     "=COUNTIFS(Investors!$G:$G,FALSE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron Fields\")+COUNTIFS(Investors!$G:$G,TRUE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron Fields\")+COUNTIFS(Investors!$G:$G,FALSE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron View\")+COUNTIFS(Investors!$G:$G,TRUE,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron View\")"])
                # "=COUNTIFS(Investors!$G:$G,TRUE,Investors!$K:$K,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron Fields")+COUNTIFS(Investors!$G:$G,FALSE,Investors!$K:$K,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron Fields")+COUNTIFS(Investors!$G:$G,TRUE,Investors!$K:$K,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron View")+COUNTIFS(Investors!$G:$G,FALSE,Investors!$K:$K,"<="&'NSST Print'!$B$3,Investors!$D:$D,"Heron View")"
                ws9.append(["No. Investments exited to date",
                            "=COUNTIFS(Investors!$G:$G,TRUE,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron Fields\")+COUNTIFS(Investors!$G:$G,FALSE,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron Fields\")+COUNTIFS(Investors!$G:$G,TRUE,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron View\")+COUNTIFS(Investors!$G:$G,FALSE,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron View\")"])
            else:
                # "=COUNTIFS(Investors!$G:$G,FALSE,Investors!$D:$D,"Goodwood",Investors!$I:$I,"<="&'NSST Print - Goodwood'!B3)+COUNTIFS(Investors!$G:$G,TRUE,Investors!$D:$D,"Goodwood",Investors!$I:$I,"<="&'NSST Print - Goodwood'!B3)"
                ws9.append(
                    ["No. of Capital Investments received",
                     f"=COUNTIFS(Investors!$G:$G,FALSE,Investors!$D:$D,\"{project}\",Investors!$I:$I,\"<=\"&'NSST Print - {project}'!B3)+COUNTIFS(Investors!$G:$G,TRUE,Investors!$D:$D,\"{project}\",Investors!$I:$I,\"<=\"&'NSST Print - {project}'!B3)"])
                # "=COUNTIFS(Investors!$G:$G,TRUE,Investors!$D:$D,"Goodwood",Investors!$K:$K,"<="&'NSST Print - Goodwood'!$B$3)+COUNTIFS(Investors!$G:$G,FALSE,Investors!$O:$O,TRUE,Investors!$P:$P,FALSE,Investors!$D:$D,"Goodwood",Investors!$K:$K,"<="&'NSST Print - Goodwood'!$B$3)"
                ws9.append(["No. Investments exited to date",
                            f"=COUNTIFS(Investors!$G:$G,TRUE,Investors!$D:$D,\"{project}\",Investors!$K:$K,\"<=\"&'NSST Print - {project}'!B3)+COUNTIFS(Investors!$G:$G,FALSE,Investors!$O:$O,TRUE,Investors!$P:$P,FALSE,Investors!$D:$D,\"{project}\",Investors!$K:$K,\"<=\"&'NSST Print - {project}'!B3)"])

            ws9.append(["No. Investments still in Development", "=B15-B16"])
            ws9.append(["SALES INCOME"])
            ws9.append(["", "Total", "Transferred", "Sold", "Remaining"])
            if project == "Consolidated":
                ws9.append(["Units",
                            "=COUNTIFS(Sales!$E:$E,TRUE)+COUNTIFS(Sales!$E:$E,FALSE)",
                            "=COUNTIFS(Sales!$E:$E,TRUE)",
                            "=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)",
                            "=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)"])
                ws9.append(["Sales Income",
                            "=SUMIFS(Sales!$K:$K,Sales!$D:$D,TRUE)+SUMIFS(Sales!$K:$K,Sales!$D:$D,FALSE)+C52",
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,TRUE)",
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)",
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)+C52"])
                # "=(SUMIFS(Sales!$O:$O,Sales!$D:$D,TRUE)+SUMIFS(Sales!$O:$O,Sales!$D:$D,FALSE))/1.15-C52"
                ws9.append(["Commission",
                            "=(SUMIFS(Sales!$O:$O,Sales!$D:$D,TRUE)+SUMIFS(Sales!$O:$O,Sales!$D:$D,FALSE))/1.15-C58",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,TRUE)/1.15",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)/1.15",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)/1.15-C58"])

                ws9.append(["Transfer Fees",
                            "=SUMIFS(Sales!$L:$L,Sales!$D:$D,TRUE)+SUMIFS(Sales!$L:$L,Sales!$D:$D,FALSE)",
                            "=SUMIFS(Sales!$L:$L,Sales!$E:$E,TRUE)",
                            "=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)",
                            "=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)"])

                ws9.append(["Bond Registration",
                            "=SUMIFS(Sales!$P:$P,Sales!$D:$D,TRUE)+SUMIFS(Sales!$P:$P,Sales!$D:$D,FALSE)",
                            "=SUMIFS(Sales!$P:$P,Sales!$E:$E,TRUE)",
                            "=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)",
                            "=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)"])

                ws9.append(["Security Release Fee",
                            "=SUMIFS(Sales!$M:$M,Sales!$D:$D,TRUE)+SUMIFS(Sales!$M:$M,Sales!$D:$D,FALSE)",
                            "=SUMIFS(Sales!$M:$M,Sales!$E:$E,TRUE)",
                            "=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)",
                            "=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)"])

                ws9.append(["Unforseen (0.05%)",
                            "=SUMIFS(Sales!$N:$N,Sales!$D:$D,TRUE)+SUMIFS(Sales!$N:$N,Sales!$D:$D,FALSE)-C59",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,TRUE)",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE)",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE)"])

            elif project == "Heron":
                ws9.append(["Units",
                            "=COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+COUNTIFS(Sales!$E:$E,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")+COUNTIFS(Sales!$E:$E,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")"])
                ws9.append(["Sales Income",
                            "=SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,FALSE)+SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,FALSE)+C52",
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$K:$K,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")+C52"])

                ws9.append(["Commission",
                            "=(SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,FALSE)+SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,FALSE))/1.15-C58",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")/1.15+SUMIFS(Sales!$O:$O,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")/1.15",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")/1.15+SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")/1.15",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")/1.15+SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")/1.15-C58"])

                ws9.append(["Transfer Fees",
                            "=SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,FALSE)+SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,FALSE)",
                            "=SUMIFS(Sales!$L:$L,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$L:$L,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")"])

                ws9.append(["Bond Registration",
                            "=SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,FALSE)+SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,FALSE)",
                            "=SUMIFS(Sales!$P:$P,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$P:$P,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")"])

                ws9.append(["Security Release Fee",
                            "=SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,FALSE)+SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,FALSE)",
                            "=SUMIFS(Sales!$M:$M,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$M:$M,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")"])

                ws9.append(["Unforseen (0.05%)",
                            "=SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"Heron Fields\",Sales!$D:$D,FALSE)+SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"Heron View\",Sales!$D:$D,FALSE)-C59",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$N:$N,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"Heron View\")",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron Fields\")+SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"Heron View\")-C59"])



            elif project == "Goodwood":

                ws9.append(["Units",
                            f"=COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")+COUNTIFS(Sales!$E:$E,FALSE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])
                ws9.append(["Sales Income",
                            f"=SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)+C46",
                            f"=SUMIFS(Sales!$K:$K,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")+C46"])

                ws9.append(["Commission",
                            f"=(SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE))-C52",
                            f"=SUMIFS(Sales!$O:$O,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")-C52"])

                ws9.append(["Transfer Fees",
                            f"=SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)",
                            f"=SUMIFS(Sales!$L:$L,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])

                ws9.append(["Bond Registration",
                            f"=SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)",
                            f"=SUMIFS(Sales!$P:$P,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])

                ws9.append(["Security Release Fee",
                            f"=SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)",
                            f"=SUMIFS(Sales!$M:$M,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])

                ws9.append(["Unforseen (0.05%)",
                            f"=SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)-C53",
                            f"=SUMIFS(Sales!$N:$N,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")-C53"])

            else:
                ws9.append(["Units",
                            f"=COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")+COUNTIFS(Sales!$E:$E,FALSE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=COUNTIFS(Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=COUNTIFS(Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])
                ws9.append(["Sales Income",
                            f"=SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$K:$K,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)+C46",
                            f"=SUMIFS(Sales!$K:$K,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")+C46"])

                ws9.append(["Commission",
                            f"=(SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$O:$O,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE))/1.15-C52",
                            f"=SUMIFS(Sales!$O:$O,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")/1.15",
                            f"=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")/1.15",
                            f"=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")/1.15-C52"])

                ws9.append(["Transfer Fees",
                            f"=SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$L:$L,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)",
                            f"=SUMIFS(Sales!$L:$L,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$L:$L,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])

                ws9.append(["Bond Registration",
                            f"=SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$P:$P,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)",
                            f"=SUMIFS(Sales!$P:$P,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$P:$P,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])

                ws9.append(["Security Release Fee",
                            f"=SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$M:$M,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)",
                            f"=SUMIFS(Sales!$M:$M,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$M:$M,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")"])

                ws9.append(["Unforseen (0.05%)",
                            f"=SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,TRUE)+SUMIFS(Sales!$N:$N,Sales!$A:$A,\"=\"&\"{project}\",Sales!$D:$D,FALSE)-C53",
                            f"=SUMIFS(Sales!$N:$N,Sales!$E:$E,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,TRUE,Sales!$A:$A,\"=\"&\"{project}\")",
                            f"=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"=\"&\"{project}\")-C53"])

            ws9.append(
                ["Transfer Income", "=B21-SUM(B22:B26)", "=C21-SUM(C22:C26)", "=D21-SUM(D22:D26)", "=E21-SUM(E22:E26)"])
            ws9.append(["INTEREST"])
            if project == "Consolidated":


                ws9.append(["Total Estimated Interest",
                            "=SUMIFS(Investors!$S:$S,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)"])

                ws9.append(["Interest Paid to Date",
                            "=SUMIFS(Investors!$S:$S,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3)"])
                ws9.append(["Remaining Interest to Exit", "=B29-B30"])

                ws9.append(["Interest Due to date",
                            "=SUMIFS(Investors!$T:$T,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE)+SUMIFS(Investors!$Q:$Q,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE)"])
                ws9.append(["Interest Due to be Earned to Exit", "=B31-B32"])

            elif project == "Heron":

                ws9.append(["Total Estimated Interest",
                            f"=SUMIFS(Investors!$S:$S,Investors!$D:$D,\"Heron Fields\",Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)+SUMIFS(Investors!$S:$S,Investors!$D:$D,\"Heron View\",Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)"])
                # "=SUMIFS(Investors!$S:$S,Investors!$K:$K,"<="&$B$3,Investors!$D:$D,"Heron Fields")+SUMIFS(Investors!$S:$S,Investors!$K:$K,"<="&$B$3,Investors!$D:$D,"Heron View")"
                ws9.append(["Interest Paid to Date",
                            f"=SUMIFS(Investors!$S:$S,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron Fields\")+SUMIFS(Investors!$S:$S,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"Heron View\")"])
                ws9.append(["Remaining Interest to Exit", "=B29-B30"])

                ws9.append(["Interest Due to date",
                            "=SUMIFS(Investors!$T:$T,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$D:$D,\"Heron Fields\")+SUMIFS(Investors!$Q:$Q,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$D:$D,\"Heron Fields\")+SUMIFS(Investors!$T:$T,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$D:$D,\"Heron View\")+SUMIFS(Investors!$Q:$Q,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$D:$D,\"Heron View\")"])
                ws9.append(["Interest Due to be Earned to Exit", "=B31-B32"])

            else:

                ws9.append(["Total Estimated Interest",
                            f"=SUMIFS(Investors!$S:$S,Investors!$D:$D,\"{project}\",Investors!$I:$I,\"<=\"&'NSST Print'!$B$3)"])

                ws9.append(["Interest Paid to Date",
                            f"=SUMIFS(Investors!$S:$S,Investors!$K:$K,\"<=\"&'NSST Print'!$B$3,Investors!$D:$D,\"{project}\")"])
                ws9.append(["Remaining Interest to Exit", "=B29-B30"])

                ws9.append(["Interest Due to date",
                            f"=SUMIFS(Investors!$T:$T,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$D:$D,\"{project}\")+SUMIFS(Investors!$Q:$Q,Investors!$G:$G,FALSE,Investors!$O:$O,FALSE,Investors!$D:$D,\"{project}\")"])
                ws9.append(["Interest Due to be Earned to Exit", "=B31-B32"])

            ws9.append(["FUNDING AVAILABLE"])
            if project == "Consolidated":
                ws9.append(["Funding", "=Cashflow!B21"])
                ws9.append(["Projected Heron Projects Income", "=Cashflow!B11+C46"])
                ws9.append(["Available Funding", "=B35+B36"])
                ws9.append(["Current Construction Cost", "=-Cashflow!B29+C47+C48+C49+C50+C51"])
                ws9.append(["Total funds (required)/Surplus - To Dec 2024", "=B37-B38"])
            elif project == "Heron":
                ws9.append(["Funding", f"='Cashflow - {project}'!B21"])
                ws9.append(["Projected Heron Projects Income", f"='Cashflow - {project}'!B11+C52"])
                ws9.append(["Available Funding", "=B35+B36"])
                ws9.append(["Current Construction Cost", f"=-'Cashflow - {project}'!B29+C53+C54+C55+C56+C57"])
                ws9.append(["Total funds (required)/Surplus - To Dec 2024", "=B37-B38"])
            elif project == "Goodwood":
                ws9.append(["Funding", f"='Cashflow - {project}'!B21"])
                ws9.append(["Projected Purple Blok Projects Income", f"='Cashflow - {project}'!B11+C46"])
                ws9.append(["Available Funding", "=B35+B36"])
                ws9.append(["Current Construction Cost", f"=-'Cashflow - {project}'!B29+C47+C48+C49+C50+C51"])
                ws9.append(["Total funds (required)/Surplus - To Dec 2024", "=B37-B38"])
            else:
                ws9.append(["Funding", f"='Cashflow - {project}'!B21"])
                ws9.append(["Projected Heron Projects Income", f"='Cashflow - {project}'!B11+C46"])
                ws9.append(["Available Funding", "=B35+B36"])
                ws9.append(["Current Construction Cost", f"=-'Cashflow - {project}'!B29+C47+C48+C49+C50+C51"])
                ws9.append(["Total funds (required)/Surplus - To Dec 2024", "=B37-B38"])
            ws9.append(["PROJECTED PROFIT"])
            if project == "Goodwood":
                ws9.append(["Projected Nett Income", "=+B27"])
                ws9.append(
                    ["Other Income (interest received)",
                     0])
                ws9.append(["Total Estimated Development and Sales Costs", 0])
                # fill in yellow
                estimated_costs_goodwood = ws9[f'B{ws9.max_row}']
                estimated_costs_goodwood.fill = PatternFill(start_color="EBF400", end_color="EBF400", fill_type="solid")
                ws9.append(["Capital and Interest Expense", "=B29-C51+B7"])
                ws9.append(["Profit", "=B41-B43-B44"])

            elif project == "Heron":
                ws9.append(["Projected Nett Income", "=+B21+SUM(B56:B63)"])
                ws9.append(
                    ["Other Income (interest received)",
                     "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,\"Interest Received - Momentum\")"])
                ws9.append(["Total Estimated Development and Sales Costs", "=+B84-C53-C54-C55-C56+SUM(D22:E26)"])
                ws9.append(["Interest Expense", "=B29-B30-C51"])
                ws9.append(["Profit", "=B41+B42-B43-B44"])
                # print("XXXXX",ws9.max_row)
                ws9.append(["TRANSACTION SUMMARY","GROSS","","NETT",""])
                ws9.append(["Gross Income still due from Sales income", "=(SUM(D21:E21)*1.15)","","=SUM(D27:E27)",""])
                ws9.append(["Cost to complete project (CPC)", "=0*115/100","","=B48/1.15",""])
                cpc = ws9[f'B{ws9.max_row}']
                cpc.fill = PatternFill(start_color="EBF400", end_color="EBF400", fill_type="solid")
                ws9.append(["Gross Income", "=B47-B48","","=D47-D48",""])
                ws9.append(["Investor Capital & Interest due", "='Investor Exit List - Heron'!Q4","","=B50",""])
                ws9.append(["Surplus / Shortfall", "=B49-B50","","=D49-D50",""])


            else:
                ws9.append(["Projected Nett Income", "=+B21+SUM(B56:B63)"])
                ws9.append(
                    ["Other Income (interest received)",
                     "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,\"Interest Received - Momentum\")"])
                ws9.append(["Total Estimated Development and Sales Costs", "=+B78-C47-C48-C49-C50+SUM(D22:E26)"])
                ws9.append(["Interest Expense", "=B29-B30-C51"])
                ws9.append(["Profit", "=B41+B42-B43-B44"])

            if project == "Heron":
                ws9.append(["Sales Increase", 0,
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")*B52",
                            "",
                            "=C52/$E$20"])
                ws9.append(
                    ["CPC Construction", 0, "=SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$D:$D,FALSE)*B53",
                     "=SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$D:$D,FALSE)-C53", "=C53/$E$20"])
                ws9.append(["Rent Salaries and wages", 0, "=400000*B54",
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A48,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)-C54",
                            "=C54/$E$20"])
                ws9.append(["CPSD", 0,
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A49,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)*B55",
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A49,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)-C55",
                            "=C55/$E$20"])
                ws9.append(["OppInvest", 0,
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A75,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)*B56",
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A75,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)-C56",
                            "=C56/$E$20"])
                ws9.append(["investor interest", 0, "=(B29-B30)*B57", "=(B29-B30)-C57", "=C57/$E$20"])
                ws9.append(["Commissions", 0,
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")/1.15*B58",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")/1.15-C58",
                            "=C58/$E$20"])
                ws9.append(["Unforseen", 0,
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")*B59",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")-C59",
                            "=C53/$E$20"])
            else:
                ws9.append(["Sales Increase", 0,
                            "=SUMIFS(Sales!$K:$K,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")*B46",
                            "",
                            "=C46/$E$20"])
                ws9.append(
                    ["CPC Construction", 0,
                     "=SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$D:$D,FALSE)*B47",
                     "=SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$D:$D,FALSE)-C47", "=C47/$E$20"])
                ws9.append(["Rent Salaries and wages", 0, "=400000*B48",
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A48,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)-C48",
                            "=C48/$E$20"])
                ws9.append(["CPSD", 0,
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A49,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)*B49",
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A49,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)-C49",
                            "=C49/$E$20"])
                ws9.append(["OppInvest", 0,
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A75,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)*B50",
                            "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A75,'Other Costs'!$C:$C,\">\"&'NSST Print'!$B$3)-C50",
                            "=C50/$E$20"])
                ws9.append(["investor interest", 0, "=(B29-B30)*B51", "=(B29-B30)-C51", "=C51/$E$20"])
                ws9.append(["Commissions", 0,
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")/1.15*B52",
                            "=SUMIFS(Sales!$O:$O,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")/1.15-C52",
                            "=C52/$E$20"])
                ws9.append(["Unforseen", 0,
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")*B53",
                            "=SUMIFS(Sales!$N:$N,Sales!$E:$E,FALSE,Sales!$D:$D,FALSE,Sales!$A:$A,\"<>\"&\"Endulini\")-C53",
                            "=C53/$E$20"])
            ws9.append([])
            ws9.append([])
            ws9.append(["Rental Income", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A56)"])
            ws9.append(["Sales - Heron Fields occupational rent", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A57)"])
            ws9.append(["Sales - Heron View Occupational Rent", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A58)"])
            ws9.append(["Bond Origination", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A59)"])
            ws9.append(["Interest Received - Deposit Grey Heron", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A60)"])
            ws9.append(["Interest Received - Deposits", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A61)"])
            ws9.append(["Interest Received - STBB Trust Account", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A62)"])
            ws9.append(["Revaluation of Land (Not Incl)", "=-SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A63)"])
            ws9.append([])
            ws9.append([])
            # "=SUMIFS(Xero!$F:$F,Xero!$C:$C,"Expenses",Xero!$A:$A,"="&Heron!$A$4)+SUMIFS(Xero!$F:$F,Xero!$C:$C,"Expenses",Xero!$A:$A,"="&Heron!$A$5)"
            ws9.append(["Costs",
                        "=SUMIFS(Xero!$F:$F,Xero!$C:$C,\"Expenses\",Xero!$A:$A,\"=\"&Heron!$A$4)+SUMIFS(Xero!$F:$F,Xero!$C:$C,\"Expenses\",Xero!$A:$A,\"=\"&Heron!$A$5)"])
            ws9.append(["Heron View Land", "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A67)"])
            ws9.append(["COS - Heron Fields - Construction", "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A68)"])
            ws9.append(["COS - Heron Fields - P & G", "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A69)"])
            ws9.append(
                ["COS - Heron Fields - Printing & Stationary", "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A70)"])
            ws9.append(["COS - Heron View - Construction",
                        "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A71)+SUMIFS('Updated Construction'!$F:$F,'Updated Construction'!$D:$D,FALSE)"])
            ws9.append(["COS - Heron View - P&G", "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A72)"])
            ws9.append(["COS - Heron View - Printing & Stationary", "=SUMIFS(Xero!$F:$F,Xero!$E:$E,'NSST Print'!$A73)"])
            ws9.append(["CPSD", "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A74)"])
            ws9.append(["OppInvest", "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A75)"])
            ws9.append(["Professional Fees", "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A76)"])
            ws9.append(
                ["Rent Salaries and Wages", "=SUMIFS('Other Costs'!$D:$D,'Other Costs'!$A:$A,'NSST Print'!$A77)"])
            ws9.append(["Total Costs", "=SUM(B66:B77)"])

            # make the first column 55 and the others 20
            for i in range(1, ws9.max_column + 1):
                if i == 1:
                    ws9.column_dimensions[get_column_letter(i)].width = 60
                else:
                    ws9.column_dimensions[get_column_letter(i)].width = 25

            # format all cells as font size 14
            for i in range(1, ws9.max_column + 1):
                for j in range(1, ws9.max_row + 1):
                    ws9[f"{get_column_letter(i)}{j}"].font = Font(size=18)
                    # After row format all cells as currency except the first column and except rows 15, 16 and 17
                    if j < 15 and j > 5 and i > 1:
                        ws9[f"{get_column_letter(i)}{j}"].number_format = 'R#,##0.00'

                    if j > 17 and j < 20 and i > 1:
                        ws9[f"{get_column_letter(i)}{j}"].number_format = 'R#,##0.00'

                    if j > 20 and i > 1:
                        ws9[f"{get_column_letter(i)}{j}"].number_format = 'R#,##0.00'
                    # After row 2, put a border around all cells until row 46
                    if j > 1 and j < 46:
                        ws9[f"{get_column_letter(i)}{j}"].border = Border(top=Side(style='thin'),
                                                                          bottom=Side(style='thin'),
                                                                          left=Side(style='thin'),
                                                                          right=Side(style='thin'))

            rows_for_full_merge = [2, 5, 14, 18, 28, 34, 40]
            for i in rows_for_full_merge:
                ws9.merge_cells(f"A{i}:E{i}")
                # make the cell bold and color it light olive and make the text white
                ws9[f"A{i}"].font = Font(bold=True, size=18, color="FFFFFF")
                ws9[f"A{i}"].fill = PatternFill(start_color="7F9F80", end_color="7F9F80", fill_type="solid")
                ws9[f"A{i}"].alignment = Alignment(horizontal='center', vertical='center')

            ws9['A2'].font = Font(bold=True, size=26, color="FFFFFF")
            ws9['E1'].font = Font(bold=True, size=22)

            if project == 'Heron':
                cols_to_format = ['A', 'B', 'C', 'D', 'E']
                for col in cols_to_format:
                    ws9[f"{col}46"].alignment = Alignment(horizontal='center', vertical='center')
                    ws9[f"{col}46"].font = Font(bold=True, size=18, color="FFFFFF")
                    ws9[f"{col}46"].fill = PatternFill(start_color="7F9F80", end_color="7F9F80", fill_type="solid")



            rows_for_half_merge = [46,47,48,49,50,51]

            if project == 'Heron':


                for i in rows_for_half_merge:
                    ws9[f"A{i}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                                 right=Side(style='thin'))
                    ws9[f"B{i}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                                 right=Side(style='thin'))
                    ws9[f"C{i}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                                 right=Side(style='thin'))
                    ws9[f"D{i}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                                 right=Side(style='thin'))
                    ws9[f"E{i}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                                 right=Side(style='thin'))
                    ws9.merge_cells(f"B{i}:C{i}")
                    ws9.merge_cells(f"D{i}:E{i}")

                # put a thin border around all the cells






            rows_for_part_merger = [3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 29, 30, 31, 32, 33, 35, 36, 37, 38,
                                    39,
                                    41, 42, 43, 44, 45]
            for i in rows_for_part_merger:
                ws9.merge_cells(f"B{i}:E{i}")
                ws9[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')



            for i in range(52, 60):

                if i == 54:
                    # format the cell in row 48 as an integer
                    ws9[f"B{i}"].number_format = '0'
                else:
                    # format as a percentage
                    ws9[f"B{i}"].number_format = '0%'

            # ws9['F43'].value = "<< SALES COSTS DO NOT APPEAR TO INCLUDED PREVIOUSLY"
            # Hide row 10, 31,32,33
            if project != "Heron":
                ws9.row_dimensions[10].hidden = True
            ws9.row_dimensions[31].hidden = True
            ws9.row_dimensions[32].hidden = True
            ws9.row_dimensions[33].hidden = True

            if project != 'Goodwood':# Hide rows 46 to 51
                for i in range(40, 46):
                    ws9.row_dimensions[i].hidden = True
            if project != "Heron":
                for i in range(56, 79):
                    ws9.row_dimensions[i].hidden = True
            else:
                for i in range(62, 85):
                    ws9.row_dimensions[i].hidden = True

            # in row 19 and 20, center the text from column B to E
            for i in range(2, 6):
                ws9[f"{get_column_letter(i)}19"].alignment = Alignment(horizontal='center', vertical='center')
                ws9[f"{get_column_letter(i)}20"].alignment = Alignment(horizontal='center', vertical='center')

            if project != "Consolidated":
                "=B12+B13-B7"
                ws9['F12'] = f"=B12+B13-B7"
                ws9['F12'].number_format = 'R#,##0.00'
                ws9['G12'] = "<CHECK"
                "=B13-'Investor Exit List - Heron'!F4"
                ws9['F13'] = f"=B13-'Investor Exit List - {project}'!F4"
                ws9['F13'].number_format = 'R#,##0.00'
                ws9['G13'] = "<CHECK"
                "=B17-'Investor Exit List - Heron'!C4"
                ws9['F17'] = f"=B17-'Investor Exit List - {project}'!C4"
                ws9['F17'].number_format = '0'
                ws9['G17'] = "<CHECK"
                "=E20+D20-'Investor Exit List - Heron'!A4"
                ws9['F20'] = f"=E20+D20-'Investor Exit List - {project}'!A4"
                ws9['F20'].number_format = '0'
                ws9['G20'] = "<CHECK"
                "=B29-B30-('Investor Exit List - Heron'!Q4-'Investor Exit List - Heron'!F4)"
                ws9['F29'] = f"=B29-B30-('Investor Exit List - {project}'!Q4-'Investor Exit List - {project}'!F4)"
                ws9['F29'].number_format = 'R#,##0.00'
                ws9['G29'] = "<CHECK"
                "=B39-'Cashflow Projection - Heron'!D69"
                variable = variables[f"{project}_running"]
                ws9['F39'] = f"=B39-'Cashflow Projection - {project}'!D{variable}"
                # if project == "Heron":
                #     ws9['F39'] = f"=B39-'Cashflow Projection - {project}'!D{heron_running}"
                # elif project == "Goodwood":
                #     ws9['F39'] = f"=B39-'Cashflow Projection - {project}'!D{goodwood_running}"
                ws9['F39'].number_format = 'R#,##0.00'
                ws9['G39'] = "<CHECK"

            if idx == 0:
                ws10 = wb.create_sheet('Investor Exit List')
            else:
                ws10 = wb.create_sheet(f'Investor Exit List - {project}')
            # make tab color red
            ws10.sheet_properties.tabColor = tab_colors[idx]
            if project != "Goodwood":
                ws10['W1'] = "C.3_e"
            else:
                ws10['W1'] = "C.5.4"
            ws10['W1'].font = Font(bold=True, size=22)
            ws10['W1'].alignment = Alignment(horizontal='center', vertical='center')
            ws10['W1'].border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'),
                                        right=Side(style='medium'))
            ws10['A2'] = f"Investor Exit List - {report_date}"
            ws10['A2'].font = Font(bold=True, color="FFFFFF", size=22)
            ws10['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws10['A2'].border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'),
                                       right=Side(style='medium'))
            # fill in blue
            ws10['A2'].fill = PatternFill(start_color="008DDA", end_color="008DDA", fill_type="solid")
            exit_headers = []
            for item in investor_exit[0]:
                exit_headers.append(item.replace("_", " ").title())
            ws10.append(exit_headers)
            # print(exit_headers)
            # merge the first row
            ws10.merge_cells(f"A2:{get_column_letter(len(investor_exit[0]))}2")
            ws10.append([])
            # make all cells in row 2 bold with a light grey background and borders all around
            for i in range(1, len(investor_exit[0]) + 1):
                ws10[f"{get_column_letter(i)}3"].font = Font(bold=True, size=13)
                ws10[f"{get_column_letter(i)}3"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3",
                                                                    fill_type="solid")
                ws10[f"{get_column_letter(i)}3"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                                 left=Side(style='thin'), right=Side(style='thin'))
                # wrap the text

                # double the row height
                ws10.row_dimensions[2].height = 55

                # ws10[f"{get_column_letter(i)}2"].alignment = Alignment(horizontal='center', vertical='center')
                ws10[f"{get_column_letter(i)}3"].alignment = Alignment(wrap_text=True)
            # make all cells in row 3 bold with a red background and font color white and borders all around
            for i in range(1, len(investor_exit[0]) + 1):
                ws10[f"{get_column_letter(i)}4"].font = Font(bold=True, size=12, color="FFFFFF")
                ws10[f"{get_column_letter(i)}4"].fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                                    fill_type="solid")
                ws10[f"{get_column_letter(i)}4"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                                 left=Side(style='thin'), right=Side(style='thin'))

            ws10.row_dimensions[3].height = 30
            # center the text

            # make all columns 20 wide except the 5th column, make it 50 wide
            for i in range(1, len(investor_exit[0]) + 1):
                if i == 5:
                    ws10.column_dimensions[get_column_letter(i)].width = 33
                else:
                    ws10.column_dimensions[get_column_letter(i)].width = 15

            # If project == "Heron" then filter out of investor_exit those which unit_number do not begin with "H"
            if project == "Heron" and project != "Consolidated":
                investor_exit_use = []
                for item in investor_exit:
                    invest_filtered = [itemI for itemI in invest if itemI['opportunity_code'] == item['unit_number']]
                    if len(invest_filtered) > 0:
                        # if the first 5 characters of invest_filtered[0]['Category'] are "Heron"
                        if invest_filtered[0]['Category'].startswith("Heron"):
                            investor_exit_use.append(item)

            elif project != "Consolidated" and project != "Heron":
                investor_exit_use = []
                for item in investor_exit:
                    invest_filtered = [itemI for itemI in invest if itemI['opportunity_code'] == item['unit_number']]
                    if len(invest_filtered) > 0:
                        if invest_filtered[0]['Category'] == project:
                            investor_exit_use.append(item)

            else:
                investor_exit_use = investor_exit

            for index, item in enumerate(investor_exit_use):

                row = []
                for key in item:
                    row.append(item[key])

                ws10.append(row)

                # print(row)
                if index == 0:
                    exit_first_row_number = ws10.max_row
            exit_last_row_number = ws10.max_row

            column_formulae = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                               'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']
            for i in column_formulae:
                for row in range(exit_first_row_number, exit_last_row_number + 1):
                    ws10[f"{i}{row}"].font = Font(bold=False, size=12)
                    if i == 'F':
                        if project == "Consolidated":
                            # "=SUMIFS(Investors!$M:$M, Investors!$I:$I, "<=" & 'NSST Print'!$B$3, Investors!$K:$K, ">" & 'NSST Print'!$B$3, Investors!$A:$A, 'Investor Exit List'!A4, Investors!$E:$E, 'Investor Exit List'!C4, Investors!$P:$P,FALSE,Investors!$L:$L,'Investor Exit List'!B4)"
                            ws10[
                                f"{i}{row}"].value = (
                                f"=SUMIFS(Investors!$M:$M, Investors!$I:$I, \"<=\" & 'NSST Print'!$B$3, Investors!$K:$K, \">\" & 'NSST Print'!$B$3, Investors!$A:$A, 'Investor Exit List'!A{row}, Investors!$E:$E, 'Investor Exit List'!C{row}, Investors!$P:$P,FALSE,Investors!$L:$L,'Investor Exit List'!B{row})")
                        else:
                            ws10[
                                f"{i}{row}"].value = (
                                f"=SUMIFS(Investors!$M:$M, Investors!$I:$I, \"<=\" & 'NSST Print'!$B$3, Investors!$K:$K, \">\" & 'NSST Print'!$B$3, Investors!$A:$A, 'Investor Exit List - {project}'!A{row}, Investors!$E:$E, 'Investor Exit List - {project}'!C{row},Investors!$P:$P,FALSE,Investors!$L:$L,'Investor Exit List - {project}'!B{row})")

                        ws10[f"{i}{row}"].number_format = 'R#,##0.00'
                    elif i == 'G' or i == 'M':
                        ws10[f"{i}{row}"].number_format = 'yyyy-mm-dd'
                    elif i == 'H':
                        # "=VLOOKUP($C4,Sales!$C:$E,2,FALSE)"
                        ws10[f"{i}{row}"].value = f"=VLOOKUP($C{row},Sales!$C:$E,2,FALSE)"
                        # if ws10[f"{i}{row}"].value in H is TRUE then fill the row in light grey
                        # print(ws10[f"C{row}"].value)



                    elif i == 'I':
                        ws10[f"{i}{row}"].value = f"=VLOOKUP($C{row},Sales!$C:$E,3,FALSE)"
                    elif i == 'J':
                        if project == "Consolidated":
                            ws10[f"{i}{row}"].value = f"=SUMIFS(Sales!$H:$H,Sales!$C:$C,'Investor Exit List'!$C{row})"
                        else:
                            ws10[
                                f"{i}{row}"].value = f"=SUMIFS(Sales!$H:$H,Sales!$C:$C,'Investor Exit List - {project}'!$C{row})"
                        ws10[f"{i}{row}"].number_format = 'yyyy-mm-dd'
                    elif i == 'K':
                        ws10[f"{i}{row}"].value = f"=IF(I{row}=TRUE,J{row},\"\")"
                        ws10[f"{i}{row}"].number_format = 'yyyy-mm-dd'
                    elif i == 'L':  # SHOULD THIS BE 731
                        ws10[f"{i}{row}"].value = f"=G{row}+731"
                        ws10[f"{i}{row}"].number_format = 'yyyy-mm-dd'
                    elif i == 'N':
                        ws10[
                            f"{i}{row}"].value = f"=IFERROR(IF(G{row}<>\"\",IF(I{row}=TRUE,0,IF(U{row}=TRUE,0,L{row}-M{row})),0),0)"
                        ws10[f"{i}{row}"].number_format = '0'
                    elif i == 'O':
                        ws10[f"{i}{row}"].value = f"=IF(N{row}>0,IF(N{row}<=180,N{row},0),0)"
                        ws10[f"{i}{row}"].number_format = '0'
                    elif i == 'P':
                        ws10[f"{i}{row}"].value = f"=IF(AND(N{row}<90,N{row}>0),Q{row},0)"
                        ws10[f"{i}{row}"].number_format = 'R#,##0.00'
                    elif i == 'Q':
                        ",Investors!$L:$L,B5"
                        # if project == "Consolidated":
                        #     "=IF(F5<>0,SUMIFS(Investors!$S:$S,Investors!$A:$A,A5,Investors!$E:$E,C5)+F5,0)"
                        ws10[
                            f"{i}{row}"].value = f"=IF(F{row}<>0,SUMIFS(Investors!$S:$S,Investors!$A:$A,A{row},Investors!$E:$E,C{row},Investors!$L:$L,B{row})+F{row},0)"
                        # else:
                        #     ws10[
                        #         f"{i}{row}"].value = f"=SUMIFS(Investors!$M:$M,Investors!$O:$O,FALSE,Investors!$P:$P,FALSE,Investors!$K:$K,\">\"&'NSST Print'!$B$3,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row})-(SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$J:$J,\">\"&'NSST Print'!$B$3,Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row})+SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$J:$J,\"\",Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row}))+SUMIFS(Investors!$S:$S,Investors!$O:$O,FALSE,Investors!$P:$P,FALSE,Investors!$K:$K,\">\"&'NSST Print'!$B$3,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row})-(SUMIFS(Investors!$S:$S,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$J:$J,\">\"&'NSST Print'!$B$3,Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row})+SUMIFS(Investors!$S:$S,Investors!$I:$I,\"<=\"&'NSST Print'!$B$3,Investors!$J:$J,\"\",Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row}))"
                        ws10[f"{i}{row}"].number_format = 'R#,##0.00'
                    elif i == 'R':
                        ws10[f"{i}{row}"].value = f"=IF(AND(H{row}=TRUE,I{row}=FALSE),Q{row},0)"
                        ws10[f"{i}{row}"].number_format = 'R#,##0.00'
                    elif i == 'S':
                        if project == "Consolidated":
                            ws10[
                                f"{i}{row}"].value = f"=SUMIFS(Investors!$M:$M,Investors!$O:$O,TRUE,Investors!$A:$A,'Investor Exit List'!$A{row},Investors!$E:$E,'Investor Exit List'!$C{row},Investors!$L:$L,'Investor Exit List'!$B{row})+SUMIFS(Investors!$S:$S,Investors!$O:$O,TRUE,Investors!$A:$A,'Investor Exit List'!$A{row},Investors!$E:$E,'Investor Exit List'!$C{row},Investors!$L:$L,'Investor Exit List'!$B{row})"
                        else:
                            ws10[
                                f"{i}{row}"].value = f"=SUMIFS(Investors!$M:$M,Investors!$O:$O,TRUE,Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row})+SUMIFS(Investors!$S:$S,Investors!$O:$O,TRUE,Investors!$A:$A,'Investor Exit List - {project}'!$A{row},Investors!$E:$E,'Investor Exit List - {project}'!$C{row},Investors!$L:$L,'Investor Exit List - {project}'!$B{row})"
                        ws10[f"{i}{row}"].number_format = 'R#,##0.00'
                    elif i == 'T':
                        ws10[f"{i}{row}"].value = f"=IF(U{row}=TRUE,$J{row},\"\")"
                        ws10[f"{i}{row}"].number_format = 'yyyy-mm-dd'
                    elif i == 'V':
                        ws10[f"{i}{row}"].value = f"=IF(AND(H{row}=TRUE,OR(K{row}=\"\",K{row}>M{row})),S{row},0)"
                        ws10[f"{i}{row}"].number_format = 'R#,##0.00'
                    elif i == 'Y':
                        # "=IF(H4=FALSE,VLOOKUP($C4,Sales!$C:$F,4,FALSE),1)"
                        ws10[f"{i}{row}"].value = f"=IF(H{row}=FALSE,VLOOKUP($C{row},Sales!$C:$F,4,FALSE),1)"
                        ws10[f"{i}{row}"].number_format = '0'
                    elif i == 'Z':
                        '=IF(Y4=0,"Release",IF(AND(L4<J4,Y4=1,H4=FALSE,U4=FALSE,MONTH(L4)&"/"&YEAR(L4)<>MONTH(J4)&"/"&YEAR(J4)),"Release",""))'
                        ws10[
                            f"{i}{row}"].value = f"=IF(Y{row}=0,\"Release\",IF(AND(L{row}<J{row},Y{row}=1,H{row}=FALSE,U{row}=FALSE,MONTH(L{row})&\"/\"&YEAR(L{row})<>MONTH(J{row})&\"/\"&YEAR(J{row})),\"Release\",\"\"))"
                    elif i == 'AA':

                        ws10[f"{i}{row}"].data_type = 'f'
                        ws10[f"{i}{row}"].value = f"=_xlfn.minifs($L:$L,$C:$C,$C{row},$F:$F,\"<>0\")"
                        # ws10[f"{i}{row}"].value = f"=+IF(Z{row}=\"Release\",AA{row-1}+30,\"\")"
                        ws10[f"{i}{row}"].number_format = "YYYY-MM-DD"

                    # put a thin border on all sides of each ro
                    elif i == 'AB':
                        # '=C4&"-"&A4'
                        ws10[f"{i}{row}"].value = f"=C{row}&\"-\"&A{row}"
                        # ws10[f"{i}{row}"].value = f""
                        # ws10[f"{i}{row}"].value = f"=+IF(Z{row}=\"Release\",AA{row-1}+30,\"\")"
                        ws10[f"{i}{row}"].number_format = "YYYY-MM-DD"

                    elif i == 'AC':
                        # '=L4'
                        ws10[f"{i}{row}"].value = f"=L{row}"
                        # ws10[f"{i}{row}"].value = f""
                        # ws10[f"{i}{row}"].value = f"=+IF(Z{row}=\"Release\",AA{row-1}+30,\"\")"
                        ws10[f"{i}{row}"].number_format = "YYYY-MM-DD"

                    elif i == 'AD':
                        "=EOMONTH(EDATE($J4, 0), 0)"
                        ws10[f"{i}{row}"].value = f"=EOMONTH(EDATE($J{row}, 0), 0)"
                        ws10[f"{i}{row}"].number_format = "YYYY-MM-DD"

                    elif i == 'AE':

                        "=IF(AND(AD505>EOMONTH(EDATE(AA505, 0), 0)),1,0)"
                        ws10[f"{i}{row}"].value = f"=IF(AND(AD{row}>EOMONTH(EDATE(AA{row}, 0), 0)),1,0)"
                        ws10[f"{i}{row}"].number_format = "0"

                    # put a thin border on all sides of each row
                    ws10[f"{i}{row}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                      left=Side(style='thin'),
                                                      right=Side(style='thin'))


            rows_to_fill = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R','S', 'T', 'U', 'V', 'W']

            for row in range(5, ws10.max_row + 1):
                # print(ws10[f"C{row}"].value)
                filtered_sales = [item for item in sales if item['opportunity_code'] == ws10[f"C{row}"].value]
                # print("filtered_sales",filtered_sales)
                if len(filtered_sales) > 0:
                    if filtered_sales[0]['sold'] == True and filtered_sales[0]['transferred'] == True:
                        for i in rows_to_fill:
                            ws10[f"{i}{row}"].fill = PatternFill(start_color="D2E0FB", end_color="D2E0FB",
                                                                fill_type="solid")
                    elif filtered_sales[0]['sold'] == True and filtered_sales[0]['transferred'] == False:
                        for i in rows_to_fill:
                            ws10[f"{i}{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                                                fill_type="solid")





            # column_letter = 'AA'
            #
            # # Iterate through the cells in the column
            # for cell in ws10[column_letter]:
            #     if cell.data_type == 'f':  # Check if the cell contains a formula
            #         # Modify the formula as needed
            #         # For example, you can update the formula to replace a character
            #         cell.value = cell.value.replace('@', '')
            #     else:
            #         # Replace the character in the cell's value
            #         cell.value = cell.value.replace('@', '')

            ws10['Y2'].value = "Build"
            ws10['Z2'].value = "Calculate Release"
            ws10['AA2'].value = "Release Date"
            ws10['AB2'].value = "UNIQUE ID"
            if project == "Consolidated" or project == "Heron":
                dev1 = ["ZZUN01", 0, "HFB215", "B", "Dev Unit", 0, "", False, False, "", "", "", "", "", "", "", "", "",
                        "", "",
                        False,
                        "", 1330640.05, False]
                ws10.append(dev1)
            # dev2 = ["ZZUN01", 0, "HFB315", "B", "Dev Unit", 0, "", False, False, "", "", "", "", "", "", "", "", "", "", "",
            #         False,
            #         "", 1311740.05, False]
            # ws10.append(dev2)
            final_row = ws10.max_row

            for i in range(final_row - 1, final_row + 1):
                for x in range(1, ws10.max_column + 1):

                    # put a border around each cell
                    ws10[f"{get_column_letter(x)}{i}"].border = Border(top=Side(style='thin'),
                                                                       bottom=Side(style='thin'),
                                                                       left=Side(style='thin'),
                                                                       right=Side(style='thin'))
                    # in columns, F & W format as currency
                    if i == 6 or i == 23:
                        ws10[f"{get_column_letter(x)}{i}"].number_format = 'R#,##0.00'

            total_columns = ['A', 'C', 'F', 'P', 'Q', 'R', 'S', 'V', 'W']

            for i in total_columns:

                if i == 'S':
                    ws10[
                        f"{i}{exit_first_row_number - 1}"].value = f"=SUBTOTAL(109,{i}{exit_first_row_number}:{i}{exit_last_row_number + 3})"
                    ws10[f"{i}{exit_first_row_number - 1}"].number_format = 'R#,##0.00'
                elif i == 'C':
                    if project == "Consolidated":
                        ws10[
                            f"{i}{exit_first_row_number - 1}"].value = f"='NSST Print'!B17"
                    elif project == "Heron":
                        ws10[
                            f"{i}{exit_first_row_number - 1}"].value = f"='NSST Print - Heron'!B17"
                    else:
                        ws10[
                            f"{i}{exit_first_row_number - 1}"].value = f"='NSST Print - {project}'!B17"

                    ws10[f"{i}{exit_first_row_number - 1}"].number_format = '0'
                elif i == 'A':
                    if project == "Consolidated":
                        ws10[
                            f"{i}{exit_first_row_number - 1}"].value = f"='NSST Print'!E20+'NSST Print'!D20"
                    elif project == "Heron":
                        ws10[
                            f"{i}{exit_first_row_number - 1}"].value = f"='NSST Print - Heron'!E20+'NSST Print - Heron'!D20"
                    else:
                        ws10[
                            f"{i}{exit_first_row_number - 1}"].value = f"='NSST Print - {project}'!E20+'NSST Print - {project}'!D20"

                    ws10[f"{i}{exit_first_row_number - 1}"].number_format = '0'

                else:
                    ws10[
                        f"{i}{exit_first_row_number - 1}"].value = f"=SUM({i}{exit_first_row_number}:{i}{exit_last_row_number + 3})"
                    ws10[f"{i}{exit_first_row_number - 1}"].number_format = 'R#,##0.00'

            # iterate through rows from exit_first_row_number to exit_last_row_number + 3 and if column I value is TRUE, color the entire row light red
            for row in range(exit_first_row_number, exit_last_row_number + 3):
                if ws10[f"H{row}"].value == True and ws10[f"I{row}"].value == False:
                    # color in a light green
                    for i in range(1, ws10.max_column + 1):
                        ws10[f"{get_column_letter(i)}{row}"].fill = PatternFill(start_color="C6EFCE",
                                                                                end_color="C6EFCE",
                                                                                fill_type="solid")

                if ws10[f"I{row}"].value == True:
                    for i in range(1, ws10.max_column + 1):
                        ws10[f"{get_column_letter(i)}{row}"].fill = PatternFill(start_color="FFC7CE",
                                                                                end_color="FFC7CE",
                                                                                fill_type="solid")
                    # hide the row
                    ws10.row_dimensions[row].hidden = True

            # freeze panes at A4
            ws10.freeze_panes = ws10['A4']

            columns_to_hide = [2, 7, 8, 9, 10, 11, 12, 13, 15, 20, 21, 24]
            for i in columns_to_hide:
                ws10.column_dimensions[get_column_letter(i)].hidden = True

        # for i in range(block_cost_range_start[0], block_cost_range_end[0] + 1):
        #     formula = "="
        #     print(i)
        #     for r in range(1, len(block_cost_range_start)):
        #         formula += f"+SUMIFS('Cashflow Projection - {reports_by_project[r]}'!$B${block_cost_range_start[r]}:$B${block_cost_range_end[r]}, 'Cashflow Projection - {reports_by_project[r]}'!$A${block_cost_range_start[r]}:$A${block_cost_range_end[r]}, A{i}, 'Cashflow Projection - {reports_by_project[r]}'!$E${block_cost_range_start[r]}:$E${block_cost_range_end[r]}, E{i})"
        #     print(formula)
        #     print()
        #     # ws6[f"F{i}"].value = formula
        #     ws6[f"Cashflow Projection!B{i}"].value = formula

        # for i in range(block_cost_range_start[0], block_cost_range_end[0] + 1):
        #     formula = "="
        #     print(i)
        #     for r in range(1, len(block_cost_range_start)):
        #         formula += f"+SUMIFS('Cashflow Projection - {reports_by_project[r]}'!$B${block_cost_range_start[r]}:$B${block_cost_range_end[r]}, 'Cashflow Projection - {reports_by_project[r]}'!$A${block_cost_range_start[r]}:$A${block_cost_range_end[r]}, A{i}, 'Cashflow Projection - {reports_by_project[r]}'!$E${block_cost_range_start[r]}:$E${block_cost_range_end[r]}, E{i})"
        #     print(formula)
        #     print()
        #     # Directly assign the formula to the cell in ws6
        #     ws6[f"B{i}"].value = formula  # Assigning to the current worksheet

        # Get the 'Cashflow Projection' worksheet
        ws_cashflow = wb['Cashflow Projection']

        for i in range(block_cost_range_start[0], block_cost_range_end[0] + 1):
            formula = "="
            # print(f"Row: {i}")
            for r in range(1, len(block_cost_range_start)):
                formula += f"+SUMIFS('Cashflow Projection - {reports_by_project[r]}'!$B${block_cost_range_start[r]}:$B${block_cost_range_end[r]}, 'Cashflow Projection - {reports_by_project[r]}'!$A${block_cost_range_start[r]}:$A${block_cost_range_end[r]}, A{i}, 'Cashflow Projection - {reports_by_project[r]}'!$E${block_cost_range_start[r]}:$E${block_cost_range_end[r]}, E{i})"

            # print(f"Formula for B{i}: {formula}")
            ws_cashflow[f"B{i}"].value = formula  # Assign the formula to the 'Cashflow Projection' worksheet

        # for i in range(refinanced_units_start_range[0], refinanced_units_end_range[0] + 1):
        #     formula = ""
        #     formula2 = "="
        #     # print(f"Row: {i}")
        #     for r in range(1, len(refinanced_units_start_range)):
        #         formula += f"+SUMIFS('Cashflow Projection - {reports_by_project[r]}'!$C${refinanced_units_start_range[r]}:$C${refinanced_units_start_range[r]}, 'Cashflow Projection - {reports_by_project[r]}'!$D${refinanced_units_start_range[r]}:$D${refinanced_units_start_range[r]}, D{i})"
        #         formula2 += f"+SUMIFS('Cashflow Projection - {reports_by_project[r]}'!$E${refinanced_units_start_range[r]}:$E${refinanced_units_start_range[r]}, 'Cashflow Projection - {reports_by_project[r]}'!$D${refinanced_units_start_range[r]}:$D${refinanced_units_start_range[r]}, D{i})"
        #     formula = f"=IF({formula}=0, \"\", {formula})"
        #     ws_cashflow[f"C{i}"].value = formula
        #     ws_cashflow[f"E{i}"].value = formula2# Assign the formula to the 'Cashflow Projection' worksheet




# # 565656
#
#         print("refinance_units_start_range", refinanced_units_start_range)
#         print("refinance_units_end_range", refinanced_units_end_range)
#         print("refinanced Heron Start:", refinanced_units_start_heron + 1)
#         print("refinanced Heron End:", refinanced_units_end_heron)
        # print(sales[0])

        for row in ws3.iter_rows(min_row=5, max_row=ws3.max_row, min_col=18, max_col=18):
            for cell in row:
                # "=IF(SUMIFS('Cashflow Projection - Heron'!$E$120:$E$199,'Cashflow Projection - Heron'!$D$120:$D$199,Sales!C158)=1,0,IF(AND($F158=1,T158=FALSE),((SUMIFS(Investors!$M:$M,Investors!$E:$E,Sales!$C158,Investors!$O:$O,FALSE,Investors!$K:$K,">="&'NSST Print'!$B$3)+SUMIFS(Investors!$S:$S,Investors!$E:$E,Sales!$C158,Investors!$O:$O,FALSE,Investors!$K:$K,">="&'NSST Print'!$B$3))*$F158)-SUMIFS('Investor Exit List'!$Q:$Q,'Investor Exit List'!$Y:$Y,1,'Investor Exit List'!$Z:$Z,"Release",'Investor Exit List'!$C:$C,Sales!$C158),0)*$F158)"
                cell.value = f"=IF(SUMIFS('Cashflow Projection - Heron'!$E${refinanced_units_start_heron + 1}:$E${refinanced_units_end_heron },'Cashflow Projection - Heron'!$D${refinanced_units_start_heron + 1}:$D${refinanced_units_end_heron},Sales!C{cell.row})>1,0,IF(AND($F{cell.row}=1,T{cell.row}=FALSE),((SUMIFS(Investors!$M:$M,Investors!$E:$E,Sales!C{cell.row},Investors!O:O,FALSE,Investors!K:K,\">=\"&'NSST Print'!$B$3)+SUMIFS(Investors!$S:$S,Investors!$E:$E,Sales!C{cell.row},Investors!O:O,FALSE,Investors!K:K,\">=\"&'NSST Print'!$B$3))*$F{cell.row})-SUMIFS('Investor Exit List'!$Q:$Q,'Investor Exit List'!$Y:$Y,1,'Investor Exit List'!$Z:$Z,\"Release\",'Investor Exit List'!$C:$C,Sales!C{cell.row}),0)*$F{cell.row})"

                cell.number_format = '#,##0.00'
        # for sale in sales:
        #     if sale['opportunity_code'] == "HVK206":
        #         print(sale)

        for i in range(5, ws3.max_row + 1):
            # if ws3[f"A{i}"].value == "Heron Fields" or ws3[f"A{i}" == "Heron View"]:
            # print("Heron Fields or Heron View")
            unit_number = ws3[f"C{i}"].value
            # print("unit_number", unit_number)
            # filter sales where opportunity_code is equal to unit_number
            sales_filtered = [item for item in sales if item['opportunity_code'] == unit_number]
            sales_price = sales_filtered[0]['sale_price']
            # "=IF(SUMIFS('Cashflow Projection - Heron'!$E$120:$E$199,'Cashflow Projection - Heron'!$D$120:$D$199,Sales!C158)=1,1200000,1499000)"
            ws3[f"I{i}"].value = f"=IF(SUMIFS('Cashflow Projection - Heron'!$E${refinanced_units_start_heron + 1}:$E${refinanced_units_end_heron},'Cashflow Projection - Heron'!$D${refinanced_units_start_heron + 1}:$D${refinanced_units_end_heron},Sales!C{i})=1,1200000,{sales_price})"











        ws11 = wb.create_sheet('Momentum')
        # make tab color red
        ws11.sheet_properties.tabColor = "FF204E"

        ws11['E1'] = '7'
        # make above bold with a border
        ws11['E1'].font = Font(bold=True, size=14)
        ws11['E1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                   right=Side(style='thin'))
        ws11['E1'].alignment = Alignment(horizontal='center', vertical='center')

        ws11['D2'] = 'C.3.b'
        # make bold 22, with borders all around and center the text
        ws11['D2'].font = Font(bold=True, size=22)
        ws11['D2'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                   right=Side(style='thin'))
        ws11['D2'].alignment = Alignment(horizontal='center', vertical='center')

        ws11['A3'] = f"Momentum Interest"
        ws11['A3'].font = Font(bold=True, color="FFFFFF", size=22)
        ws11['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws11['A3'].border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'),
                                   right=Side(style='medium'))

        # merge row three from A to E
        ws11.merge_cells('A3:D3')

        # fill in blue
        ws11['A3'].fill = PatternFill(start_color="008DDA", end_color="008DDA", fill_type="solid")
        ws11.append([])
        ws11.append([])
        momentum_headers = []
        for item in momentum[0]:
            momentum_headers.append(item.replace("_", " ").title())
        ws11.append(momentum_headers)
        header_row_momentum = ws11.max_row

        # Cells light grey with a border, center the text and make it bold.
        for i in range(1, ws11.max_column):
            ws11[f"{get_column_letter(i)}{header_row_momentum}"].font = Font(bold=True, size=13)
            ws11[f"{get_column_letter(i)}{header_row_momentum}"].fill = PatternFill(start_color="D3D3D3",
                                                                                    end_color="D3D3D3",
                                                                                    fill_type="solid")
            ws11[f"{get_column_letter(i)}{header_row_momentum}"].border = Border(top=Side(style='thin'),
                                                                                 bottom=Side(style='thin'),
                                                                                 left=Side(style='thin'),
                                                                                 right=Side(style='thin'))
            ws11[f"{get_column_letter(i)}{header_row_momentum}"].alignment = Alignment(horizontal='center',
                                                                                       vertical='center')
            # make each column 55 wide
            ws11.column_dimensions[get_column_letter(i)].width = 39

        for index, item in enumerate(momentum):
            row = []
            for key in item:
                row.append(item[key])
            ws11.append(row)
            if index == 0:
                momentum_first_row_number = ws11.max_row
        momentum_last_row_number = ws11.max_row

        for i in range(1, ws11.max_column):
            for row in range(momentum_first_row_number, momentum_last_row_number + 1):
                ws11[f"{get_column_letter(i)}{row}"].font = Font(bold=False, size=14)
                # put a border around each cell
                ws11[f"{get_column_letter(i)}{row}"].border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                                     left=Side(style='thin'), right=Side(style='thin'))
                if i == 1:
                    ws11[f"{get_column_letter(i)}{row}"].number_format = 'yyyy-mm-dd'
                    # ws11[f"{get_column_letter(i)}{row}"].number_format = 'R#,##0.00'
                else:
                    ws11[f"{get_column_letter(i)}{row}"].number_format = 'R#,##0.00'
                # elif i == 8:
                #     ws11[f"{get_column_letter(i)}{row}"].value = f"=IF(E{row}=TRUE,F{row},\"\

        ws11.append(["TOTAL", f"=SUM(B{momentum_first_row_number}:B{momentum_last_row_number})",
                     f"=SUM(C{momentum_first_row_number}:C{momentum_last_row_number})",
                     f"=SUM(D{momentum_first_row_number}:D{momentum_last_row_number})"])
        momentum_totals_row = ws11.max_row
        for i in range(1, ws11.max_column):
            ws11[f"{get_column_letter(i)}{momentum_totals_row}"].font = Font(bold=True, size=14)
            ws11[f"{get_column_letter(i)}{momentum_totals_row}"].number_format = 'R#,##0.00'
            if i == 1:
                ws11[f"{get_column_letter(i)}{momentum_totals_row}"].border = Border(top=Side(style='medium'),
                                                                                     bottom=Side(style='medium'),
                                                                                     left=Side(style='medium'))
                # Center the test
                ws11[f"{get_column_letter(i)}{momentum_totals_row}"].alignment = Alignment(horizontal='center',
                                                                                           vertical='center')
            elif i == ws11.max_column:
                ws11[f"{get_column_letter(i)}{momentum_totals_row}"].border = Border(top=Side(style='medium'),
                                                                                     bottom=Side(style='medium'),
                                                                                     right=Side(style='medium'))
            else:
                ws11[f"{get_column_letter(i)}{momentum_totals_row}"].border = Border(top=Side(style='medium'),
                                                                                     bottom=Side(style='medium'))

        ws11.append([])
        ws11.append(["MOMENTUM INVESTOR CAPITAL",
                     f"=(SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$J:$J,\">\"&'NSST Print'!B3)+SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$J:$J,\"\"))",
                     "",
                     "per App"])
        recon_start = ws11.max_row
        ws11.append(["MOMENTUM INTEREST EARNED SINCE INCEPTION", f"=+B{momentum_totals_row}", "", "from Momentum"])
        ws11.append(["MOMENTUM ADVICE FEES ", f"=+D{momentum_totals_row}", "", "from Momentum"])
        ws11.append(
            ["NET INTEREST", f"=+B{momentum_totals_row}-D{momentum_totals_row}", "",
             "Interest less Ongoing Advice Fees"])
        ws11.append(["ADVICE FEES", f"=+C{momentum_totals_row}", "", ""])
        ws11.append(
            ["TOTAL INTEREST IN ACCOUNT", f"=+B{momentum_totals_row}-D{momentum_totals_row}-C{momentum_totals_row}",
             "", ""])
        ws11.append(["MOMENTUM INTEREST DRAWN TO DATE", 1861191.23, "", "Interest Draw on 3 October 2023"])
        ws11.append(
            ["TOTAL MOMENTUM DRAWS", f"=SUMIFS(Investors!$M:$M,Investors!$J:$J,\"<=\"&'NSST Print'!B3)+137832.56", "",
             ""])
        total_momentum_draws = ws11.max_row
        ws11.append(["GLC NOT ON DRAW", 4183000.82, "", ""])
        anomolies_start = ws11.max_row
        ws11.append(["C Oberholzer incorrect amount", -455, "", ""])
        ws11.append(["Draw 10 (Rudlynn Trust)", -150000, "", ""])
        anomolies_end = ws11.max_row
        ws11.append(["MOMENTUM BALANCE", 0, "", "Momentum Statement Balance"])
        momentum_balance = ws11.max_row
        ws11.append([])
        ws11.append(["", "RECONCILIATION", "", ""])
        ws11.append([])
        ws11.append(["Momentum Statement", f"=+B{ws11.max_row - 2}", "", "per Statement"])
        ws11.append(["Momentum Xero", f"=Cashflow!B13", "", "Per App"])
        ws11.append(["Difference", f"=+B{ws11.max_row - 1}-B{ws11.max_row}", "", ""])
        ws11.append([])
        ws11.append(["Momentum Xero", f"=Cashflow!B13", "", "per Xero"])
        momentum_xero = ws11.max_row
        ws11.append(["Momentum - App",
                     f"=(SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$J:$J,\">\"&'NSST Print'!B3)+SUMIFS(Investors!$M:$M,Investors!$I:$I,\"<=\"&'NSST Print'!B3,Investors!$J:$J,\"\"))",
                     "", "per App"])
        ws11.append(["Difference", f"=+B{ws11.max_row - 1}-B{ws11.max_row}", "", "Reconciled Below"])
        ws11.append([])
        ws11.append(
            ["TOTAL INTEREST IN ACCOUNT", f"=+B{momentum_totals_row}-D{momentum_totals_row}-C{momentum_totals_row}", "",
             ""])
        ws11.append(["MOMENTUM INTEREST DRAWN TO DATE", 1861191.23, "", "Interest Draw on 3 October 2023"])
        ws11.append(["", f"=+B{ws11.max_row - 1}-B{ws11.max_row}", "", ""])
        ws11.append(["", 137832.56, "", "Interest STBB"])
        interest_stbb = ws11.max_row
        subtotal = ws11.max_row
        ws11.append(["", f"=B{subtotal - 1}-B{subtotal}-B{subtotal + 2}", "", "to be reconciled"])
        ws11.append(["", f"=B{ws11.max_row - 6}", "", ""])
        recon_end = ws11.max_row

        for i in range(recon_start, recon_end + 1):
            # Merge B and C of each row
            ws11.merge_cells(f"B{i}:C{i}")
            # center the text
            ws11[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')

        # put a medium border around the area from recon_start to recon_end and columns A to D
        for i in range(recon_start, recon_end + 3):
            for x in range(1, 5):
                if x == 1 and i == recon_start:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(top=Side(style='medium'),
                                                                       left=Side(style='medium'))

                elif x == 4 and i == recon_start:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(top=Side(style='medium'),
                                                                       right=Side(style='medium'))
                elif x > 1 and x < 5 and i == recon_start:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(top=Side(style='medium'))

                elif x == 1 and i > recon_start and i < recon_end + 2:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(left=Side(style='medium'))

                elif x == 4 and i > recon_start and i < recon_end + 2:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(right=Side(style='medium'))

                elif x == 1 and i == recon_end + 2:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(left=Side(style='medium'),
                                                                       bottom=Side(style='medium'))
                elif x == 4 and i == recon_end + 2:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(right=Side(style='medium'),
                                                                       bottom=Side(style='medium'))
                elif x > 1 and x < 5 and i == recon_end + 2:
                    ws11[f"{get_column_letter(x)}{i}"].border = Border(bottom=Side(style='medium'))

                if x == 2:
                    # format as currency
                    ws11[f"{get_column_letter(x)}{i}"].number_format = 'R#,##0.00'
                    ws11[f"{get_column_letter(x)}{i}"].font = Font(bold=True, size=14)

        for i in range(anomolies_start, anomolies_end + 1):
            # fill column B with a light grey
            ws11[f"B{i}"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            # fill column B at anomolies + 1 in a light yellow

        ws11[f"B{anomolies_end + 1}"].fill = PatternFill(start_color="FFFDD7", end_color="FFFDD7", fill_type="solid")

        ws12 = wb.create_sheet('Checklist')

        # Make the tab color purple
        ws12.sheet_properties.tabColor = "551A8B"

        ws12['A1'] = 'Checklist'
        ws12['A1'].font = Font(bold=True, size=22)
        ws12['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws12.append([])
        ws12.append([])
        ws12.append([])

        ws12.append(["Check", "Explanation", " Option 1", "Option 2", "Difference", "Reason for Difference"])
        ws12_last_row = ws12.max_row
        ws12.append(
            ["Funds Drawn Down", f"Momentum vs NSST", f"=Momentum!B{total_momentum_draws}-Momentum!B{interest_stbb}",
             "='NSST Print'!B8", "=C6-D6", ""])
        ws12.append(["Momentum Investment Account", "Momentum Xero vs NSST Balance", f"=Momentum!B{momentum_xero}",
                     "='NSST Print'!B9", "=C7-D7", ""])
        ws12.append(
            ["Momentum Balance", "Momentum vs Xero", f"=Momentum!B{momentum_balance}", f"=Momentum!B{momentum_xero}",
             "=C8-D8", ""])
        ws12.append(["Capital Still in Investments", "NSST vs Investor Exit List", "='NSST Print'!B13",
                     "='Investor Exit List'!F3", "=C9-D9", ""])
        ws12.append(
            ["Funds Available", "Cashflow vs Projection", "='Cashflow Projection'!D61", "=Cashflow!B20", "=C10-D10",
             ""])
        ws12.append(["Units Unsold and Remaining", "NSST vs Investor Exit Report", "='NSST Print'!D20+'NSST Print'!E20",
                     "='Investor Exit List'!A3", "=C11-D11", ""])
        ws12.append(["Investments Still in Development", "NSST vs Investor Exit Report", "='NSST Print'!B17",
                     "='NSST Print'!B17", "=C12-D12", ""])
        ws12.append(
            ["Total Income", "NSST vs Investor Exit Report", "=Cashflow!B11", "='Cashflow Projection'!D6", "=C13-D13",
             ""])
        ws12.append(
            ["Total Cash", "NSST vs Investor Exit Report", "=Cashflow!B20", "='Cashflow Projection'!D61", "=C14-D14",
             ""])
        ws12.append(
            ["Total Costs to Complete", "Cashflow vs Projection", "=Cashflow!B24", "='Cashflow Projection'!D38",
             "=C15-D15", ""])
        ws12.append(["Total Running Costs", "Cashflow vs Projection", "=Cashflow!B24", "='Cashflow Projection'!D38",
                     "=C16-D16", ""])
        ws12.append(["Total VAT Payable / Receivable", "Cashflow vs Projection", "=Cashflow!B26",
                     "='Cashflow Projection'!D26+'Cashflow Projection'!D55", "=C17-D17", ""])
        ws12.append(["Total Net Effect", "Cashflow vs Projection", "=Cashflow!B30",
                     "='Cashflow Projection'!D35+'Cashflow Projection'!D26+'Cashflow Projection'!D38+'Cashflow Projection'!D55+'Cashflow Projection'!D57",
                     "=C18-D18", ""])

        # get a list of sheets in the workbook
        sheet_names = wb.sheetnames
        # print(sheet_names)
        sheet_names_to_hide = ['Updated Construction', 'Operational Costs', 'Xero', 'Xero_bank', 'Other Costs', 'Investors', 'Sales',
                               'Construction', 'Opportunities', 'Heron', 'Cashflow Projection', 'Cashflow','NSST Print', 'Investor Exit List','Momentum', 'Checklist']
        # loop through the sheets and hide the ones in the list

        for sheet in sheet_names_to_hide:
            ws = wb[sheet]
            ws.sheet_state = 'hidden'


        # filter column D in ws5 to only show 8400/000 and 8440/000 and 8400/010 and hide the rest using filter from row 4 to the last row








        # move sheet 'NSST Print' to the beginning of the workbook

        # sheets_to_move = ['Investors', 'Construction', 'Sales']
        # # move the sheets to the nd of the workbook
        # for sheet in sheets_to_move:
        #     ws = wb[sheet]
        #     sheet_names.move_sheet(ws, -1)

        # remove the default sheet
        # ws6.column_dimensions.group('F', 'H', hidden=False)
        # ws6.row_dimensions.group(toggles_start, toggles_end, hidden=False)
        # ws6.row_dimensions.group(vat_income - 1, vat_payable_on_sales - 1, hidden=False)
        # ws6.row_dimensions.group(funds_available_start, funds_available_end - 1, hidden=False)
        # ws6.row_dimensions.group(block_costs_start, block_costs_end + 1, hidden=False)
        # # ws6.row_dimensions.group(block_costs_start + 1, block_costs_end + 1, hidden=True)
        # # ws6.row_dimensions.group(block_costs_start, block_costs_end, hidden=True)
        #
        # ws6.print_area = 'A1:AE78'
        #
        # # Set page orientation to landscape
        # ws6.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        #
        # # Set scale to fit 1 page wide by 1 pages tall
        # ws6.page_setup.scale = 50
        # ws6.page_setup.fitToWidth = 1
        # ws6.page_setup.fitToHeight = 1
        #
        # # Set page margins
        # ws6.page_margins.left = 0.5
        # ws6.page_margins.right = 0.5
        # ws6.page_margins.top = 2.75
        # ws6.page_margins.bottom = 0.75

        wb.remove(wb['Sheet'])
        wb.save('cashflow_p&l_files/cashflow_projection.xlsx')
        end_time = time.time()
        print("Report Took:", end_time - start_time)
        return "cashflow_p&l_files/cashflow_projection.xlsx"
    except Exception as e:
        print("Error XX", e)
        return "Error"


other_costs = [
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2023-10-31",
        "Amount": 5125504.85
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2023-11-30",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2023-12-31",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-01-31",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-02-29",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-03-31",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-04-30",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-05-31",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-06-30",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-07-31",
        "Amount": 314037.88
    },
    {
        "AccountName": "CPSD",
        "Category": "Expenses",
        "ReportDate": "2024-08-31",
        "Amount": 314037.88
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2023-10-31",
        "Amount": 5676905
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2023-11-30",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2023-12-31",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-01-31",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-02-29",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-03-31",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-04-30",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-05-31",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-06-30",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-07-31",
        "Amount": 392914.302
    },
    {
        "AccountName": "OppInvest",
        "Category": "Expenses",
        "ReportDate": "2024-08-31",
        "Amount": 392914.302
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2023-10-31",
        "Amount": 17272544
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2023-11-30",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2023-12-31",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-01-31",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-02-29",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-03-31",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-04-30",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-05-31",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-06-30",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-07-31",
        "Amount": 800000
    },
    {
        "AccountName": "Rent Salaries and Wages",
        "Category": "Expenses",
        "ReportDate": "2024-08-31",
        "Amount": 800000
    },
    {
        "AccountName": "Professional Fees",
        "Category": "Expenses",
        "ReportDate": "2023-10-31",
        "Amount": 5361342.92
    }
]
