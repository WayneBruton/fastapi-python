from openpyxl import Workbook, formatting, formula
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def create_workbook(data):
    if os.path.exists("excel_files/rollovers_consultants.xlsx"):
        os.remove("excel_files/rollovers_consultants.xlsx")
        print("File deleted")
    wb = Workbook()
    ws = wb.active
    ws.title = "ROLLOVERS_CONSULTANTS"
    # set column width
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    # post data to excel
    ws['A1'] = "Investor Account Number"
    ws['B1'] = "Investor Name"
    ws['C1'] = "Development"
    ws['D1'] = "Unit Number"
    ws['E1'] = "Investment Amount"
    ws['F1'] = "Rollover From"
    ws['G1'] = "Consultant"

    # set font style
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)

    for item in data:
        row = []
        #loop through the dictionary and append to the row list
        for key in item:
            row.append(item[key])
        #append the row to the worksheet
        ws.append(row)

    # save workbook
    wb.save("excel_files/rollovers_consultants.xlsx")
    return "excel_files/rollovers_consultants.xlsx"





