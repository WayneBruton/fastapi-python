from openpyxl import Workbook, formatting, formula
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.utils import get_column_letter


def daily_cashflow(sales, investor_exit, report_date):
    wb = Workbook()
    ws1 = wb.active
    ws1.tabColor = "1072BA"
    ws1.title = "Sales"
    ws1['A1'] = "Sales"
    ws1['A1'].font = Font(size=20, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1['A2'] = "Date"
    ws1['A2'].font = Font(bold=True)
    ws1['B2'] = report_date
    ws1['B2'].font = Font(bold=True)
    ws1['B2'].alignment = Alignment(horizontal='center')
    ws1.append([])
    row = []
    for item in sales[0]:
        # print(item)
        item = item.replace("_", " ")
        item = item.title()
        row.append(item)
    ws1.append(row)
    max_row = ws1.max_row
    # make max row bold
    for i in range(1, len(row) + 1):
        ws1.cell(row=max_row, column=i).font = Font(bold=True)
    # make all columns 20 width
    for i in range(1, len(row) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 20
    for sale in sales:
        row = []
        for item in sale:
            row.append(sale[item])
        ws1.append(row)
    sales_date_format_columns = ['G', 'H', 'U']
    for column in sales_date_format_columns:
        for i in range(max_row + 1, ws1.max_row + 1):
            if column == 'H':
                # fill in yellow
                ws1[f'{column}{i}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws1[f'{column}{i}'].number_format = 'yyyy-mm-dd'
    sales_currency_format_columns = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
    for column in sales_currency_format_columns:
        for i in range(max_row + 1, ws1.max_row + 1):
            ws1[f'{column}{i}'].number_format = '"R" #,##0.00'

    for i in range(0, len(sales_currency_format_columns)):
        ws1[
            f"{sales_currency_format_columns[i]}2"] = f"=subtotal(9,{sales_currency_format_columns[i]}{max_row + 1}:{sales_currency_format_columns[i]}{ws1.max_row})"
        ws1[f"{sales_currency_format_columns[i]}2"].number_format = '"R" #,##0.00'

    for i in range(5, ws1.max_row + 1):
        # "=I5/115*15"
        ws1[f"J{i}"] = f"=I{i}/115*15"
        # ws1[f"J{i}"].number_fo/rmat = '"R" #,##0.00'
        # "=I5-J5"
        ws1[f"K{i}"] = f"=I{i}-J{i}"
        # "=K5-SUM(L5:P5)"
        ws1[f"Q{i}"] = f"=K{i}-SUM(L{i}:P{i})"
        # "=IF(E5=FALSE,SUMIFS(Investors!$R:$R,Investors!$G:$G,Sales!C5),0)"
        ws1[f"R{i}"] = f"=IF(E{i}=FALSE,SUMIFS(Investors!$R:$R,Investors!$G:$G,Sales!C{i}),0)"
        # "=Q5-R5"
        ws1[f"S{i}"] = f"=Q{i}-R{i}"
        "=IF(MOD(MONTH(H5), 2) <> 0, EOMONTH(H5, 2), EOMONTH(H5, 1))"
        ws1[f'U{i}'] = f"=IF(MOD(MONTH(H{i}), 2) <> 0, EOMONTH(H{i}, 2), EOMONTH(H{i}, 1))"


    # freeze panes at row 4
    ws1.freeze_panes = "A5"
    # filter the data from row 4 to the last row
    ws1.auto_filter.ref = f"A4:{get_column_letter(ws1.max_column)}{ws1.max_row}"

    ws2 = wb.create_sheet("Investors")
    ws2.tabColor = "1072BA"
    ws2.title = "Investors"
    ws2['A1'] = "Investors"
    ws2['A1'].font = Font(size=20, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2['A2'] = "Date"
    ws2['A2'].font = Font(bold=True)
    ws2['B2'] = report_date
    ws2['B2'].font = Font(bold=True)
    ws2['B2'].alignment = Alignment(horizontal='center')
    ws2.append([])
    row = []
    for item in investor_exit[0]:
        # print(item)
        item = item.replace("_", " ")
        item = item.title()
        row.append(item)
    ws2.append(row)
    max_row = ws2.max_row

    # make max row bold
    for i in range(1, len(row) + 1):
        ws2.cell(row=max_row, column=i).font = Font(bold=True)
    # make all columns 20 width
    for i in range(1, len(row) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 20
    # print("Got Here!!!!!!!!")
    for index, investor in enumerate(investor_exit):
        row = []
        for item in investor:
            row.append(investor[item])
        # print("Got Here!!!!!!!!", index, investor)

        ws2.append(row)
    date_format_columns = ['H', 'I', 'J', 'P']
    for column in date_format_columns:
        for i in range(max_row + 1, ws2.max_row + 1):
            ws2[f'{column}{i}'].number_format = 'yyyy-mm-dd'
            if column == 'J':
                # fill in yellow
                ws2[f'{column}{i}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    currency_format_columns = ['K', 'M', 'N', 'O']
    percent_format_columns = ['L']
    for column in currency_format_columns:
        for i in range(max_row + 1, ws2.max_row + 1):
            ws2[f'{column}{i}'].number_format = '"R" #,##0.00'

    for i in range(0, len(currency_format_columns)):

        ws2[
            f"{currency_format_columns[i]}2"] = f"=subtotal(9,{currency_format_columns[i]}{max_row + 1}:{currency_format_columns[i]}{ws2.max_row})"

        ws2[f"{currency_format_columns[i]}2"].number_format = '"R" #,##0.00'
    for column in percent_format_columns:
        for i in range(max_row + 1, ws2.max_row + 1):
            ws2[f'{column}{i}'].number_format = '0.00%'

    for i in range(5, ws2.max_row + 1):
        "=IF(I5="",K5/365*0.11*((H5+30)-H5),K5/365*0.11*(I5-H5))"
        ws2[f"M{i}"] = f"=IF(I{i}=\"\",K{i}/365*0.11*((H{i}+30)-H{i}),K{i}/365*0.11*(I{i}-H{i}))"
        "=K5*L5/365*(P5-I5)"
        ws2[f"N{i}"] = f"=K{i}*L{i}/365*(P{i}-I{i})"
        "=M5+N5"
        ws2[f"O{i}"] = f"=M{i}+N{i}"
        "=IF(J5>SUMIFS(Sales!$H:$H,Sales!$C:$C,Investors!G5),SUMIFS(Sales!$H:$H,Sales!$C:$C,Investors!G5),Investors!J5)"
        ws2[
            f"P{i}"] = f"=IF(J{i}>SUMIFS(Sales!$H:$H,Sales!$C:$C,Investors!G{i}),SUMIFS(Sales!$H:$H,Sales!$C:$C,Investors!G{i}),Investors!J{i})"
        "=K10+O10"
        ws2[f"Q{i}"] = f"=K{i}+O{i}"
        "=IF(J6<SUMIFS(Sales!$H:$H,Sales!$C:$C,Investors!G6),0,Investors!Q6)"
        ws2[
            f"R{i}"] = f"=IF(J{i}<SUMIFS(Sales!$H:$H,Sales!$C:$C,Investors!G{i}),0,Investors!Q{i})"


    # freeze panes at row 4
    ws2.freeze_panes = "A5"
    # filter the data from row 4 to the last row
    ws2.auto_filter.ref = f"A4:{get_column_letter(ws2.max_column)}{ws2.max_row}"

    ws3 = wb.create_sheet("Exits")
    ws3.tabColor = "1072BA"
    ws3.title = "Exits"
    ws3['A1'] = "Exits"
    ws3['A1'].font = Font(size=20, bold=True)
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3['A2'] = "Date"
    ws3['A2'].font = Font(bold=True)
    ws3['B2'] = report_date
    ws3['B2'].font = Font(bold=True)
    ws3['B2'].alignment = Alignment(horizontal='center')
    ws3.append([])
    header_row = ["Investor Acc Number", "Unit Number", "Total", 0]
    # -365
    for i in range(30,780,30):
        header_row.append(i)
    ws3.append(header_row)
    # make all columns 20 width
    for i in range(1, len(header_row) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 20
    # all cells bold and centrered
    for i in range(1, len(header_row) + 1):
        ws3.cell(row=4, column=i).font = Font(bold=True)
        ws3.cell(row=4, column=i).alignment = Alignment(horizontal='center')
    for i in range(5, len(header_row) + 1):
        ws3.cell(row=3, column=i).value = f"< Days"
        ws3.cell(row=3, column=i).font = Font(bold=True)
        ws3.cell(row=3, column=i).alignment = Alignment(horizontal='center')



    for investor in investor_exit:
        row = []
        row.append(investor['investor_acc_number'])
        row.append(investor['unit_number'])
        ws3.append(row)


    formula_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
                       'X', 'Y', 'Z', 'AA', 'AB', 'AC']
    for index, column in enumerate(formula_columns):
        if index == 0:
            for i in range(5, ws3.max_row + 1):
                ws3[
                    f"{column}{i}"] = f"=IF(AND(SUMIFS(Investors!$P:$P,Investors!$A:$A,$A{i},Investors!$G:$G,$B{i})-$B$2<={column}$4,SUMIFS(Investors!$P:$P,Investors!$A:$A,$A{i},Investors!$G:$G,$B{i})-$B$2>D$4),SUMIFS(Investors!$Q:$Q,Investors!$A:$A,$A{i},Investors!$G:$G,$B{i}),0)"
                # format as currency
                ws3[f"{column}{i}"].number_format = '"R" #,##0.00'
        else:
            for i in range(5, ws3.max_row + 1):
                ws3[
                    f"{column}{i}"] = f"=IF(AND(SUMIFS(Investors!$P:$P,Investors!$A:$A,$A{i},Investors!$G:$G,$B{i})-$B$2<={column}$4,SUMIFS(Investors!$P:$P,Investors!$A:$A,$A{i},Investors!$G:$G,$B{i})-$B$2>{formula_columns[index - 1]}$4),SUMIFS(Investors!$Q:$Q,Investors!$A:$A,$A{i},Investors!$G:$G,$B{i}),0)"
                ws3[f"{column}{i}"].number_format = '"R" #,##0.00'
    for i in range(5, ws3.max_row + 1):
        ws3[f"C{i}"] = f"=SUM(E{i}:AC{i})"
        ws3[f"C{i}"].number_format = '"R" #,##0.00'

    for column in formula_columns:
            ws3[f'{column}2'].number_format = '"R" #,##0.00'
            ws3[f'{column}2'] = f"=subtotal(9,{column}5:{column}{ws3.max_row})"

    # freeze panes at row 4
    ws3.freeze_panes = "A5"
    # filter the data from row 4 to the last row
    ws3.auto_filter.ref = f"A4:{get_column_letter(ws3.max_column)}{ws3.max_row}"

    # hide column D
    ws3.column_dimensions['D'].hidden = True

    ws4 = wb.create_sheet("General Expenses")
    ws4.tabColor = "1072BA"
    ws4.title = "General Expenses"
    ws4['A1'] = "General Expenses"
    ws4['A1'].font = Font(size=20, bold=True)
    ws4['A1'].alignment = Alignment(horizontal='center')
    row = ["Date", "Description", "Amount"]
    ws4.append(row)
    # make all columns 20 width
    for i in range(1, len(row) + 1):
        ws4.column_dimensions[get_column_letter(i)].width = 20
    # all cells bold and centrered
    for i in range(1, len(row) + 1):
        ws4.cell(row=2, column=i).font = Font(bold=True)
        ws4.cell(row=2, column=i).alignment = Alignment(horizontal='center')
    # format column A as date and column C as currency
    date_format_columns = ['A']
    for column in date_format_columns:
        for i in range(3, 100):
            ws4[f'{column}{i}'].number_format = 'yyyy-mm-dd'
    currency_format_columns = ['C']
    for column in currency_format_columns:
        for i in range(3, 100):
            ws4[f'{column}{i}'].number_format = '"R" #,##0.00'

    ws5 = wb.create_sheet("Daily")
    ws5.tabColor = "1072BA"
    ws5.title = "Daily"
    ws5['A1'] = "Daily"
    ws5['A1'].font = Font(size=20, bold=True)
    ws5['A1'].alignment = Alignment(horizontal='center')
    row = ["Date", "Opening Balance", "Transfer","VAT Sales", "Early Exit", "General Expenses", "Daily Balance", "Cumulative Balance"]
    ws5.append(row)
    # make all columns 20 width
    for i in range(1, len(row) + 1):
        ws5.column_dimensions[get_column_letter(i)].width = 20
    # all cells bold and centrered
    for i in range(1, len(row) + 1):
        ws5.cell(row=2, column=i).font = Font(bold=True)
        ws5.cell(row=2, column=i).alignment = Alignment(horizontal='center')
    # format column A as date the rest as currency
    date_format_columns = ['A']
    for column in date_format_columns:
        for i in range(3, 731):
            ws5[f'{column}{i}'].number_format = 'yyyy-mm-dd'
    currency_format_columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    for column in currency_format_columns:
        for i in range(3, 731):
            ws5[f'{column}{i}'].number_format = '"R" #,##0.00'


            "=SUMIFS(Sales!$S:$S,Sales!$H:$H,A3)"
            ws5[f"C{i}"] = f"=SUMIFS(Sales!$S:$S,Sales!$H:$H,A{i})+SUMIFS(Sales!$J:$J,Sales!$H:$H,A{i})"
            "=SUMIFS(Sales!$J:$J,Sales!$U:$U,A3)"
            ws5[f"D{i}"] = f"=SUMIFS(Sales!$J:$J,Sales!$U:$U,A{i})"

            "=IF(C3<>0,SUMIFS(Investors!$Q:$Q,Investors!$P:$P,A3)-SUMIFS(Sales!$R:$R,Sales!$H:$H,A3),SUMIFS(Investors!$Q:$Q,Investors!$P:$P,A3))"

            ws5[f"E{i}"] = f"=IF(C{i}<>0,SUMIFS(Investors!$Q:$Q,Investors!$P:$P,A{i})-SUMIFS(Sales!$R:$R,Sales!$H:$H,A{i}),SUMIFS(Investors!$Q:$Q,Investors!$P:$P,A{i}))"
            "=SUMIFS('General Expenses'!$C:$C,'General Expenses'!$A:$A,A3)"
            ws5[f"F{i}"] = f"=SUMIFS('General Expenses'!$C:$C,'General Expenses'!$A:$A,A{i})"
            "=B3+C3-D3-E3"
            ws5[f"G{i}"] = f"=B{i}+C{i}-D{i}-E{i}-F{i}"
            "=G2+F3"
            ws5[f"H{i}"] = f"=H{i-1}+G{i}"
            if i == 3:
                ws5[f'A{i}'] = 0
                "=F3"
                ws5[f'H{i}'] = f"=G{i}"
                # fill in yellow
                ws5[f'B{i}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws5[f'A{i}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    subtotal_rows = ['C', 'D', 'E', 'F','G']
    for column in subtotal_rows:
        ws5[f'{column}1' ] = f"=subtotal(9,{column}3:{column}{ws5.max_row})"
        ws5[f'{column}1'].number_format = '"R" #,##0.00'


    for i in range(3, 731):
        if i == 3:
            ws5[f"A{i}"] = f"{report_date}"
            ws5[f"A{i}"].font = "yyyy-mm-dd"
        else:
            ws5[f"A{i}"] = f"=A{i-1}+1"
            ws5[f"A{i}"].font = "yyyy-mm-dd"

    # freeze panes at row 4
    ws5.freeze_panes = "A3"







    # save the workbook
    filename = "cashflow_p&l_files/daily_cashflow.xlsx"
    wb.save(filename)
    # return "hello"
    return {"filename": filename}
    # ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    # ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # ws.page_setup.fitToWidth = 1
    # ws.page_setup.fitToHeight = 1
    # ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    # ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True

    # Set column widths
    # ws.column_dimensions['A'].width = 20
    # ws.column_dimensions['B'].width = 20
    # ws.column_dimensions['C'].width = 20
    # ws.column_dimensions['D'].width = 20
    # ws.column_dimensions['E'].width = 20
    # ws.column_dimensions['F'].width = 20
    # ws.column_dimensions['G'].width = 20
    # ws.column_dimensions['H'].width = 20
    # ws.column_dimensions['I'].width = 20
    # ws.column_dimensions['J'].width = 20
    # ws.column_dimensions['K'].width = 20
    # ws.column_dimensions['L'].width = 20
    # ws.column_dimensions['M'].width = 20
    # ws.column_dimensions['N'].width = 20
    # ws.column_dimensions['O'].width = 20
    # ws.column_dimensions['P'].width = 20
    # ws.column_dimensions['Q'].width = 20
    # ws.column_dimensions['R'].width = 20
    # ws.column_dimensions['S'].width = 20
    # ws.column_dimensions['T'].width = 20
    # ws.column_dimensions['U'].width = 20
    # ws.column_dimensions['V'].width = 20
    # ws.column_dimensions['W'].width = 20
    # ws.column_dimensions['X'].width = 20
    # ws.column_dimensions['Y'].width = 20
    # ws.column_dimensions['Z'].width = 20
    # ws.column_dimensions['AA'].width = 20
    # ws.column_dimensions['AB'].width = 20
    # ws.column_dimensions['AC'].width = 20
    # ws.column_dimensions['AD'].width =
