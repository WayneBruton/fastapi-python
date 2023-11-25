# from config.db import db
from datetime import datetime, timedelta

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter, column_index_from_string

from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def cashflow_hf_hv(data, data2, report_date):
    global trading_income_start_row, trading_income_end_row, other_income_start_row, other_income_end_row, cos_start_row, cos_end_row, operating_expenses_start_row, operating_expenses_end_row, gross_profit_start_row, nett_profit_start_row
    try:
        print("report_date", report_date)
        report_date = datetime.strptime(report_date, '%Y-%m-%d')
        print("report_date", report_date)

        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }

        # loop through data and convert 'Month' to a datetime object on the last day of the month formatted as dd MMM
        # yyyy
        for item in data:
            del item['_id']
            date_string = item['Month']
            month_str, year_str = date_string.split('-')
            month_num = month_map[month_str]
            last_day = (datetime(int(year_str) + 2000, month_num, 1) + timedelta(days=32)).replace(day=1) - timedelta(
                days=1)
            formatted_date = last_day.strftime('%Y-%m-%d')
            item['Month'] = formatted_date
            # print("Hello")
            formatted_date = datetime.strptime(formatted_date, '%Y-%m-%d')
            if formatted_date > report_date:
                # print("Good Bye")
                item['Actual'] = 0
            if item['Actual'] == '0' or item['Actual'] == 0:
                item['use'] = item['Forecast']
            else:
                item['use'] = item['Actual']

        # print(data[0])
        # print()
        # for item in data:
        #     print(item)
        #     print()

        # sort data by 'Month' in ascending order, then by 'Applicable_dev', then by 'Account'
        data = sorted(data, key=lambda k: (k['Month'], k['Applicable_dev'], k['Account']))
        worksheet_input = []
        headers = []
        # loop through data[0] to get the headers and insert into a list and then insert into worksheet_input
        for key in data[0]:
            # key['use'] = 0.0
            headers.append(key)
        worksheet_input.append(headers)
        # print(worksheet_input)
        # print()
        # loop through data and insert into worksheet_input
        for item in data:
            row = []
            if item['Actual'] == '0' or item['Actual'] == 0:
                item['use'] = item['Forecast']
            else:
                item['use'] = item['Actual']
            row.append(item['Account'])
            row.append(item['Category'])
            row.append(item['Development'])
            row.append(item['Applicable_dev'])
            row.append(item['Month'])
            row.append(float(item['Actual']))
            row.append(float(item['Forecast']))
            row.append(float(item['use']))
            worksheet_input.append(row)

        # create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        # name tab data
        ws.title = 'data'
        # make tab color green
        ws.sheet_properties.tabColor = "1072BA"

        # insert worksheet_input into worksheet
        # CREATE DATA SHEET
        for row in worksheet_input:
            ws.append(row)
        # set column width
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15




        # get last row and last column
        last_row = ws.max_row
        last_column = ws.max_column

        # in colum H, from row 2 to the last row, set the value to the following formula
        # =IF(F2=0, G2, F2)
        for row in ws.iter_rows(min_row=2, min_col=last_column, max_row=last_row, max_col=last_column):
            for cell in row:
                # print(cell)
                cell.value = f"=IF(F{cell.row}=0, G{cell.row}, F{cell.row})"
                cell.number_format = 'R #,##0.00'
        #format column E as a date with the format yyyy-mm-dd
        for row in ws.iter_rows(min_row=2, min_col=5, max_row=last_row, max_col=5):
            for cell in row:
                cell.number_format = 'yyyy-mm-dd'

        # in column I, from row 2 add the following formula +E2+0
        for row in ws.iter_rows(min_row=2, min_col=9, max_row=last_row, max_col=9):
            for cell in row:
                cell.value = f"=E{cell.row}+0"
                cell.number_format = 'yyyy-mm-dd'

        applicable_dev = ['Heron Fields', 'Heron View', 'Heron']

        for index, dev in enumerate(applicable_dev):
            # filter data by 'Applicable_dev'
            if index < 2:
                filtered_data = [item for item in data if item['Applicable_dev'] == dev]
            else:
                filtered_data = data
            months = []
            accounts = []
            categories = ['Trading Income', 'Other Income', 'COS', 'Operating Expenses']

            for item in filtered_data:
                months.append(item['Month'])
                insert = {'Account': item['Account'], 'Category': item['Category']}
                accounts.append(insert)
            months = sorted(list(set(months)))
            accounts = sorted(accounts, key=lambda k: (k['Category'], k['Account']))
            seen = set()
            accounts = [item for item in accounts if
                        item['Account'] not in seen and not seen.add(item['Account'])]

            print("accounts", accounts[0])
            # remove any duplicates from accounts which is a list of dictionaries
            # accounts = list(dict.fromkeys(accounts))

            # if dev == 'Heron View':
            #     for account in accounts:
            #         if account['Account'] == 'COS - Legal Fees':
            #             print("account - Line 143", account)




            ws = wb.create_sheet(dev)
            # make tab color red
            ws.sheet_properties.tabColor = "FF0000"

            # in A1 insert 'profit and loss'
            ws['A1'] = 'Profit and Loss'
            # in A2 insert dev
            ws['A2'] = dev
            # insert 2 empty strings at the beginning of months
            months.insert(0, '')
            months.insert(0, '')
            ws['A3'] = ''
            ws['A4'] = ''
            # insert months into worksheet at row 5
            ws.append(months)

            # get max_column and set to variable l_column
            l_column = ws.max_column

            columns = []
            for col in ws.iter_cols(min_row=5, min_col=1, max_row=5, max_col=l_column):
                columns.append(col[0].column_letter)
            # print("columns", columns)
            for col in columns:
                ws.column_dimensions[col].width = 15



            # set font and alignment for months
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for index1, category in enumerate(categories):
                p_max_row = ws.max_row

                # add 1 to p_max_row
                p_max_row += 1
                ws['A' + str(p_max_row)] = category
                # set font and alignment for category
                ws['A' + str(p_max_row)].font = Font(bold=True)
                ws['A' + str(p_max_row)].alignment = Alignment(horizontal='left')
                # set variable called p_max_row to the max row in the worksheet
                p_max_row = ws.max_row

                filtered_accounts = [item for item in accounts if item['Category'] == category]
                # print("filtered_accounts", filtered_accounts)
                # print()
                for acc in filtered_accounts:
                    # add 1 to p_max_row
                    p_max_row += 1

                    # "Account" in column A
                    ws['A' + str(p_max_row)] = acc['Account']
                # add 3 empty rows after each category
                ws['A' + str(p_max_row + 1)] = f"Total {category}"

                # make font bold for total
                ws['A' + str(p_max_row + 1)].font = Font(bold=True)

                if index1 == 2:
                    ws['A' + str(p_max_row + 2)] = ''
                    ws['A' + str(p_max_row + 3)] = ''
                    ws['A' + str(p_max_row + 4)] = 'Gross Profit'
                    ws['A' + str(p_max_row + 4)].font = Font(bold=True)
                    ws['A' + str(p_max_row + 5)] = ''
                    ws['A' + str(p_max_row + 6)] = ''


                else:
                    ws['A' + str(p_max_row + 2)] = ''
                    ws['A' + str(p_max_row + 3)] = ''

            # get Max row of worksheet
            p_max_row = ws.max_row
            # get Max column of worksheet
            p_max_column = ws.max_column

            ws['A' + str(p_max_row + 1)] = 'Nett Profit'
            ws['A' + str(p_max_row + 1)].font = Font(bold=True)

            # loop through column A and if the cell value is 'Trading Income' then get the row number and print it
            # in column B
            for cell in ws['A']:
                if cell.value == 'Trading Income':
                    row_num = cell.row + 1
                    trading_income_start_row = row_num
                elif cell.value == 'Total Trading Income':
                    row_num = cell.row - 1
                    trading_income_end_row = row_num
                elif cell.value == 'Other Income':
                    row_num = cell.row + 1
                    other_income_start_row = row_num
                elif cell.value == 'Total Other Income':
                    row_num = cell.row - 1
                    other_income_end_row = row_num
                elif cell.value == 'COS':
                    row_num = cell.row + 1
                    cos_start_row = row_num
                elif cell.value == 'Total COS':
                    row_num = cell.row - 1
                    cos_end_row = row_num
                elif cell.value == 'Operating Expenses':
                    row_num = cell.row + 1
                    operating_expenses_start_row = row_num
                elif cell.value == 'Total Operating Expenses':
                    row_num = cell.row - 1
                    operating_expenses_end_row = row_num
                elif cell.value == 'Gross Profit':
                    row_num = cell.row
                    gross_profit_start_row = row_num
                elif cell.value == 'Nett Profit':
                    row_num = cell.row
                    nett_profit_start_row = row_num

            formula_values = [
                {"start": trading_income_start_row, "end": trading_income_end_row},
                {"start": other_income_start_row, "end": other_income_end_row},
                {"start": cos_start_row, "end": cos_end_row},
                {"start": operating_expenses_start_row, "end": operating_expenses_end_row},
                # { "start": gross_profit_start_row, "end": gross_profit_start_row },
                # { "start": nett_profit_start_row, "end": nett_profit_start_row }
            ]

            for value in formula_values:  # in columns B to the last column, set the value to 0
                for row in ws.iter_rows(min_row=value['start'], min_col=3, max_row=value['end'],
                                        max_col=3):
                    for cell in row:
                        if index < 2:

                            cell.value = f"=SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, '{dev}'!$A{cell.row}, data!$D$1:$D${last_row}, '{dev}'!$A$2, data!$E$1:$E${last_row}, '{dev}'!{get_column_letter(cell.column)}$5)"
                        else:
                            cell.value = f"=SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, '{dev}'!$A{cell.row}, data!$E$1:$E${last_row}, '{dev}'!{get_column_letter(cell.column)}$5)"

                        cell.number_format = 'R #,##0.00'

                for row in ws.iter_rows(min_row=value['start'], min_col=4, max_row=value['end'],
                                        max_col=p_max_column):
                    for cell in row:
                        if index < 2:

                            cell.value = f"={get_column_letter(cell.column - 1)}{cell.row}+SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, '{dev}'!$A{cell.row}, data!$D$1:$D${last_row}, '{dev}'!$A$2, data!$E$1:$E${last_row}, '{dev}'!{get_column_letter(cell.column)}$5)"
                            # cell.value = f"=SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, '{dev}'!$A{cell.row}, data!$D$1:$D${last_row}, '{dev}'!$A$2, data!$E$1:$E${last_row}, '{dev}'!{get_column_letter(cell.column)}$5)"

                        else:
                            cell.value = f"={get_column_letter(cell.column - 1)}{cell.row}+SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, '{dev}'!$A{cell.row},  data!$E$1:$E${last_row}, '{dev}'!{get_column_letter(cell.column)}$5)"
                            # cell.value = f"=SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, '{dev}'!$A{cell.row}, data!$E$1:$E${last_row}, '{dev}'!{get_column_letter(cell.column)}$5)"

                        cell.number_format = 'R #,##0.00'

                for row in ws.iter_rows(min_row=value['end'] + 1, min_col=3, max_row=value['end'] + 1,
                                        max_col=p_max_column):
                    for cell in row:
                        cell.value = f"=SUM({get_column_letter(cell.column)}{value['start']}:{get_column_letter(cell.column)}{value['end']})"
                        # the values bold, font size 12, format as currency with 2 decimal places and R for South African and a top border
                        cell.font = Font(bold=True, size=12)
                        cell.number_format = 'R #,##0.00'
                        cell.border = Border(top=Side(border_style='thin', color='FF000000'))

            for row in ws.iter_rows(min_row=gross_profit_start_row, min_col=3, max_row=gross_profit_start_row,
                                    max_col=p_max_column):
                for cell in row:
                    cell.value = (f"=+{get_column_letter(cell.column)}{trading_income_end_row + 1}+"
                                  f"{get_column_letter(cell.column)}{other_income_end_row + 1}-"
                                  f"({get_column_letter(cell.column)}{cos_end_row + 1})")
                    # format as currency with 2 decimal places and R for South African, make bold and font size 12 and a top and bottom border
                    cell.number_format = 'R #,##0.00'
                    cell.font = Font(bold=True, size=12)
                    cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))

            for row in ws.iter_rows(min_row=nett_profit_start_row, min_col=3, max_row=nett_profit_start_row,
                                    max_col=p_max_column):
                for cell in row:
                    cell.value = (f"=+{get_column_letter(cell.column)}{gross_profit_start_row}-"
                                  f"{get_column_letter(cell.column)}{operating_expenses_end_row + 1}")
                    # format as currency with 2 decimal places and R for South African, make bold and font size 12 and a top border and double bottom border
                    cell.number_format = 'R #,##0.00'
                    cell.font = Font(bold=True, size=12)
                    cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='double', color='FF000000'))

            ws.column_dimensions['A'].width = 38
            ws.column_dimensions['B'].width = 4
            # freeze panes at B6
            ws.freeze_panes = 'B6'

        # create a new sheet called 'NSST Print'
        ws = wb.create_sheet('NSST Print')
        # make tab color blue
        ws.sheet_properties.tabColor = "00B0F0"

        # insert worksheet_input into worksheet
        print("data2", data2[20])

        for row in data2:
            ws.append(row)

        # format all cells as currency with 2 decimal places and R for South African
        for col in ws.iter_cols(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in col:
                cell.number_format = 'R #,##0.00'

        # make font lager
        for col in ws.iter_cols(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in col:
                cell.font = Font(size=12)

        # set width of column A to 38 pixels
        # create a list of column letters
        columns = []
        for col in ws.iter_cols(min_row=1, min_col=1, max_row=1, max_col=ws.max_column):
            columns.append(col[0].column_letter)
        # print("columns", columns)
        for col in columns:
            ws.column_dimensions[col].width = 18

        # a border around cells E1
        ws['E1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'))
        # make the font 20 and bold
        ws['E1'].font = Font(size=20, bold=True)
        # do the same for J1
        ws['J1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'))
        ws['J1'].font = Font(size=20, bold=True)

        # do the same for O1
        ws['O1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'))
        ws['O1'].font = Font(size=20, bold=True)
        # do the same for T1
        ws['T1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'))
        ws['T1'].font = Font(size=20, bold=True)
        # put a border around all cells from row 2 onwards
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=42, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(top=Side(border_style='thin', color='FF000000'),
                                     bottom=Side(border_style='thin', color='FF000000'),
                                     left=Side(border_style='thin', color='FF000000'),
                                     right=Side(border_style='thin', color='FF000000'))

        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['F'].width = 42
        ws.column_dimensions['K'].width = 42
        ws.column_dimensions['P'].width = 42

        total_cost_to_complete_columns = ['B', 'G', 'L', 'Q']

        # =SUMIFS(data!$H$1:$H$1644, data!$A$1:$A$1644, "Cost To Complete Project", data!$E$1:$E$1644, 'NSST Print'!B3)
        # in row 2, in columns inside total_cost_to_complete_columns insert a formula as per above and format as currency with 2 decimal places and R for South African
        for col in total_cost_to_complete_columns:
            #f"=SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, \"Cost To Complete Project\", data!$E$1:$E${last_row}, 'NSST Print'!{col}3)-+SUM(D44:D50)"
            ws[f'{col}35'].value = f"=SUMIFS(data!$H$1:$H${last_row}, data!$A$1:$A${last_row}, \"Cost To Complete Project\", data!$E$1:$E${last_row}, 'NSST Print'!{col}3)-SUM(C44:C50)"
            ws[f'{col}35'].number_format = 'R #,##0.00'
            # =B9+SUMIFS(data!$H$1:$H$1648,data!$A$1:$A$1648,'NSST Print'!$A$32,data!$E$1:$E$1648,'NSST Print'!B3) in row 32
            ws[f'{col}32'].value = f"={col}9+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},'NSST Print'!$A$32,data!$E$1:$E${last_row},'NSST Print'!{col}3)"
            ws[f'{col}32'].number_format = 'R #,##0.00'
            # in row 33 ==SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,'NSST Print'!$A$33,data!$E$1:$E$1683,'NSST Print'!B3)+C43+C51+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"Momentum Interest")++SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"Attorneys Deposit")++SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"FNB Bank Account")
            ws[f'{col}33'].value = f"=SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},'NSST Print'!$A$33,data!$E$1:$E${last_row},'NSST Print'!{col}3)+C43+C51+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Momentum Interest\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Attorneys Deposit\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"FNB Bank Account\")"
            ws[f'{col}33'].number_format = 'R #,##0.00'
            # in row 39 =SUMIFS(data!$H$1:$H$1656,data!$A$1:$A$1656,"Interest Received - Momentum")
            ws[f'{col}39'].value = f"=SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Interest Received - Momentum\")"
            ws[f'{col}39'].number_format = 'R #,##0.00'
            # in row 40 =SUMIFS(data!$H$1:$H$1680,data!$B$1:$B$1680,"Operating Expenses")+SUMIFS(data!$H$1:$H$1680,data!$B$1:$B$1680,"COS")-SUM(D44:D50)
            ws[f'{col}40'].value = f"=SUMIFS(data!$H$1:$H${last_row},data!$B$1:$B${last_row},\"Operating Expenses\")+SUMIFS(data!$H$1:$H${last_row},data!$B$1:$B${last_row},\"COS\")-SUM(C44:C50)"
            ws[f'{col}40'].number_format = 'R #,##0.00'
            # in row 41 =SUMPRODUCT(ISNUMBER(SEARCH("Interest Paid - Investors",data!$A$2:$A$1684))*(data!$H$2:$H$1684))
            # ws[f'{col}41'].value = f"=SUMPRODUCT(ISNUMBER(SEARCH(\"Interest Paid - Investors\",data!$A$2:$A${last_row}))*(data!$H$2:$H${last_row}))"
            # ws[f'{col}41'].number_format = 'R #,##0.00'




        toggle_rows_to_format = [43, 44, 45, 46, 47, 48, 49, 50, 51]





        # format rows in toggle_rows_to_format as percent with 2 decimal places and a % sign for columns in total_cost_to_complete_columns, except for 45, format those cells a 'General'
        for row in toggle_rows_to_format:
            for col in total_cost_to_complete_columns:
                if row == 45:
                    ws[f'{col}{row}'].number_format = 'General'
                else:
                    ws[f'{col}{row}'].number_format = '0.00%'

        # copy row 21 to row 53
        # row_to_copy = 21  # Replace with the row number you want to copy
        #
        # # Copy the row
        # new_row = ws.copy_row(row_to_copy)
        #
        # # Specify the row where you want to insert the copied row
        # row_to_insert = 53  # Replace with the row number where you want to insert the copied row
        #
        # # Insert the copied row at the specified location
        # ws.insert_rows(new_row, row_to_insert)


        columns_to_sum = ['B', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'Q', 'R', 'S', 'T']
        # =B21 - SUM(B22: B26)
        # in row 27, in columns inside columns_to_sum insert a formula as per above
        for col in columns_to_sum:
            # =B21*0.05
            # put the formula in row 22
            ws[f'{col}22'].value = f"={col}21*0.05"
            ws[f'{col}27'].value = f"={col}21-SUM({col}22:{col}26)"
            # format as currency with 2 decimal places and R for South African
            ws[f'{col}27'].number_format = 'R #,##0.00'
            # make the font bold
            # ws[f'{col}27'].font = Font(bold=True)

        # =B32 + B33
        # in row 34, in columns inside columns_to_sum insert a formula as per above
        for col in columns_to_sum:
            ws[f'{col}34'].value = f"={col}32+{col}33"
            # format as currency with 2 decimal places and R for South African
            ws[f'{col}34'].number_format = 'R #,##0.00'
            # make the font bold
            # ws[f'{col}34'].font = Font(bold=True)

            # =+B21
            # in row 38, in columns inside columns_to_sum insert a formula as per above
        for col in columns_to_sum:
            ws[f'{col}38'].value = f"={col}21"
            # format as currency with 2 decimal places and R for South African
            ws[f'{col}38'].number_format = 'R #,##0.00'
            # make the font bold
            # ws[f'{col}38'].font = Font(bold=True)

        # =B34 - B35
        # in row 36, in columns inside columns_to_sum insert a formula as per above
        for col in columns_to_sum:
            ws[f'{col}36'].value = f"={col}34-{col}35"
            # format as currency with 2 decimal places and R for South African
            ws[f'{col}36'].number_format = 'R #,##0.00'
            # make the font bold
            # ws[f'{col}36'].font = Font(bold=True)

        # =B38 + B39 - B40 - B41
        # in row 42, in columns inside columns_to_sum insert a formula as per above
        for col in columns_to_sum:
            ws[f'{col}42'].value = f"={col}38+{col}39-{col}40-{col}41"
            # format as currency with 2 decimal places and R for South African
            ws[f'{col}42'].number_format = 'R #,##0.00'
            # make the font bold
            # ws[f'{col}42'].font = Font(bold=True)

        # add the following formula to cell B21 =SUMIFS(data!$H$1:$H$1681,data!$A$1:$A$1681,"Sales - Heron View Sales")+SUMIFS(data!$H$1:$H$1681,data!$A$1:$A$1681,"Sales - Heron Fields")+D43
        ws['B21'].value = f"=SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Sales - Heron View Sales\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Sales - Heron Fields\")+C43"

        # add the following formul to cell E21 =E53+D43
        ws['E21'].value = f"=E53+C43"
        # add the following formula to cell D43 =E53*B43
        ws['C43'].value = f"=E21*B43"

        # Add the folowing formula to E43 =E53*B43
        ws['C43'].value = f"=E53*B43"


        # add this formula to cell d44 =(SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Heron Fields -
        # Construction", data!$D$1:$D$1680,"Heron Fields")+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Heron
        # View - Construction", data!$D$1:$D$1680,"Heron View")+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,
        # "COS - Heron Fields - P & G", data!$D$1:$D$1680,"Heron Fields")+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,
        # "COS - Heron View - P&G", data!$D$1:$D$1680,"Heron View"))-(SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,
        # "COS - Heron Fields - Construction", data!$D$1:$D$1680,"Heron Fields",data!$I$1:$I$1680, B3)+SUMIFS(
        # data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Heron View - Construction", data!$D$1:$D$1680,"Heron View",
        # data!$I$1:$I$1680, B3)+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Heron Fields - P & G",
        # data!$D$1:$D$1680,"Heron Fields",data!$I$1:$I$1680, B3)+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,
        # "COS - Heron View - P&G", data!$D$1:$D$1680,"Heron View",data!$I$1:$I$1680, B3))*B44

        ws['C44'].value = f"=((SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - Construction\", data!$D$1:$D${last_row},\"Heron Fields\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - Construction\", data!$D$1:$D${last_row},\"Heron View\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - P & G\", data!$D$1:$D${last_row},\"Heron Fields\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - P&G\", data!$D$1:$D${last_row},\"Heron View\"))-(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - Construction\", data!$D$1:$D${last_row},\"Heron Fields\",data!$I$1:$I${last_row}, \"<=\"&B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - Construction\", data!$D$1:$D${last_row},\"Heron View\",data!$I$1:$I${last_row}, \"<=\"&B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - P & G\", data!$D$1:$D${last_row},\"Heron Fields\",data!$I$1:$I${last_row}, \"<=\"&B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - P&G\", data!$D$1:$D${last_row},\"Heron View\",data!$I$1:$I${last_row}, \"<=\"&B3)))*B44"
        # in row D44 the following =((SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron Fields - Construction", data!$D$1:$D$1683,"Heron Fields")+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron View - Construction", data!$D$1:$D$1683,"Heron View")+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron Fields - P & G", data!$D$1:$D$1683,"Heron Fields")+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron View - P&G", data!$D$1:$D$1683,"Heron View"))-(SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron Fields - Construction", data!$D$1:$D$1683,"Heron Fields",data!$I$1:$I$1683, "<="&B3)+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron View - Construction", data!$D$1:$D$1683,"Heron View",data!$I$1:$I$1683, "<="&B3)+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron Fields - P & G", data!$D$1:$D$1683,"Heron Fields",data!$I$1:$I$1683, "<="&B3)+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Heron View - P&G", data!$D$1:$D$1683,"Heron View",data!$I$1:$I$1683, "<="&B3)))-C44
        ws['D44'].value = f"=((SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - Construction\", data!$D$1:$D${last_row},\"Heron Fields\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - Construction\", data!$D$1:$D${last_row},\"Heron View\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - P & G\", data!$D$1:$D${last_row},\"Heron Fields\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - P&G\", data!$D$1:$D${last_row},\"Heron View\"))-(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - Construction\", data!$D$1:$D${last_row},\"Heron Fields\",data!$I$1:$I${last_row}, \"<=\"&B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - Construction\", data!$D$1:$D${last_row},\"Heron View\",data!$I$1:$I${last_row}, \"<=\"&B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron Fields - P & G\", data!$D$1:$D${last_row},\"Heron Fields\",data!$I$1:$I${last_row}, \"<=\"&B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Heron View - P&G\", data!$D$1:$D${last_row},\"Heron View\",data!$I$1:$I${last_row}, \"<=\"&B3)))-C44"
        # in row D46 =(SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"CPSD")-SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"CPSD",data!$I$1:$I$1683,"<="&'NSST Print'!B3))-C46
        ws['D46'].value = f"=(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"CPSD\")-SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"CPSD\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3))-C46"
        # in row d47 =(SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"Opp Invest")-SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"Opp Invest",data!$I$1:$I$1683,"<="&'NSST Print'!B3))-C47
        ws['D47'].value = f"=(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Opp Invest\")-SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Opp Invest\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3))-C47"
        # in row d48 =(SUMPRODUCT(ISNUMBER(SEARCH("Interest Paid - Investors",data!$A$2:$A$1683))*(data!$H$2:$H$1683))-SUMPRODUCT( --(ISNUMBER(SEARCH("Interest Paid - Investors", data!$A$2:$A$1683))), --(data!$I$2:$I$1683 <= B3), data!$H$2:$H$1683 ))-C48
        ws['D48'].value = f"=(SUMPRODUCT(ISNUMBER(SEARCH(\"Interest Paid - Investors\",data!$A$2:$A${last_row}))*(data!$H$2:$H${last_row}))-SUMPRODUCT( --(ISNUMBER(SEARCH(\"Interest Paid - Investors\", data!$A$2:$A${last_row}))), --(data!$I$2:$I${last_row} <= B3), data!$H$2:$H${last_row} ))-C48"
        # in row D49 =((SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Commission HF Units")+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Commission HV Units"))-(SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Commission HF Units",data!$I$1:$I$1683,"<="&'NSST Print'!B3)+SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"COS - Commission HV Units",data!$I$1:$I$1683,"<="&'NSST Print'!B3)))-C49
        ws['D49'].value = f"=((SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HF Units\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HV Units\"))-(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HF Units\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HV Units\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3)))-C49"
        # in row D43 = ""
        ws['D43'].value = f""
        ws['D45'].value = f""
        ws['D50'].value = f""
        ws['D51'].value = f""
        # in row E44 =C44/$E$20
        ws['E44'].value = f"=C44/$E$20"
        # in row E45 =C45/$E$20
        ws['E45'].value = f"=C45/$E$20"
        # in row E46 =C46/$E$20
        ws['E46'].value = f"=C46/$E$20"
        # in row E47 =C47/$E$20
        ws['E47'].value = f"=C47/$E$20"
        # in row E48 =C48/$E$20
        ws['E48'].value = f"=C48/$E$20"
        # in row E49 =C49/$E$20
        ws['E49'].value = f"=C49/$E$20"
        # in row E50 =C50/$E$20
        ws['E50'].value = f"=C50/$E$20"
        # row E51 = sum(E44:E50)
        ws['E51'].value = f"=sum(E44:E50)"












        #add this formula to cell D48 =(SUMPRODUCT(ISNUMBER(SEARCH("Interest Paid - Investors",data!$A$2:$A$1680))*(data!$H$2:$H$1680))-SUMPRODUCT( --(ISNUMBER(SEARCH("Interest Paid - Investors", data!$A$2:$A$1680))), --(data!$I$2:$I$1680 <= B3), data!$H$2:$H$1680 ))*B48
        ws['C48'].value = f"=(SUMPRODUCT(ISNUMBER(SEARCH(\"Interest Paid - Investors\",data!$A$2:$A${last_row}))*(data!$H$2:$H${last_row}))-SUMPRODUCT( --(ISNUMBER(SEARCH(\"Interest Paid - Investors\", data!$A$2:$A${last_row}))), --(data!$I$2:$I${last_row} <= B3), data!$H$2:$H${last_row} ))*B48"
        # cell d49 =((SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Commission HF Units")+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Commission HV Units"))-(SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Commission HF Units",data!$I$1:$I$1680,"<="&'NSST Print'!B3)+SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"COS - Commission HV Units",data!$I$1:$I$1680,"<="&'NSST Print'!B3)))*B49
        ws['C49'].value = f"=((SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HF Units\")+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HV Units\"))-(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HF Units\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3)+SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"COS - Commission HV Units\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3)))*B49"
        # D50 =B26*B50
        ws['C50'].value = f"=B26*B50"
        # row D45 ==400000*B45
        ws['C45'].value = f"=400000*B45"
        # row 46 =(SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"CPSD")-SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"CPSD",data!$I$1:$I$1680,"<="&'NSST Print'!B3))*B46
        ws['C46'].value = f"=(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"CPSD\")-SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"CPSD\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3))*B46"


        # row47 =(SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"Opp Invest")-SUMIFS(data!$H$1:$H$1680,data!$A$1:$A$1680,"Opp Invest",data!$I$1:$I$1680,"<="&'NSST Print'!B3))*B47
        ws['C47'].value = f"=(SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Opp Invest\")-SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Opp Invest\",data!$I$1:$I${last_row},\"<=\"&'NSST Print'!B3))*B47"
        # in row 51 =SUMIFS(data!$H$1:$H$1683,data!$A$1:$A$1683,"Early Exit Loan")*B51
        ws['C51'].value = f"=SUMIFS(data!$H$1:$H${last_row},data!$A$1:$A${last_row},\"Early Exit Loan\")*B51"



        # ws.insert_rows(data2[2])

        # grey_cells = ['A27', 'A29', 'A34', 'A38', 'A42']
        # for cell in grey_cells:
        #     ws[cell].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        columns_to_center = [ 'B', 'C', 'D', 'E',  'G', 'H', 'I', 'J',
                              'L', 'M', 'N', 'O', 'Q', 'R', 'S', 'T']

        # make the cells in row 19 and 20 center aligned for columns in columns_to_center
        for col in columns_to_center:
            ws[f'{col}19'].alignment = Alignment(horizontal='center')
            ws[f'{col}20'].alignment = Alignment(horizontal='center')

        # make row 42 fill color grey
        for cell in ws[42]:
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # in row1 merge cells A1 to D1, F1 to I1, K1 to M1, O1 to R1
        cols_to_merge_full = [{"start": 'A',
                               'end': 'E'}, {"start": 'F', 'end': 'J'}, {"start": 'K', 'end': 'O'},
                              {"start": 'P', 'end': 'T'}]
        rows_to_merge_full = [2, 5, 14, 18, 28, 31, 37]

        for row in rows_to_merge_full:

            for index, col in enumerate(cols_to_merge_full):
                if index == 0:
                    color = '99B080'
                elif index == 1:
                    color = 'AF2655'
                elif index == 2:
                    color = '6527BE'
                elif index == 3:
                    color = 'FE0000'
                # # make font white

                ws.merge_cells(f'{col["start"]}{row}:{col["end"]}{row}')
                # make fill color of merged cells green
                ws[f'{col["start"]}{row}'].fill = PatternFill(start_color=color, end_color=color,
                                                              fill_type='solid')
                # make alignment center
                ws[f'{col["start"]}{row}'].alignment = Alignment(horizontal='center')

        # make the font white in rows_to_merge_full
        for row in rows_to_merge_full:
            for col in cols_to_merge_full:
                ws[f'{col["start"]}{row}'].font = Font(size=16, color='FFFBF5', bold=True)
                # and 18 size

        cols_to_merge_partial = [{"start": 'B',
                               'end': 'E'}, {"start": 'G', 'end': 'J'}, {"start": 'L', 'end': 'O'},
                              {"start": 'Q', 'end': 'T'}]

        rows_to_merge_partial = [3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16,17,29,30,32,33,34,35,36, 38, 39, 40, 41,42]

        for row in rows_to_merge_partial:
            for col in cols_to_merge_partial:
                ws.merge_cells(f'{col["start"]}{row}:{col["end"]}{row}')
                ws[f'{col["start"]}{row}'].alignment = Alignment(horizontal='center')

        # format row 20 as general
        for cell in ws[20]:
            cell.number_format = 'General'
        for cell in ws[15]:
            cell.number_format = 'General'
        for cell in ws[16]:
            cell.number_format = 'General'
        for cell in ws[17]:
            cell.number_format = 'General'


        ws.merge_cells('A1:D1')
        ws.merge_cells('F1:I1')
        ws.merge_cells('K1:N1')
        ws.merge_cells('P1:S1')

        cells_to_format_as_date = ['B3', 'G3', 'L3', 'Q3']
        for cell in cells_to_format_as_date:
            ws[cell].number_format = 'dd MMM yy'

        # hide row 41 and 53
        ws.row_dimensions[41].hidden = True
        ws.row_dimensions[53].hidden = True



        wb.save('cashflow_p&l_files/cashflow_hf_hv.xlsx')

        return "Awesome"
    except Exception as e:
        print(e)
        return "Error"
