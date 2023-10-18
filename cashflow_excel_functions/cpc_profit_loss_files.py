import os

import openpyxl


def insert_data_from_xero_profit_loss(base_data, base_data2, base_data3, hf_tb, hv_tb, report_date):
    # get all excel file names from the folder cashflow_p&l_files and insert filenames into a list called files
    files = []
    for filename in os.listdir('cashflow_p&l_files'):
        files.append(filename)

    files = [file for file in files if not file.startswith('~$')]
    files = [file for file in files if file.endswith('.xlsx')]

    try:
        for file in files:
            # Load the existing Excel workbook
            workbook = openpyxl.load_workbook(f'cashflow_p&l_files/{file}')

            # Select the desired sheet by name

            # Insert data into the sheet
            for entry in base_data:
                sheet = workbook['CPC 24']
                account_name = entry['Account']
                amounts = entry['Amount']

                # Find the row in the sheet where the account matches
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        if cell.value == account_name:
                            row_idx = cell.row

                            # Insert amount values in columns B to M
                            for col_idx, amount in enumerate(amounts, start=2):
                                sheet.cell(row=row_idx, column=col_idx, value=amount)

            # Insert data into the sheet
            if file == '4. Heron Fields Profit & Loss.xlsx':
                sheet = workbook['2024 Xero HF']
                for entry in base_data2:
                    account_name = entry['Account']
                    amounts = entry['Amount']

                    # Find the row in the sheet where the account matches
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value == account_name:
                                row_idx = cell.row

                                # Insert amount values in columns B to M
                                for col_idx, amount in enumerate(amounts, start=4):
                                    sheet.cell(row=row_idx, column=col_idx, value=amount)

                sheet = workbook['TB HF 2024']
                sheet.delete_rows(sheet.min_row, sheet.max_row)

                # Cell A1 = 'Trial Balance'
                sheet.cell(row=1, column=1, value='Trial Balance')
                sheet.cell(row=2, column=1, value='Heron Fields (Pty) Ltd')
                sheet.cell(row=3, column=1, value=f'As at {report_date}')
                # Let Headers = ['Account Code', 'Account, '', 'Debit - Year to date', 'Credit - Year to date']
                headers = ['Account Code', 'Account', '', 'Debit - Year to date', 'Credit - Year to date']
                # Insert headers into the sheet in row 5
                for col_idx, header in enumerate(headers, start=1):
                    sheet.cell(row=5, column=col_idx, value=header)
                insert_data = []
                for entry in hf_tb:
                    insert_list = []
                    for key in entry:
                        insert_list.append(entry[key])
                    insert_data.append(insert_list)
                # print(insert_data)
                for row_idx, row in enumerate(insert_data, start=6):
                    for col_idx, value in enumerate(row, start=1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                # format columns D and E as currency
                for col in ['D', 'E']:
                    for row in range(6, sheet.max_row + 1):
                        sheet[f'{col}{row}'].number_format = '#,##0.00'
                # rows 1 to 5 bold
                for row in range(1, 6):
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).font = openpyxl.styles.Font(bold=True)
                # add a total row and sum columns D and E and format as currency and bold
                sheet.cell(row=sheet.max_row + 2, column=4, value=f'=SUM(D6:D{sheet.max_row})')
                sheet.cell(row=sheet.max_row, column=4).number_format = '#,##0.00'
                sheet.cell(row=sheet.max_row, column=4).font = openpyxl.styles.Font(bold=True)
                sheet.cell(row=sheet.max_row, column=5, value=f'=SUM(E6:E{sheet.max_row - 2})')
                sheet.cell(row=sheet.max_row, column=5).number_format = '#,##0.00'
                sheet.cell(row=sheet.max_row, column=5).font = openpyxl.styles.Font(bold=True)

            if file == '5. Heron View Profit & Loss.xlsx':
                sheet = workbook['2024 Xero HV']
                # Insert data into the sheet
                for entry in base_data3:
                    account_name = entry['Account']
                    amounts = entry['Amount']

                    # Find the row in the sheet where the account matches
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value == account_name:
                                row_idx = cell.row

                                # Insert amount values in columns B to M
                                for col_idx, amount in enumerate(amounts, start=4):
                                    sheet.cell(row=row_idx, column=col_idx, value=amount)

                sheet = workbook['TB HV 24']
                sheet.delete_rows(sheet.min_row, sheet.max_row)

                # Cell A1 = 'Trial Balance'
                sheet.cell(row=1, column=1, value='Trial Balance')
                sheet.cell(row=2, column=1, value='Heron View (Pty) Ltd')
                sheet.cell(row=3, column=1, value=f'As at {report_date}')
                # Let Headers = ['Account Code', 'Account, '', 'Debit - Year to date', 'Credit - Year to date']
                headers = ['Account Code', 'Account', '', 'Debit - Year to date', 'Credit - Year to date']
                # Insert headers into the sheet in row 5
                for col_idx, header in enumerate(headers, start=1):
                    sheet.cell(row=5, column=col_idx, value=header)
                insert_data = []
                for entry in hv_tb:
                    insert_list = []
                    for key in entry:
                        insert_list.append(entry[key])
                    insert_data.append(insert_list)
                # print(insert_data)
                for row_idx, row in enumerate(insert_data, start=6):
                    for col_idx, value in enumerate(row, start=1):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                # format columns D and E as currency
                for col in ['D', 'E']:
                    for row in range(6, sheet.max_row + 1):
                        sheet[f'{col}{row}'].number_format = '#,##0.00'
                # rows 1 to 5 bold
                for row in range(1, 6):
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).font = openpyxl.styles.Font(bold=True)
                # add a total row and sum columns D and E and format as currency and bold
                sheet.cell(row=sheet.max_row + 2, column=4, value=f'=SUM(D6:D{sheet.max_row})')
                sheet.cell(row=sheet.max_row, column=4).number_format = '#,##0.00'
                sheet.cell(row=sheet.max_row, column=4).font = openpyxl.styles.Font(bold=True)
                sheet.cell(row=sheet.max_row, column=5, value=f'=SUM(E6:E{sheet.max_row - 2})')
                sheet.cell(row=sheet.max_row, column=5).number_format = '#,##0.00'
                sheet.cell(row=sheet.max_row, column=5).font = openpyxl.styles.Font(bold=True)

            workbook.save(f'cashflow_p&l_files/{file}')
        return {"Success": True}
    except Exception as e:
        print("Error", e)
        return {"Success": False, "Error": str(e)}
